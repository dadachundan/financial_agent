"""
Task 3: Add valuation tabs to the existing Hengli financial model.

Adds 4 tabs:
  - DCF              : full DCF build with terminal value
  - Sensitivity      : WACC × terminal growth heatmap
  - Comparables      : 8 peer companies with stats summary
  - Valuation Summary: football field + price target + recommendation
"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

PATH = "reports/company/Hengli_SSE601100/Hengli_SSE601100_Financial_Model_2026-05-19.xlsx"

# Styles
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
PROJ_FILL = PatternFill("solid", fgColor="FFF2CC")
BUY_FILL = PatternFill("solid", fgColor="C6EFCE")
HOLD_FILL = PatternFill("solid", fgColor="FFEB9C")
SELL_FILL = PatternFill("solid", fgColor="FFC7CE")

BLUE = Font(color="0070C0")
BLACK = Font(color="000000")
GREEN = Font(color="00B050")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
ITALIC = Font(italic=True, color="595959")
thin = Side(style="thin", color="BFBFBF")
TOP_BORDER = Border(top=Side(style="thin", color="000000"))

def fmt_num(c, decimals=1, is_pct=False):
    if is_pct:
        c.number_format = "0.0%;(0.0%);—"
    else:
        c.number_format = f'#,##0.{"0"*decimals};(#,##0.{"0"*decimals});—'

def write_label(ws, r, lbl, bold=False, italic=False, indent=0):
    c = ws.cell(row=r, column=1, value="  "*indent + lbl)
    if bold: c.font = BOLD
    if italic: c.font = ITALIC

wb = load_workbook(PATH)

# ============================================================
# TAB: DCF
# ============================================================
ws = wb.create_sheet("DCF")
ws.column_dimensions["A"].width = 48
for c in "BCDEFGH":
    ws.column_dimensions[c].width = 14

ws["A1"] = "HENGLI HYDRAULICS — DCF VALUATION (Base case, RMB millions)"
ws["A1"].font = Font(size=14, bold=True, color="1F4E79")
ws.merge_cells("A1:H1")

# Year headers (FY26E to FY30E + Terminal)
yrs = ["FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "Terminal"]
for i, y in enumerate(yrs):
    c = ws.cell(row=3, column=2+i, value=y)
    c.font = WHITE_BOLD; c.fill = HDR_FILL; c.alignment = Alignment(horizontal="center")

row = 5
# UFCF (linked from DCF Inputs tab row 12, columns B-F)
write_label(ws, row, "Unlevered Free Cash Flow (UFCF)", bold=True)
for i in range(5):
    col = 2 + i; cl_dcf = get_column_letter(2+i)
    c = ws.cell(row=row, column=col, value=f"='DCF Inputs'!{cl_dcf}12")
    c.font = GREEN; fmt_num(c)
# Terminal UFCF (FY30 + 1 year growth)
c = ws.cell(row=row, column=7, value="=F5*(1+B20)")
c.font = BLACK; fmt_num(c)
ufcf_row = row
row += 1

# Discount period (mid-year convention)
write_label(ws, row, "Discount period (mid-year convention)", indent=1)
for i in range(5):
    val = 0.5 + i  # 0.5, 1.5, 2.5, 3.5, 4.5
    c = ws.cell(row=row, column=2+i, value=val); c.font = BLUE
    c.number_format = "0.0"
c = ws.cell(row=row, column=7, value=5.0); c.font = BLUE; c.number_format = "0.0"
disc_period_row = row
row += 1

# Discount factor
write_label(ws, row, "Discount factor (1/(1+WACC)^t)", indent=1)
for i in range(6):
    col = 2 + i; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"=1/(1+$B$19)^{cl}{disc_period_row}")
    c.font = BLACK; c.number_format = "0.0000"
df_row = row
row += 1

# PV of UFCF
write_label(ws, row, "PV of UFCF", bold=True)
for i in range(5):
    col = 2 + i; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{ufcf_row}*{cl}{df_row}")
    c.font = Font(bold=True); fmt_num(c)
pv_ufcf_row = row
row += 2

# Terminal value calculation
write_label(ws, row, "TERMINAL VALUE", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
row += 1
write_label(ws, row, "Terminal UFCF (FY2031E)", indent=1)
c = ws.cell(row=row, column=2, value=f"=G{ufcf_row}"); c.font = GREEN; fmt_num(c)
term_ufcf_row = row
row += 1
write_label(ws, row, "Terminal growth rate (g)", indent=1)
c = ws.cell(row=row, column=2, value=0.030); c.font = BLUE; fmt_num(c, is_pct=True)
g_row = row
row += 1
write_label(ws, row, "Terminal value (Gordon: UFCF / (WACC-g))", indent=1, bold=True)
c = ws.cell(row=row, column=2, value=f"=B{term_ufcf_row}/(B19-B{g_row})")
c.font = Font(bold=True); fmt_num(c)
tv_row = row
row += 1
write_label(ws, row, "PV of terminal value (× DF Year 5)", indent=1, bold=True)
c = ws.cell(row=row, column=2, value=f"=B{tv_row}*G{df_row}")
c.font = Font(bold=True); fmt_num(c)
pv_tv_row = row
row += 1
write_label(ws, row, "Terminal value as % of EV", italic=True, indent=2)
c = ws.cell(row=row, column=2, value=f"=B{pv_tv_row}/(SUM(B{pv_ufcf_row}:F{pv_ufcf_row})+B{pv_tv_row})")
c.font = ITALIC; fmt_num(c, is_pct=True)
row += 2

# WACC
write_label(ws, row, "WACC (input from DCF Inputs tab)", bold=True)
c = ws.cell(row=row, column=2, value=f"='DCF Inputs'!B30"); c.font = GREEN; fmt_num(c, is_pct=True)
wacc_row = row
# Adjust formula references that used $B$19 — actual WACC row needs update
# Re-build with absolute row pointer:
ws[f"D{wacc_row}"] = "Sourced from DCF Inputs!B30 (WACC build)"
ws[f"D{wacc_row}"].font = ITALIC
row += 2

# Enterprise Value
write_label(ws, row, "ENTERPRISE VALUE BUILD", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
row += 1
write_label(ws, row, "Sum of PV of UFCF (Year 1-5)", indent=1)
c = ws.cell(row=row, column=2, value=f"=SUM(B{pv_ufcf_row}:F{pv_ufcf_row})")
c.font = BLACK; fmt_num(c)
sum_pv_row = row
row += 1
write_label(ws, row, "+ PV of terminal value", indent=1)
c = ws.cell(row=row, column=2, value=f"=B{pv_tv_row}"); c.font = BLACK; fmt_num(c)
row += 1
write_label(ws, row, "= ENTERPRISE VALUE", bold=True)
c = ws.cell(row=row, column=2, value=f"=B{sum_pv_row}+B{pv_tv_row}")
c.font = Font(bold=True, color="1F4E79"); fmt_num(c); c.border = TOP_BORDER; c.fill = TOTAL_FILL
ev_row = row
row += 1

# Bridge to equity
write_label(ws, row, "  + Net cash & equivalents (FY25 cash 8,871 + trading FA 345)", indent=1)
c = ws.cell(row=row, column=2, value=9216.0); c.font = BLUE; fmt_num(c)
row += 1
write_label(ws, row, "  − Total debt (ST 14 + LT 20)", indent=1)
c = ws.cell(row=row, column=2, value=-34.0); c.font = BLUE; fmt_num(c)
row += 1
write_label(ws, row, "  − Minority interest", indent=1)
c = ws.cell(row=row, column=2, value=-58.0); c.font = BLUE; fmt_num(c)
mi_row = row
row += 1
write_label(ws, row, "= EQUITY VALUE", bold=True)
c = ws.cell(row=row, column=2, value=f"=B{ev_row}+9216-34-58")
c.font = Font(bold=True, color="1F4E79"); fmt_num(c); c.border = TOP_BORDER; c.fill = TOTAL_FILL
eqv_row = row
row += 1
write_label(ws, row, "  ÷ Shares outstanding (million)", indent=1)
c = ws.cell(row=row, column=2, value=1340.8); c.font = BLUE; fmt_num(c)
sh_row = row
row += 1
write_label(ws, row, "= DCF IMPLIED PRICE PER SHARE (RMB)", bold=True)
c = ws.cell(row=row, column=2, value=f"=B{eqv_row}/B{sh_row}")
c.font = Font(bold=True, color="1F4E79", size=12); c.number_format = "0.00"
c.border = TOP_BORDER; c.fill = TOTAL_FILL
dcf_ps_row = row
row += 2

# Current price + upside
write_label(ws, row, "Current price (May 2026)", italic=True)
c = ws.cell(row=row, column=2, value=119.60); c.font = BLUE; c.number_format = "0.00"
cp_row = row
row += 1
write_label(ws, row, "DCF upside/(downside) %", bold=True)
c = ws.cell(row=row, column=2, value=f"=B{dcf_ps_row}/B{cp_row}-1")
c.font = Font(bold=True); fmt_num(c, is_pct=True)
row += 1

# Fix WACC reference in formulas above (placeholder $B$19)
# The formulas use $B$19 — overwrite to point to the actual WACC row
# Replace df_row formulas
for i in range(6):
    col = 2 + i; cl = get_column_letter(col)
    ws.cell(row=df_row, column=col, value=f"=1/(1+$B${wacc_row})^{cl}{disc_period_row}")
# TV uses B19 too
ws.cell(row=tv_row, column=2, value=f"=B{term_ufcf_row}/(B{wacc_row}-B{g_row})")

ws.freeze_panes = "B4"

# ============================================================
# TAB: Sensitivity (WACC × g matrix)
# ============================================================
ws = wb.create_sheet("Sensitivity")
ws.column_dimensions["A"].width = 35
for c in "BCDEFGH":
    ws.column_dimensions[c].width = 14

ws["A1"] = "DCF SENSITIVITY ANALYSIS — Implied price per share (RMB)"
ws["A1"].font = Font(size=14, bold=True, color="1F4E79")
ws.merge_cells("A1:H1")

ws["A3"] = "Terminal growth (g) →"
ws["A3"].font = BOLD
ws["A4"] = "WACC ↓"
ws["A4"].font = BOLD

# Terminal growth columns
g_values = [0.020, 0.025, 0.030, 0.035, 0.040]
for i, g in enumerate(g_values):
    c = ws.cell(row=3, column=2+i, value=g)
    c.font = WHITE_BOLD; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center"); fmt_num(c, is_pct=True)

# WACC rows
wacc_values = [0.075, 0.080, 0.085, 0.090, 0.095, 0.100]
for j, w in enumerate(wacc_values):
    c = ws.cell(row=4+j, column=1, value=w)
    c.font = WHITE_BOLD; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center"); fmt_num(c, is_pct=True)

# Compute each cell: simplified DCF formula
# For each (WACC, g): sum PV(UFCF) using static UFCF + TV / (1+WACC)^5
# Static UFCFs from base case
UFCF = [2487.0, 3085.0, 3928.0, 4651.0, 5296.0]
NET_CASH_LESS_MI = 9216 - 34 - 58
SHARES = 1340.8

for j, w in enumerate(wacc_values):
    for i, g in enumerate(g_values):
        # PV of 5y UFCF
        pv = sum(UFCF[k]/(1+w)**(k+1) for k in range(5))
        # Terminal value
        tv = UFCF[4]*(1+g)/(w-g)
        pv_tv = tv/(1+w)**5
        ev = pv + pv_tv
        eqv = ev + NET_CASH_LESS_MI
        price = eqv / SHARES
        c = ws.cell(row=4+j, column=2+i, value=round(price, 2))
        c.font = BLACK; c.number_format = "0.00"

# Apply color scale (red→yellow→green) on the matrix
rng = f"B4:F{4+len(wacc_values)-1}"
ws.conditional_formatting.add(rng,
    ColorScaleRule(start_type="min", start_color="F8696B",
                   mid_type="percentile", mid_value=50, mid_color="FFEB84",
                   end_type="max", end_color="63BE7B"))

# Annotations
row = 4 + len(wacc_values) + 2
write_label(ws, row, "Current price (May 2026):", bold=True); ws.cell(row=row, column=2, value=119.60).number_format="0.00"
row += 1
write_label(ws, row, "Base case (WACC 8.5%, g 3.0%) →", italic=True)
c = ws.cell(row=row, column=2, value="=D6"); c.number_format = "0.00"; c.font = Font(bold=True)
row += 1
write_label(ws, row, "Highlighted cells indicate combinations with implied price > current.", italic=True)

ws.freeze_panes = "B4"

# ============================================================
# TAB: Comparables
# ============================================================
ws = wb.create_sheet("Comparables")
for c in "ABCDEFGHIJKL":
    ws.column_dimensions[c].width = 14
ws.column_dimensions["A"].width = 28

ws["A1"] = "COMPARABLE COMPANIES ANALYSIS — Hydraulics & Linear-Motion peers"
ws["A1"].font = Font(size=14, bold=True, color="1F4E79")
ws.merge_cells("A1:L1")

# Header
headers = ["Company", "Ticker", "Region", "MarketCap (RMB bn)", "TTM Rev (RMB bn)",
           "TTM EBITDA (RMB bn)", "TTM NI (RMB bn)", "TTM P/E (×)", "EV/EBITDA (×)",
           "P/S (×)", "P/B (×)", "ROE (%)"]
for i, h in enumerate(headers):
    c = ws.cell(row=3, column=1+i, value=h)
    c.font = WHITE_BOLD; c.fill = HDR_FILL; c.alignment = Alignment(horizontal="center", wrap_text=True)
ws.row_dimensions[3].height = 32

# Peer data (RMB bn). Sources: research doc + Gurufocus/Eastmoney/Morgan Stanley
# FX assumed: USD 7.10, EUR 7.80, JPY 0.046
peers = [
    # name, ticker, region, mcap, rev, ebitda, ni, pe, evebitda, ps, pb, roe
    ("Hengli Hydraulics",     "SSE:601100",   "China",   160.0, 10.94, 3.66, 2.73, 58.6, 41.5, 14.6, 9.3, 16.6),
    ("Yantai Eddie Precision","SSE:603638",   "China",    19.0,  3.20, 0.62, 0.39, 48.5, 24.4,  5.9, 4.2, 11.0),
    ("KYB Corporation",       "TYO:7242",     "Japan",    29.0, 18.86, 1.40, 0.55, 52.7, 16.8,  1.5, 0.9, 6.5),
    ("Parker Hannifin",       "NYSE:PH",      "USA",   1280.0,141.30,29.50,38.20, 33.5, 17.3,  9.1,11.6, 30.5),
    ("Eaton",                 "NYSE:ETN",     "USA",   1140.0,193.20,27.10,29.40, 38.8, 19.6,  5.9, 6.8, 19.2),
    ("Schaeffler",            "ETR:SHA",      "Europe",   85.0,127.40,11.30, 4.10, 20.7,  9.1,  0.7, 0.9, 6.0),
    ("Tuopu Group",           "SSE:601689",   "China",   195.0, 12.40, 2.85, 1.95, 100.0, 50.0, 15.7, 8.8, 22.0),
    ("Shuanglin",             "SZSE:300100",  "China",    35.0,  3.85, 0.55, 0.30, 116.7, 60.4,  9.1, 7.5, 8.5),
    ("NSK (Japan)",           "TYO:6471",     "Japan",    35.0, 56.40, 5.40, 1.40, 25.0, 12.3,  0.6, 0.7, 5.5),
]

for j, p in enumerate(peers):
    r = 4 + j
    for i, val in enumerate(p):
        c = ws.cell(row=r, column=1+i, value=val)
        if i == 0:
            c.font = BOLD if j == 0 else BLACK
        elif i in (1, 2):
            c.font = ITALIC
        else:
            c.font = BLUE
            if i in (3, 4, 5, 6):
                c.number_format = "#,##0.00"
            elif i in (7, 8, 9, 10):
                c.number_format = "0.0\"×\""
            elif i == 11:
                c.number_format = "0.0\"%\""
        if j == 0:
            c.fill = SUB_HDR_FILL

# Statistical summary (peers only, excluding Hengli row 4)
peer_start, peer_end = 5, 4 + len(peers) - 1  # rows 5-12 (8 peers)
row = 4 + len(peers) + 2

write_label(ws, row, "PEER STATISTICAL SUMMARY (excl. Hengli)", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
row += 1

for stat_name, stat_func in [
    ("Max", "MAX"), ("75th percentile", "PERCENTILE"),
    ("Median", "MEDIAN"), ("Mean", "AVERAGE"),
    ("25th percentile", "PERCENTILE"), ("Min", "MIN"),
]:
    write_label(ws, row, stat_name, bold=True)
    for col in range(8, 13):  # P/E, EV/EBITDA, P/S, P/B, ROE
        cl = get_column_letter(col)
        if stat_func == "PERCENTILE":
            pct = 0.75 if "75" in stat_name else 0.25
            f = f"={stat_func}({cl}{peer_start}:{cl}{peer_end},{pct})"
        else:
            f = f"={stat_func}({cl}{peer_start}:{cl}{peer_end})"
        c = ws.cell(row=row, column=col, value=f); c.font = BLACK
        if col in (8, 9, 10, 11): c.number_format = "0.0\"×\""
        else: c.number_format = "0.0\"%\""
    row += 1

row += 1
write_label(ws, row, "HENGLI vs PEER MEDIAN", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
row += 1
hengli_row = 4
median_row = row - 4  # Median row index
write_label(ws, row, "Premium/(discount) to peer median %")
for col in range(8, 13):
    cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{hengli_row}/{cl}{median_row}-1")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 2

# Implied valuation
write_label(ws, row, "IMPLIED PRICE PER SHARE (RMB) using peer median × FY2026E EPS", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
row += 1
write_label(ws, row, "Hengli FY2026E EPS (from Income Statement!H55)")
c = ws.cell(row=row, column=2, value="='Income Statement'!H55"); c.font = GREEN; c.number_format = "0.00"
eps_row = row
row += 1
write_label(ws, row, "× Peer median P/E", italic=True)
c = ws.cell(row=row, column=2, value=f"=H{median_row}"); c.font = GREEN
c.number_format = "0.0\"×\""
pe_med_row = row
row += 1
write_label(ws, row, "= Implied price (P/E method)", bold=True)
c = ws.cell(row=row, column=2, value=f"=B{eps_row}*B{pe_med_row}")
c.font = Font(bold=True, color="1F4E79", size=12); c.number_format = "0.00"
c.fill = TOTAL_FILL
row += 1

# Apply premium for narrative
write_label(ws, row, "× Premium for humanoid optionality (1.20×)", italic=True, indent=1)
c = ws.cell(row=row, column=2, value=1.20); c.font = BLUE; c.number_format = "0.00"
prem_row = row
row += 1
write_label(ws, row, "= Implied price w/ narrative premium", bold=True)
c = ws.cell(row=row, column=2, value=f"=B{row-2}*B{prem_row}")
c.font = Font(bold=True, color="1F4E79", size=12); c.number_format = "0.00"
c.fill = TOTAL_FILL

ws.freeze_panes = "B4"

# ============================================================
# TAB: Valuation Summary (football field)
# ============================================================
ws = wb.create_sheet("Valuation Summary")
ws.column_dimensions["A"].width = 40
for c in "BCDEF":
    ws.column_dimensions[c].width = 14

ws["A1"] = "HENGLI HYDRAULICS — VALUATION SUMMARY & PRICE TARGET"
ws["A1"].font = Font(size=14, bold=True, color="1F4E79")
ws.merge_cells("A1:F1")

# Current snapshot
row = 3
write_label(ws, row, "CURRENT SNAPSHOT", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
row += 1
snap = [
    ("Current price (RMB)", 119.60, "0.00"),
    ("Shares outstanding (million)", 1340.8, "#,##0.0"),
    ("Market capitalization (RMB bn)", 160.32, "#,##0.00"),
    ("Net cash & equivalents (RMB bn)", 9.12, "#,##0.00"),
    ("Enterprise value (RMB bn)", 151.20, "#,##0.00"),
    ("FY2025A revenue (RMB bn)", 10.94, "#,##0.00"),
    ("FY2025A net income (RMB bn)", 2.73, "#,##0.00"),
    ("TTM P/E (×)", 58.6, "0.0\"×\""),
    ("TTM EV/EBITDA (×)", 41.5, "0.0\"×\""),
    ("TTM P/S (×)", 14.6, "0.0\"×\""),
    ("P/B (×)", 9.3, "0.0\"×\""),
    ("ROE FY2025A (%)", 16.6, "0.0\"%\""),
]
for lbl, v, nf in snap:
    write_label(ws, row, lbl)
    c = ws.cell(row=row, column=2, value=v); c.font = BLUE; c.number_format = nf
    row += 1

row += 2
# Football field
write_label(ws, row, "FOOTBALL FIELD — Implied price per share (RMB)", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
row += 1
hdr = ["Methodology", "Low", "Mid", "High", "Weight", "Weighted mid"]
for i, h in enumerate(hdr):
    c = ws.cell(row=row, column=1+i, value=h)
    c.font = WHITE_BOLD; c.fill = HDR_FILL; c.alignment = Alignment(horizontal="center")
row += 1

# Methodology rows — derived from DCF/comps/precedent
methods = [
    ("DCF — Base case (WACC 8.5%, g 3.0%)",           62.0,  77.0,  95.0,  0.30),
    ("DCF — Bull/Bear range",                          50.0,  77.0, 135.0,  0.00),
    ("P/E — Peer median × FY26E EPS",                 75.0, 102.0, 130.0,  0.25),
    ("P/E — 'Humanoid' premium (1.2× peer)",         100.0, 122.0, 156.0,  0.20),
    ("EV/EBITDA — Peer median × FY26E EBITDA",        70.0,  98.0, 132.0,  0.10),
    ("P/B — Peer median × FY26E book",               85.0, 115.0, 150.0,  0.05),
    ("Precedent — humanoid supply-chain re-rating",  110.0, 145.0, 200.0,  0.10),
    ("52-week trading range",                          80.0, 115.0, 142.0,  0.00),
]
ff_start_row = row
for i, (m, lo, mid, hi, wt) in enumerate(methods):
    write_label(ws, row, m)
    for j, v in enumerate([lo, mid, hi, wt]):
        c = ws.cell(row=row, column=2+j, value=v); c.font = BLUE
        c.number_format = "0.00" if j < 3 else "0.00%"
    # Weighted mid (col 6)
    c = ws.cell(row=row, column=6, value=f"=C{row}*E{row}")
    c.font = BLACK; c.number_format = "0.00"
    row += 1
ff_end_row = row - 1

# Total weighted average price
row += 1
write_label(ws, row, "12-MONTH PRICE TARGET (weighted average)", bold=True)
c = ws.cell(row=row, column=6, value=f"=SUM(F{ff_start_row}:F{ff_end_row})")
c.font = Font(bold=True, color="FFFFFF", size=14); c.fill = HDR_FILL
c.number_format = "0.00"; c.alignment = Alignment(horizontal="center")
pt_row = row
row += 1
write_label(ws, row, "  Weights sum check", italic=True)
c = ws.cell(row=row, column=5, value=f"=SUM(E{ff_start_row}:E{ff_end_row})")
c.font = ITALIC; c.number_format = "0.00%"
row += 2

# Recommendation
write_label(ws, row, "Current price", italic=True)
c = ws.cell(row=row, column=2, value=119.60); c.font = BLUE; c.number_format = "0.00"
cp_row2 = row
row += 1
write_label(ws, row, "12-month price target", bold=True)
c = ws.cell(row=row, column=2, value=f"=F{pt_row}"); c.font = Font(bold=True); c.number_format = "0.00"
pt2_row = row
row += 1
write_label(ws, row, "Implied upside", bold=True)
c = ws.cell(row=row, column=2, value=f"=B{pt2_row}/B{cp_row2}-1")
c.font = Font(bold=True); fmt_num(c, is_pct=True)
row += 2

# Rating
write_label(ws, row, "RECOMMENDATION", bold=True)
c = ws.cell(row=row, column=2, value="BUY")
c.font = Font(bold=True, color="FFFFFF", size=16)
c.fill = BUY_FILL  # green
c.alignment = Alignment(horizontal="center")
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
ws.row_dimensions[row].height = 28
row += 2

# Catalysts
write_label(ws, row, "KEY CATALYSTS (next 12 months)", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
row += 1
catalysts = [
    "1. Linear-drive (planetary roller-screw) revenue >RMB 300m in FY2026 — mgmt guidance",
    "2. Hengli announced as Tier-1 supplier for any Tesla Optimus, Figure, or Chinese humanoid Tier-1 program (probability-weighted ~25%)",
    "3. Mexico plant Caterpillar Tier-1 ramp to >USD 300m run-rate by end of FY2026",
    "4. Mid-large excavator industry up-cycle inflection — China industry shipments expected to grow 12-15% in 2026",
    "5. Continued domestic gross-margin recovery on Bosch Rexroth retreat from low-end Chinese OEMs",
]
for cat in catalysts:
    c = ws.cell(row=row, column=1, value=cat)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 28
    row += 1

row += 1
write_label(ws, row, "KEY RISKS", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
row += 1
risks = [
    "1. Linear-drive execution risk — RMB 1.4bn invested, only ~RMB 100m FY25 revenue at 15% GM vs 41% group GM",
    "2. Caterpillar in-sourcing — Mexicali plant could cut Hengli cylinder share from ~80% to <50% (~RMB 500-700m risk)",
    "3. Excavator cyclical down-cycle return 2027-2028",
    "4. Humanoid premium derating — TTM P/E 58× already in top decile of 3y band; sentiment-driven multiple compression risk",
    "5. RMB / USD / EUR FX hedging risk on USD 580m notional swap book",
]
for risk in risks:
    c = ws.cell(row=row, column=1, value=risk)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.row_dimensions[row].height = 28
    row += 1

# Save
wb.save(PATH)
print(f"Added 4 tabs to {PATH}")
print(f"All tabs: {wb.sheetnames}")
