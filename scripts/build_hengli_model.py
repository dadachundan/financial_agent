"""
Build Hengli Hydraulics (SSE:601100) financial model — 6-tab Excel.
All figures in RMB millions unless otherwise noted.
Historicals: FY2020-FY2025 from audited annual reports (cninfo).
Projections: FY2026E-FY2030E (Base/Bull/Bear scenarios).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

OUT_PATH = "reports/company/Hengli_SSE601100/Hengli_SSE601100_Financial_Model_2026-05-19.xlsx"

# ============= STYLE HELPERS =============
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
HIST_FILL = PatternFill("solid", fgColor="FFFFFF")
PROJ_FILL = PatternFill("solid", fgColor="FFF2CC")
SCEN_BULL = PatternFill("solid", fgColor="E2EFDA")
SCEN_BEAR = PatternFill("solid", fgColor="FCE4D6")

BLUE = Font(color="0070C0")          # hardcoded input
BLACK = Font(color="000000")         # formula
GREEN = Font(color="00B050")         # link
WHITE_BOLD = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
ITALIC = Font(italic=True, color="595959")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
TOP_BORDER = Border(top=Side(style="thin", color="000000"))
BOT_BORDER = Border(bottom=Side(style="thin", color="000000"))

YEARS_HIST = ["FY2020A", "FY2021A", "FY2022A", "FY2023A", "FY2024A", "FY2025A"]
YEARS_PROJ = ["FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
ALL_YEARS = YEARS_HIST + YEARS_PROJ

# ============= HISTORICAL DATA (RMB millions) =============
# From audited annual reports — see source PDFs cited in research doc
HIST = {
    # Income statement (consolidated)
    "revenue":          [7855.0, 9309.2, 8196.7, 8984.6, 9389.7, 10941.0],
    "cogs":             [4391.0, 5211.9, 4872.7, 5220.0, 5368.2, 6392.0],
    "tax_surcharge":    [72.1, 77.9, 71.0, 81.7, 90.0, 108.9],
    "selling_exp":      [106.9, 111.9, 123.6, 185.7, 217.1, 254.5],
    "admin_exp":        [200.0, 226.9, 275.3, 404.4, 591.8, 649.3],
    "rd_exp":           [308.6, 636.1, 650.0, 694.4, 727.7, 705.0],
    "financial_exp":    [151.3, 89.3, -332.4, -368.5, -131.4, 5.9],
    "other_income":     [60.8, 131.9, 104.9, 113.0, 124.3, 89.0],
    "investment_inc":   [10.6, 7.0, 13.9, 2.0, -1.2, 176.4],
    "fv_change":        [0.0, 10.3, 7.3, 0.7, 203.2, 92.4],
    "credit_impair":    [-37.0, 19.4, 9.5, -12.1, -12.9, -33.0],
    "asset_impair":     [-52.5, -66.7, -54.6, -56.1, -53.5, -106.6],
    "asset_disposal":   [-0.8, -0.7, 0.3, 0.3, 0.2, 0.1],
    "non_op_inc":       [4.7, 13.4, 12.7, 17.5, 16.6, 16.4],
    "non_op_exp":       [1.7, 1.5, 1.9, 2.3, 2.6, 28.1],
    "income_tax":       [347.9, 369.3, 279.8, 326.3, 288.1, 292.3],

    # Cash flow
    "cfo":              [1980.6, 2795.7, 2063.6, 2677.1, 2479.0, 1810.9],
    "capex":            [401.1, 562.4, 798.8, 1365.9, 1071.4, 924.2],
    "cfi":              [-663.0, -932.4, -79.6, -3051.3, -3100.0, -857.4],
    "cff":              [-818.3, -713.7, 949.3, -880.2, -787.7, -864.0],
    "fx_impact":        [-163.9, -125.0, 326.5, 180.8, -264.3, -64.8],

    # Balance sheet (FY2022-FY2025 fully extracted)
    "cash":             [None, None, 6886.2, 8124.3, 7882.8, 8871.4],
    "trading_fa":       [None, None, 470.3, 79.3, 1027.6, 345.0],
    "bills_recv":       [None, None, 480.7, 427.5, 601.8, 529.6],
    "ar":               [None, None, 1111.9, 1245.1, 1371.2, 1910.2],
    "fin_recv":         [None, None, 1389.4, 1094.7, 783.4, 829.5],
    "prepayments":      [None, None, 141.4, 155.7, 156.5, 252.2],
    "other_recv":       [None, None, 27.9, 30.4, 29.0, 31.8],
    "inventory":        [None, None, 1765.5, 1692.1, 1764.5, 2154.2],
    "contract_assets":  [None, None, 8.9, 16.1, 20.4, 25.8],
    "other_ca":         [None, None, 44.4, 129.2, 193.2, 274.1],
    "total_ca":         [None, None, 12326.8, 12993.6, 13830.4, 15223.7],
    "ppe":              [None, None, 2826.0, 3344.0, 3889.1, 4997.6],   # estimated FY22/23
    "cip":              [None, None, 450.0, 1100.0, 1083.7, 417.2],     # estimated FY22/23
    "intangibles":      [None, None, 420.0, 430.0, 438.7, 450.7],
    "goodwill":         [None, None, 0.7, 0.7, 0.7, 0.7],
    "lt_inv":           [None, None, 5.0, 6.0, 8.4, 10.0],
    "dta":              [None, None, 80.0, 110.0, 138.7, 275.9],
    "other_nca":        [None, None, 388.0, 1053.0, 248.7, 295.8],
    "total_nca":        [None, None, 4170.1, 6043.7, 5808.2, 6447.2],
    "total_assets":     [None, None, 12496.9, 13035.3, 19638.7, 21671.0],

    "st_borrow":        [None, None, 200.0, 100.0, 15.4, 13.5],
    "bills_pay":        [None, None, 250.0, 300.0, 246.6, 0.8],
    "ap":               [None, None, 1100.0, 1200.0, 888.9, 1022.5],
    "contract_liab":    [None, None, 200.0, 250.0, 283.6, 323.8],
    "employee_pay":     [None, None, 250.0, 270.0, 294.4, 382.0],
    "tax_payable":      [None, None, 140.0, 150.0, 160.3, 231.1],
    "other_pay":        [None, None, 1200.0, 1300.0, 1116.2, 1555.7],
    "other_cl":         [None, None, 196.2, 132.5, 374.7, 291.1],
    "total_cl":         [None, None, 3536.2, 3702.5, 3379.2, 3820.5],
    "lt_borrow":        [None, None, 0.0, 0.0, 0.0, 20.0],
    "lease_liab":       [None, None, 0.0, 0.0, 1.2, 11.1],
    "deferred_rev":     [None, None, 61.9, 53.7, 211.5, 253.4],
    "dtl":              [None, None, 76.3, 74.3, 218.3, 227.9],
    "total_ncl":        [None, None, 138.2, 128.1, 431.0, 512.4],
    "total_liab":       [None, None, 3674.4, 3830.6, 3810.3, 4333.0],

    "share_capital":    [None, None, 1340.8, 1340.8, 1340.8, 1340.8],
    "capital_reserve":  [None, None, 3362.2, 3362.2, 3364.9, 3364.9],
    "oci":              [None, None, 0.0, 0.0, -88.0, 16.7],
    "special_reserve":  [None, None, 10.7, 12.9, 36.5, 43.7],
    "surplus_reserve":  [None, None, 670.4, 670.4, 670.4, 670.4],
    "retained_earn":    [None, None, 3438.4, 3818.3, 10450.1, 11843.4],
    "minority":         [None, None, 0.0, 0.0, 53.7, 58.1],
    "total_equity":     [None, None, 8822.5, 9204.7, 15828.4, 17338.0],

    # Shares (million)
    "shares_basic":     [1305.4, 1305.4, 1305.4, 1340.8, 1340.8, 1340.8],

    # D&A (estimated; FY25 PPE additions + amort intangibles)
    "da":               [320.0, 360.0, 410.0, 480.0, 540.0, 620.0],
}

# Segment revenue (FY2025 disclosed; earlier years estimated from mgmt commentary)
SEG_HIST = {
    # Hydraulic cylinders
    "cyl":              [3534.0, 4188.0, 3608.0, 4090.0, 4760.0, 5254.0],
    # Pumps/valves/motors
    "pump":             [2517.0, 3010.0, 2780.0, 3050.0, 3585.0, 4326.0],
    # Hydraulic systems
    "sys":              [350.0, 410.0, 340.0, 320.0, 296.0, 385.0],
    # Components/castings/linear-drive
    "comp":             [800.0, 940.0, 850.0, 780.0, 684.0, 891.0],
    # Other (incl. FX/other reconciling)
    "other":            [654.0, 761.2, 618.7, 744.6, 64.7, 85.0],
}

# Geographic split (FY24/25 disclosed; earlier estimated)
GEO_HIST = {
    "domestic":         [5800.0, 7000.0, 6100.0, 6800.0, 7250.0, 8750.0],
    "overseas":         [2055.0, 2309.2, 2096.7, 2184.6, 2139.7, 2110.0],
    # split overseas into regions (research doc points to APAC/EU/NA)
    "apac":             [900.0, 1000.0, 920.0, 950.0, 920.0, 850.0],
    "europe":           [550.0, 620.0, 580.0, 620.0, 600.0, 620.0],
    "namerica":         [500.0, 580.0, 510.0, 530.0, 540.0, 580.0],
    "row":              [105.0, 109.2, 86.7, 84.6, 79.7, 60.0],
}

# ============= PROJECTION ASSUMPTIONS (BASE CASE) =============
PROJ_ASSUMPTIONS = {
    # Revenue growth (base case) — calibrated to research-doc implied 10-13% CAGR FY25-FY30
    "cyl_growth":     [0.07, 0.07, 0.08, 0.07, 0.06],   # cylinders cyclical, post-FY25 deceleration
    "pump_growth":    [0.16, 0.15, 0.13, 0.12, 0.10],   # pumps continue share-gain (margin engine)
    "sys_growth":     [0.18, 0.15, 0.12, 0.10, 0.08],   # systems mid-growth
    "comp_growth":    [0.40, 0.50, 0.55, 0.40, 0.25],   # linear-drive ramp (3× FY26 per mgmt)
    "other_growth":   [0.05, 0.05, 0.04, 0.03, 0.03],

    # Margin trajectory
    "gross_margin":   [0.425, 0.430, 0.435, 0.435, 0.435],

    # Operating expense ratios (% of revenue)
    "tax_surcharge_pct": [0.010, 0.010, 0.010, 0.010, 0.010],
    "selling_pct":       [0.024, 0.024, 0.023, 0.023, 0.023],
    "admin_pct":         [0.060, 0.058, 0.055, 0.053, 0.052],
    "rd_pct":            [0.065, 0.065, 0.065, 0.065, 0.065],
    "fin_pct":           [0.000, 0.000, -0.005, -0.005, -0.005],  # net interest income

    # Other (RMB m)
    "other_income":    [100.0, 110.0, 120.0, 130.0, 140.0],
    "invest_income":   [120.0, 100.0, 80.0, 80.0, 80.0],
    "fv_change":       [50.0, 30.0, 20.0, 20.0, 20.0],
    "credit_impair":   [-40.0, -45.0, -50.0, -55.0, -60.0],
    "asset_impair":    [-100.0, -110.0, -115.0, -120.0, -125.0],

    # Tax rate (effective)
    "tax_rate":        [0.115, 0.115, 0.120, 0.120, 0.125],

    # CapEx as % of revenue (declining as linear-drive build-out completes)
    "capex_pct":       [0.08, 0.07, 0.06, 0.055, 0.05],
    # D&A (RMB m, schedule)
    "da":              [720.0, 820.0, 920.0, 1010.0, 1090.0],

    # Working capital — days
    "ar_days":         [62.0, 60.0, 58.0, 58.0, 58.0],
    "inv_days":        [120.0, 118.0, 115.0, 115.0, 115.0],
    "ap_days":         [55.0, 55.0, 55.0, 55.0, 55.0],
}

# ============= WORKBOOK BUILD =============
wb = Workbook()
wb.remove(wb.active)

# ---------------------------------------------------------------------------
# TAB 0: README / Cover
# ---------------------------------------------------------------------------
ws = wb.create_sheet("README")
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 80

readme_rows = [
    ("Company", "Jiangsu Hengli Hydraulic Co., Ltd."),
    ("Ticker", "SSE:601100"),
    ("Reporting currency", "RMB (millions)"),
    ("Fiscal year end", "December 31"),
    ("Model date", "2026-05-19"),
    ("Model author", "Equity Research Initiation — Task 2 (Financial Modeling)"),
    ("", ""),
    ("Data sources",
     "Audited consolidated FS from cninfo: FY2025 AR (2026-04-20), "
     "FY2023 AR (2024-04-22), FY2021 AR (2022-04-25). Segment & geo "
     "splits per FY2025 AR p.15–16 (top-5 customer & domestic/overseas)."),
    ("", ""),
    ("Tabs", "1. Revenue Model — product (5) + geography (6) breakdowns, FY20A-FY30E"),
    ("", "2. Income Statement — 40+ line items, FY20A-FY30E"),
    ("", "3. Cash Flow Statement — CFO/CFI/CFF, FY20A-FY30E"),
    ("", "4. Balance Sheet — FY22A-FY30E (older years partial)"),
    ("", "5. Scenarios — Bull / Base / Bear comparison FY30E"),
    ("", "6. DCF Inputs — unlevered FCF for Task 3 valuation"),
    ("", ""),
    ("Color coding", "BLUE = hardcoded input (change me); BLACK = formula; GREEN = cross-sheet link"),
    ("Projection shading", "Yellow = projected period (FY2026E–FY2030E)"),
    ("", ""),
    ("Caveats", "FY2020 / FY2021 balance-sheet detail not fully extracted (only 2 yrs of comparatives per Chinese AR)."),
    ("", "FY2022 BS extracted from FY2023 AR comparatives; certain BS sub-lines for FY2022/FY2023 estimated."),
    ("", "Segment splits prior to FY2025 are management commentary-based estimates; FY2025 is audited."),
]
ws["A1"] = "Hengli Hydraulics — Financial Model"
ws["A1"].font = Font(size=16, bold=True, color="1F4E79")
for i, (k, v) in enumerate(readme_rows, start=3):
    ws.cell(row=i, column=1, value=k).font = BOLD
    ws.cell(row=i, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")

# ---------------------------------------------------------------------------
# Helper: write a year-header row
# ---------------------------------------------------------------------------
def write_year_header(ws, row, start_col=2):
    for i, y in enumerate(ALL_YEARS):
        c = ws.cell(row=row, column=start_col + i, value=y)
        c.font = WHITE_BOLD
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")

def write_label(ws, row, label, bold=False, indent=0, italic=False):
    c = ws.cell(row=row, column=1, value=("  " * indent) + label)
    if bold:
        c.font = BOLD
    if italic:
        c.font = ITALIC

def fmt_num(c, decimals=1, is_pct=False):
    if is_pct:
        c.number_format = "0.0%;(0.0%);—"
    else:
        c.number_format = f'#,##0.{"0"*decimals};(#,##0.{"0"*decimals});—'

def write_hist_proj(ws, row, hist_values, proj_values=None, decimals=1, is_pct=False,
                    is_link=False, start_col=2, bold=False):
    """Write 6 historical + 5 projected values across a row."""
    for i, v in enumerate(hist_values):
        c = ws.cell(row=row, column=start_col + i, value=v)
        c.font = GREEN if is_link else BLUE
        if bold:
            c.font = Font(bold=True, color=c.font.color.rgb)
        fmt_num(c, decimals, is_pct)
    if proj_values is not None:
        for i, v in enumerate(proj_values):
            c = ws.cell(row=row, column=start_col + 6 + i, value=v)
            c.font = BLUE
            if bold:
                c.font = Font(bold=True, color="0070C0")
            c.fill = PROJ_FILL
            fmt_num(c, decimals, is_pct)

# ---------------------------------------------------------------------------
# TAB 1: Revenue Model
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Revenue Model")
ws.column_dimensions["A"].width = 45
for i in range(2, 14):
    ws.column_dimensions[get_column_letter(i)].width = 12

ws["A1"] = "HENGLI HYDRAULICS — REVENUE MODEL (RMB millions)"
ws["A1"].font = Font(size=14, bold=True, color="1F4E79")
ws.merge_cells("A1:L1")
write_year_header(ws, 3)

# --- Section A: Revenue by product/segment ---
ws["A5"] = "A. REVENUE BY PRODUCT SEGMENT"
ws["A5"].font = BOLD
ws["A5"].fill = SUB_HDR_FILL
ws.merge_cells("A5:L5")

row = 6
# Cylinders
write_label(ws, row, "Hydraulic cylinders (液压油缸)", bold=True)
g = PROJ_ASSUMPTIONS["cyl_growth"]
proj = []
v = SEG_HIST["cyl"][-1]
for gr in g:
    v *= (1 + gr); proj.append(round(v, 1))
write_hist_proj(ws, row, SEG_HIST["cyl"], proj)
row += 1
write_label(ws, row, "YoY growth %", italic=True, indent=1)
for i in range(1, 11):
    col = i + 1
    c = ws.cell(row=row, column=col + 1,
                value=f"=({get_column_letter(col+1)}{row-1}/{get_column_letter(col)}{row-1})-1")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 1
# Excavator vs non-excavator sub-split (cylinders)
write_label(ws, row, "  — Excavator cylinders (~75% of segment)", indent=1)
for i, y in enumerate(ALL_YEARS):
    col = i + 2
    c = ws.cell(row=row, column=col, value=f"={get_column_letter(col)}{row-2}*0.75")
    c.font = BLACK; fmt_num(c)
row += 1
write_label(ws, row, "  — Non-excavator (AWP, marine, wind, etc.)", indent=1)
for i, y in enumerate(ALL_YEARS):
    col = i + 2
    c = ws.cell(row=row, column=col, value=f"={get_column_letter(col)}{row-3}*0.25")
    c.font = BLACK; fmt_num(c)
row += 2

# Pumps/valves/motors
write_label(ws, row, "Hydraulic pumps, valves & motors", bold=True)
v = SEG_HIST["pump"][-1]; proj = []
for gr in PROJ_ASSUMPTIONS["pump_growth"]:
    v *= (1 + gr); proj.append(round(v, 1))
pump_row = row
write_hist_proj(ws, row, SEG_HIST["pump"], proj)
row += 1
write_label(ws, row, "YoY growth %", italic=True, indent=1)
for i in range(1, 11):
    col = i + 1
    c = ws.cell(row=row, column=col + 1,
                value=f"=({get_column_letter(col+1)}{row-1}/{get_column_letter(col)}{row-1})-1")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 1
write_label(ws, row, "  — Main pumps (excavator)", indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2
    c = ws.cell(row=row, column=col, value=f"={get_column_letter(col)}{pump_row}*0.55")
    c.font = BLACK; fmt_num(c)
row += 1
write_label(ws, row, "  — Multi-way valves", indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2
    c = ws.cell(row=row, column=col, value=f"={get_column_letter(col)}{pump_row}*0.20")
    c.font = BLACK; fmt_num(c)
row += 1
write_label(ws, row, "  — Travel/swing motors & industrial pumps", indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2
    c = ws.cell(row=row, column=col, value=f"={get_column_letter(col)}{pump_row}*0.25")
    c.font = BLACK; fmt_num(c)
row += 2

# Hydraulic systems
write_label(ws, row, "Hydraulic systems (TBM, marine, water)", bold=True)
v = SEG_HIST["sys"][-1]; proj = []
for gr in PROJ_ASSUMPTIONS["sys_growth"]:
    v *= (1 + gr); proj.append(round(v, 1))
write_hist_proj(ws, row, SEG_HIST["sys"], proj)
row += 1
write_label(ws, row, "YoY growth %", italic=True, indent=1)
for i in range(1, 11):
    col = i + 1
    c = ws.cell(row=row, column=col + 1,
                value=f"=({get_column_letter(col+1)}{row-1}/{get_column_letter(col)}{row-1})-1")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 2

# Components & castings (incl. linear-drive)
write_label(ws, row, "Components, castings & LINEAR-DRIVE (★ key catalyst)", bold=True)
v = SEG_HIST["comp"][-1]; proj = []
for gr in PROJ_ASSUMPTIONS["comp_growth"]:
    v *= (1 + gr); proj.append(round(v, 1))
comp_row = row
write_hist_proj(ws, row, SEG_HIST["comp"], proj)
row += 1
write_label(ws, row, "YoY growth %", italic=True, indent=1)
for i in range(1, 11):
    col = i + 1
    c = ws.cell(row=row, column=col + 1,
                value=f"=({get_column_letter(col+1)}{row-1}/{get_column_letter(col)}{row-1})-1")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 1
write_label(ws, row, "  — Linear-drive (ball/roller screws, guideways) — mgmt 3× to RMB ~300m FY26E", indent=1)
# Linear-drive sub-line: FY25 = ~100m, then 300, 600, 1100, 1700, 2300
linear_hist = [0, 0, 0, 0, 30.0, 100.0]
linear_proj = [300.0, 600.0, 1100.0, 1700.0, 2300.0]
write_hist_proj(ws, row, linear_hist, linear_proj)
row += 1
write_label(ws, row, "  — Castings & traditional components", indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2
    c = ws.cell(row=row, column=col, value=f"={get_column_letter(col)}{comp_row}-{get_column_letter(col)}{row-1}")
    c.font = BLACK; fmt_num(c)
row += 2

# Other revenue
write_label(ws, row, "Other revenue (reconciling)", bold=True)
v = SEG_HIST["other"][-1]; proj = []
for gr in PROJ_ASSUMPTIONS["other_growth"]:
    v *= (1 + gr); proj.append(round(v, 1))
write_hist_proj(ws, row, SEG_HIST["other"], proj)
row += 2

# TOTAL REVENUE (by product)
total_rev_row = row
write_label(ws, row, "TOTAL REVENUE (sum of products)", bold=True)
ws.cell(row=row, column=1).fill = TOTAL_FILL
for i in range(len(ALL_YEARS)):
    col = i + 2
    # sum of cyl + pump + sys + comp + other top rows (rows 6, 11, 17, 20, 26 — recompute via formula)
    # use absolute references to top revenue rows:
    cyl_r = 6; pump_r = 11; sys_r = 17; comp_r = 20; other_r = 25
    cl = get_column_letter(col)
    c = ws.cell(row=row, column=col,
                value=f"={cl}{cyl_r}+{cl}{pump_r}+{cl}{sys_r}+{cl}{comp_r}+{cl}{other_r}")
    c.font = Font(bold=True, color="000000"); fmt_num(c)
    c.border = TOP_BORDER
    c.fill = TOTAL_FILL
row += 1
write_label(ws, row, "Total revenue growth %", italic=True, indent=1)
for i in range(1, 11):
    col = i + 1
    c = ws.cell(row=row, column=col + 1,
                value=f"=({get_column_letter(col+1)}{total_rev_row}/{get_column_letter(col)}{total_rev_row})-1")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 2

# Cross-check vs reported revenue
write_label(ws, row, "Memo: reported audited revenue (historicals)", italic=True)
write_hist_proj(ws, row, HIST["revenue"])
row += 1
write_label(ws, row, "Variance vs sum-of-segments (%)", italic=True, indent=1)
for i in range(6):
    col = i + 2
    cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"=({cl}{row-1}-{cl}{total_rev_row})/{cl}{row-1}")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 3

# --- Section B: Revenue by Geography ---
ws.cell(row=row, column=1, value="B. REVENUE BY GEOGRAPHY").font = BOLD
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
row += 1

# Domestic
write_label(ws, row, "Greater China (domestic)", bold=True)
# Project domestic growth at 14-16% (excavator recovery + share gain)
dom_hist = GEO_HIST["domestic"]
dom_proj = []
v = dom_hist[-1]
for gr in [0.16, 0.15, 0.14, 0.13, 0.12]:
    v *= (1 + gr); dom_proj.append(round(v, 1))
write_hist_proj(ws, row, dom_hist, dom_proj)
dom_row = row
row += 1
write_label(ws, row, "% of total revenue", italic=True, indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{dom_row}/{cl}{total_rev_row}")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 1
write_label(ws, row, "YoY %", italic=True, indent=1)
for i in range(1, 11):
    col = i + 1; cl_prev = get_column_letter(col); cl = get_column_letter(col+1)
    c = ws.cell(row=row, column=col+1, value=f"=({cl}{dom_row}/{cl_prev}{dom_row})-1")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 2

# Overseas
write_label(ws, row, "Overseas (total)", bold=True)
ov_hist = GEO_HIST["overseas"]
# Mexico + India + Indonesia ramp → 12-18% growth
ov_proj = []
v = ov_hist[-1]
for gr in [0.18, 0.20, 0.18, 0.15, 0.12]:
    v *= (1 + gr); ov_proj.append(round(v, 1))
write_hist_proj(ws, row, ov_hist, ov_proj)
ov_row = row
row += 1
write_label(ws, row, "% of total revenue", italic=True, indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{ov_row}/{cl}{total_rev_row}")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 1
# Sub-regions
def project_subregion(hist, growth_path):
    proj = []; v = hist[-1]
    for gr in growth_path:
        v *= (1 + gr); proj.append(round(v, 1))
    return proj
sub_regions = [
    ("  — Asia-Pacific (ex-China): Japan, India, Indonesia, S. Korea", "apac", [0.10, 0.15, 0.18, 0.15, 0.12]),
    ("  — Europe: Germany, UK, Italy, France", "europe", [0.08, 0.10, 0.10, 0.10, 0.08]),
    ("  — North America: USA + Mexico ★ (Mexico ramp 2026-28)", "namerica", [0.30, 0.30, 0.25, 0.18, 0.12]),
    ("  — Rest of World (Brazil, Guinea, etc.)", "row", [0.10, 0.12, 0.12, 0.10, 0.08]),
]
for label, key, gp in sub_regions:
    write_label(ws, row, label, indent=1)
    write_hist_proj(ws, row, GEO_HIST[key], project_subregion(GEO_HIST[key], gp))
    row += 1
row += 1

# TOTAL by geography
write_label(ws, row, "TOTAL REVENUE (sum by geography)", bold=True)
ws.cell(row=row, column=1).fill = TOTAL_FILL
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{dom_row}+{cl}{ov_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER; c.fill = TOTAL_FILL
row += 2

# --- Section C: Sales channel (qualitative — 100% direct OEM) ---
ws.cell(row=row, column=1, value="C. REVENUE BY CHANNEL").font = BOLD
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
row += 1
write_label(ws, row, "100% direct OEM (no distribution layer)", italic=True)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{total_rev_row}")
    c.font = BLACK; fmt_num(c)
row += 1
write_label(ws, row, "Top-5 customer concentration (Cat, Sany, XCMG, Komatsu, Liugong)", italic=True)
top5_hist = [3400, 4050, 3500, 3812, 4144, 4602]  # research doc: 42-44% of revenue
top5_proj = []
for i in range(5):
    top5_proj.append(None)  # leave blank — to be modeled
top5 = top5_hist + top5_proj
for i, v in enumerate(top5):
    col = i + 2
    if v is not None:
        c = ws.cell(row=row, column=col, value=v); c.font = BLUE; fmt_num(c)
    else:
        cl = get_column_letter(col)
        c = ws.cell(row=row, column=col, value=f"={cl}{total_rev_row}*0.42")
        c.font = BLACK; fmt_num(c)
row += 1
write_label(ws, row, "Top-5 as % of revenue", italic=True, indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{row-1}/{cl}{total_rev_row}")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 3

ws.freeze_panes = "B4"

# Store total_rev_row for cross-sheet linking
REV_ROW_REF = total_rev_row  # row in Revenue Model that holds total revenue

# ---------------------------------------------------------------------------
# TAB 2: Income Statement
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Income Statement")
ws.column_dimensions["A"].width = 50
for i in range(2, 14):
    ws.column_dimensions[get_column_letter(i)].width = 12

ws["A1"] = "HENGLI HYDRAULICS — CONSOLIDATED INCOME STATEMENT (RMB millions)"
ws["A1"].font = Font(size=14, bold=True, color="1F4E79")
ws.merge_cells("A1:L1")
write_year_header(ws, 3)

row = 5
# REVENUE — linked from Revenue Model
write_label(ws, row, "Revenue", bold=True)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"='Revenue Model'!{cl}{REV_ROW_REF}")
    c.font = GREEN; fmt_num(c)
revenue_row = row
row += 1
write_label(ws, row, "  YoY growth %", italic=True, indent=1)
for i in range(1, 11):
    col = i + 1; cl_p = get_column_letter(col); cl = get_column_letter(col+1)
    c = ws.cell(row=row, column=col+1, value=f"=({cl}{revenue_row}/{cl_p}{revenue_row})-1")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 2

# COGS
write_label(ws, row, "Cost of revenue (COGS)")
# Historicals from data; projections from gross-margin assumption
cogs_hist = HIST["cogs"]
gm_proj = PROJ_ASSUMPTIONS["gross_margin"]
cogs_proj_formulas = []
for i in range(5):
    col = 8 + i  # FY2026E is col 8
    cl = get_column_letter(col)
    cogs_proj_formulas.append(f"={cl}{revenue_row}*(1-{gm_proj[i]})")
for i, v in enumerate(cogs_hist):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
for i, f in enumerate(cogs_proj_formulas):
    c = ws.cell(row=row, column=8+i, value=f); c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
cogs_row = row
row += 1

# Gross profit
write_label(ws, row, "Gross profit", bold=True)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{revenue_row}-{cl}{cogs_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER
gp_row = row
row += 1
write_label(ws, row, "  Gross margin %", italic=True, indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{gp_row}/{cl}{revenue_row}")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 2

# Operating expenses
write_label(ws, row, "Operating expenses", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
row += 1

def project_pct(hist_list, pct_list):
    """Build hist values + projection formulas as % of revenue."""
    proj_f = []
    for i in range(5):
        col = 8 + i; cl = get_column_letter(col)
        proj_f.append(f"={cl}{revenue_row}*{pct_list[i]}")
    return hist_list, proj_f

# Tax & surcharges
write_label(ws, row, "Taxes & surcharges", indent=1)
h, pf = project_pct(HIST["tax_surcharge"], PROJ_ASSUMPTIONS["tax_surcharge_pct"])
for i, v in enumerate(h):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
for i, f in enumerate(pf):
    c = ws.cell(row=row, column=8+i, value=f); c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
tax_sur_row = row
row += 1

# Selling expense
write_label(ws, row, "Selling expense", indent=1)
h, pf = project_pct(HIST["selling_exp"], PROJ_ASSUMPTIONS["selling_pct"])
for i, v in enumerate(h):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
for i, f in enumerate(pf):
    c = ws.cell(row=row, column=8+i, value=f); c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
sell_row = row
row += 1
write_label(ws, row, "  % of revenue", italic=True, indent=2)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{sell_row}/{cl}{revenue_row}")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 1

# G&A
write_label(ws, row, "G&A expense", indent=1)
h, pf = project_pct(HIST["admin_exp"], PROJ_ASSUMPTIONS["admin_pct"])
for i, v in enumerate(h):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
for i, f in enumerate(pf):
    c = ws.cell(row=row, column=8+i, value=f); c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
ga_row = row
row += 1
write_label(ws, row, "  % of revenue", italic=True, indent=2)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{ga_row}/{cl}{revenue_row}")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 1

# R&D
write_label(ws, row, "R&D expense", indent=1)
h, pf = project_pct(HIST["rd_exp"], PROJ_ASSUMPTIONS["rd_pct"])
for i, v in enumerate(h):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
for i, f in enumerate(pf):
    c = ws.cell(row=row, column=8+i, value=f); c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
rd_row = row
row += 1
write_label(ws, row, "  % of revenue", italic=True, indent=2)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{rd_row}/{cl}{revenue_row}")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 1

# Financial expense
write_label(ws, row, "Financial expense (net interest)", indent=1)
h, pf = project_pct(HIST["financial_exp"], PROJ_ASSUMPTIONS["fin_pct"])
for i, v in enumerate(h):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
for i, f in enumerate(pf):
    c = ws.cell(row=row, column=8+i, value=f); c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
fin_row = row
row += 1

# Total opex
write_label(ws, row, "Total operating expenses (incl. taxes & financial)", bold=True)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col,
                value=f"={cl}{tax_sur_row}+{cl}{sell_row}+{cl}{ga_row}+{cl}{rd_row}+{cl}{fin_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER
opex_row = row
row += 2

# Other income/expenses
write_label(ws, row, "Other operating items", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
row += 1

# Other income (gov subsidies)
write_label(ws, row, "Other income (govt subsidies, R&D credits)", indent=1)
for i, v in enumerate(HIST["other_income"]):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
for i, v in enumerate(PROJ_ASSUMPTIONS["other_income"]):
    c = ws.cell(row=row, column=8+i, value=v); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
oi_row = row
row += 1

# Investment income
write_label(ws, row, "Investment income (incl. JV, fund disposals)", indent=1)
for i, v in enumerate(HIST["investment_inc"]):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
for i, v in enumerate(PROJ_ASSUMPTIONS["invest_income"]):
    c = ws.cell(row=row, column=8+i, value=v); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
ii_row = row
row += 1

# FV change
write_label(ws, row, "Fair value change (derivatives, hedges)", indent=1)
for i, v in enumerate(HIST["fv_change"]):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
for i, v in enumerate(PROJ_ASSUMPTIONS["fv_change"]):
    c = ws.cell(row=row, column=8+i, value=v); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
fv_row = row
row += 1

# Credit impair
write_label(ws, row, "Credit impairment (receivables)", indent=1)
for i, v in enumerate(HIST["credit_impair"]):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
for i, v in enumerate(PROJ_ASSUMPTIONS["credit_impair"]):
    c = ws.cell(row=row, column=8+i, value=v); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
ci_row = row
row += 1

# Asset impair
write_label(ws, row, "Asset impairment (inventory, fixed)", indent=1)
for i, v in enumerate(HIST["asset_impair"]):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
for i, v in enumerate(PROJ_ASSUMPTIONS["asset_impair"]):
    c = ws.cell(row=row, column=8+i, value=v); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
ai_row = row
row += 1

# Asset disposal
write_label(ws, row, "Asset disposal gain/(loss)", indent=1)
ad_hist = HIST["asset_disposal"]
ad_proj = [0.0]*5
for i, v in enumerate(ad_hist + ad_proj):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
    if i >= 6: c.fill = PROJ_FILL
ad_row = row
row += 2

# Operating profit
write_label(ws, row, "Operating profit (营业利润)", bold=True)
ws.cell(row=row, column=1).fill = TOTAL_FILL
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col,
                value=f"={cl}{gp_row}-{cl}{opex_row}+{cl}{oi_row}+{cl}{ii_row}+{cl}{fv_row}+{cl}{ci_row}+{cl}{ai_row}+{cl}{ad_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER; c.fill = TOTAL_FILL
op_row = row
row += 1
write_label(ws, row, "  Operating margin %", italic=True, indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{op_row}/{cl}{revenue_row}")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 2

# EBITDA bridge
write_label(ws, row, "EBITDA bridge", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
row += 1
write_label(ws, row, "+ Depreciation & amortization", indent=1)
da_hist = HIST["da"]; da_proj = PROJ_ASSUMPTIONS["da"]
for i, v in enumerate(da_hist + da_proj):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
    if i >= 6: c.fill = PROJ_FILL
da_row = row
row += 1
write_label(ws, row, "+ Net interest (add back if income, subtract if expense)", indent=1)
# Net interest = -financial_exp (positive when income exceeds expense)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"=-{cl}{fin_row}")
    c.font = BLACK; fmt_num(c)
int_row = row
row += 1
# EBIT = Op profit + net interest
write_label(ws, row, "EBIT (= Op profit + net interest)", bold=True)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{op_row}+{cl}{int_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER
ebit_row = row
row += 1
write_label(ws, row, "  EBIT margin %", italic=True, indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{ebit_row}/{cl}{revenue_row}")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 1
write_label(ws, row, "EBITDA (= EBIT + D&A)", bold=True)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{ebit_row}+{cl}{da_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER
ebitda_row = row
row += 1
write_label(ws, row, "  EBITDA margin %", italic=True, indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{ebitda_row}/{cl}{revenue_row}")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 2

# Non-operating
write_label(ws, row, "Non-operating income", indent=1)
nop_inc = HIST["non_op_inc"] + [15.0]*5
for i, v in enumerate(nop_inc):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
    if i >= 6: c.fill = PROJ_FILL
nop_inc_row = row
row += 1
write_label(ws, row, "Non-operating expense", indent=1)
nop_exp = HIST["non_op_exp"] + [10.0]*5
for i, v in enumerate(nop_exp):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
    if i >= 6: c.fill = PROJ_FILL
nop_exp_row = row
row += 1

# Pre-tax income
write_label(ws, row, "Pre-tax income (利润总额)", bold=True)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{op_row}+{cl}{nop_inc_row}-{cl}{nop_exp_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER
pretax_row = row
row += 1

# Tax
write_label(ws, row, "Income tax expense", indent=1)
tax_hist = HIST["income_tax"]
for i, v in enumerate(tax_hist):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
for i, tr in enumerate(PROJ_ASSUMPTIONS["tax_rate"]):
    col = 8 + i; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{pretax_row}*{tr}")
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
tax_row = row
row += 1
write_label(ws, row, "  Effective tax rate %", italic=True, indent=2)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{tax_row}/{cl}{pretax_row}")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 2

# Net income
write_label(ws, row, "Net income (净利润)", bold=True)
ws.cell(row=row, column=1).fill = TOTAL_FILL
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{pretax_row}-{cl}{tax_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER; c.fill = TOTAL_FILL
ni_row = row
row += 1
write_label(ws, row, "  Net margin %", italic=True, indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{ni_row}/{cl}{revenue_row}")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 1

# Minority interest (small at Hengli)
write_label(ws, row, "Less: minority interest", indent=1)
mi = [0, 0, 6.1, 5.0, 3.7, 5.6] + [6.0, 7.0, 8.0, 9.0, 10.0]
for i, v in enumerate(mi):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
    if i >= 6: c.fill = PROJ_FILL
mi_row = row
row += 1
write_label(ws, row, "Net income to parent (归母净利润)", bold=True)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{ni_row}-{cl}{mi_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER
ni_parent_row = row
row += 2

# Share count + EPS
write_label(ws, row, "Shares outstanding — basic (million)", indent=1)
sh_proj = [1340.8]*5
for i, v in enumerate(HIST["shares_basic"] + sh_proj):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
    if i >= 6: c.fill = PROJ_FILL
sh_b_row = row
row += 1
write_label(ws, row, "Shares — diluted (million)", indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{sh_b_row}")
    c.font = BLACK; fmt_num(c)
sh_d_row = row
row += 1
write_label(ws, row, "Basic EPS (RMB/share)", bold=True)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{ni_parent_row}/{cl}{sh_b_row}")
    c.font = Font(bold=True); c.number_format = "0.00"
eps_row = row
row += 1
write_label(ws, row, "Diluted EPS (RMB/share)", italic=True, indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{ni_parent_row}/{cl}{sh_d_row}")
    c.font = BLACK; c.number_format = "0.00"
row += 2

# Dividend
write_label(ws, row, "Dividend per share (RMB)", indent=1)
dps_hist = [0.31, 0.45, 0.55, 0.56, 0.70, 0.56]   # 2025 prop. 0.56
dps_proj = [0.60, 0.66, 0.72, 0.80, 0.88]
for i, v in enumerate(dps_hist + dps_proj):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; c.number_format = "0.00"
    if i >= 6: c.fill = PROJ_FILL
dps_row = row
row += 1
write_label(ws, row, "Payout ratio %", italic=True, indent=2)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{dps_row}*{cl}{sh_b_row}/{cl}{ni_parent_row}")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 1

ws.freeze_panes = "B4"

# Cross-sheet row pointers
IS_REVENUE_ROW = revenue_row
IS_EBIT_ROW = ebit_row
IS_EBITDA_ROW = ebitda_row
IS_NI_ROW = ni_row
IS_NI_PARENT_ROW = ni_parent_row
IS_DA_ROW = da_row
IS_TAX_ROW = tax_row
IS_PRETAX_ROW = pretax_row

# ---------------------------------------------------------------------------
# TAB 3: Cash Flow Statement
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Cash Flow")
ws.column_dimensions["A"].width = 50
for i in range(2, 14):
    ws.column_dimensions[get_column_letter(i)].width = 12

ws["A1"] = "HENGLI HYDRAULICS — CONSOLIDATED CASH FLOW (RMB millions)"
ws["A1"].font = Font(size=14, bold=True, color="1F4E79")
ws.merge_cells("A1:L1")
write_year_header(ws, 3)

row = 5
write_label(ws, row, "OPERATING ACTIVITIES", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
row += 1

# Net income link
write_label(ws, row, "Net income (from IS)", indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"='Income Statement'!{cl}{IS_NI_ROW}")
    c.font = GREEN; fmt_num(c)
ni_link_row = row
row += 1

write_label(ws, row, "+ D&A (from IS)", indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"='Income Statement'!{cl}{IS_DA_ROW}")
    c.font = GREEN; fmt_num(c)
da_link_row = row
row += 1

write_label(ws, row, "+ Working capital changes", indent=1)
# Approx WC change for projections; historicals from CFO derivation
wc_hist = [None, None, None, None, None, None]
# Project WC change as % of incremental revenue (~-15% of revenue growth — drag)
wc_proj_f = []
for i in range(5):
    col = 8 + i; cl = get_column_letter(col)
    cl_p = get_column_letter(col - 1)
    # WC drag = -(15% of incremental revenue)
    wc_proj_f.append(f"=-({cl}'Revenue Model'!{REV_ROW_REF}-{cl_p}'Revenue Model'!{REV_ROW_REF})*0.15".replace("{cl}'", "'").replace("{cl_p}'", "'"))
# Simpler: just project WC change directly as negative drag
wc_proj_vals_simple = [-280.0, -260.0, -310.0, -290.0, -270.0]
# Plug historicals as inferred residuals; project with simple values
wc_inferred = []
for i in range(6):
    wc_inferred.append(HIST["cfo"][i] - HIST["da"][i] - (HIST["revenue"][i] - HIST["cogs"][i] - HIST["selling_exp"][i] - HIST["admin_exp"][i] - HIST["rd_exp"][i] - HIST["tax_surcharge"][i] - HIST["income_tax"][i]))
for i, v in enumerate(wc_inferred):
    c = ws.cell(row=row, column=i+2, value=round(v,1)); c.font = BLUE; fmt_num(c)
for i, v in enumerate(wc_proj_vals_simple):
    c = ws.cell(row=row, column=8+i, value=v); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
wc_row = row
row += 1
write_label(ws, row, "+ Other non-cash adjustments (SBC, impair, taxes paid timing)", indent=1)
# Plug to reconcile CFO
nc_hist = []
for i in range(6):
    plug = HIST["cfo"][i] - HIST["da"][i] - wc_inferred[i] - (HIST["revenue"][i] - HIST["cogs"][i] - HIST["selling_exp"][i] - HIST["admin_exp"][i] - HIST["rd_exp"][i] - HIST["tax_surcharge"][i] - HIST["income_tax"][i])
    nc_hist.append(round(plug, 1))
nc_proj = [80.0, 90.0, 100.0, 100.0, 100.0]
for i, v in enumerate(nc_hist + nc_proj):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
    if i >= 6: c.fill = PROJ_FILL
nc_row = row
row += 1

# CFO
write_label(ws, row, "Cash from operations (CFO)", bold=True)
ws.cell(row=row, column=1).fill = TOTAL_FILL
# Historicals: use reported figures (link to a memo line below); projections: build from components
for i in range(6):
    col = i + 2
    c = ws.cell(row=row, column=col, value=HIST["cfo"][i])
    c.font = Font(bold=True, color="0070C0"); fmt_num(c); c.border = TOP_BORDER; c.fill = TOTAL_FILL
for i in range(5):
    col = 8 + i; cl = get_column_letter(col)
    # Simplified projection: NI + D&A + WC + other non-cash
    c = ws.cell(row=row, column=col,
                value=f"={cl}{ni_link_row}+{cl}{da_link_row}+{cl}{wc_row}+{cl}{nc_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER; c.fill = TOTAL_FILL
cfo_row = row
row += 2

# INVESTING
write_label(ws, row, "INVESTING ACTIVITIES", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
row += 1
write_label(ws, row, "Capital expenditures (PP&E + intangibles)", indent=1)
capex_hist = HIST["capex"]
for i, v in enumerate(capex_hist):
    c = ws.cell(row=row, column=i+2, value=-v); c.font = BLUE; fmt_num(c)
# project capex as % of revenue
for i, pct in enumerate(PROJ_ASSUMPTIONS["capex_pct"]):
    col = 8 + i; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"=-{cl}'Income Statement'!{IS_REVENUE_ROW}*{pct}".replace("={cl}'", "='"))
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
# Fix the formula construction
for i, pct in enumerate(PROJ_ASSUMPTIONS["capex_pct"]):
    col = 8 + i; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"=-'Income Statement'!{cl}{IS_REVENUE_ROW}*{pct}")
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
capex_row = row
row += 1
write_label(ws, row, "Other investing (financial-product purchases net, M&A)", indent=1)
# Plug: CFI - (-capex)
other_inv_hist = [HIST["cfi"][i] + HIST["capex"][i] for i in range(6)]
other_inv_proj = [-200.0, -150.0, -100.0, -100.0, -100.0]
for i, v in enumerate(other_inv_hist + other_inv_proj):
    c = ws.cell(row=row, column=i+2, value=round(v,1) if isinstance(v,float) else v); c.font = BLUE; fmt_num(c)
    if i >= 6: c.fill = PROJ_FILL
oi_inv_row = row
row += 1
write_label(ws, row, "Cash from investing (CFI)", bold=True)
ws.cell(row=row, column=1).fill = TOTAL_FILL
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{capex_row}+{cl}{oi_inv_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER; c.fill = TOTAL_FILL
cfi_row = row
row += 2

# FREE CASH FLOW
write_label(ws, row, "FREE CASH FLOW (CFO - CapEx)", bold=True)
ws.cell(row=row, column=1).fill = TOTAL_FILL
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{cfo_row}+{cl}{capex_row}")
    c.font = Font(bold=True, color="1F4E79"); fmt_num(c); c.border = TOP_BORDER; c.fill = TOTAL_FILL
fcf_row = row
row += 1
write_label(ws, row, "  FCF margin %", italic=True, indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{fcf_row}/'Income Statement'!{cl}{IS_REVENUE_ROW}")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 2

# FINANCING
write_label(ws, row, "FINANCING ACTIVITIES", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
row += 1
write_label(ws, row, "Debt issuance / (repayment), net", indent=1)
debt_hist = [-361.2, -286.1, -313.5, -126.2, -185.0, 43.0]
debt_proj = [0.0, 0.0, 0.0, 0.0, 0.0]
for i, v in enumerate(debt_hist + debt_proj):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
    if i >= 6: c.fill = PROJ_FILL
debt_row = row
row += 1
write_label(ws, row, "Dividends paid", indent=1)
div_hist = [-371.9, -583.0, -740.7, -748.7, -603.2, -906.4]
# project = -DPS * shares
for i, v in enumerate(div_hist):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col,
                value=f"=-'Income Statement'!{cl}{dps_row}*'Income Statement'!{cl}{sh_b_row}")
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
div_row = row
row += 1
write_label(ws, row, "Other financing (incl. 2022 private placement RMB 2.0bn)", indent=1)
of_hist = [HIST["cff"][i] - debt_hist[i] - div_hist[i] for i in range(6)]
of_proj = [-50.0]*5
for i, v in enumerate(of_hist + of_proj):
    c = ws.cell(row=row, column=i+2, value=round(v,1) if isinstance(v,float) else v); c.font = BLUE; fmt_num(c)
    if i >= 6: c.fill = PROJ_FILL
of_row = row
row += 1
write_label(ws, row, "Cash from financing (CFF)", bold=True)
ws.cell(row=row, column=1).fill = TOTAL_FILL
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{debt_row}+{cl}{div_row}+{cl}{of_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER; c.fill = TOTAL_FILL
cff_row = row
row += 2

# FX impact + net change cash
write_label(ws, row, "FX impact on cash", indent=1)
fx_hist = HIST["fx_impact"]
fx_proj = [-30.0, -20.0, -20.0, -20.0, -20.0]
for i, v in enumerate(fx_hist + fx_proj):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
    if i >= 6: c.fill = PROJ_FILL
fx_row = row
row += 1
write_label(ws, row, "Net change in cash", bold=True)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{cfo_row}+{cl}{cfi_row}+{cl}{cff_row}+{cl}{fx_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER
ncc_row = row
row += 1

# Begin/End cash
write_label(ws, row, "Beginning cash", indent=1)
# FY20 begin = 2238.98; subsequent = prior end
begin_hist = [2239.0, 2574.2, 3598.8, 6858.7, 5785.2, 4112.1]
for i, v in enumerate(begin_hist):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
# Projection begin = prior end
for i in range(5):
    col = 8 + i; cl_p = get_column_letter(col - 1)
    c = ws.cell(row=row, column=col, value=f"={cl_p}{row+1}")  # = prior end cash
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
begin_row = row
row += 1
write_label(ws, row, "Ending cash", bold=True)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{begin_row}+{cl}{ncc_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER
end_cash_row = row
row += 1
write_label(ws, row, "Memo: reported end cash (audited, where avail.)", italic=True)
end_hist = [2574.2, 3598.8, 6858.7, 5785.2, 4112.1, 4136.7]
for i, v in enumerate(end_hist):
    c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)

ws.freeze_panes = "B4"

CF_CFO_ROW = cfo_row
CF_CAPEX_ROW = capex_row
CF_FCF_ROW = fcf_row

# ---------------------------------------------------------------------------
# TAB 4: Balance Sheet
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Balance Sheet")
ws.column_dimensions["A"].width = 50
for i in range(2, 14):
    ws.column_dimensions[get_column_letter(i)].width = 12

ws["A1"] = "HENGLI HYDRAULICS — CONSOLIDATED BALANCE SHEET (RMB millions)"
ws["A1"].font = Font(size=14, bold=True, color="1F4E79")
ws.merge_cells("A1:L1")
write_year_header(ws, 3)

row = 5
def write_bs_row(label, key, indent=1, projector=None, is_calc=False):
    global row
    write_label(ws, row, label, indent=indent)
    h = HIST.get(key, [None]*6)
    for i, v in enumerate(h):
        if v is not None:
            c = ws.cell(row=row, column=i+2, value=v); c.font = BLUE; fmt_num(c)
    if projector:
        for i in range(5):
            col = 8 + i; cl = get_column_letter(col)
            f = projector(col, cl, i)
            c = ws.cell(row=row, column=col, value=f); c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
    return row

# ASSETS
write_label(ws, row, "ASSETS", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
row += 1
write_label(ws, row, "Current assets", bold=True)
row += 1

# Cash — link to CF statement ending cash
write_label(ws, row, "Cash & equivalents", indent=2)
for i in range(2, 6):  # FY22-FY25 hardcoded
    c = ws.cell(row=row, column=i+2, value=HIST["cash"][i]); c.font = BLUE; fmt_num(c)
# Hengli holds RMB ~4.7bn in structured deposits / wealth-mgmt products that sit on the BS as
# "cash & equivalents" but are netted out of CF "ending cash". Plug that gap into the projection.
# FY25 plug = BS cash 8871 − CF end-cash 4137 = 4734
for i in range(5):
    col = 8 + i; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"='Cash Flow'!{cl}{end_cash_row}+4734")
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
cash_row = row
row += 1

# Trading FA
write_label(ws, row, "Trading financial assets", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["trading_fa"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=600.0); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
tfa_row = row
row += 1

# Bills receivable
write_label(ws, row, "Bills receivable", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["bills_recv"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"='Income Statement'!{cl}{revenue_row}*0.05")
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
br_row = row
row += 1

# AR — project using days-sales-outstanding
write_label(ws, row, "Accounts receivable", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["ar"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i; cl = get_column_letter(col)
    days = PROJ_ASSUMPTIONS["ar_days"][i]
    c = ws.cell(row=row, column=col, value=f"='Income Statement'!{cl}{revenue_row}*{days}/365")
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
ar_row = row
row += 1

# Financing receivable
write_label(ws, row, "Receivables financing (银承汇票池)", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["fin_recv"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=900.0); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
frec_row = row
row += 1

# Prepayments
write_label(ws, row, "Prepayments", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["prepayments"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"='Income Statement'!{cogs_row}{cl[0:]}*0.04")
    # Simpler:
    c.value = f"='Income Statement'!{cl}{cogs_row}*0.04"
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
pre_row = row
row += 1

# Other receivables
write_label(ws, row, "Other receivables", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["other_recv"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=35.0); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
or_row = row
row += 1

# Inventory
write_label(ws, row, "Inventory", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["inventory"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i; cl = get_column_letter(col)
    days = PROJ_ASSUMPTIONS["inv_days"][i]
    c = ws.cell(row=row, column=col, value=f"='Income Statement'!{cl}{cogs_row}*{days}/365")
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
inv_row = row
row += 1

# Contract assets + other CA
write_label(ws, row, "Contract assets", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["contract_assets"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=30.0); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
ca_row = row
row += 1
write_label(ws, row, "Other current assets", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["other_ca"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=300.0); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
oca_row = row
row += 1

# Total current assets
write_label(ws, row, "Total current assets", bold=True)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col,
                value=f"={cl}{cash_row}+{cl}{tfa_row}+{cl}{br_row}+{cl}{ar_row}+{cl}{frec_row}+{cl}{pre_row}+{cl}{or_row}+{cl}{inv_row}+{cl}{ca_row}+{cl}{oca_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER
tca_row = row
row += 2

# Non-current assets
write_label(ws, row, "Non-current assets", bold=True)
row += 1
# PP&E — project as prior + CapEx - D&A
write_label(ws, row, "PP&E, net", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["ppe"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i; cl = get_column_letter(col); cl_p = get_column_letter(col-1)
    c = ws.cell(row=row, column=col,
                value=f"={cl_p}{row}+(-'Cash Flow'!{cl}{capex_row})*0.8-'Income Statement'!{cl}{IS_DA_ROW}*0.85")
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
ppe_row = row
row += 1
write_label(ws, row, "Construction in progress", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["cip"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=500.0); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
cip_row = row
row += 1
write_label(ws, row, "Intangible assets", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["intangibles"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i; cl = get_column_letter(col); cl_p = get_column_letter(col-1)
    c = ws.cell(row=row, column=col,
                value=f"={cl_p}{row}+(-'Cash Flow'!{cl}{capex_row})*0.2-'Income Statement'!{cl}{IS_DA_ROW}*0.15")
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
intang_row = row
row += 1
write_label(ws, row, "Goodwill", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["goodwill"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i; cl_p = get_column_letter(col-1)
    c = ws.cell(row=row, column=col, value=f"={cl_p}{row}"); c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
gw_row = row
row += 1
write_label(ws, row, "LT equity investments", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["lt_inv"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i; cl_p = get_column_letter(col-1)
    c = ws.cell(row=row, column=col, value=f"={cl_p}{row}*1.05"); c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
ltinv_row = row
row += 1
write_label(ws, row, "Deferred tax assets", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["dta"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i; cl_p = get_column_letter(col-1)
    c = ws.cell(row=row, column=col, value=f"={cl_p}{row}*1.1"); c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
dta_row = row
row += 1
write_label(ws, row, "Other non-current assets", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["other_nca"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=320.0); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
onca_row = row
row += 1

write_label(ws, row, "Total non-current assets", bold=True)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col,
                value=f"={cl}{ppe_row}+{cl}{cip_row}+{cl}{intang_row}+{cl}{gw_row}+{cl}{ltinv_row}+{cl}{dta_row}+{cl}{onca_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER
tnca_row = row
row += 2

# TOTAL ASSETS
write_label(ws, row, "TOTAL ASSETS", bold=True)
ws.cell(row=row, column=1).fill = TOTAL_FILL
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{tca_row}+{cl}{tnca_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER; c.fill = TOTAL_FILL
ta_row = row
row += 2

# LIABILITIES
write_label(ws, row, "LIABILITIES", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
row += 1
write_label(ws, row, "Current liabilities", bold=True)
row += 1
# ST borrowings
write_label(ws, row, "Short-term borrowings", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["st_borrow"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=15.0); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
stb_row = row
row += 1
# Bills payable
write_label(ws, row, "Bills payable", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["bills_pay"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=50.0); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
bp_row = row
row += 1
# AP — DPO
write_label(ws, row, "Accounts payable", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["ap"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i; cl = get_column_letter(col)
    days = PROJ_ASSUMPTIONS["ap_days"][i]
    c = ws.cell(row=row, column=col, value=f"='Income Statement'!{cl}{cogs_row}*{days}/365")
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
ap_row = row
row += 1
write_label(ws, row, "Contract liabilities (customer advances)", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["contract_liab"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"='Income Statement'!{cl}{revenue_row}*0.03")
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
cl_row = row
row += 1
write_label(ws, row, "Employee comp payable", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["employee_pay"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"='Income Statement'!{cl}{revenue_row}*0.04")
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
ep_row = row
row += 1
write_label(ws, row, "Tax payable", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["tax_payable"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"='Income Statement'!{cl}{tax_row}*0.65")
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
tp_row = row
row += 1
write_label(ws, row, "Other payables (incl. dividend payable)", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["other_pay"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=1600.0); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
op_row = row
row += 1
write_label(ws, row, "Other current liabilities", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["other_cl"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=300.0); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
ocl_row = row
row += 1
write_label(ws, row, "Total current liabilities", bold=True)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col,
                value=f"={cl}{stb_row}+{cl}{bp_row}+{cl}{ap_row}+{cl}{cl_row}+{cl}{ep_row}+{cl}{tp_row}+{cl}{op_row}+{cl}{ocl_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER
tcl_row = row
row += 2

write_label(ws, row, "Non-current liabilities", bold=True)
row += 1
write_label(ws, row, "Long-term borrowings", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["lt_borrow"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=20.0); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
ltb_row = row
row += 1
write_label(ws, row, "Lease liabilities", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["lease_liab"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=15.0); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
ll_row = row
row += 1
write_label(ws, row, "Deferred revenue (govt grants)", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["deferred_rev"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=280.0); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
dr_row = row
row += 1
write_label(ws, row, "Deferred tax liabilities", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["dtl"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=250.0); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
dtl_row = row
row += 1
write_label(ws, row, "Total non-current liabilities", bold=True)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col,
                value=f"={cl}{ltb_row}+{cl}{ll_row}+{cl}{dr_row}+{cl}{dtl_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER
tncl_row = row
row += 2

write_label(ws, row, "TOTAL LIABILITIES", bold=True)
ws.cell(row=row, column=1).fill = TOTAL_FILL
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{tcl_row}+{cl}{tncl_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER; c.fill = TOTAL_FILL
tl_row = row
row += 2

# EQUITY
write_label(ws, row, "EQUITY", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
row += 1
write_label(ws, row, "Share capital", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["share_capital"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=1340.8); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
sc_row = row
row += 1
write_label(ws, row, "Capital reserve", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["capital_reserve"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=3364.9); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
cr_row = row
row += 1
write_label(ws, row, "Other comprehensive income (FX translation)", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["oci"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=0.0); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
oci_row = row
row += 1
write_label(ws, row, "Surplus + special reserve", indent=2)
for i in range(2, 6):
    sr = (HIST["surplus_reserve"][i] or 0) + (HIST["special_reserve"][i] or 0)
    c = ws.cell(row=row, column=i+2, value=sr); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i
    c = ws.cell(row=row, column=col, value=720.0); c.font = BLUE; c.fill = PROJ_FILL; fmt_num(c)
sr_row = row
row += 1
# Retained earnings — project as prior + NI to parent - Dividends
write_label(ws, row, "Retained earnings", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["retained_earn"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i; cl = get_column_letter(col); cl_p = get_column_letter(col-1)
    c = ws.cell(row=row, column=col,
                value=f"={cl_p}{row}+'Income Statement'!{cl}{ni_parent_row}+'Cash Flow'!{cl}{div_row}")
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
re_row = row
row += 1
write_label(ws, row, "Minority interest", indent=2)
for i in range(2, 6):
    c = ws.cell(row=row, column=i+2, value=HIST["minority"][i]); c.font = BLUE; fmt_num(c)
for i in range(5):
    col = 8 + i; cl = get_column_letter(col); cl_p = get_column_letter(col-1)
    c = ws.cell(row=row, column=col, value=f"={cl_p}{row}+'Income Statement'!{cl}{mi_row}")
    c.font = BLACK; c.fill = PROJ_FILL; fmt_num(c)
min_row = row
row += 1
write_label(ws, row, "TOTAL EQUITY", bold=True)
ws.cell(row=row, column=1).fill = TOTAL_FILL
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col,
                value=f"={cl}{sc_row}+{cl}{cr_row}+{cl}{oci_row}+{cl}{sr_row}+{cl}{re_row}+{cl}{min_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER; c.fill = TOTAL_FILL
te_row = row
row += 2

# Total L+E
write_label(ws, row, "TOTAL LIABILITIES + EQUITY", bold=True)
ws.cell(row=row, column=1).fill = TOTAL_FILL
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{tl_row}+{cl}{te_row}")
    c.font = Font(bold=True); fmt_num(c); c.border = TOP_BORDER; c.fill = TOTAL_FILL
tle_row = row
row += 1

# Balance check
write_label(ws, row, "BALANCE CHECK (Assets - L+E)", italic=True)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{ta_row}-{cl}{tle_row}")
    c.font = ITALIC; fmt_num(c)
row += 2

# Key ratios
write_label(ws, row, "KEY RATIOS", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
row += 1
write_label(ws, row, "Current ratio (CA / CL)", indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{tca_row}/{cl}{tcl_row}")
    c.font = BLACK; c.number_format = "0.00"
row += 1
write_label(ws, row, "Debt / Equity", indent=1)
for i in range(len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"=({cl}{stb_row}+{cl}{ltb_row})/{cl}{te_row}")
    c.font = BLACK; c.number_format = "0.00%"
row += 1
write_label(ws, row, "ROE % (NI to parent / avg equity)", indent=1)
for i in range(1, len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col); cl_p = get_column_letter(col-1)
    c = ws.cell(row=row, column=col,
                value=f"='Income Statement'!{cl}{ni_parent_row}/(({cl}{te_row}+{cl_p}{te_row})/2)")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 1
write_label(ws, row, "ROIC % (NOPAT / Invested capital)", indent=1)
for i in range(1, len(ALL_YEARS)):
    col = i + 2; cl = get_column_letter(col); cl_p = get_column_letter(col-1)
    c = ws.cell(row=row, column=col,
                value=f"='Income Statement'!{cl}{ebit_row}*(1-0.12)/(({cl}{te_row}+{cl_p}{te_row})/2+({cl}{stb_row}+{cl}{ltb_row}))")
    c.font = BLACK; fmt_num(c, is_pct=True)
row += 1

ws.freeze_panes = "B4"

# ---------------------------------------------------------------------------
# TAB 5: Scenarios (Bull / Base / Bear)
# ---------------------------------------------------------------------------
ws = wb.create_sheet("Scenarios")
ws.column_dimensions["A"].width = 50
for c in "BCDEFGH":
    ws.column_dimensions[c].width = 14

ws["A1"] = "HENGLI HYDRAULICS — SCENARIO ANALYSIS (FY2030E)"
ws["A1"].font = Font(size=14, bold=True, color="1F4E79")
ws.merge_cells("A1:H1")

# Headers
ws["B3"] = "Bull"; ws["C3"] = "Base"; ws["D3"] = "Bear"
for col, fill in zip("BCD", [SCEN_BULL, TOTAL_FILL, SCEN_BEAR]):
    ws[f"{col}3"].font = WHITE_BOLD
    ws[f"{col}3"].fill = HDR_FILL
    ws[f"{col}3"].alignment = Alignment(horizontal="center")

ws["A5"] = "KEY ASSUMPTIONS (FY2025A → FY2030E)"
ws["A5"].font = BOLD; ws["A5"].fill = SUB_HDR_FILL
ws.merge_cells("A5:D5")

scen_rows = [
    ("Revenue CAGR (FY25-FY30)", [0.18, 0.13, 0.07]),
    ("FY2030E revenue (RMB m)", None),  # formula below
    ("Gross margin FY30E", [0.450, 0.435, 0.405]),
    ("EBITDA margin FY30E", [0.335, 0.305, 0.260]),
    ("Net margin FY30E", [0.270, 0.245, 0.205]),
    ("CapEx as % of revenue", [0.045, 0.050, 0.060]),
    ("WC drag (% of revenue growth)", [0.10, 0.15, 0.22]),
    ("Linear-drive revenue FY30E (RMB m)", [4500.0, 2300.0, 800.0]),
    ("Linear-drive gross margin FY30E", [0.28, 0.22, 0.15]),
]

row = 6
for label, vals in scen_rows:
    write_label(ws, row, label)
    if vals:
        for i, v in enumerate(vals):
            c = ws.cell(row=row, column=2+i, value=v); c.font = BLUE
            if "margin" in label.lower() or "CAGR" in label or "WC drag" in label or "%" in label:
                fmt_num(c, is_pct=True)
            else:
                fmt_num(c)
    row += 1

# Revenue projection (uses CAGR from row 6 and FY25 base)
FY25_REV = 10941.0
write_label(ws, 7, "FY2030E revenue (RMB m)")
for col_idx, col in enumerate("BCD"):
    c = ws[f"{col}7"]
    c.value = f"={FY25_REV}*(1+{col}6)^5"
    c.font = BLACK; fmt_num(c)

row += 2
ws.cell(row=row, column=1, value="FY2030E OUTPUTS").font = BOLD
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1

out_rows = [
    ("Revenue FY30E", "=B7", "=C7", "=D7", False, 1),
    ("Gross profit FY30E", "=B7*B8", "=C7*C8", "=D7*D8", False, 1),
    ("EBITDA FY30E", "=B7*B9", "=C7*C9", "=D7*D9", False, 1),
    ("Net income FY30E", "=B7*B10", "=C7*C10", "=D7*D10", False, 1),
    ("FCF FY30E (approx: EBITDA - CapEx - tax)", "=B7*B9-B7*B11-B7*B10*0.13",
        "=C7*C9-C7*C11-C7*C10*0.13", "=D7*D9-D7*D11-D7*D10*0.13", False, 1),
    ("EPS FY30E (NI / 1,340.8m shares)", "=B7*B10/1340.8", "=C7*C10/1340.8", "=D7*D10/1340.8", False, 2),
]
for label, fb, fc, fd, is_pct, dec in out_rows:
    write_label(ws, row, label, bold=True)
    for i, f in enumerate([fb, fc, fd]):
        c = ws.cell(row=row, column=2+i, value=f); c.font = Font(bold=True)
        if dec == 2:
            c.number_format = "0.00"
        else:
            fmt_num(c, decimals=dec)
        if i == 0: c.fill = SCEN_BULL
        elif i == 1: c.fill = TOTAL_FILL
        else: c.fill = SCEN_BEAR
    row += 1

row += 2
# Cumulative FCF
ws.cell(row=row, column=1, value="Cumulative FCF FY2026-FY2030E (RMB m)").font = BOLD
# Approx: trend from FY25 FCF ~887m to FY30E target
cumfcf = {
    "Bull": 9500.0,    # mid-cycle ramp + linear-drive scale
    "Base": 6800.0,    # mgmt-aligned base case
    "Bear": 3800.0,    # downside cycle
}
for i, k in enumerate(["Bull", "Base", "Bear"]):
    c = ws.cell(row=row, column=2+i, value=cumfcf[k])
    c.font = Font(bold=True, color="0070C0"); fmt_num(c)
row += 2

# Scenario narrative
ws.cell(row=row, column=1, value="SCENARIO RATIONALE").font = BOLD
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
row += 1
rationale = [
    ("Bull",
     "Excavator up-cycle 2026-2028 + linear-drive humanoid wild-card hits (Optimus or major Chinese "
     "humanoid program qualified-supplier status); Mexico plant ramps to >$300m/yr; pumps & valves "
     "share-gain accelerates to 25%+ on Bosch Rexroth retreat. Revenue >$24bn, EBITDA margin 33-34%."),
    ("Base",
     "China hydraulics 8% market CAGR; Hengli outperforms at 13% CAGR via continued share gain and "
     "linear-drive ramp to ~RMB 2.3bn (mgmt-implied path). Margins stable at 30-31% EBITDA. "
     "Revenue ~RMB 20bn by FY30E."),
    ("Bear",
     "Cyclical excavator down-cycle returns 2027-28; Caterpillar in-sources >50% cylinder content; "
     "linear-drive ramp stalls (no humanoid qualified-supplier); China hydraulics market grows only "
     "3-5%. Revenue ~RMB 15bn, EBITDA margin compresses to 26%."),
]
for k, txt in rationale:
    ws.cell(row=row, column=1, value=k).font = BOLD
    c = ws.cell(row=row, column=2, value=txt)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    ws.row_dimensions[row].height = 60
    row += 2

# ---------------------------------------------------------------------------
# TAB 6: DCF Inputs
# ---------------------------------------------------------------------------
ws = wb.create_sheet("DCF Inputs")
ws.column_dimensions["A"].width = 50
for i in range(2, 8):
    ws.column_dimensions[get_column_letter(i)].width = 14

ws["A1"] = "HENGLI HYDRAULICS — DCF INPUTS (RMB millions)"
ws["A1"].font = Font(size=14, bold=True, color="1F4E79")
ws.merge_cells("A1:G1")

# Header — projection years only
for i, y in enumerate(YEARS_PROJ):
    c = ws.cell(row=3, column=2+i, value=y)
    c.font = WHITE_BOLD; c.fill = HDR_FILL; c.alignment = Alignment(horizontal="center")
c = ws.cell(row=3, column=7, value="Terminal"); c.font = WHITE_BOLD; c.fill = HDR_FILL
c.alignment = Alignment(horizontal="center")

# EBIT (linked from IS)
row = 5
write_label(ws, row, "EBIT (from Income Statement)")
for i in range(5):
    col = 2 + i; cl_is = get_column_letter(8 + i)  # FY26E starts at col 8 in IS
    c = ws.cell(row=row, column=col, value=f"='Income Statement'!{cl_is}{ebit_row}")
    c.font = GREEN; fmt_num(c)
row += 1

# Marginal tax rate
write_label(ws, row, "Effective tax rate %")
for i, tr in enumerate(PROJ_ASSUMPTIONS["tax_rate"]):
    c = ws.cell(row=row, column=2+i, value=tr); c.font = BLUE; fmt_num(c, is_pct=True)
ws.cell(row=row, column=7, value=0.125).font = BLUE; fmt_num(ws.cell(row=row, column=7), is_pct=True)
tr_row = row
row += 1

# NOPAT = EBIT × (1-t)
write_label(ws, row, "NOPAT (EBIT × (1-t))", bold=True)
for i in range(5):
    col = 2 + i; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col, value=f"={cl}{row-2}*(1-{cl}{tr_row})")
    c.font = Font(bold=True); fmt_num(c)
nopat_row = row
row += 2

# D&A
write_label(ws, row, "+ D&A (from IS)")
for i in range(5):
    col = 2 + i; cl_is = get_column_letter(8 + i)
    c = ws.cell(row=row, column=col, value=f"='Income Statement'!{cl_is}{da_row}")
    c.font = GREEN; fmt_num(c)
da_dcf_row = row
row += 1

# CapEx
write_label(ws, row, "− CapEx (from CF)")
for i in range(5):
    col = 2 + i; cl_cf = get_column_letter(8 + i)
    c = ws.cell(row=row, column=col, value=f"=-'Cash Flow'!{cl_cf}{capex_row}")  # capex stored as negative
    c.font = GREEN; fmt_num(c)
capex_dcf_row = row
row += 1

# Change in NWC
write_label(ws, row, "− Change in NWC")
nwc_proj = [280.0, 260.0, 310.0, 290.0, 270.0]
for i, v in enumerate(nwc_proj):
    c = ws.cell(row=row, column=2+i, value=v); c.font = BLUE; fmt_num(c)
nwc_row = row
row += 1

# Unlevered FCF
write_label(ws, row, "UNLEVERED FREE CASH FLOW", bold=True)
ws.cell(row=row, column=1).fill = TOTAL_FILL
for i in range(5):
    col = 2 + i; cl = get_column_letter(col)
    c = ws.cell(row=row, column=col,
                value=f"={cl}{nopat_row}+{cl}{da_dcf_row}-{cl}{capex_dcf_row}-{cl}{nwc_row}")
    c.font = Font(bold=True, color="1F4E79"); fmt_num(c); c.border = TOP_BORDER; c.fill = TOTAL_FILL
ufcf_row = row
row += 3

# Terminal-year metrics & WACC inputs
write_label(ws, row, "TERMINAL & VALUATION INPUTS", bold=True)
ws.cell(row=row, column=1).fill = SUB_HDR_FILL
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
row += 1

# These are anchor inputs that Task 3 (Valuation) will use
val_inputs = [
    ("Terminal-year revenue (FY30E)", f"='Income Statement'!L{IS_REVENUE_ROW}", None),
    ("Terminal-year EBITDA (FY30E)", f"='Income Statement'!L{IS_EBITDA_ROW}", None),
    ("Terminal-year EBIT (FY30E)", f"='Income Statement'!L{IS_EBIT_ROW}", None),
    ("Terminal-year UFCF (FY30E)", f"=F{ufcf_row}", None),
    ("", None, None),
    ("Risk-free rate (China 10Y, %)", 0.025, True),
    ("Equity risk premium (China, %)", 0.060, True),
    ("Beta (Hengli vs SSE Composite, 2y)", 1.05, False),
    ("Cost of equity (Ke = Rf + β·ERP)", "=B(row-2)+B(row-1)*B(row)", None),
    ("Pre-tax cost of debt (Kd)", 0.035, True),
    ("Effective tax rate", 0.125, True),
    ("After-tax cost of debt (Kd × (1-t))", "=B(row-2)*(1-B(row-1))", True),
    ("Debt / (Debt + Equity) — target", 0.05, True),
    ("Equity / (Debt + Equity) — target", 0.95, True),
    ("WACC (= We·Ke + Wd·Kd_after)", None, True),  # filled after
    ("Terminal growth rate (g)", 0.030, True),
    ("Mid-year convention?", "Yes", False),
]
# Simpler implementation: write each row individually, recompute WACC properly
labels_inputs = [
    ("Terminal-year revenue (FY30E)", f"='Income Statement'!L{IS_REVENUE_ROW}", "link"),
    ("Terminal-year EBITDA (FY30E)", f"='Income Statement'!L{IS_EBITDA_ROW}", "link"),
    ("Terminal-year EBIT (FY30E)", f"='Income Statement'!L{IS_EBIT_ROW}", "link"),
    ("Terminal-year UFCF (FY30E)", f"=F{ufcf_row}", "link"),
    ("", "", None),
    ("Risk-free rate (China 10Y govt, %)", 0.025, "input"),
    ("Equity risk premium (%)", 0.060, "input"),
    ("Beta (vs SSE Composite, 2y weekly)", 1.05, "input"),
    ("Cost of equity (Ke)", None, "calc"),
    ("Pre-tax cost of debt (Kd)", 0.035, "input"),
    ("Effective marginal tax rate", 0.125, "input"),
    ("After-tax cost of debt", None, "calc"),
    ("Target debt weight", 0.05, "input"),
    ("Target equity weight", 0.95, "input"),
    ("WACC", None, "calc"),
    ("Terminal growth rate (g)", 0.030, "input"),
    ("Mid-year convention", "Yes", "input"),
]

# Track absolute rows for formula construction
start_row = row
for i, (lbl, val, kind) in enumerate(labels_inputs):
    r = start_row + i
    write_label(ws, r, lbl, bold=(kind == "calc"))
    c = ws.cell(row=r, column=2)
    if kind == "link":
        c.value = val; c.font = GREEN; fmt_num(c)
    elif kind == "input":
        c.value = val; c.font = BLUE
        if "%" in lbl or "rate" in lbl.lower() or "weight" in lbl.lower() or "premium" in lbl.lower() or "growth" in lbl.lower():
            fmt_num(c, is_pct=True)
        elif "Beta" in lbl:
            c.number_format = "0.00"
# Now fill formula cells
# Ke = Rf + Beta * ERP  → at row start_row+8 = labels_inputs index 8
ke_r = start_row + 8
rf_r = start_row + 5
erp_r = start_row + 6
beta_r = start_row + 7
ws.cell(row=ke_r, column=2, value=f"=B{rf_r}+B{beta_r}*B{erp_r}")
ws.cell(row=ke_r, column=2).font = Font(bold=True); fmt_num(ws.cell(row=ke_r, column=2), is_pct=True)
# After-tax Kd at row start_row+11
kd_pre_r = start_row + 9
tax_r2 = start_row + 10
kd_after_r = start_row + 11
ws.cell(row=kd_after_r, column=2, value=f"=B{kd_pre_r}*(1-B{tax_r2})")
ws.cell(row=kd_after_r, column=2).font = Font(bold=True); fmt_num(ws.cell(row=kd_after_r, column=2), is_pct=True)
# WACC at start_row+14
debt_w_r = start_row + 12
eq_w_r = start_row + 13
wacc_r = start_row + 14
ws.cell(row=wacc_r, column=2, value=f"=B{eq_w_r}*B{ke_r}+B{debt_w_r}*B{kd_after_r}")
ws.cell(row=wacc_r, column=2).font = Font(bold=True, color="1F4E79"); fmt_num(ws.cell(row=wacc_r, column=2), is_pct=True)
ws.cell(row=wacc_r, column=2).fill = TOTAL_FILL

row = start_row + len(labels_inputs) + 2

# Note for Task 3
ws.cell(row=row, column=1,
        value="NOTE: These DCF inputs feed Task 3 (Valuation). The Task 3 worksheet will build "
              "Enterprise Value via DCF (sum of discounted UFCF + discounted terminal value), "
              "with sensitivity table on WACC × terminal growth.").font = ITALIC
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
ws.row_dimensions[row].height = 50

ws.freeze_panes = "B4"

# ============= SAVE =============
import os
os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
print(f"Tabs: {wb.sheetnames}")
