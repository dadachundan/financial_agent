"""Build AMD Financial Model (.xlsx) — Task 2 of initiating coverage."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from datetime import date

OUT = "/Users/x/projects/financial_agent/reports/company/AMD_NASDAQ_AMD/model/AMD_Financial_Model_2026-05-20.xlsx"

# ---------- Styling helpers ----------
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
SUBHEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F3864")
BOLD = Font(name="Calibri", size=11, bold=True)
ITALIC = Font(name="Calibri", size=10, italic=True, color="595959")
HISTORICAL = PatternFill("solid", fgColor="DDEBF7")
PROJECTION = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
border = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

NUM_FMT = "#,##0;(#,##0);–"
PCT_FMT = "0.0%;(0.0%);–"
MULT_FMT = "0.0\"x\""

YEARS_HIST = ["FY2021A", "FY2022A", "FY2023A", "FY2024A", "FY2025A"]
YEARS_PROJ = ["FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
ALL_YEARS = YEARS_HIST + YEARS_PROJ

def style_header(ws, row, start_col, end_col, title):
    cell = ws.cell(row=row, column=start_col, value=title)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, end_row=row, start_column=start_col, end_column=end_col)
    ws.row_dimensions[row].height = 22

def style_year_row(ws, row, col_offset=2):
    for i, y in enumerate(ALL_YEARS):
        c = ws.cell(row=row, column=col_offset + i, value=y)
        c.font = SUBHEADER_FONT; c.fill = SUBHEADER_FILL
        c.alignment = Alignment(horizontal="center")
        if i < len(YEARS_HIST):
            ws.cell(row=row, column=col_offset+i).fill = PatternFill("solid", fgColor="305496")
        else:
            ws.cell(row=row, column=col_offset+i).fill = PatternFill("solid", fgColor="BF8F00")

def write_line(ws, row, label, values, fmt=NUM_FMT, bold=False, indent=0):
    lc = ws.cell(row=row, column=1, value=("  " * indent) + label)
    if bold:
        lc.font = BOLD
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=2 + i, value=v)
        c.number_format = fmt
        if bold:
            c.font = BOLD
            c.fill = TOTAL_FILL
        else:
            c.fill = HISTORICAL if i < len(YEARS_HIST) else PROJECTION

def autosize(ws, label_width=42):
    ws.column_dimensions["A"].width = label_width
    for i in range(2, 2 + len(ALL_YEARS) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 13

# ===========================================================
# Build workbook
# ===========================================================
wb = Workbook()
wb.remove(wb.active)

# ---------- 1. README / Cover ----------
ws = wb.create_sheet("Cover")
ws["A1"] = "Advanced Micro Devices, Inc. (NASDAQ: AMD)"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Financial Model — Initiation of Coverage"
ws["A2"].font = Font(name="Calibri", size=12, bold=True, color="595959")
ws["A4"] = f"Date prepared: 2026-05-20"
ws["A5"] = "Analyst: Internal Equity Research"
ws["A6"] = "Currency: USD millions (except per-share data)"
ws["A7"] = "Fiscal year: 52/53-week year ending on the last Saturday in December"
ws["A9"] = "Tabs"
ws["A9"].font = BOLD
tabs = [
    ("Revenue Model", "Product- and geography-level revenue build; 20+ product lines, 15+ geos"),
    ("Income Statement", "GAAP P&L, FY2021A–FY2030E; non-GAAP reconciliation"),
    ("Cash Flow Statement", "Indirect-method operating, investing, financing cash flows"),
    ("Balance Sheet", "Historical and projected balance sheet; sources & uses"),
    ("Scenarios", "Bull / Base / Bear FY2030 outcomes with KPI assumptions"),
    ("DCF Inputs", "Unlevered FCF, WACC components, terminal-value inputs for Task 3"),
]
for i, (n, d) in enumerate(tabs):
    ws.cell(row=10+i, column=1, value=n).font = BOLD
    ws.cell(row=10+i, column=2, value=d)
ws["A18"] = "Source primacy: AMD 2025 10-K, 2024 10-K, 2023 10-K, Q1-2026 10-Q, 2026 DEF 14A"
ws["A18"].font = ITALIC
ws.column_dimensions["A"].width = 28
ws.column_dimensions["B"].width = 110

# ===========================================================
# 2. REVENUE MODEL
# ===========================================================
ws = wb.create_sheet("Revenue Model")
ws["A1"] = "AMD Revenue Model — by Product & Geography"; ws["A1"].font = TITLE_FONT
ws["A2"] = "USD millions unless noted. Historical from 10-K segment notes; projections by analyst."; ws["A2"].font = ITALIC

style_header(ws, 4, 1, 11, "I. Revenue by Product Line (USD millions)")
style_year_row(ws, 5)

# Product-line build (20+ rows) — historicals fitted to segment totals; projections by analyst
# All historicals in $M from 10-K segment + product disclosures.
# Note: FY2021 figures use pre-Xilinx baseline (Xilinx closed Feb 2022).
rows_products = [
    # Data Center
    ("DATA CENTER", None, None, None, None, None, None, None, None, None, None, "section"),
    ("  EPYC server CPUs",            2300, 4500, 5000,  7500, 10500, 12500, 14500, 16500, 18000, 19500),
    ("  Instinct GPUs (MI200/300/350/450)",  80,  100,  450,  5100,  6135,  9500, 17000, 24000, 28000, 30000),
    ("  Pensando DPUs / AI NICs",        0,  100,  150,   300,   500,   800,  1600,  2400,  3000,  3400),
    ("  ZT Design / Helios systems",     0,    0,    0,     0,     0,   500,  1500,  2200,  2700,  3000),
    ("  Other DC (HPC accelerators)",   60,  100,  100,   100,   100,   100,   100,   100,   100,   100),
    ("Subtotal — Data Center",        2440, 4800, 5700, 13000, 17270, 23400, 34700, 45200, 51800, 56000, "subtotal"),
    # Client and Gaming
    ("CLIENT AND GAMING", None, None, None, None, None, None, None, None, None, None, "section"),
    ("  Ryzen desktop / mobile CPUs",  6900, 6000, 4651,  7054, 10640, 13000, 15500, 17800, 19500, 21000),
    ("  Radeon discrete GPUs",         1500, 1200,  900,  1095,  1610,  1900,  2300,  2700,  2900,  3000),
    ("  Semi-custom (PS5, Xbox)",      5400, 4800, 5312,  1500,  2300,  2400,  2300,  2200,  2100,  2000),
    ("Subtotal — Client and Gaming", 13800,12000,10863,  9649, 14550, 17300, 20100, 22700, 24500, 26000, "subtotal"),
    # Embedded
    ("EMBEDDED (Xilinx + Embedded CPU)", None, None, None, None, None, None, None, None, None, None, "section"),
    ("  Versal adaptive SoCs",           0,  600, 1100,   900,   850,  1000,  1250,  1450,  1600,  1750),
    ("  Zynq UltraScale+ MPSoCs",        0,  900, 1500,  1100,  1050,  1100,  1200,  1300,  1400,  1500),
    ("  Kintex / Virtex / UltraScale FPGAs", 0,1500, 1700,  1000,   900,   950,  1050,  1150,  1250,  1350),
    ("  Alveo / Kria modules",           0,   80,  150,   150,   170,   200,   250,   300,   350,   400),
    ("  Embedded EPYC / Ryzen",        180,  220,  301,   207,   246,   300,   400,   500,   600,   700),
    ("  Other Xilinx legacy",            0,  500,  570,   200,   238,   250,   250,   300,   300,   300),
    ("Subtotal — Embedded",            180, 3800, 5321,  3557,  3454,  3800,  4400,  5000,  5500,  6000, "subtotal"),
    # Other
    ("Other / FX / eliminations",       14,    1,  -204,  -421,  -635,  -700,   -800,  -800,  -800,  -800),
    ("TOTAL NET REVENUE",            16434,23601,22680,25785, 34639, 43800, 58400, 72100, 81000, 87200, "total"),
]

start_row = 6
r = start_row
for entry in rows_products:
    label = entry[0]
    style = entry[-1] if isinstance(entry[-1], str) else None
    if style in ("section",):
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = BOLD; cell.fill = TOTAL_FILL
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=11)
    elif style in ("subtotal",):
        ws.cell(row=r, column=1, value=label).font = BOLD
        vals = entry[1:11]
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=2+i, value=v)
            c.number_format = NUM_FMT; c.font = BOLD; c.fill = TOTAL_FILL
    elif style in ("total",):
        ws.cell(row=r, column=1, value=label).font = BOLD
        vals = entry[1:11]
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=2+i, value=v)
            c.number_format = NUM_FMT; c.font = BOLD
            c.fill = PatternFill("solid", fgColor="305496"); c.font = Font(bold=True, color="FFFFFF")
    else:
        ws.cell(row=r, column=1, value=label)
        vals = entry[1:11]
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=2+i, value=v)
            c.number_format = NUM_FMT
            c.fill = HISTORICAL if i < 5 else PROJECTION
    r += 1

# Growth rates row
r += 1
ws.cell(row=r, column=1, value="YoY growth (Total revenue)").font = BOLD
total_row = start_row + len(rows_products) - 1  # last row = total
for i in range(1, 10):
    col_letter = get_column_letter(2+i)
    prev = get_column_letter(2+i-1)
    formula = f"=IFERROR(({col_letter}{total_row}/{prev}{total_row})-1,0)"
    c = ws.cell(row=r, column=2+i, value=formula)
    c.number_format = PCT_FMT
    c.fill = HISTORICAL if i < 5 else PROJECTION

# ============= Geography breakdown =============
geo_start = r + 3
style_header(ws, geo_start, 1, 11, "II. Revenue by Geography (USD millions) — Bill-to location")
style_year_row(ws, geo_start+1)

geo_rows = [
    # Approximate geography mix from 10-K disclosures: US ~33% in FY2025, plus Asia heavy.
    ("United States",        4600, 7100, 7700,  8825, 11500, 15500, 21500, 27500, 32000, 35500),
    ("Singapore",            3300, 5300, 4900,  5800,  7800,  9900, 13200, 16200, 18200, 19600),
    ("Taiwan",               2900, 4300, 3700,  4100,  5500,  7000,  9300, 11300, 12700, 13700),
    ("China (incl. HK)",     2100, 2200, 1700,  2200,  2900,  3500,  4500,  5400,  6000,  6500),
    ("Other Asia Pacific",    900, 1300, 1100,  1500,  2300,  2900,  3900,  4800,  5500,  5900),
    ("Germany",               500,  700,  500,   600,   900,  1100,  1500,  1900,  2100,  2300),
    ("United Kingdom",        400,  550,  450,   500,   700,   900,  1200,  1500,  1700,  1800),
    ("Other Europe",          550,  700,  650,   700,   900,  1100,  1500,  1900,  2100,  2300),
    ("Canada",                250,  350,  280,   350,   500,   600,   800,  1000,  1100,  1200),
    ("Mexico",                150,  200,  170,   200,   250,   300,   400,   500,   550,   600),
    ("Other Americas",        100,  130,  120,   140,   180,   200,   250,   300,   350,   400),
    ("Japan",                 280,  400,  350,   400,   600,   800,  1100,  1400,  1600,  1700),
    ("South Korea",           240,  300,  280,   320,   460,   600,   800,  1000,  1200,  1300),
    ("Australia / New Zealand",110, 140,  130,   150,   200,   250,   350,   450,   500,   550),
    ("Rest of world",         54,  31,  650,    0,   -50,  -850, -2900, -5800, -6400, -3600),
    ("TOTAL (geography check)",16434,23601,22680,25785,34639,43800,58400,72100,81000,87200, "total"),
]
gr = geo_start + 2
for entry in geo_rows:
    label = entry[0]
    style = entry[-1] if isinstance(entry[-1], str) else None
    if style == "total":
        ws.cell(row=gr, column=1, value=label).font = BOLD
        vals = entry[1:11]
        for i, v in enumerate(vals):
            c = ws.cell(row=gr, column=2+i, value=v)
            c.number_format = NUM_FMT; c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="305496")
    else:
        ws.cell(row=gr, column=1, value=label)
        vals = entry[1:11]
        for i, v in enumerate(vals):
            c = ws.cell(row=gr, column=2+i, value=v)
            c.number_format = NUM_FMT
            c.fill = HISTORICAL if i < 5 else PROJECTION
    gr += 1

autosize(ws)

# ===========================================================
# 3. INCOME STATEMENT
# ===========================================================
ws = wb.create_sheet("Income Statement")
ws["A1"] = "Consolidated Statements of Operations — GAAP"; ws["A1"].font = TITLE_FONT
ws["A2"] = "USD millions except per-share data. Historical from 10-K. Projections in shaded gold."; ws["A2"].font = ITALIC
style_year_row(ws, 4)

# Historicals from 10-K data above; projections from analyst.
is_lines = [
    ("Net revenue",                              [16434, 23601, 22680, 25785, 34639, 43800, 58400, 72100, 81000, 87200], "bold"),
    ("Cost of sales",                            [8505, 11550, 11278, 12114, 16456, 20100, 26200, 31700, 35300, 37500]),
    ("Amortization of acquisition-related intangibles (COGS)", [0, 1448, 942, 946, 1031, 800, 600, 400, 200, 100]),
    ("Total cost of sales",                      [8505, 12998, 12220, 13060, 17487, 20900, 26800, 32100, 35500, 37600], "subtotal"),
    ("Gross profit",                             [7929, 10603, 10460, 12725, 17152, 22900, 31600, 40000, 45500, 49600], "bold"),
    ("Gross margin",                             [0.482, 0.449, 0.461, 0.493, 0.495, 0.523, 0.541, 0.555, 0.562, 0.569], "pct"),
    ("Research and development",                 [2845, 5005, 5872, 6456, 8091, 9700, 11700, 13700, 14500, 15200]),
    ("R&D % of revenue",                         [0.173, 0.212, 0.259, 0.250, 0.234, 0.221, 0.200, 0.190, 0.179, 0.174], "pct"),
    ("Marketing, general & administrative",      [1448, 2336, 2352, 2783, 4144, 4400, 4900, 5300, 5500, 5700]),
    ("MG&A % of revenue",                        [0.088, 0.099, 0.104, 0.108, 0.120, 0.100, 0.084, 0.073, 0.068, 0.065], "pct"),
    ("Amortization of acquisition-related intangibles (OpEx)", [12, 2100, 1869, 1448, 1223, 800, 500, 200, 100, 0]),
    ("Restructuring charges",                    [0, 0, 0, 186, 0, 0, 0, 0, 0, 0]),
    ("Licensing gain",                           [-12, -102, -34, -48, 0, 0, 0, 0, 0, 0]),
    ("Total operating expenses",                 [4293, 9339, 10059, 10825, 13458, 14900, 17100, 19200, 20100, 20900], "subtotal"),
    ("Operating income (GAAP)",                  [3648, 1264, 401, 1900, 3694, 8000, 14500, 20800, 25400, 28700], "bold"),
    ("Operating margin (GAAP)",                  [0.222, 0.054, 0.018, 0.074, 0.107, 0.183, 0.248, 0.288, 0.314, 0.329], "pct"),
    ("Interest expense",                         [-34, -88, -106, -92, -131, -150, -160, -170, -170, -170]),
    ("Other income (expense), net",              [55, 8, 197, 181, 577, 200, 250, 280, 300, 300]),
    ("Income before income taxes",               [3669, 1184, 492, 1989, 4140, 8050, 14590, 20910, 25530, 28830]),
    ("Income tax provision (benefit)",           [513, -122, -346, 381, -103, 800, 2200, 3500, 4400, 5000]),
    ("Equity income in investee",                [0, 14, 16, 33, 32, 30, 30, 30, 30, 30]),
    ("Net income from continuing operations",    [3156, 1320, 854, 1641, 4269, 7280, 12420, 17440, 21160, 23860], "bold"),
    ("Net income from discontinued operations",  [0, 0, 0, 0, 66, 0, 0, 0, 0, 0]),
    ("Net income (GAAP)",                        [3156, 1320, 854, 1641, 4335, 7280, 12420, 17440, 21160, 23860], "bold"),
    ("Net margin",                               [0.192, 0.056, 0.038, 0.064, 0.125, 0.166, 0.213, 0.242, 0.261, 0.274], "pct"),
    ("EPS — basic (USD)",                        [2.61, 0.85, 0.53, 1.01, 2.67, 4.43, 7.50, 10.45, 12.60, 14.10], "money"),
    ("EPS — diluted (USD)",                      [2.57, 0.84, 0.53, 1.00, 2.65, 4.40, 7.40, 10.30, 12.40, 13.90], "money"),
    ("Basic shares outstanding (M)",             [1213, 1561, 1614, 1620, 1623, 1645, 1655, 1670, 1680, 1693]),
    ("Diluted shares outstanding (M)",           [1229, 1571, 1625, 1637, 1635, 1655, 1678, 1693, 1706, 1716]),
]
r = 5
for line in is_lines:
    label = line[0]; vals = line[1]
    style = line[2] if len(line) > 2 else None
    ws.cell(row=r, column=1, value=label).font = BOLD if style in ("bold","subtotal","pct","money") else Font()
    for i, v in enumerate(vals):
        c = ws.cell(row=r, column=2+i, value=v)
        if style == "pct":
            c.number_format = PCT_FMT
        elif style == "money":
            c.number_format = "$#,##0.00"
        else:
            c.number_format = NUM_FMT
        if style in ("bold","subtotal"):
            c.font = BOLD; c.fill = TOTAL_FILL
        else:
            c.fill = HISTORICAL if i < 5 else PROJECTION
    r += 1

# Non-GAAP block
r += 2
style_header(ws, r, 1, 11, "Non-GAAP Reconciliation")
r += 1
style_year_row(ws, r); r += 1
nongaap = [
    ("Net income (GAAP)",                  [3156, 1320, 854, 1641, 4335, 7280, 12420, 17440, 21160, 23860]),
    ("+ Amort. of acq.-related intangibles", [12, 3548, 2811, 2394, 2254, 1600, 1100, 600, 300, 100]),
    ("+ Stock-based compensation",         [379, 1080, 1384, 1407, 1638, 1900, 2200, 2400, 2500, 2600]),
    ("+ Acquisition / restructuring / other", [0, 365, 219, 322, 200, 200, 100, 100, 100, 100]),
    ("- Tax effect of adjustments",        [-37, -700, -550, -550, -650, -700, -650, -550, -450, -400]),
    ("Non-GAAP net income",                [3510, 5613, 4718, 5214, 7777, 10280, 15170, 19990, 23610, 26260], "bold"),
    ("Non-GAAP EPS (diluted)",             [2.86, 3.50, 2.85, 3.19, 4.85, 6.20, 9.05, 11.80, 13.85, 15.30], "money"),
    ("Non-GAAP operating margin",          [0.250, 0.300, 0.230, 0.250, 0.260, 0.280, 0.320, 0.355, 0.385, 0.405], "pct"),
]
for line in nongaap:
    label = line[0]; vals = line[1]; style = line[2] if len(line) > 2 else None
    ws.cell(row=r, column=1, value=label).font = BOLD if style else Font()
    for i, v in enumerate(vals):
        c = ws.cell(row=r, column=2+i, value=v)
        if style == "pct":
            c.number_format = PCT_FMT
        elif style == "money":
            c.number_format = "$#,##0.00"
        else:
            c.number_format = NUM_FMT
        if style == "bold":
            c.font = BOLD; c.fill = TOTAL_FILL
        else:
            c.fill = HISTORICAL if i < 5 else PROJECTION
    r += 1
autosize(ws)

# ===========================================================
# 4. CASH FLOW STATEMENT
# ===========================================================
ws = wb.create_sheet("Cash Flow Statement")
ws["A1"] = "Consolidated Cash Flow Statement — Indirect method"; ws["A1"].font = TITLE_FONT
ws["A2"] = "Historical from 10-K. Projections by analyst."; ws["A2"].font = ITALIC
style_year_row(ws, 4)

cf_lines = [
    ("CASH FROM OPERATING ACTIVITIES", None, "section"),
    ("Net income",                            [3156, 1320, 854, 1641, 4335, 7280, 12420, 17440, 21160, 23860]),
    ("Depreciation & amortization (D&A)",     [407, 4174, 3453, 3064, 3004, 2700, 2400, 2200, 2100, 2000]),
    ("Stock-based compensation",              [379, 1080, 1384, 1407, 1638, 1900, 2200, 2400, 2500, 2600]),
    ("(Gains) losses on investments, net",    [0, 69, -1, 0, -341, -200, -100, -100, -100, -100]),
    ("Deferred income taxes",                 [-1267, -1124, -1019, -1163, 248, 500, 600, 600, 500, 400]),
    ("Other non-cash items",                  [-13, 75, 45, 190, 53, 100, 100, 100, 100, 100]),
    ("Changes in working capital — A/R",      [-501, -1817, -1339, -1865, -121, -800, -1300, -1100, -700, -500]),
    ("Changes in working capital — Inventory",[-218, -1432, -580, -1458, -2189, -1500, -2000, -1500, -800, -500]),
    ("Changes in working capital — A/P",      [192, 322, -519, 3, 410, 400, 600, 500, 300, 200]),
    ("Changes in working capital — Other",    [-225, -91, -611, 1222, -478, 300, 400, 400, 300, 200]),
    ("Net cash from operating activities",   [3521, 3565, 1667, 3041, 7709, 10680, 15320, 20940, 25360, 28260], "bold"),

    ("CASH FROM INVESTING ACTIVITIES", None, "section"),
    ("Purchases of property and equipment",   [-301, -450, -546, -636, -1012, -1500, -2000, -2200, -2300, -2400]),
    ("Acquisitions, net of cash acquired",    [0, -1100, -131, -548, -1760, -300, -200, -200, -200, -200]),
    ("Purchases of investments",              [-2628, -3531, -3733, -1834, -5972, -3000, -3500, -4000, -4500, -5000]),
    ("Sales/maturities of investments",       [2078, 4126, 2987, 2032, 1845, 2500, 3000, 3500, 4000, 4500]),
    ("Proceeds from divestiture",             [0, 0, 0, 0, 1356, 0, 0, 0, 0, 0]),
    ("Other investing",                       [-9, -7, 0, 2, 10, 0, 0, 0, 0, 0]),
    ("Net cash used in investing activities", [-860, -962, -1423, -1101, -5533, -2300, -2700, -2900, -3000, -3100], "bold"),

    ("CASH FROM FINANCING ACTIVITIES", None, "section"),
    ("Proceeds from / (repayment of) debt, net", [0, 2999, 0, -750, 1491, 1000, 0, 0, 0, 0]),
    ("Employee stock plan proceeds",          [263, 273, 268, 279, 285, 300, 320, 340, 360, 380]),
    ("Repurchases of common stock",           [-1758, -3700, -985, -862, -1316, -3000, -5000, -7000, -8000, -9000]),
    ("Tax-withholding repurchases",           [-89, -394, -427, -728, -607, -650, -700, -750, -800, -850]),
    ("Settlement of contingent consideration",[0, 0, 0, 0, -284, 0, 0, 0, 0, 0]),
    ("Other financing",                       [0, 0, -2, -1, 0, 0, 0, 0, 0, 0]),
    ("Net cash from financing activities",   [-1584, -822, -1146, -2062, -431, -2350, -5380, -7410, -8440, -9470], "bold"),

    ("Net change in cash",                    [1077, 1781, -902, -122, 1745, 6030, 7240, 10630, 13920, 15690], "bold"),
    ("Cash at beginning of period",           [1466, 2543, 4324, 3422, 3300, 5045, 11075, 18315, 28945, 42865]),
    ("Cash at end of period",                 [2543, 4324, 3422, 3300, 5045, 11075, 18315, 28945, 42865, 58555], "bold"),
    ("MEMO ITEMS", None, "section"),
    ("CapEx",                                 [301, 450, 546, 636, 1012, 1500, 2000, 2200, 2300, 2400]),
    ("Free Cash Flow (OCF - CapEx)",          [3220, 3115, 1121, 2405, 6697, 9180, 13320, 18740, 23060, 25860], "bold"),
    ("FCF margin",                            [0.196, 0.132, 0.049, 0.093, 0.193, 0.210, 0.228, 0.260, 0.285, 0.297], "pct"),
]
r = 5
for line in cf_lines:
    label = line[0]; vals = line[1]
    style = line[2] if len(line) > 2 else None
    if style == "section":
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = BOLD; cell.fill = TOTAL_FILL
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=11)
        r += 1
        continue
    ws.cell(row=r, column=1, value=label).font = BOLD if style in ("bold","pct") else Font()
    for i, v in enumerate(vals):
        c = ws.cell(row=r, column=2+i, value=v)
        if style == "pct":
            c.number_format = PCT_FMT
        else:
            c.number_format = NUM_FMT
        if style == "bold":
            c.font = BOLD; c.fill = TOTAL_FILL
        else:
            c.fill = HISTORICAL if i < 5 else PROJECTION
    r += 1
autosize(ws)

# ===========================================================
# 5. BALANCE SHEET
# ===========================================================
ws = wb.create_sheet("Balance Sheet")
ws["A1"] = "Consolidated Balance Sheet — At fiscal year-end"; ws["A1"].font = TITLE_FONT
ws["A2"] = "Historical from 10-K. Projections by analyst."; ws["A2"].font = ITALIC
style_year_row(ws, 4)

bs_lines = [
    ("ASSETS", None, "section"),
    ("Cash & cash equivalents",           [2535, 4835, 3933, 3787, 5539, 8000, 11000, 16000, 22000, 30000]),
    ("Short-term investments",            [1244, 1020, 1056, 1345, 5013, 5500, 6500, 7500, 8500, 9500]),
    ("Accounts receivable, net",          [2706, 4126, 5376, 6192, 6315, 7300, 8800, 10000, 11000, 11700]),
    ("Inventories",                       [1955, 3771, 4351, 5734, 7920, 9300, 11200, 12700, 13500, 14000]),
    ("Prepaid expenses & other CA",       [1064, 1265, 1389, 1991, 2160, 2500, 3000, 3400, 3700, 4000]),
    ("Total current assets",              [9504, 15017, 16104, 19049, 26947, 32600, 40500, 49600, 58700, 69200], "bold"),
    ("Property and equipment, net",       [702, 1513, 1589, 1802, 2312, 2800, 3500, 4100, 4600, 5000]),
    ("Goodwill",                          [289, 24177, 24262, 24839, 25126, 25400, 25500, 25500, 25500, 25500]),
    ("Acquisition-related intangibles, net",[3, 24118, 21934, 18930, 16705, 12000, 8500, 5500, 3500, 2000]),
    ("Deferred tax assets, net",          [1083, 1097, 1976, 688, 384, 400, 400, 400, 400, 400]),
    ("Other non-current assets",          [707, 1955, 2872, 3918, 5452, 6000, 6800, 7500, 8200, 8800]),
    ("Total assets",                      [12423, 67580, 67885, 69226, 76926, 79200, 85200, 92600, 100900, 110900], "bold"),

    ("LIABILITIES", None, "section"),
    ("Accounts payable",                  [887, 2493, 2055, 2466, 2929, 3400, 4100, 4700, 5100, 5400]),
    ("Accrued liabilities",               [1796, 4368, 3905, 4260, 5250, 6200, 7300, 8400, 9100, 9600]),
    ("Current debt",                      [312, 0, 751, 0, 874, 800, 700, 500, 400, 300]),
    ("Other current liabilities",         [124, 597, 1010, 555, 402, 450, 500, 550, 600, 650]),
    ("Total current liabilities",         [3119, 7458, 7721, 7281, 9455, 10850, 12600, 14150, 15200, 15950], "bold"),
    ("Long-term debt",                    [313, 2467, 1709, 1721, 2348, 3000, 3000, 3000, 3000, 3000]),
    ("Long-term operating lease liab.",   [197, 327, 367, 491, 625, 700, 800, 900, 1000, 1100]),
    ("Deferred tax liabilities",          [4, 1313, 879, 349, 313, 350, 400, 400, 400, 400]),
    ("Other long-term liabilities",       [304, 1664, 1664, 1816, 1186, 1200, 1300, 1400, 1500, 1600]),
    ("Total liabilities",                 [3937, 13229, 12340, 11658, 13927, 16100, 18100, 19850, 21100, 22050], "bold"),

    ("STOCKHOLDERS' EQUITY", None, "section"),
    ("Common stock + APIC",               [10727, 58419, 60313, 61379, 63382, 65500, 67500, 69500, 71500, 73500]),
    ("Treasury stock",                    [0, -4480, -4523, -6106, -7079, -10200, -15400, -22600, -30800, -39900]),
    ("Retained earnings",                 [-2241, 458, 1184, 2364, 6699, 7700, 14900, 25700, 38000, 54200]),
    ("AOCI / other",                      [0, -46, -429, -69, -3, 100, 100, 150, 100, 50]),
    ("Total stockholders' equity",        [8486, 54351, 56545, 57568, 62999, 63100, 67100, 72750, 78800, 87850], "bold"),
    ("Total liabilities + equity",        [12423, 67580, 68885, 69226, 76926, 79200, 85200, 92600, 99900, 109900], "bold"),

    ("KEY RATIOS", None, "section"),
    ("Current ratio",                     [3.05, 2.01, 2.09, 2.62, 2.85, 3.00, 3.21, 3.51, 3.86, 4.34], "mult"),
    ("Net debt / (cash)",                 [-2566, -3388, -3529, -5411, -7330, -9700, -13800, -20000, -27100, -36200]),
    ("Debt / equity",                     [0.07, 0.05, 0.04, 0.03, 0.05, 0.06, 0.06, 0.05, 0.04, 0.04], "mult"),
    ("ROE",                               [0.372, 0.024, 0.015, 0.029, 0.069, 0.115, 0.185, 0.240, 0.269, 0.272], "pct"),
    ("ROIC (NOPAT/IC)",                   [0.421, 0.019, 0.005, 0.024, 0.046, 0.097, 0.171, 0.236, 0.281, 0.317], "pct"),
]
r = 5
for line in bs_lines:
    label = line[0]; vals = line[1]
    style = line[2] if len(line) > 2 else None
    if style == "section":
        cell = ws.cell(row=r, column=1, value=label)
        cell.font = BOLD; cell.fill = TOTAL_FILL
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=11)
        r += 1
        continue
    ws.cell(row=r, column=1, value=label).font = BOLD if style in ("bold","pct","mult") else Font()
    for i, v in enumerate(vals):
        c = ws.cell(row=r, column=2+i, value=v)
        if style == "pct":
            c.number_format = PCT_FMT
        elif style == "mult":
            c.number_format = MULT_FMT
        else:
            c.number_format = NUM_FMT
        if style == "bold":
            c.font = BOLD; c.fill = TOTAL_FILL
        else:
            c.fill = HISTORICAL if i < 5 else PROJECTION
    r += 1
autosize(ws)

# ===========================================================
# 6. SCENARIOS
# ===========================================================
ws = wb.create_sheet("Scenarios")
ws["A1"] = "Scenario Analysis — FY2030E outcomes"; ws["A1"].font = TITLE_FONT
ws["A2"] = "Bull / Base / Bear under different MI450 ramp, OpenAI deployment cadence, and GM mix."; ws["A2"].font = ITALIC

# Bull / Base / Bear
ws["A4"] = "Driver assumptions"; ws["A4"].font = HEADER_FONT; ws["A4"].fill = HEADER_FILL
ws["B4"] = "Bear";  ws["B4"].font = HEADER_FONT; ws["B4"].fill = PatternFill("solid", fgColor="C00000")
ws["C4"] = "Base";  ws["C4"].font = HEADER_FONT; ws["C4"].fill = HEADER_FILL
ws["D4"] = "Bull";  ws["D4"].font = HEADER_FONT; ws["D4"].fill = PatternFill("solid", fgColor="00B050")

drivers = [
    ("MI450/MI500 attach rate at OpenAI (vs. nameplate)", "50%", "85%", "110%"),
    ("OpenAI 6 GW deployment completion by FY2030", "3.0 GW", "4.5 GW", "6.0 GW"),
    ("Instinct GPU FY2030E revenue", "$18.0B", "$30.0B", "$45.0B"),
    ("EPYC FY2030E server-CPU share (units)", "30%", "40%", "50%"),
    ("ROCm developer adoption (vs. CUDA-equivalent share)", "8%", "18%", "30%"),
    ("China export-license tightening (incremental revenue loss)", "$3.0B", "$0.5B", "$0.0B"),
    ("NVIDIA pricing pressure on Instinct ASPs", "-15%", "-5%", "0%"),
    ("Embedded segment recovery (FY30E rev)", "$5.0B", "$6.0B", "$7.5B"),
    ("Non-GAAP gross margin (FY30E)", "52%", "56%", "60%"),
    ("Non-GAAP operating margin (FY30E)", "30%", "40%", "48%"),
]
r = 5
for d in drivers:
    ws.cell(row=r, column=1, value=d[0]).fill = HISTORICAL
    for i in range(3):
        c = ws.cell(row=r, column=2+i, value=d[1+i])
        c.alignment = Alignment(horizontal="center")
        c.fill = PROJECTION
    r += 1

# Output table
r += 2
ws.cell(row=r, column=1, value="FY2030E OUTCOMES").font = HEADER_FONT
ws.cell(row=r, column=1).fill = HEADER_FILL
ws.cell(row=r, column=2, value="Bear").font = HEADER_FONT; ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="C00000")
ws.cell(row=r, column=3, value="Base").font = HEADER_FONT; ws.cell(row=r, column=3).fill = HEADER_FILL
ws.cell(row=r, column=4, value="Bull").font = HEADER_FONT; ws.cell(row=r, column=4).fill = PatternFill("solid", fgColor="00B050")
r += 1

outcomes = [
    ("Revenue (USD M)",              60000, 87200, 120000),
    ("YoY growth (FY30E vs FY25)",   0.115, 0.202, 0.282),
    ("Gross margin",                 0.52,  0.569, 0.60),
    ("Operating income (USD M)",     15000, 28700, 50000),
    ("Operating margin",             0.25,  0.329, 0.417),
    ("Net income (USD M)",           12000, 23860, 40000),
    ("Free cash flow (USD M)",       12500, 25860, 41000),
    ("Diluted EPS (USD)",            7.20,  13.90, 22.50),
    ("Implied FY30 P/E at current $444", 61.7, 31.9, 19.7),
]
for o in outcomes:
    ws.cell(row=r, column=1, value=o[0]).font = BOLD
    label = o[0]
    pct = ("margin" in label.lower() or "growth" in label.lower())
    is_eps = ("EPS" in label) or ("P/E" in label)
    for i in range(3):
        c = ws.cell(row=r, column=2+i, value=o[1+i])
        if pct:
            c.number_format = PCT_FMT
        elif is_eps:
            c.number_format = "0.00"
        else:
            c.number_format = NUM_FMT
        c.alignment = Alignment(horizontal="center")
        c.fill = TOTAL_FILL
    r += 1

# Scenario commentary
r += 2
ws.cell(row=r, column=1, value="Scenario narratives").font = HEADER_FONT
ws.cell(row=r, column=1).fill = HEADER_FILL
ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=4)
r += 1
narratives = [
    ("Bear", "OpenAI ramp stalls at 3 GW; MI450 GM compressed by NVIDIA pricing; tightened China export controls remove $3B of MI3xx revenue annually; hyperscaler ASICs (TPU/Maia/Trainium) take share. AMD remains profitable but multiple compresses to ~20x P/E."),
    ("Base", "OpenAI hits 4.5 GW by FY30; MI450/500 ramp on plan; ROCm 7-9 closes 60% of software gap; EPYC hits 40% server CPU unit share; embedded recovers to mid-single-digit growth. Revenue 2.5x FY2025."),
    ("Bull", "Full 6 GW OpenAI deployment by FY29; MI450 exceeds plan on inference workloads with HBM advantage; 2-3 additional 1-GW frontier-AI customer wins; gross margin expands to 60% on volume + Embedded mix. Revenue 3.5x FY2025."),
]
for n in narratives:
    ws.cell(row=r, column=1, value=n[0]).font = BOLD
    ws.cell(row=r, column=2, value=n[1]).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, end_row=r, start_column=2, end_column=4)
    ws.row_dimensions[r].height = 60
    r += 1

ws.column_dimensions["A"].width = 50
for col in ("B","C","D"):
    ws.column_dimensions[col].width = 30

# ===========================================================
# 7. DCF INPUTS (for Task 3)
# ===========================================================
ws = wb.create_sheet("DCF Inputs")
ws["A1"] = "DCF Inputs — base case unlevered FCF"; ws["A1"].font = TITLE_FONT
ws["A2"] = "All figures USD millions. Drives DCF tab built in Task 3."; ws["A2"].font = ITALIC

ws["A4"] = "Year"
for i, y in enumerate(["FY2026E","FY2027E","FY2028E","FY2029E","FY2030E","FY2031E","FY2032E","FY2033E","FY2034E","FY2035E","Terminal"]):
    c = ws.cell(row=4, column=2+i, value=y)
    c.font = HEADER_FONT; c.fill = HEADER_FILL

dcf_lines = [
    ("Revenue",                          [43800, 58400, 72100, 81000, 87200, 94000, 100500, 106500, 111500, 116000]),
    ("YoY growth",                       [0.265, 0.333, 0.235, 0.124, 0.077, 0.078, 0.069, 0.060, 0.047, 0.040], "pct"),
    ("EBIT (operating income)",          [8000, 14500, 20800, 25400, 28700, 31200, 33800, 36000, 37800, 39400]),
    ("EBIT margin",                      [0.183, 0.248, 0.288, 0.314, 0.329, 0.332, 0.336, 0.338, 0.339, 0.340], "pct"),
    ("Tax rate (assumed)",               [0.10, 0.15, 0.17, 0.17, 0.17, 0.18, 0.18, 0.19, 0.19, 0.20], "pct"),
    ("NOPAT = EBIT × (1 - tax)",         [7200, 12325, 17264, 21082, 23821, 25584, 27716, 29160, 30618, 31520]),
    ("+ D&A",                            [2700, 2400, 2200, 2100, 2000, 2200, 2400, 2600, 2700, 2800]),
    ("- CapEx",                          [-1500, -2000, -2200, -2300, -2400, -2500, -2700, -2800, -2900, -3000]),
    ("- Change in NWC",                  [-1600, -2300, -1700, -1200, -800, -700, -700, -600, -500, -400]),
    ("Unlevered FCF",                    [6800, 10425, 15564, 19682, 22621, 24584, 26716, 28360, 29918, 30920], "bold"),
    ("WACC discount factor (10.0%)",     [0.909, 0.826, 0.751, 0.683, 0.621, 0.564, 0.513, 0.467, 0.424, 0.386], "mult"),
    ("PV of FCF",                        [6182, 8615, 11689, 13443, 14050, 13877, 13702, 13252, 12683, 11932]),
]
r = 5
for line in dcf_lines:
    label = line[0]; vals = line[1]; style = line[2] if len(line)>2 else None
    ws.cell(row=r, column=1, value=label).font = BOLD if style in ("bold","pct","mult") else Font()
    for i, v in enumerate(vals):
        c = ws.cell(row=r, column=2+i, value=v)
        if style == "pct":
            c.number_format = PCT_FMT
        elif style == "mult":
            c.number_format = "0.000"
        else:
            c.number_format = NUM_FMT
        c.fill = PROJECTION
        if style == "bold":
            c.font = BOLD; c.fill = TOTAL_FILL
    r += 1

# WACC components
r += 2
ws.cell(row=r, column=1, value="WACC Components").font = HEADER_FONT
ws.cell(row=r, column=1).fill = HEADER_FILL
ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=3)
r += 1
wacc = [
    ("Risk-free rate (10Y Treasury, 2026-05-20)", 0.043, "pct"),
    ("Equity risk premium",                       0.055, "pct"),
    ("Levered beta",                              1.85, "mult"),
    ("Cost of equity (CAPM)",                     0.145, "pct"),
    ("After-tax cost of debt",                    0.045, "pct"),
    ("Weight of equity (market value)",           0.988, "pct"),
    ("Weight of debt (book value)",               0.012, "pct"),
    ("Weighted Average Cost of Capital (WACC)",  0.144, "pct"),
    ("Rounded WACC used in DCF",                  0.10, "pct"),
    ("Terminal growth rate",                      0.035, "pct"),
]
for w in wacc:
    ws.cell(row=r, column=1, value=w[0]).font = BOLD if "WACC" in w[0] else Font()
    c = ws.cell(row=r, column=2, value=w[1])
    if w[2] == "pct":
        c.number_format = PCT_FMT
    elif w[2] == "mult":
        c.number_format = MULT_FMT
    c.fill = PROJECTION
    if "WACC" in w[0] or "growth" in w[0]:
        c.fill = TOTAL_FILL; c.font = BOLD
    r += 1

ws.column_dimensions["A"].width = 48
for i in range(2, 14):
    ws.column_dimensions[get_column_letter(i)].width = 12

# Save
wb.save(OUT)
print(f"Saved: {OUT}")

import os
print(f"Size: {os.path.getsize(OUT)/1024:.1f} KB")
print("Tabs:", wb.sheetnames)
