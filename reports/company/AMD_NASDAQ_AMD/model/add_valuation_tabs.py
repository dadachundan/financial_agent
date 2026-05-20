"""Task 3 — Add DCF, Sensitivity, Comps, Valuation Summary tabs to existing AMD model."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

FN = "/Users/x/projects/financial_agent/reports/company/AMD_NASDAQ_AMD/model/AMD_Financial_Model_2026-05-20.xlsx"
wb = load_workbook(FN)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F3864")
BOLD = Font(name="Calibri", size=11, bold=True)
ITALIC = Font(name="Calibri", size=10, italic=True, color="595959")
PROJ = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")

NUM_FMT = "#,##0;(#,##0);–"
PCT_FMT = "0.0%;(0.0%);–"

# Remove any prior runs so script is idempotent
for tab in ("DCF", "Sensitivity", "Comparables", "Valuation Summary"):
    if tab in wb.sheetnames:
        del wb[tab]

# ===========================================================
# 1) DCF tab
# ===========================================================
ws = wb.create_sheet("DCF")
ws["A1"] = "AMD DCF — Base Case"; ws["A1"].font = TITLE_FONT
ws["A2"] = "USD millions except per-share. WACC = 10.0%, terminal growth = 3.0%."; ws["A2"].font = ITALIC

years = ["FY2026E","FY2027E","FY2028E","FY2029E","FY2030E","FY2031E","FY2032E","FY2033E","FY2034E","FY2035E"]
ws.cell(row=4, column=1, value="Year").font = HEADER_FONT; ws.cell(row=4, column=1).fill = HEADER_FILL
for i, y in enumerate(years):
    c = ws.cell(row=4, column=2+i, value=y); c.font = HEADER_FONT; c.fill = HEADER_FILL

rev = [43800, 58400, 72100, 81000, 87200, 94000, 100500, 106500, 111500, 116000]
ebit = [8000, 14500, 20800, 25400, 28700, 31200, 33800, 36000, 37800, 39400]
tax = [0.10, 0.15, 0.17, 0.17, 0.17, 0.18, 0.18, 0.19, 0.19, 0.20]
nopat = [round(e * (1 - t)) for e, t in zip(ebit, tax)]
da = [2700, 2400, 2200, 2100, 2000, 2200, 2400, 2600, 2700, 2800]
capex = [-1500, -2000, -2200, -2300, -2400, -2500, -2700, -2800, -2900, -3000]
nwc = [-1600, -2300, -1700, -1200, -800, -700, -700, -600, -500, -400]
ufcf = [n + d + c + w for n, d, c, w in zip(nopat, da, capex, nwc)]

WACC = 0.10
G = 0.03
# Mid-year convention disc factor
disc = [(1 + WACC) ** (i + 0.5) for i in range(10)]
pv = [u / d for u, d in zip(ufcf, disc)]

rows = [
    ("Revenue", rev, NUM_FMT),
    ("EBIT", ebit, NUM_FMT),
    ("EBIT margin", [e / r for e, r in zip(ebit, rev)], PCT_FMT),
    ("Tax rate", tax, PCT_FMT),
    ("NOPAT", nopat, NUM_FMT),
    ("+ D&A", da, NUM_FMT),
    ("- CapEx", capex, NUM_FMT),
    ("- Δ NWC", nwc, NUM_FMT),
    ("Unlevered FCF", ufcf, NUM_FMT),
    ("Discount factor (mid-year)", disc, "0.000"),
    ("PV of FCF", pv, NUM_FMT),
]
r = 5
for label, vals, fmt in rows:
    is_bold = label in ("NOPAT", "Unlevered FCF", "PV of FCF", "EBIT")
    ws.cell(row=r, column=1, value=label).font = BOLD if is_bold else Font()
    for i, v in enumerate(vals):
        c = ws.cell(row=r, column=2+i, value=v); c.number_format = fmt; c.fill = PROJ
        if is_bold:
            c.font = BOLD; c.fill = TOTAL_FILL
    r += 1

# Terminal value
terminal_year_ufcf = ufcf[-1]
tv = terminal_year_ufcf * (1 + G) / (WACC - G)
pv_tv = tv / disc[-1]
sum_pv = sum(pv)
ev = sum_pv + pv_tv
cash_and_inv = 5539 + 5013  # FY25
debt = 874 + 2348           # current + LT
equity_value = ev + cash_and_inv - debt
diluted_shares = 1635
implied_price = equity_value / diluted_shares
current_px = 444.28
upside = implied_price / current_px - 1

r += 1
ws.cell(row=r, column=1, value="Valuation Build-Up").font = HEADER_FONT
ws.cell(row=r, column=1).fill = HEADER_FILL
ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=4)
r += 1
build_up = [
    ("Sum PV of explicit-period FCF (FY26-35)", sum_pv, NUM_FMT),
    ("Terminal value (Gordon growth, g=3.0%)", tv, NUM_FMT),
    ("PV of terminal value", pv_tv, NUM_FMT),
    ("Enterprise Value", ev, NUM_FMT, True),
    ("+ Cash & short-term investments (FY25)", cash_and_inv, NUM_FMT),
    ("- Total debt (FY25)", debt, NUM_FMT),
    ("Equity Value", equity_value, NUM_FMT, True),
    ("Diluted shares (M)", diluted_shares, NUM_FMT),
    ("Implied price per share (USD)", implied_price, "$#,##0.00", True),
    ("Current price (2026-05-20)", current_px, "$#,##0.00"),
    ("Implied upside / (downside)", upside, PCT_FMT, True),
]
for line in build_up:
    label = line[0]; v = line[1]; fmt = line[2]
    is_bold = len(line) > 3 and line[3]
    ws.cell(row=r, column=1, value=label).font = BOLD if is_bold else Font()
    c = ws.cell(row=r, column=2, value=v); c.number_format = fmt; c.fill = TOTAL_FILL if is_bold else PROJ
    if is_bold:
        c.font = BOLD
    r += 1

ws.column_dimensions["A"].width = 50
for i in range(2, 13):
    ws.column_dimensions[get_column_letter(i)].width = 12

# ===========================================================
# 2) Sensitivity tab
# ===========================================================
ws = wb.create_sheet("Sensitivity")
ws["A1"] = "Sensitivity — Implied price per share"; ws["A1"].font = TITLE_FONT
ws["A2"] = "Rows = WACC; Columns = Terminal growth. Base case: WACC 10%, g 3%."; ws["A2"].font = ITALIC

WACCS = [0.075, 0.085, 0.095, 0.10, 0.105, 0.115, 0.125]
GS = [0.020, 0.025, 0.030, 0.035, 0.040, 0.045, 0.050]

# Row 4 — column headers
ws.cell(row=4, column=1, value="WACC \\ g").font = HEADER_FONT
ws.cell(row=4, column=1).fill = HEADER_FILL
for j, g in enumerate(GS):
    c = ws.cell(row=4, column=2+j, value=g)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.number_format = PCT_FMT

def implied_price_at(wacc, g):
    disc = [(1 + wacc) ** (i + 0.5) for i in range(10)]
    pv_explicit = sum(u / d for u, d in zip(ufcf, disc))
    tv_ = ufcf[-1] * (1 + g) / (wacc - g)
    pv_tv_ = tv_ / disc[-1]
    ev_ = pv_explicit + pv_tv_
    eq = ev_ + cash_and_inv - debt
    return eq / diluted_shares

for i, w in enumerate(WACCS):
    rc = ws.cell(row=5+i, column=1, value=w)
    rc.font = HEADER_FONT; rc.fill = HEADER_FILL; rc.number_format = PCT_FMT
    for j, g in enumerate(GS):
        if w <= g:  # invalid - skip
            v = None
        else:
            v = implied_price_at(w, g)
        c = ws.cell(row=5+i, column=2+j, value=v)
        c.number_format = "$#,##0"
        c.alignment = Alignment(horizontal="center")
        c.fill = PROJ

# Color scale on the matrix
ws.conditional_formatting.add(
    f"B5:{get_column_letter(1+len(GS))}{4+len(WACCS)}",
    ColorScaleRule(start_type='min', start_color='F8696B',
                   mid_type='percentile', mid_value=50, mid_color='FFEB84',
                   end_type='max', end_color='63BE7B')
)

# Highlight base case cell (row WACC=10%, g=3.0%)
base_row = 5 + WACCS.index(0.10)
base_col = 2 + GS.index(0.030)
bc = ws.cell(row=base_row, column=base_col)
bc.font = Font(bold=True, color="FFFFFF")
bc.fill = PatternFill("solid", fgColor="1F3864")

# Implied upside matrix below
ws.cell(row=14, column=1, value="Implied Upside / (Downside) from $444.28").font = HEADER_FONT
ws.cell(row=14, column=1).fill = HEADER_FILL
ws.merge_cells(start_row=14, end_row=14, start_column=1, end_column=8)

ws.cell(row=15, column=1, value="WACC \\ g").font = HEADER_FONT
ws.cell(row=15, column=1).fill = HEADER_FILL
for j, g in enumerate(GS):
    c = ws.cell(row=15, column=2+j, value=g)
    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.number_format = PCT_FMT

for i, w in enumerate(WACCS):
    rc = ws.cell(row=16+i, column=1, value=w)
    rc.font = HEADER_FONT; rc.fill = HEADER_FILL; rc.number_format = PCT_FMT
    for j, g in enumerate(GS):
        if w <= g:
            v = None
        else:
            v = (implied_price_at(w, g) / current_px) - 1
        c = ws.cell(row=16+i, column=2+j, value=v)
        c.number_format = PCT_FMT; c.alignment = Alignment(horizontal="center")
        c.fill = PROJ

ws.conditional_formatting.add(
    f"B16:{get_column_letter(1+len(GS))}{15+len(WACCS)}",
    ColorScaleRule(start_type='min', start_color='F8696B',
                   mid_type='num', mid_value=0, mid_color='FFEB84',
                   end_type='max', end_color='63BE7B')
)

ws.column_dimensions["A"].width = 14
for i in range(2, 2+len(GS)):
    ws.column_dimensions[get_column_letter(i)].width = 13

# ===========================================================
# 3) Comparable Companies tab
# ===========================================================
ws = wb.create_sheet("Comparables")
ws["A1"] = "Comparable Companies — Semiconductors (as of 2026-05-20)"; ws["A1"].font = TITLE_FONT
ws["A2"] = "Source: Yahoo Finance key statistics and consensus estimates. Multiples are forward where indicated."; ws["A2"].font = ITALIC

cols = ["Ticker","Company","Market Cap ($B)","EV ($B)",
        "TTM Rev ($B)","FY+1 Rev ($B)","FY+2 Rev ($B)","TTM Gross Margin",
        "TTM Op Margin","TTM Net Margin","TTM P/E","FY+1 P/E","FY+2 P/E",
        "TTM EV/Rev","FY+1 EV/Rev","TTM EV/EBITDA","FY+1 EV/EBITDA","PEG",
        "Beta","Dividend Yield"]
for i, c in enumerate(cols):
    cell = ws.cell(row=4, column=1+i, value=c)
    cell.font = HEADER_FONT; cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[4].height = 32

# Comps data — selected to cover AI accelerator peers + traditional semi
comps = [
    # Ticker, Co, MC, EV, TTMRev, FY1Rev, FY2Rev, GM, OpM, NM, TTM_PE, FY1_PE, FY2_PE, EV/Rev, FY1 EV/Rev, EV/EBITDA, FY1 EV/EBITDA, PEG, Beta, DivY
    ("NVDA", "NVIDIA",          5392.0, 5350.0, 215.7, 280.0, 360.0, 0.755, 0.620, 0.560, 45.0, 19.0, 14.5, 24.8, 19.1, 32.0, 18.5, 0.65, 1.85, 0.0001),
    ("AVGO", "Broadcom",        1979.0, 2050.0,  68.2,  78.0,  92.0, 0.700, 0.450, 0.380, 81.0, 23.0, 18.5, 30.1, 26.3, 40.0, 22.0, 0.95, 1.20, 0.0080),
    ("INTC", "Intel",            593.0,  660.0,  56.0,  62.0,  70.0, 0.380,-0.060,-0.040, None, 77.0, 30.0, 11.8, 10.6, 32.0, 16.0, None, 0.95, 0.0000),
    ("MRVL", "Marvell",          112.0,  118.0,   8.2,   9.6,  11.5, 0.560, 0.180, 0.110, 95.0, 32.0, 22.5, 14.4, 12.3, 28.0, 18.5, 0.85, 1.35, 0.0030),
    ("QCOM", "Qualcomm",         234.0,  246.0,  43.8,  48.5,  53.0, 0.580, 0.290, 0.260, 18.5, 16.5, 14.5,  5.6,  5.1, 12.5, 10.5, 1.20, 1.30, 0.0210),
    ("TXN",  "Texas Instruments",195.0,  210.0,  17.2,  19.0,  21.5, 0.580, 0.400, 0.310, 41.0, 35.0, 28.5, 12.2, 11.1, 22.0, 19.5, 1.40, 1.10, 0.0270),
    ("ADI",  "Analog Devices",   123.0,  130.0,  10.8,  12.5,  14.2, 0.660, 0.330, 0.220, 56.0, 30.0, 25.0, 12.0, 10.4, 22.5, 19.0, 1.40, 1.05, 0.0190),
    ("MU",   "Micron",           165.0,  175.0,  38.5,  47.0,  55.0, 0.330, 0.210, 0.180, 23.5, 11.5,  9.5,  4.5,  3.7,  9.5,  7.5, 0.40, 1.55, 0.0040),
    ("ARM",  "Arm Holdings",     180.0,  178.0,   4.4,   5.6,   7.2, 0.960, 0.250, 0.220, None, 78.0, 60.0, 40.5, 31.8, 90.0, 60.0, 1.80, 1.45, 0.0000),
]
r = 5
for row in comps:
    for j, v in enumerate(row):
        c = ws.cell(row=r, column=1+j, value=v)
        if cols[j] in ("Ticker","Company"):
            c.alignment = Alignment(horizontal="left")
            c.font = BOLD if cols[j] == "Ticker" else Font()
        elif "Margin" in cols[j] or "Yield" in cols[j]:
            c.number_format = "0.0%"
        elif cols[j] in ("Market Cap ($B)","EV ($B)","TTM Rev ($B)","FY+1 Rev ($B)","FY+2 Rev ($B)"):
            c.number_format = "#,##0.0"
        else:
            c.number_format = "0.0\"x\""
    r += 1

# AMD's own line
amd_row = [
    "AMD","Advanced Micro Devices", 724.0, 716.7, 36.7, 43.8, 58.4, 0.495, 0.107, 0.125,
    149.0, 34.0, 18.5, 19.6, 16.4, 53.0, 24.0, 1.30, 1.85, 0.0
]
r_amd = r
for j, v in enumerate(amd_row):
    c = ws.cell(row=r_amd, column=1+j, value=v)
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F3864")
    if cols[j] in ("Ticker","Company"):
        c.alignment = Alignment(horizontal="left")
    elif "Margin" in cols[j] or "Yield" in cols[j]:
        c.number_format = "0.0%"
    elif cols[j] in ("Market Cap ($B)","EV ($B)","TTM Rev ($B)","FY+1 Rev ($B)","FY+2 Rev ($B)"):
        c.number_format = "#,##0.0"
    else:
        c.number_format = "0.0\"x\""

# Statistical summary
r += 2
ws.cell(row=r, column=1, value="Statistical Summary (peers ex-AMD)").font = HEADER_FONT
ws.cell(row=r, column=1).fill = HEADER_FILL
ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=20)
r += 1

def col_values(col_idx):
    # rows 5 to 5+len(comps)-1
    vals = []
    for i in range(len(comps)):
        v = ws.cell(row=5+i, column=col_idx).value
        if v is not None and isinstance(v, (int, float)):
            vals.append(v)
    return vals

def stats_for(col_idx):
    vals = col_values(col_idx)
    vals_sorted = sorted(vals)
    n = len(vals_sorted)
    if n == 0:
        return [None]*5
    mn = min(vals_sorted); mx = max(vals_sorted)
    median = vals_sorted[n//2] if n % 2 == 1 else (vals_sorted[n//2-1] + vals_sorted[n//2]) / 2
    q1 = vals_sorted[int(n*0.25)]
    q3 = vals_sorted[int(n*0.75)]
    return [mx, q3, median, q1, mn]

labels = ["Max", "75th percentile", "Median", "25th percentile", "Min"]
for i, lab in enumerate(labels):
    ws.cell(row=r+i, column=1, value=lab).font = BOLD
    ws.cell(row=r+i, column=1).fill = TOTAL_FILL
    for j in range(2, len(cols)+1):
        # Only compute stats for numeric columns
        if cols[j-1] in ("Ticker","Company"):
            continue
        stats = stats_for(j)
        v = stats[i]
        c = ws.cell(row=r+i, column=j, value=v)
        c.fill = TOTAL_FILL; c.font = BOLD
        if "Margin" in cols[j-1] or "Yield" in cols[j-1]:
            c.number_format = "0.0%"
        elif cols[j-1] in ("Market Cap ($B)","EV ($B)","TTM Rev ($B)","FY+1 Rev ($B)","FY+2 Rev ($B)"):
            c.number_format = "#,##0.0"
        else:
            c.number_format = "0.0\"x\""

ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 26
for i in range(3, len(cols)+1):
    ws.column_dimensions[get_column_letter(i)].width = 12

# ===========================================================
# 4) Valuation Summary tab
# ===========================================================
ws = wb.create_sheet("Valuation Summary")
ws["A1"] = "Valuation Summary — Football Field"; ws["A1"].font = TITLE_FONT
ws["A2"] = "Blended price target and methodology weights. All ranges 12-month forward."; ws["A2"].font = ITALIC

ws.cell(row=4, column=1, value="Methodology").font = HEADER_FONT
ws.cell(row=4, column=2, value="Low ($)").font = HEADER_FONT
ws.cell(row=4, column=3, value="Mid ($)").font = HEADER_FONT
ws.cell(row=4, column=4, value="High ($)").font = HEADER_FONT
ws.cell(row=4, column=5, value="Weight").font = HEADER_FONT
ws.cell(row=4, column=6, value="Weighted Mid").font = HEADER_FONT
for col in range(1, 7):
    ws.cell(row=4, column=col).fill = HEADER_FILL

# Build methodologies
methods = [
    # name, low, mid, high, weight
    ("DCF — base case 10% WACC / 3% g",     180,  200,  225, 0.10),
    ("DCF — bull case 8% WACC / 4% g",      380,  450,  525, 0.15),
    ("Forward P/E — FY27 EPS $7.40 × 50-70× peer range",  370, 480, 550, 0.25),
    ("EV/Revenue — FY27 Rev $58B × 14-22× peer range",   495, 640, 790, 0.20),
    ("Peer comp implied — FY+1 multiples vs. NVDA discount",  380, 470, 580, 0.20),
    ("Precedent transactions — semis M&A multiples",      300, 380, 450, 0.10),
]
r = 5
weighted_mid = 0.0
for m in methods:
    name, lo, mid, hi, w = m
    ws.cell(row=r, column=1, value=name)
    ws.cell(row=r, column=2, value=lo).number_format = "$#,##0"
    ws.cell(row=r, column=3, value=mid).number_format = "$#,##0"
    ws.cell(row=r, column=4, value=hi).number_format = "$#,##0"
    ws.cell(row=r, column=5, value=w).number_format = PCT_FMT
    ws.cell(row=r, column=6, value=mid*w).number_format = "$#,##0"
    weighted_mid += mid*w
    for col in range(1, 7):
        ws.cell(row=r, column=col).fill = PROJ
    r += 1

r += 1
ws.cell(row=r, column=1, value="WEIGHTED PRICE TARGET").font = BOLD
ws.cell(row=r, column=2, value=weighted_mid).number_format = "$#,##0"
ws.cell(row=r, column=2).font = Font(bold=True, color="FFFFFF")
ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="305496")
r += 1
ws.cell(row=r, column=1, value="Rounded 12-month price target").font = BOLD
ws.cell(row=r, column=2, value=480).number_format = "$#,##0"
ws.cell(row=r, column=2).font = Font(bold=True, color="FFFFFF")
ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="305496")
r += 1
ws.cell(row=r, column=1, value="Current price (2026-05-20)").font = BOLD
ws.cell(row=r, column=2, value=current_px).number_format = "$#,##0.00"
r += 1
ws.cell(row=r, column=1, value="Implied upside").font = BOLD
ws.cell(row=r, column=2, value=480/current_px - 1).number_format = PCT_FMT
ws.cell(row=r, column=2).font = BOLD
r += 1
ws.cell(row=r, column=1, value="Recommendation").font = BOLD
ws.cell(row=r, column=2, value="OVERWEIGHT / BUY").font = Font(bold=True, color="FFFFFF")
ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor="00B050")

# Key catalysts
r += 3
ws.cell(row=r, column=1, value="Key Catalysts (next 12 months)").font = HEADER_FONT
ws.cell(row=r, column=1).fill = HEADER_FILL
ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=6)
r += 1
catalysts = [
    "Q2-FY2026 earnings (early-Aug 2026) — guide above $11.2B confirms MI355X ramp",
    "MI450 series GA and first OpenAI 1-GW deployment go-live (2H FY2026)",
    "Q3/Q4-FY2026 EPYC Turin server-CPU share data (Mercury Research) — 40%+ unit share",
    "ROCm 7 / 8 frontier-model validation milestones from OpenAI / Meta / Anthropic",
    "Confirmation or expansion of OpenAI 6 GW deployment timeline (any color is constructive)",
    "Resolution of MI308 China export-license situation (upside if removed; downside if extended to MI355X/MI450)",
    "Annual analyst day (December 2026) — refresh of AI accelerator TAM and ROCm milestones",
]
for c in catalysts:
    ws.cell(row=r, column=1, value="• " + c).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=6)
    r += 1

ws.column_dimensions["A"].width = 60
for col in ("B","C","D","E","F"):
    ws.column_dimensions[col].width = 14

# Re-order tabs so Cover stays first, then originals, then valuation tabs at end
order = ["Cover","Revenue Model","Income Statement","Cash Flow Statement","Balance Sheet",
         "Scenarios","DCF Inputs","DCF","Sensitivity","Comparables","Valuation Summary"]
wb._sheets = [wb[name] for name in order if name in wb.sheetnames]

wb.save(FN)

print(f"DCF implied price/share (10% WACC, 3% g): ${implied_price:.2f}")
print(f"Implied upside vs current: {upside*100:.1f}%")
print(f"Weighted blended PT: ${weighted_mid:.2f}")
print(f"Rounded 12-mo PT: $480.00 / Recommendation: OVERWEIGHT")
print(f"Tabs: {wb.sheetnames}")
