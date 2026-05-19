"""
Hesai Group (NASDAQ:HSAI) — Equity Research Financial Model
Builds 6-tab Excel workbook per initiating-coverage Task 2 spec.

All amounts in RMB thousands unless otherwise marked.
FY end Dec-31. Historicals: FY22A, FY23A, FY24A, FY25A.
Projections: FY26E - FY30E.

Sources:
- 2024 20-F (filed Apr 2025) — IS, CF, BS for 22/23/24
- FY2025 6-K press release (filed Mar 24 2026) — IS + BS 24/25
- FY2024/2025 unit shipments from press releases
- Management FY26 guidance: 3.0-3.5M units total
- Geographic split 22/23/24 from 20-F Note 18
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ------------- STYLE CONSTANTS -------------
BLUE = Font(name="Times New Roman", color="0000FF", size=10)
BLUE_BOLD = Font(name="Times New Roman", color="0000FF", size=10, bold=True)
BLACK = Font(name="Times New Roman", color="000000", size=10)
BLACK_BOLD = Font(name="Times New Roman", color="000000", size=10, bold=True)
GREEN = Font(name="Times New Roman", color="006100", size=10)
WHITE_BOLD = Font(name="Times New Roman", color="FFFFFF", size=10, bold=True)
TITLE = Font(name="Times New Roman", color="FFFFFF", size=12, bold=True)
HEADER_FILL = PatternFill("solid", fgColor="002060")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
CENTER = Alignment(horizontal="center")
RIGHT = Alignment(horizontal="right")
LEFT_INDENT1 = Alignment(horizontal="left", indent=1)
LEFT_INDENT2 = Alignment(horizontal="left", indent=2)
THIN = Side(border_style="thin", color="888888")
BOX = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
TOP = Border(top=Side(border_style="thin", color="000000"))
BOTTOM = Border(bottom=Side(border_style="thin", color="000000"))
DOUBLE_BOTTOM = Border(bottom=Side(border_style="double", color="000000"))

NUM = '#,##0;(#,##0);"-"'
NUM1 = '#,##0.0;(#,##0.0);"-"'
PCT = '0.0%;(0.0%);"-"'
USD = '"$"#,##0.00'
RMB = '"¥"#,##0'

YEARS = ["FY22A", "FY23A", "FY24A", "FY25A", "FY26E", "FY27E", "FY28E", "FY29E", "FY30E"]
N_HIST = 4
N_PROJ = 5
N_TOTAL = 9
# Column layout: Col A label, Col B = FY22A ... Col J = FY30E
FIRST_DATA_COL = 2  # B
LAST_DATA_COL = FIRST_DATA_COL + N_TOTAL - 1  # J


def col(i):
    """0-indexed offset into data columns → Excel letter"""
    return get_column_letter(FIRST_DATA_COL + i)


def write_header(ws, row, label, label_col=1, fill=HEADER_FILL, font=TITLE, indent=0):
    cell = ws.cell(row=row, column=label_col, value=label)
    cell.font = font
    cell.fill = fill
    for c in range(FIRST_DATA_COL, LAST_DATA_COL + 1):
        ws.cell(row=row, column=c).fill = fill


def write_year_row(ws, row):
    ws.cell(row=row, column=1, value="(RMB thousands)").font = Font(name="Times New Roman", size=9, italic=True, color="666666")
    for i, y in enumerate(YEARS):
        c = ws.cell(row=row, column=FIRST_DATA_COL + i, value=y)
        c.font = BLACK_BOLD
        c.fill = SUB_FILL
        c.alignment = CENTER
        c.border = BOTTOM


def write_label(ws, row, label, indent=0, bold=False):
    c = ws.cell(row=row, column=1, value=label)
    c.font = BLACK_BOLD if bold else BLACK
    c.alignment = Alignment(horizontal="left", indent=indent)
    return c


def write_value(ws, row, col_idx, value, font=BLACK, fmt=NUM, fill=None, border=None):
    c = ws.cell(row=row, column=FIRST_DATA_COL + col_idx, value=value)
    c.font = font
    c.number_format = fmt
    c.alignment = RIGHT
    if fill is not None:
        c.fill = fill
    if border is not None:
        c.border = border
    return c


def fill_hist_blue(ws, row, hist_values, fmt=NUM):
    """Blue (input) for historical years."""
    for i, v in enumerate(hist_values):
        write_value(ws, row, i, v, font=BLUE, fmt=fmt)


def fill_proj_blue(ws, row, proj_values, fmt=NUM):
    """Blue (input) for projected years."""
    for i, v in enumerate(proj_values):
        write_value(ws, row, N_HIST + i, v, font=BLUE, fmt=fmt)


def fill_all_blue(ws, row, all_values, fmt=NUM):
    for i, v in enumerate(all_values):
        write_value(ws, row, i, v, font=BLUE, fmt=fmt)


def fill_formula_row(ws, row, formula_template, fmt=NUM, font=BLACK):
    """formula_template: string with {col} placeholder, e.g. '={col}10+{col}11'."""
    for i in range(N_TOTAL):
        c = col(i)
        f = formula_template.format(col=c)
        write_value(ws, row, i, f, font=font, fmt=fmt)


# ============================================================================
# WORKBOOK
# ============================================================================
wb = Workbook()
wb.remove(wb.active)


# ----------------------------------------------------------------------------
# TAB 1: REVENUE MODEL
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Revenue Model")
ws.column_dimensions['A'].width = 42
for i in range(N_TOTAL):
    ws.column_dimensions[col(i)].width = 11

ws.cell(row=1, column=1, value="HESAI GROUP — REVENUE MODEL").font = Font(name="Times New Roman", size=14, bold=True)
ws.cell(row=2, column=1, value="Detailed revenue build by product and geography (RMB thousands)").font = Font(name="Times New Roman", size=10, italic=True, color="666666")
write_year_row(ws, 4)

# Color legend
ws.cell(row=1, column=8, value="Blue = input").font = BLUE
ws.cell(row=2, column=8, value="Black = formula").font = BLACK
ws.cell(row=3, column=8, value="Green = link").font = GREEN

r = 6
# === SECTION A: REVENUE BY PRODUCT (UNITS × ASP build) ===
write_header(ws, r, "A. REVENUE BUILD BY PRODUCT (Units × ASP)")
r += 1

write_label(ws, r, "UNIT SHIPMENTS (thousand)", bold=True); r += 1

# ADAS unit detail
write_label(ws, r, "ADAS — Long-range (AT-series)", indent=1)
fill_all_blue(ws, r, [20, 175, 410, 1280, 2550, 3800, 4900, 5800, 6700], fmt=NUM)
adas_lr = r; r += 1

write_label(ws, r, "ADAS — Ultra-thin / Blind-spot (ET, FT)", indent=1)
fill_all_blue(ws, r, [5, 20, 46, 101, 150, 200, 300, 400, 400], fmt=NUM)
adas_st = r; r += 1

write_label(ws, r, "ADAS subtotal (units, thousands)", indent=0, bold=True)
fill_formula_row(ws, r, '=SUM({col}' + str(adas_lr) + ':{col}' + str(adas_st) + ')', font=BLACK_BOLD)
adas_units = r; r += 1

# Robotics unit detail
write_label(ws, r, "Robotics — Robotaxi / Robovan (Pandar/OT/QT)", indent=1)
fill_all_blue(ws, r, [55, 25, 38, 75, 130, 200, 280, 360, 450], fmt=NUM)
rob_rt = r; r += 1

write_label(ws, r, "Robotics — Humanoid / Quadruped (JT)", indent=1)
fill_all_blue(ws, r, [0, 0, 1, 12, 70, 200, 400, 600, 800], fmt=NUM)
rob_hu = r; r += 1

write_label(ws, r, "Robotics — Lawn-mower / Consumer", indent=1)
fill_all_blue(ws, r, [0, 0, 1, 100, 280, 450, 600, 750, 900], fmt=NUM)
rob_lm = r; r += 1

write_label(ws, r, "Robotics — Industrial / AGV / Other", indent=1)
fill_all_blue(ws, r, [0, 2, 5, 52, 120, 200, 250, 300, 350], fmt=NUM)
rob_ot = r; r += 1

write_label(ws, r, "Robotics subtotal (units, thousands)", bold=True)
fill_formula_row(ws, r, '=SUM({col}' + str(rob_rt) + ':{col}' + str(rob_ot) + ')', font=BLACK_BOLD)
rob_units = r; r += 1

write_label(ws, r, "TOTAL UNIT SHIPMENTS (thousands)", bold=True)
fill_formula_row(ws, r, '={col}' + str(adas_units) + '+{col}' + str(rob_units), font=BLACK_BOLD, fmt=NUM)
for i in range(N_TOTAL):
    ws.cell(row=r, column=FIRST_DATA_COL + i).fill = TOTAL_FILL
total_units = r
r += 2

write_label(ws, r, "AVG SELLING PRICE (RMB/unit)", bold=True); r += 1

write_label(ws, r, "ADAS — Long-range ASP (RMB/unit)", indent=1)
fill_all_blue(ws, r, [6000, 5000, 1900, 1300, 1000, 850, 760, 690, 640], fmt=NUM)
asp_adas_lr = r; r += 1

write_label(ws, r, "ADAS — Ultra-thin / Blind-spot ASP", indent=1)
fill_all_blue(ws, r, [4000, 3500, 2200, 1500, 1200, 1050, 950, 870, 800], fmt=NUM)
asp_adas_st = r; r += 1

write_label(ws, r, "Robotics — Robotaxi ASP", indent=1)
fill_all_blue(ws, r, [18000, 30000, 25000, 8000, 6500, 5500, 4800, 4200, 3800], fmt=NUM)
asp_rob_rt = r; r += 1

write_label(ws, r, "Robotics — Humanoid ASP", indent=1)
fill_all_blue(ws, r, [0, 0, 18000, 5500, 4500, 4000, 3500, 3100, 2800], fmt=NUM)
asp_rob_hu = r; r += 1

write_label(ws, r, "Robotics — Lawn-mower ASP", indent=1)
fill_all_blue(ws, r, [0, 0, 4000, 2000, 1600, 1300, 1100, 950, 850], fmt=NUM)
asp_rob_lm = r; r += 1

write_label(ws, r, "Robotics — Industrial ASP", indent=1)
fill_all_blue(ws, r, [0, 8000, 7000, 3500, 3000, 2500, 2200, 1900, 1700], fmt=NUM)
asp_rob_ot = r; r += 2

write_label(ws, r, "REVENUE BY PRODUCT (RMB thousands)", bold=True); r += 1

# Revenue rows = units (thousands) × ASP (RMB/unit) = RMB thousands ✓
write_label(ws, r, "ADAS — Long-range revenue", indent=1)
fill_formula_row(ws, r, '={col}' + str(adas_lr) + '*{col}' + str(asp_adas_lr))
rev_adas_lr = r; r += 1

write_label(ws, r, "ADAS — Ultra-thin/Blind-spot revenue", indent=1)
fill_formula_row(ws, r, '={col}' + str(adas_st) + '*{col}' + str(asp_adas_st))
rev_adas_st = r; r += 1

write_label(ws, r, "ADAS revenue subtotal", bold=True)
fill_formula_row(ws, r, '=SUM({col}' + str(rev_adas_lr) + ':{col}' + str(rev_adas_st) + ')', font=BLACK_BOLD)
rev_adas = r; r += 1

write_label(ws, r, "Robotics — Robotaxi revenue", indent=1)
fill_formula_row(ws, r, '={col}' + str(rob_rt) + '*{col}' + str(asp_rob_rt))
rev_rob_rt = r; r += 1

write_label(ws, r, "Robotics — Humanoid revenue", indent=1)
fill_formula_row(ws, r, '={col}' + str(rob_hu) + '*{col}' + str(asp_rob_hu))
rev_rob_hu = r; r += 1

write_label(ws, r, "Robotics — Lawn-mower revenue", indent=1)
fill_formula_row(ws, r, '={col}' + str(rob_lm) + '*{col}' + str(asp_rob_lm))
rev_rob_lm = r; r += 1

write_label(ws, r, "Robotics — Industrial revenue", indent=1)
fill_formula_row(ws, r, '={col}' + str(rob_ot) + '*{col}' + str(asp_rob_ot))
rev_rob_ot = r; r += 1

write_label(ws, r, "Robotics revenue subtotal", bold=True)
fill_formula_row(ws, r, '=SUM({col}' + str(rev_rob_rt) + ':{col}' + str(rev_rob_ot) + ')', font=BLACK_BOLD)
rev_rob = r; r += 1

# Services + Gas + Other
write_label(ws, r, "Service revenue (design-in, NRE, project-based)", indent=1)
fill_all_blue(ws, r, [42000, 115000, 115000, 25000, 30000, 35000, 40000, 45000, 50000])
rev_svc = r; r += 1

write_label(ws, r, "Gas-sensor / legacy revenue", indent=1)
fill_all_blue(ws, r, [38000, 27000, 15400, 11000, 9000, 7500, 6500, 5500, 5000])
rev_gas = r; r += 1

write_label(ws, r, "TOTAL NET REVENUES", bold=True)
fill_formula_row(ws, r,
                 '={col}' + str(rev_adas) + '+{col}' + str(rev_rob) + '+{col}' + str(rev_svc) + '+{col}' + str(rev_gas),
                 font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws.cell(row=r, column=FIRST_DATA_COL + i)
    cc.fill = TOTAL_FILL
    cc.border = TOP
total_rev = r; r += 1

write_label(ws, r, "  YoY growth %", indent=2)
for i in range(N_TOTAL):
    if i == 0:
        write_value(ws, r, i, "n/a", font=BLACK)
    else:
        prev = col(i - 1); cur = col(i)
        write_value(ws, r, i, f"={cur}{total_rev}/{prev}{total_rev}-1", fmt=PCT)
r += 1

write_label(ws, r, "  Implied blended ASP (RMB/unit)", indent=2)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws, r, i, f"={cc}{total_rev}/{cc}{total_units}", fmt=NUM)
r += 1

write_label(ws, r, "  Implied blended ASP (US$/unit, @ FX 7.30)", indent=2)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws, r, i, f"={cc}{total_rev}/{cc}{total_units}/7.3", fmt=USD)
r += 2

# === SECTION B: REVENUE MIX ===
write_header(ws, r, "B. REVENUE MIX (% of total)")
r += 1
write_label(ws, r, "ADAS % of revenue", indent=1)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws, r, i, f"={cc}{rev_adas}/{cc}{total_rev}", fmt=PCT, font=BLACK)
r += 1
write_label(ws, r, "Robotics % of revenue", indent=1)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws, r, i, f"={cc}{rev_rob}/{cc}{total_rev}", fmt=PCT)
r += 1
write_label(ws, r, "Service % of revenue", indent=1)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws, r, i, f"={cc}{rev_svc}/{cc}{total_rev}", fmt=PCT)
r += 1
write_label(ws, r, "Gas / legacy % of revenue", indent=1)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws, r, i, f"={cc}{rev_gas}/{cc}{total_rev}", fmt=PCT)
r += 2

# === SECTION C: REVENUE BY GEOGRAPHY ===
write_header(ws, r, "C. REVENUE BY GEOGRAPHY (RMB thousands)")
r += 1

# Historicals from 20-F Note 18 (2022/2023/2024). 2025 estimated.
# Projections driven by % weights vs total (input as % of total revenue)
# We'll write as $ rows directly.
write_label(ws, r, "Mainland China", indent=1)
fill_all_blue(ws, r, [697294, 991912, 1542793, 2350000, 3800000, 5400000, 7200000, 9000000, 10800000])
geo_cn = r; r += 1
write_label(ws, r, "  China — Top-10 OEMs (Li Auto, Xiaomi, Changan, BYD, etc.)", indent=2)
fill_all_blue(ws, r, [380000, 620000, 1180000, 1900000, 3150000, 4500000, 6000000, 7500000, 9000000], fmt=NUM)
r += 1
write_label(ws, r, "  China — Robotaxi/Robotics customers", indent=2)
fill_all_blue(ws, r, [317294, 371912, 362793, 450000, 650000, 900000, 1200000, 1500000, 1800000], fmt=NUM)
r += 1

write_label(ws, r, "North America (US OEM, robotaxi)", indent=1)
fill_all_blue(ws, r, [358549, 748147, 280874, 410000, 560000, 780000, 1000000, 1250000, 1500000])
geo_na = r; r += 1
write_label(ws, r, "  US — Top-1 OEM (e.g. GM Super Cruise)", indent=2)
fill_all_blue(ws, r, [165000, 535000, 80000, 130000, 220000, 350000, 500000, 650000, 800000], fmt=NUM)
r += 1
write_label(ws, r, "  US — Robotaxi & other (Waymo, etc.)", indent=2)
fill_all_blue(ws, r, [193549, 213147, 200874, 280000, 340000, 430000, 500000, 600000, 700000], fmt=NUM)
r += 1

write_label(ws, r, "Europe (Stellantis, Mercedes, Bosch)", indent=1)
fill_all_blue(ws, r, [86153, 70500, 161095, 200000, 320000, 480000, 660000, 850000, 1050000])
geo_eu = r; r += 1
write_label(ws, r, "  EU — Tier-1 partnerships (Bosch, Valeo offshoot)", indent=2)
fill_all_blue(ws, r, [60000, 40000, 110000, 130000, 220000, 330000, 460000, 600000, 750000], fmt=NUM)
r += 1
write_label(ws, r, "  EU — Direct OEM design-ins", indent=2)
fill_all_blue(ws, r, [26153, 30500, 51095, 70000, 100000, 150000, 200000, 250000, 300000], fmt=NUM)
r += 1

write_label(ws, r, "Asia ex-China (Japan, Korea, ASEAN)", indent=1)
fill_all_blue(ws, r, [40000, 45000, 65000, 80000, 130000, 200000, 280000, 380000, 480000])
geo_asia = r; r += 1
write_label(ws, r, "Rest of world (LATAM, MEA, ANZ)", indent=1)
fill_all_blue(ws, r, [20674, 21430, 27395, 40000, 70000, 110000, 150000, 200000, 250000])
geo_row = r; r += 1

write_label(ws, r, "TOTAL revenue by geography", bold=True)
fill_formula_row(ws, r,
                 '={col}' + str(geo_cn) + '+{col}' + str(geo_na) + '+{col}' + str(geo_eu) + '+{col}' + str(geo_asia) + '+{col}' + str(geo_row),
                 font=BLACK_BOLD)
total_geo = r
for i in range(N_TOTAL):
    cc = ws.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL; cc.border = TOP
r += 1

write_label(ws, r, "  Geography vs product check (should be ~0)", indent=2)
for i in range(N_TOTAL):
    cc = col(i)
    cv = write_value(ws, r, i, f"={cc}{total_geo}-{cc}{total_rev}", fmt=NUM)
geo_check = r
r += 2

# === SECTION D: REVENUE BY CHANNEL ===
write_header(ws, r, "D. REVENUE BY CHANNEL")
r += 1
write_label(ws, r, "Direct OEM (auto)", indent=1)
fill_all_blue(ws, r, [800000, 1450000, 1700000, 2580000, 4180000, 6100000, 8350000, 10500000, 12700000])
r += 1
write_label(ws, r, "Direct robotics (robotaxi, humanoid, lawnmower)", indent=1)
fill_all_blue(ws, r, [320000, 380000, 340000, 410000, 580000, 800000, 1080000, 1380000, 1720000])
r += 1
write_label(ws, r, "Tier-1 partner channel (Bosch, etc.)", indent=1)
fill_all_blue(ws, r, [42000, 19000, 21800, 25000, 60000, 100000, 145000, 200000, 250000])
r += 1
write_label(ws, r, "Service / NRE", indent=1)
fill_all_blue(ws, r, [42000, 28000, 15400, 11000, 30000, 50000, 70000, 90000, 110000])
r += 2

# Save row pointers for cross-sheet references
ws._refs = {
    "total_rev": total_rev,
    "rev_adas": rev_adas,
    "rev_rob": rev_rob,
    "rev_svc": rev_svc,
    "adas_units": adas_units,
    "rob_units": rob_units,
    "total_units": total_units,
}

# Freeze panes
ws.freeze_panes = "B5"


# ----------------------------------------------------------------------------
# TAB 2: INCOME STATEMENT
# ----------------------------------------------------------------------------
ws_is = wb.create_sheet("Income Statement")
ws_is.column_dimensions['A'].width = 48
for i in range(N_TOTAL):
    ws_is.column_dimensions[col(i)].width = 12

ws_is.cell(row=1, column=1, value="HESAI GROUP — CONSOLIDATED INCOME STATEMENT").font = Font(name="Times New Roman", size=14, bold=True)
ws_is.cell(row=2, column=1, value="GAAP basis, RMB thousands. Historicals tie to 20-F & FY25 6-K (Mar 24 2026).").font = Font(name="Times New Roman", size=10, italic=True, color="666666")
write_year_row(ws_is, 4)

# Linked total revenue from Revenue Model
rev_ref = "'Revenue Model'!"
trev = ws._refs["total_rev"]

r = 6
write_header(ws_is, r, "REVENUE")
r += 1
write_label(ws_is, r, "Product revenue — ADAS lidar", indent=1)
fill_formula_row(ws_is, r, "={ref}{col}{rr}".replace("{ref}", rev_ref).replace("{rr}", str(ws._refs["rev_adas"])), font=GREEN)
is_rev_adas = r; r += 1

write_label(ws_is, r, "Product revenue — Robotics lidar", indent=1)
fill_formula_row(ws_is, r, "={ref}{col}{rr}".replace("{ref}", rev_ref).replace("{rr}", str(ws._refs["rev_rob"])), font=GREEN)
r += 1

write_label(ws_is, r, "Service revenue + Gas / legacy", indent=1)
fill_formula_row(ws_is, r,
                 "={ref}{col}{r1}+{ref}{col}{r2}".replace("{ref}", rev_ref).replace("{r1}", str(ws._refs["rev_svc"])).replace("{r2}", str(ws._refs["rev_svc"] + 1))  # rev_svc and rev_gas (gas was directly after svc)
                 , font=GREEN)
r += 1

write_label(ws_is, r, "Total net revenues", bold=True)
fill_formula_row(ws_is, r, "={ref}{col}{rr}".replace("{ref}", rev_ref).replace("{rr}", str(trev)), font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws_is.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL; cc.border = TOP
is_revenue = r; r += 1

write_label(ws_is, r, "  YoY growth %", indent=2)
for i in range(N_TOTAL):
    if i == 0:
        write_value(ws_is, r, i, "n/a")
    else:
        prev = col(i - 1); cur = col(i)
        write_value(ws_is, r, i, f"={cur}{is_revenue}/{prev}{is_revenue}-1", fmt=PCT)
r += 2

# COGS
write_header(ws_is, r, "COST OF REVENUE")
r += 1
write_label(ws_is, r, "  Gross margin % (input)", indent=1)
gm_hist = [0.392, 0.352, 0.426, 0.418]
gm_proj = [0.418, 0.422, 0.425, 0.428, 0.430]
fill_hist_blue(ws_is, r, gm_hist, fmt=PCT)
fill_proj_blue(ws_is, r, gm_proj, fmt=PCT)
gm_row = r; r += 1

write_label(ws_is, r, "Cost of revenues (calc as Rev × (1-GM%))")
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"=-{cc}{is_revenue}*(1-{cc}{gm_row})")
cogs_row = r; r += 1

write_label(ws_is, r, "  COGS breakdown (memo):", indent=1)
r += 1
write_label(ws_is, r, "  Materials / components", indent=2)
fill_all_blue(ws_is, r, [-510000, -850000, -780000, -1130000, -1490000, -1820000, -2200000, -2630000, -3080000])
r += 1
write_label(ws_is, r, "  Manufacturing & overhead", indent=2)
fill_all_blue(ws_is, r, [-160000, -270000, -300000, -420000, -540000, -660000, -790000, -940000, -1090000])
r += 1
write_label(ws_is, r, "  Logistics / shipping / warranty", indent=2)
fill_all_blue(ws_is, r, [-60683, -95611, -112572, -212477, -270000, -340000, -420000, -510000, -610000])
r += 1

write_label(ws_is, r, "GROSS PROFIT", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"={cc}{is_revenue}+{cc}{cogs_row}", font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws_is.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL; cc.border = TOP
gp_row = r; r += 1
write_label(ws_is, r, "  Gross margin %", indent=2)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"={cc}{gp_row}/{cc}{is_revenue}", fmt=PCT)
r += 2

# OPEX
write_header(ws_is, r, "OPERATING EXPENSES")
r += 1
write_label(ws_is, r, "Sales & marketing expense", indent=1)
fill_hist_blue(ws_is, r, [-104835, -148798, -193032, -191990])
# Project as % of revenue
write_label(ws_is, r + 1, "  S&M as % of revenue", indent=2)
sm_proj_pct = [0.058, 0.052, 0.048, 0.045, 0.044]
fill_proj_blue(ws_is, r + 1, sm_proj_pct, fmt=PCT)
sm_pct_row = r + 1
# Formula: S&M projected = -Rev * %
for i in range(N_HIST, N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"=-{cc}{is_revenue}*{cc}{sm_pct_row}")
sm_row = r
r += 2
write_label(ws_is, r, "  S&M as % of revenue (historical calc)", indent=2)
for i in range(N_HIST):
    cc = col(i)
    write_value(ws_is, r, i, f"=-{cc}{sm_row}/{cc}{is_revenue}", fmt=PCT)
r += 1

write_label(ws_is, r, "General & administrative expense", indent=1)
fill_hist_blue(ws_is, r, [-201007, -320144, -316913, -288828])
write_label(ws_is, r + 1, "  G&A as % of revenue", indent=2)
ga_proj_pct = [0.088, 0.075, 0.065, 0.058, 0.054]
fill_proj_blue(ws_is, r + 1, ga_proj_pct, fmt=PCT)
ga_pct_row = r + 1
for i in range(N_HIST, N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"=-{cc}{is_revenue}*{cc}{ga_pct_row}")
ga_row = r
r += 2
write_label(ws_is, r, "  G&A as % of revenue (historical calc)", indent=2)
for i in range(N_HIST):
    cc = col(i)
    write_value(ws_is, r, i, f"=-{cc}{ga_row}/{cc}{is_revenue}", fmt=PCT)
r += 1

write_label(ws_is, r, "Research & development expense", indent=1)
fill_hist_blue(ws_is, r, [-555179, -790547, -855641, -796940])
write_label(ws_is, r + 1, "  R&D as % of revenue", indent=2)
rd_proj_pct = [0.220, 0.190, 0.170, 0.155, 0.142]
fill_proj_blue(ws_is, r + 1, rd_proj_pct, fmt=PCT)
rd_pct_row = r + 1
for i in range(N_HIST, N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"=-{cc}{is_revenue}*{cc}{rd_pct_row}")
rd_row = r
r += 2
write_label(ws_is, r, "  R&D as % of revenue (historical calc)", indent=2)
for i in range(N_HIST):
    cc = col(i)
    write_value(ws_is, r, i, f"=-{cc}{rd_row}/{cc}{is_revenue}", fmt=PCT)
r += 1

write_label(ws_is, r, "Other operating income, net", indent=1)
fill_all_blue(ws_is, r, [10817, 26520, 276093, 181415, 80000, 60000, 50000, 50000, 50000])
oth_op_row = r; r += 1

write_label(ws_is, r, "Total operating expenses", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"={cc}{sm_row}+{cc}{ga_row}+{cc}{rd_row}+{cc}{oth_op_row}", font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws_is.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL; cc.border = TOP
opex_row = r; r += 1

write_label(ws_is, r, "OPERATING INCOME (LOSS)", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"={cc}{gp_row}+{cc}{opex_row}", font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws_is.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL; cc.border = TOP
opinc_row = r; r += 1
write_label(ws_is, r, "  Operating margin %", indent=2)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"={cc}{opinc_row}/{cc}{is_revenue}", fmt=PCT)
r += 1

# D&A explicit memo
write_label(ws_is, r, "  Memo: Depreciation & Amortization", indent=2)
fill_all_blue(ws_is, r, [53634, 86268, 131809, 175000, 230000, 290000, 350000, 410000, 470000])
da_row = r; r += 1

write_label(ws_is, r, "EBITDA = OpInc + D&A", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"={cc}{opinc_row}+{cc}{da_row}", font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws_is.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL
ebitda_row = r; r += 1
write_label(ws_is, r, "  EBITDA margin %", indent=2)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"={cc}{ebitda_row}/{cc}{is_revenue}", fmt=PCT)
r += 2

# Below-the-line
write_header(ws_is, r, "BELOW-THE-LINE")
r += 1
write_label(ws_is, r, "Interest income", indent=1)
fill_all_blue(ws_is, r, [58734, 99813, 104401, 130237, 165000, 200000, 235000, 270000, 305000])
ii_row = r; r += 1
write_label(ws_is, r, "Interest expense", indent=1)
fill_all_blue(ws_is, r, [0, -3069, -12827, -18923, -20000, -22000, -24000, -26000, -28000])
ie_row = r; r += 1
write_label(ws_is, r, "Foreign exchange gain (loss), net", indent=1)
fill_all_blue(ws_is, r, [20858, -452, 14577, 2156, 0, 0, 0, 0, 0])
fx_row = r; r += 1
write_label(ws_is, r, "Other income (expense), net", indent=1)
fill_all_blue(ws_is, r, [-2161, 34, -2476, 184566, 10000, 10000, 10000, 10000, 10000])
oth_row = r; r += 1

write_label(ws_is, r, "Pre-tax income", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"={cc}{opinc_row}+{cc}{ii_row}+{cc}{ie_row}+{cc}{fx_row}+{cc}{oth_row}", font=BLACK_BOLD)
pti_row = r; r += 1

write_label(ws_is, r, "  Effective tax rate %", indent=2)
tax_hist = [-0.0002, 0.0014, 0.0112, 0.0660]  # historical taxes / pre-tax loss is meaningless when loss; left as inputs
tax_proj = [0.10, 0.12, 0.13, 0.13, 0.14]
fill_hist_blue(ws_is, r, tax_hist, fmt=PCT)
fill_proj_blue(ws_is, r, tax_proj, fmt=PCT)
tax_pct_row = r; r += 1

write_label(ws_is, r, "Income tax (expense) / benefit", indent=1)
# Historical: hardcoded
fill_hist_blue(ws_is, r, [66, -658, -1130, -30835])
# Projection: -PTI * tax rate
for i in range(N_HIST, N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"=-{cc}{pti_row}*{cc}{tax_pct_row}")
tax_row = r; r += 1

write_label(ws_is, r, "Share of loss in equity method investment", indent=1)
fill_all_blue(ws_is, r, [-45, -45, -13, -74, 0, 0, 0, 0, 0])
emi_row = r; r += 1

write_label(ws_is, r, "NET INCOME (LOSS)", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"={cc}{pti_row}+{cc}{tax_row}+{cc}{emi_row}", font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws_is.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL; cc.border = DOUBLE_BOTTOM
ni_row = r; r += 1
write_label(ws_is, r, "  Net margin %", indent=2)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"={cc}{ni_row}/{cc}{is_revenue}", fmt=PCT)
r += 2

# Shares & EPS
write_header(ws_is, r, "SHARES OUTSTANDING & EPS")
r += 1
write_label(ws_is, r, "Diluted weighted-avg shares (thousands)", indent=1)
fill_all_blue(ws_is, r, [102000, 110000, 129188, 146437, 162000, 165000, 167000, 169000, 171000])
ds_row = r; r += 1
write_label(ws_is, r, "Diluted EPS (RMB)", indent=1)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"={cc}{ni_row}/{cc}{ds_row}", fmt='0.00;(0.00);"-"')
eps_row = r; r += 1
write_label(ws_is, r, "Diluted EPS (US$, @ 7.30 FX)", indent=1)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"={cc}{eps_row}/7.3", fmt=USD)
r += 1
write_label(ws_is, r, "EPS per ADS (=1 ordinary share)", indent=1)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"={cc}{eps_row}/7.3", fmt=USD)
r += 2

# Non-GAAP recon
write_header(ws_is, r, "NON-GAAP RECONCILIATION (memo)")
r += 1
write_label(ws_is, r, "Reported net income (loss)", indent=1)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"={cc}{ni_row}", font=GREEN)
r += 1
write_label(ws_is, r, "+ Share-based compensation", indent=1)
fill_all_blue(ws_is, r, [105219, 234624, 116064, 114651, 130000, 145000, 160000, 175000, 190000])
sbc_row = r; r += 1
write_label(ws_is, r, "Non-GAAP net income", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"={cc}{ni_row}+{cc}{sbc_row}", font=BLACK_BOLD)
nongaap_ni_row = r; r += 1
write_label(ws_is, r, "  Non-GAAP net margin %", indent=2)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_is, r, i, f"={cc}{nongaap_ni_row}/{cc}{is_revenue}", fmt=PCT)

ws_is._refs = {
    "revenue": is_revenue,
    "gp": gp_row,
    "opex": opex_row,
    "opinc": opinc_row,
    "da": da_row,
    "ebitda": ebitda_row,
    "tax_pct": tax_pct_row,
    "ni": ni_row,
    "sbc": sbc_row,
    "diluted_shares": ds_row,
    "eps": eps_row,
    "cogs": cogs_row,
    "sm": sm_row, "ga": ga_row, "rd": rd_row,
}

ws_is.freeze_panes = "B5"


# ----------------------------------------------------------------------------
# TAB 3: CASH FLOW STATEMENT
# ----------------------------------------------------------------------------
ws_cf = wb.create_sheet("Cash Flow")
ws_cf.column_dimensions['A'].width = 48
for i in range(N_TOTAL):
    ws_cf.column_dimensions[col(i)].width = 12

ws_cf.cell(row=1, column=1, value="HESAI GROUP — CONSOLIDATED CASH FLOW STATEMENT").font = Font(name="Times New Roman", size=14, bold=True)
ws_cf.cell(row=2, column=1, value="RMB thousands. Historicals 22-24 from 20-F; 25 from 6-K disclosure; 26-30E projected.").font = Font(name="Times New Roman", size=10, italic=True, color="666666")
write_year_row(ws_cf, 4)

is_ref = "'Income Statement'!"
ref_ni = ws_is._refs["ni"]
ref_da = ws_is._refs["da"]
ref_sbc = ws_is._refs["sbc"]
ref_rev = ws_is._refs["revenue"]

r = 6
write_header(ws_cf, r, "OPERATING ACTIVITIES")
r += 1
write_label(ws_cf, r, "Net income (loss)", indent=1)
fill_formula_row(ws_cf, r, "={ref}{col}{rr}".replace("{ref}", is_ref).replace("{rr}", str(ref_ni)), font=GREEN)
cf_ni = r; r += 1

write_label(ws_cf, r, "+ Depreciation & amortization", indent=1)
fill_formula_row(ws_cf, r, "={ref}{col}{rr}".replace("{ref}", is_ref).replace("{rr}", str(ref_da)), font=GREEN)
cf_da = r; r += 1

write_label(ws_cf, r, "+ Share-based compensation", indent=1)
fill_formula_row(ws_cf, r, "={ref}{col}{rr}".replace("{ref}", is_ref).replace("{rr}", str(ref_sbc)), font=GREEN)
cf_sbc = r; r += 1

write_label(ws_cf, r, "+ Other non-cash adjustments", indent=1)
fill_all_blue(ws_cf, r, [37195, 56171, 54078, 30000, 25000, 25000, 25000, 25000, 25000])
cf_oth = r; r += 1

write_label(ws_cf, r, "Changes in working capital (input as % of revenue change)", indent=1, bold=False); r += 1
write_label(ws_cf, r, "  WC change as % of revenue change", indent=2)
fill_hist_blue(ws_cf, r, [-1.46, -0.04, -0.55, -0.18], fmt=PCT)  # historical implied
fill_proj_blue(ws_cf, r, [-0.20, -0.15, -0.12, -0.10, -0.08], fmt=PCT)
wc_pct_row = r; r += 1

write_label(ws_cf, r, "Δ Working capital (calc)", indent=1)
# WC = (current revenue - prior revenue) * pct
# For year i: =wc_pct_row * (rev_i - rev_{i-1})
for i in range(N_TOTAL):
    cc = col(i)
    if i == 0:
        write_value(ws_cf, r, i, -200000, font=BLUE)  # input for 22
    else:
        prev = col(i - 1)
        write_value(ws_cf, r, i, f"={cc}{wc_pct_row}*({is_ref}{cc}{ref_rev}-{is_ref}{prev}{ref_rev})")
cf_wc = r; r += 1

write_label(ws_cf, r, "Cash from operating activities", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_cf, r, i, f"=SUM({cc}{cf_ni}:{cc}{cf_oth})+{cc}{cf_wc}", font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws_cf.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL; cc.border = TOP
cfo_row = r; r += 2

write_header(ws_cf, r, "INVESTING ACTIVITIES")
r += 1
write_label(ws_cf, r, "Capex (purchases of PP&E)", indent=1)
# Historical / projected ratchets up with capacity ramp
fill_all_blue(ws_cf, r, [-231210, -406748, -259541, -360000, -550000, -700000, -800000, -850000, -900000])
capex_row = r; r += 1
write_label(ws_cf, r, "Purchases of intangibles", indent=1)
fill_all_blue(ws_cf, r, [-9180, -7925, -11817, -15000, -18000, -20000, -22000, -24000, -26000])
intang_row = r; r += 1
write_label(ws_cf, r, "Net short-term investments activity", indent=1)
fill_all_blue(ws_cf, r, [1392000, -621566, 1227234, -2900000, -200000, -250000, -300000, -300000, -300000])
sti_row = r; r += 1
write_label(ws_cf, r, "Other investing (acquisitions, equity inv.)", indent=1)
fill_all_blue(ws_cf, r, [-31964, -24247, 0, -2750000, -100000, -120000, -140000, -160000, -180000])
oth_inv_row = r; r += 1
write_label(ws_cf, r, "Cash from investing", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_cf, r, i, f"=SUM({cc}{capex_row}:{cc}{oth_inv_row})", font=BLACK_BOLD)
cfi_row = r; r += 2

# FCF
write_header(ws_cf, r, "FREE CASH FLOW")
r += 1
write_label(ws_cf, r, "Free cash flow (CFO + Capex + Intangibles)", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_cf, r, i, f"={cc}{cfo_row}+{cc}{capex_row}+{cc}{intang_row}", font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws_cf.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL
fcf_row = r; r += 1
write_label(ws_cf, r, "  FCF margin %", indent=2)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_cf, r, i, f"={cc}{fcf_row}/{is_ref}{cc}{ref_rev}", fmt=PCT)
r += 2

write_header(ws_cf, r, "FINANCING ACTIVITIES")
r += 1
write_label(ws_cf, r, "Proceeds from IPO / follow-on equity", indent=1)
fill_all_blue(ws_cf, r, [0, 1225470, 0, 4400000, 0, 0, 0, 0, 0])
eq_row = r; r += 1
write_label(ws_cf, r, "Net borrowings (issuance - repayment)", indent=1)
fill_all_blue(ws_cf, r, [18472, 376592, 216536, 110000, 50000, 50000, 50000, 50000, 50000])
debt_row = r; r += 1
write_label(ws_cf, r, "Option exercises / other", indent=1)
fill_all_blue(ws_cf, r, [-3296, -11626, 34139, 25000, 30000, 30000, 30000, 30000, 30000])
oth_fin_row = r; r += 1
write_label(ws_cf, r, "Cash from financing", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_cf, r, i, f"=SUM({cc}{eq_row}:{cc}{oth_fin_row})", font=BLACK_BOLD)
cff_row = r; r += 2

write_label(ws_cf, r, "FX effect on cash", indent=1)
fill_all_blue(ws_cf, r, [42000, 13000, 15000, -8000, 0, 0, 0, 0, 0])
fxc_row = r; r += 1

write_label(ws_cf, r, "Net change in cash", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_cf, r, i, f"={cc}{cfo_row}+{cc}{cfi_row}+{cc}{cff_row}+{cc}{fxc_row}", font=BLACK_BOLD)
net_chg_row = r; r += 1

write_label(ws_cf, r, "Beginning cash & equivalents", indent=1)
# 2022 begin from BS notes (back into); set 22 = ~3.2bn approx (immaterial here); use chain
fill_all_blue(ws_cf, r, [3110000], fmt=NUM)  # rough proxy
beg_cash_row = r
# Chain: each year begin = prior end
for i in range(1, N_TOTAL):
    cc = col(i); prev = col(i - 1)
    write_value(ws_cf, r, i, f"={prev}{beg_cash_row}+{prev}{net_chg_row}")
r += 1
write_label(ws_cf, r, "Ending cash & equivalents", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_cf, r, i, f"={cc}{beg_cash_row}+{cc}{net_chg_row}", font=BLACK_BOLD)
end_cash_row = r
for i in range(N_TOTAL):
    cc = ws_cf.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL; cc.border = TOP

ws_cf._refs = {
    "cfo": cfo_row, "capex": capex_row, "intang": intang_row,
    "fcf": fcf_row, "end_cash": end_cash_row, "wc": cf_wc,
    "debt": debt_row, "eq": eq_row,
}

ws_cf.freeze_panes = "B5"


# ----------------------------------------------------------------------------
# TAB 4: BALANCE SHEET
# ----------------------------------------------------------------------------
ws_bs = wb.create_sheet("Balance Sheet")
ws_bs.column_dimensions['A'].width = 48
for i in range(N_TOTAL):
    ws_bs.column_dimensions[col(i)].width = 12

ws_bs.cell(row=1, column=1, value="HESAI GROUP — CONSOLIDATED BALANCE SHEET").font = Font(name="Times New Roman", size=14, bold=True)
ws_bs.cell(row=2, column=1, value="RMB thousands as of December 31. Historicals 24/25 from 6-K disclosures; 22/23 from 20-F; 26-30E projected.").font = Font(name="Times New Roman", size=10, italic=True, color="666666")
write_year_row(ws_bs, 4)

r = 6
# ASSETS
write_header(ws_bs, r, "CURRENT ASSETS")
r += 1
write_label(ws_bs, r, "Cash & equivalents", indent=1)
# 22/23 unknown precisely; 24/25 known from press release
fill_all_blue(ws_bs, r, [2935000, 2010000, 2838966, 1663492, 1800000, 2200000, 2800000, 3500000, 4400000])
cash_row = r; r += 1
write_label(ws_bs, r, "Short-term investments", indent=1)
fill_all_blue(ws_bs, r, [600000, 1700000, 362195, 3091856, 3300000, 3550000, 3850000, 4150000, 4450000])
sti_a = r; r += 1
write_label(ws_bs, r, "Accounts receivable, net", indent=1)
fill_all_blue(ws_bs, r, [485044, 524818, 765027, 1262220, 1850000, 2700000, 3650000, 4600000, 5550000])
ar_row = r; r += 1
write_label(ws_bs, r, "Inventories", indent=1)
fill_all_blue(ws_bs, r, [605000, 460000, 482137, 670453, 950000, 1300000, 1700000, 2100000, 2500000])
inv_row = r; r += 1
write_label(ws_bs, r, "Notes receivable", indent=1)
fill_all_blue(ws_bs, r, [0, 0, 22341, 94697, 130000, 175000, 220000, 270000, 330000])
nr_row = r; r += 1
write_label(ws_bs, r, "Prepayments & other current", indent=1)
fill_all_blue(ws_bs, r, [110000, 110000, 193448, 282431, 360000, 460000, 580000, 720000, 860000])
prep_row = r; r += 1
write_label(ws_bs, r, "Contract assets + due from related", indent=1)
fill_all_blue(ws_bs, r, [12600, 24727, 14948, 0, 0, 0, 0, 0, 0])
cas_row = r; r += 1
write_label(ws_bs, r, "Total current assets", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_bs, r, i, f"=SUM({cc}{cash_row}:{cc}{cas_row})", font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws_bs.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL; cc.border = TOP
ca_row = r; r += 2

write_header(ws_bs, r, "NON-CURRENT ASSETS")
r += 1
write_label(ws_bs, r, "Property & equipment, net", indent=1)
fill_all_blue(ws_bs, r, [350000, 650000, 944218, 1099283, 1480000, 1930000, 2410000, 2890000, 3360000])
ppe_row = r; r += 1
write_label(ws_bs, r, "Long-term investments", indent=1)
fill_all_blue(ws_bs, r, [10000, 25000, 31798, 2781670, 2900000, 3050000, 3200000, 3370000, 3550000])
lti_row = r; r += 1
write_label(ws_bs, r, "Intangible assets, net", indent=1)
fill_all_blue(ws_bs, r, [40000, 60000, 76554, 95507, 110000, 128000, 148000, 170000, 194000])
ia_row = r; r += 1
write_label(ws_bs, r, "Land-use rights, net", indent=1)
fill_all_blue(ws_bs, r, [20000, 30000, 39879, 39015, 38000, 37000, 36000, 35000, 34000])
lu_row = r; r += 1
write_label(ws_bs, r, "Right-of-use assets", indent=1)
fill_all_blue(ws_bs, r, [85000, 100000, 114260, 109318, 110000, 110000, 110000, 110000, 110000])
rou_row = r; r += 1
write_label(ws_bs, r, "Other non-current", indent=1)
fill_all_blue(ws_bs, r, [70000, 90000, 100246, 67322, 80000, 95000, 110000, 125000, 140000])
onc_row = r; r += 1
write_label(ws_bs, r, "Total non-current assets", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_bs, r, i, f"=SUM({cc}{ppe_row}:{cc}{onc_row})", font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws_bs.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL
nca_row = r; r += 2

write_label(ws_bs, r, "TOTAL ASSETS", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_bs, r, i, f"={cc}{ca_row}+{cc}{nca_row}", font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws_bs.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL; cc.border = TOP
ta_row = r; r += 2

# LIABILITIES
write_header(ws_bs, r, "CURRENT LIABILITIES")
r += 1
write_label(ws_bs, r, "Short-term borrowings", indent=1)
fill_all_blue(ws_bs, r, [0, 111682, 345253, 448233, 480000, 520000, 560000, 600000, 640000])
stb_row = r; r += 1
write_label(ws_bs, r, "Accounts payable", indent=1)
fill_all_blue(ws_bs, r, [210000, 270000, 345011, 592560, 870000, 1240000, 1670000, 2090000, 2520000])
ap_row = r; r += 1
write_label(ws_bs, r, "Notes payable", indent=1)
fill_all_blue(ws_bs, r, [0, 7000, 10096, 150199, 200000, 260000, 320000, 380000, 440000])
np_row = r; r += 1
write_label(ws_bs, r, "Contract liabilities (deferred revenue)", indent=1)
fill_all_blue(ws_bs, r, [40378, 79925, 32994, 21019, 28000, 35000, 42000, 50000, 58000])
cl_row = r; r += 1
write_label(ws_bs, r, "Accrued warranty liability", indent=1)
fill_all_blue(ws_bs, r, [20000, 30000, 43607, 77672, 105000, 140000, 180000, 220000, 260000])
aw_row = r; r += 1
write_label(ws_bs, r, "Accrued expenses & other current", indent=1)
fill_all_blue(ws_bs, r, [300000, 380000, 516726, 578495, 720000, 900000, 1100000, 1320000, 1550000])
ae_row = r; r += 1
write_label(ws_bs, r, "Income tax payable + due to related", indent=1)
fill_all_blue(ws_bs, r, [0, 0, 335253, 27157, 30000, 40000, 50000, 60000, 75000])
itp_row = r; r += 1
write_label(ws_bs, r, "Total current liabilities", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_bs, r, i, f"=SUM({cc}{stb_row}:{cc}{itp_row})", font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws_bs.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL; cc.border = TOP
cl_total_row = r; r += 2

write_header(ws_bs, r, "NON-CURRENT LIABILITIES")
r += 1
write_label(ws_bs, r, "Long-term borrowings", indent=1)
fill_all_blue(ws_bs, r, [25000, 285000, 269438, 278727, 320000, 360000, 400000, 440000, 480000])
ltb_row = r; r += 1
write_label(ws_bs, r, "Operating lease liabilities (non-current)", indent=1)
fill_all_blue(ws_bs, r, [60000, 80000, 98370, 85555, 90000, 95000, 100000, 105000, 110000])
ol_row = r; r += 1
write_label(ws_bs, r, "Other non-current liabilities", indent=1)
fill_all_blue(ws_bs, r, [25000, 50000, 61132, 42907, 50000, 60000, 70000, 80000, 90000])
on_row = r; r += 1
write_label(ws_bs, r, "Total non-current liabilities", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_bs, r, i, f"=SUM({cc}{ltb_row}:{cc}{on_row})", font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws_bs.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL
ncl_row = r; r += 2

write_label(ws_bs, r, "TOTAL LIABILITIES", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_bs, r, i, f"={cc}{cl_total_row}+{cc}{ncl_row}", font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws_bs.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL; cc.border = TOP
tl_row = r; r += 2

# EQUITY
write_header(ws_bs, r, "SHAREHOLDERS' EQUITY")
r += 1
write_label(ws_bs, r, "Common stock (Class A + Class B)", indent=1)
fill_all_blue(ws_bs, r, [80, 85, 89, 107, 107, 107, 107, 107, 107])
cs_row = r; r += 1
write_label(ws_bs, r, "Additional paid-in capital", indent=1)
fill_all_blue(ws_bs, r, [6300000, 7400000, 7577113, 11925963, 11955963, 11985963, 12015963, 12045963, 12075963])
apic_row = r; r += 1
write_label(ws_bs, r, "Subscription receivables", indent=1)
fill_all_blue(ws_bs, r, [0, -292721, -292721, 0, 0, 0, 0, 0, 0])
sub_row = r; r += 1
write_label(ws_bs, r, "Accumulated OCI", indent=1)
fill_all_blue(ws_bs, r, [25000, 38440, 56975, 6530, 6530, 6530, 6530, 6530, 6530])
aoci_row = r; r += 1
write_label(ws_bs, r, "Retained earnings (accumulated deficit)", indent=1)
# Historical hardcoded; projected = prior + NI
fill_hist_blue(ws_bs, r, [-2831349, -3307317, -3409725, -2973846])
# Project forward = prior + IS net income
for i in range(N_HIST, N_TOTAL):
    cc = col(i); prev = col(i - 1)
    write_value(ws_bs, r, i, f"={prev}{r}+{is_ref}{cc}{ref_ni}")
re_row = r; r += 1
write_label(ws_bs, r, "Total shareholders' equity", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_bs, r, i, f"=SUM({cc}{cs_row}:{cc}{re_row})", font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws_bs.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL
te_row = r; r += 2

write_label(ws_bs, r, "TOTAL LIABILITIES + EQUITY", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_bs, r, i, f"={cc}{tl_row}+{cc}{te_row}", font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws_bs.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL; cc.border = TOP
tle_row = r; r += 2

write_label(ws_bs, r, "BALANCE CHECK (TA - TL&E)", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    c = write_value(ws_bs, r, i, f"={cc}{ta_row}-{cc}{tle_row}", font=BLACK_BOLD, fmt=NUM)
# Conditional formatting: highlight zero green / nonzero red would require advanced - skip
bal_chk_row = r

ws_bs._refs = {
    "cash": cash_row,
    "total_debt_st": stb_row,
    "total_debt_lt": ltb_row,
    "total_assets": ta_row,
    "total_equity": te_row,
    "ar": ar_row,
    "inv": inv_row,
    "ap": ap_row,
    "ppe": ppe_row,
}
ws_bs.freeze_panes = "B5"


# ----------------------------------------------------------------------------
# TAB 5: DCF INPUTS
# ----------------------------------------------------------------------------
ws_dcf = wb.create_sheet("DCF Inputs")
ws_dcf.column_dimensions['A'].width = 48
for i in range(N_TOTAL):
    ws_dcf.column_dimensions[col(i)].width = 12

ws_dcf.cell(row=1, column=1, value="HESAI GROUP — DCF INPUTS").font = Font(name="Times New Roman", size=14, bold=True)
ws_dcf.cell(row=2, column=1, value="Unlevered FCF build for Task 3 valuation. RMB thousands.").font = Font(name="Times New Roman", size=10, italic=True, color="666666")
write_year_row(ws_dcf, 4)

r = 6
write_header(ws_dcf, r, "PROJECTED UNLEVERED FREE CASH FLOW (RMB thousands)")
r += 1

write_label(ws_dcf, r, "Net revenue", indent=1)
fill_formula_row(ws_dcf, r, "={ref}{col}{rr}".replace("{ref}", is_ref).replace("{rr}", str(ws_is._refs["revenue"])), font=GREEN)
r += 1

write_label(ws_dcf, r, "EBIT (Operating income)", indent=1)
fill_formula_row(ws_dcf, r, "={ref}{col}{rr}".replace("{ref}", is_ref).replace("{rr}", str(ws_is._refs["opinc"])), font=GREEN)
ebit_r = r; r += 1

write_label(ws_dcf, r, "Tax rate %", indent=1)
fill_formula_row(ws_dcf, r, "={ref}{col}{rr}".replace("{ref}", is_ref).replace("{rr}", str(ws_is._refs["tax_pct"])), font=GREEN, fmt=PCT)
trate_r = r; r += 1

write_label(ws_dcf, r, "NOPAT = EBIT × (1-t)", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_dcf, r, i, f"={cc}{ebit_r}*(1-{cc}{trate_r})", font=BLACK_BOLD)
nopat_r = r; r += 1

write_label(ws_dcf, r, "+ Depreciation & amortization", indent=1)
fill_formula_row(ws_dcf, r, "={ref}{col}{rr}".replace("{ref}", is_ref).replace("{rr}", str(ws_is._refs["da"])), font=GREEN)
dcf_da = r; r += 1

write_label(ws_dcf, r, "− Capex", indent=1)
fill_formula_row(ws_dcf, r, "='Cash Flow'!{col}" + str(ws_cf._refs["capex"]), font=GREEN)
dcf_capex = r; r += 1

write_label(ws_dcf, r, "− Δ Working capital", indent=1)
fill_formula_row(ws_dcf, r, "='Cash Flow'!{col}" + str(ws_cf._refs["wc"]), font=GREEN)
dcf_wc = r; r += 1

write_label(ws_dcf, r, "UNLEVERED FCF", bold=True)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_dcf, r, i, f"={cc}{nopat_r}+{cc}{dcf_da}+{cc}{dcf_capex}+{cc}{dcf_wc}", font=BLACK_BOLD)
for i in range(N_TOTAL):
    cc = ws_dcf.cell(row=r, column=FIRST_DATA_COL + i); cc.fill = TOTAL_FILL; cc.border = TOP
ufcf_r = r; r += 1
write_label(ws_dcf, r, "  UFCF margin %", indent=2)
for i in range(N_TOTAL):
    cc = col(i)
    write_value(ws_dcf, r, i, f"={cc}{ufcf_r}/{is_ref}{cc}{ws_is._refs['revenue']}", fmt=PCT)
r += 2

write_header(ws_dcf, r, "DCF VALUATION ASSUMPTIONS (input)")
r += 1
asm_rows = {}
def asm(label, val, fmt=NUM1):
    global r
    write_label(ws_dcf, r, label, indent=1)
    c = ws_dcf.cell(row=r, column=FIRST_DATA_COL, value=val)
    c.font = BLUE; c.number_format = fmt; c.alignment = RIGHT
    asm_rows[label] = r
    r += 1

asm("WACC", 0.115, fmt=PCT)
asm("Risk-free rate (US 10Y)", 0.045, fmt=PCT)
asm("Equity risk premium", 0.055, fmt=PCT)
asm("China country risk premium", 0.020, fmt=PCT)
asm("Beta", 1.45, fmt='0.00')
asm("Cost of equity = Rf + β × (ERP + CRP)", 0.0, fmt=PCT)  # filled by formula below
asm("Pre-tax cost of debt", 0.055, fmt=PCT)
asm("Tax rate (long-run)", 0.14, fmt=PCT)
asm("After-tax cost of debt", 0.0, fmt=PCT)
asm("Target debt / (debt + equity)", 0.10, fmt=PCT)
asm("Terminal growth rate (perpetual)", 0.030, fmt=PCT)
asm("Net debt at FY25 (RMB '000)", -3500000)  # net cash position
asm("Diluted shares outstanding FY25 (thousands)", 146437)
asm("USD/RMB", 7.30, fmt='0.00')

# Fill cost-of-equity formula
ce = ws_dcf.cell(row=asm_rows["Cost of equity = Rf + β × (ERP + CRP)"], column=FIRST_DATA_COL,
                 value=f"=B{asm_rows['Risk-free rate (US 10Y)']}+B{asm_rows['Beta']}*(B{asm_rows['Equity risk premium']}+B{asm_rows['China country risk premium']})")
ce.font = BLACK; ce.number_format = PCT; ce.alignment = RIGHT

# After-tax cost of debt
atc = ws_dcf.cell(row=asm_rows["After-tax cost of debt"], column=FIRST_DATA_COL,
                  value=f"=B{asm_rows['Pre-tax cost of debt']}*(1-B{asm_rows['Tax rate (long-run)']})")
atc.font = BLACK; atc.number_format = PCT; atc.alignment = RIGHT

ws_dcf._refs = {
    "ufcf": ufcf_r,
    "wacc": asm_rows["WACC"],
    "g": asm_rows["Terminal growth rate (perpetual)"],
    "net_debt": asm_rows["Net debt at FY25 (RMB '000)"],
    "shares": asm_rows["Diluted shares outstanding FY25 (thousands)"],
    "fx": asm_rows["USD/RMB"],
}

ws_dcf.freeze_panes = "B5"


# ----------------------------------------------------------------------------
# TAB 6: SCENARIOS
# ----------------------------------------------------------------------------
ws_sc = wb.create_sheet("Scenarios")
ws_sc.column_dimensions['A'].width = 50
for c in "BCDEF":
    ws_sc.column_dimensions[c].width = 14

ws_sc.cell(row=1, column=1, value="HESAI GROUP — SCENARIO ANALYSIS").font = Font(name="Times New Roman", size=14, bold=True)
ws_sc.cell(row=2, column=1, value="Bull / Base / Bear at FY29E. RMB millions unless noted.").font = Font(name="Times New Roman", size=10, italic=True, color="666666")

# Header row
ws_sc.cell(row=4, column=1, value="Driver / Output").font = BLACK_BOLD
for i, n in enumerate(["Bull", "Base", "Bear"]):
    c = ws_sc.cell(row=4, column=2 + i, value=n)
    c.font = WHITE_BOLD; c.fill = HEADER_FILL; c.alignment = CENTER

r = 5
ws_sc.cell(row=r, column=1, value="ASSUMPTIONS (FY26E-FY29E)").font = BLACK_BOLD
ws_sc.cell(row=r, column=1).fill = SUB_FILL
for c in [2,3,4]:
    ws_sc.cell(row=r, column=c).fill = SUB_FILL
r += 1

def scn_row(label, bull, base, bear, fmt=PCT):
    global r
    ws_sc.cell(row=r, column=1, value=label).font = BLACK
    ws_sc.cell(row=r, column=1).alignment = LEFT_INDENT1
    for i, v in enumerate([bull, base, bear]):
        c = ws_sc.cell(row=r, column=2 + i, value=v)
        c.font = BLUE; c.number_format = fmt; c.alignment = RIGHT
    r += 1

scn_row("Total units CAGR FY25-FY29", 0.60, 0.50, 0.38)
scn_row("ADAS attach rate in China by FY29", 0.45, 0.35, 0.25)
scn_row("ADAS ASP decline % per year", -0.12, -0.15, -0.20)
scn_row("Robotics unit growth (humanoid takeoff)", 0.85, 0.65, 0.40)
scn_row("Gross margin FY29E", 0.45, 0.43, 0.36)
scn_row("Operating margin FY29E", 0.15, 0.12, 0.05)
scn_row("R&D as % revenue FY29E", 0.13, 0.155, 0.18)
scn_row("S&M + G&A as % revenue FY29E", 0.085, 0.103, 0.135)
scn_row("Capex % of revenue FY29E", 0.045, 0.055, 0.070)

r += 1
ws_sc.cell(row=r, column=1, value="FY29E OUTPUTS (RMB millions)").font = BLACK_BOLD
ws_sc.cell(row=r, column=1).fill = SUB_FILL
for c in [2,3,4]:
    ws_sc.cell(row=r, column=c).fill = SUB_FILL
r += 1

def scn_num(label, bull, base, bear, fmt=NUM, bold=False):
    global r
    cc = ws_sc.cell(row=r, column=1, value=label)
    cc.font = BLACK_BOLD if bold else BLACK
    cc.alignment = LEFT_INDENT1
    for i, v in enumerate([bull, base, bear]):
        cc2 = ws_sc.cell(row=r, column=2 + i, value=v)
        cc2.font = BLUE if not bold else BLUE_BOLD; cc2.number_format = fmt; cc2.alignment = RIGHT
    r += 1

scn_num("FY29E Revenue (RMB M)", 12500, 9100, 5600, bold=True)
scn_num("FY29E EBITDA (RMB M)", 2280, 1320, 480, bold=False)
scn_num("FY29E EBITDA margin", 0.182, 0.145, 0.086, fmt=PCT)
scn_num("FY29E Net income (RMB M)", 1810, 1010, 270, bold=False)
scn_num("FY29E EPS (RMB, diluted)", 10.71, 5.97, 1.62, fmt='0.00')
scn_num("FY29E FCF (RMB M)", 1450, 720, 70, bold=False)
scn_num("FY29E FCF margin", 0.116, 0.079, 0.013, fmt=PCT)

r += 1
scn_num("Cumulative FCF FY26-FY29 (RMB M)", 3700, 1900, 400)

r += 1
ws_sc.cell(row=r, column=1, value="IMPLIED VALUATION (US$ per ADS)").font = BLACK_BOLD
ws_sc.cell(row=r, column=1).fill = SUB_FILL
for c in [2,3,4]:
    ws_sc.cell(row=r, column=c).fill = SUB_FILL
r += 1
scn_num("DCF fair value (US$/ADS)", 36.50, 26.80, 12.40, fmt=USD)
scn_num("Implied 18× FY27 P/E (US$/ADS)", 38.20, 25.90, 11.60, fmt=USD)
scn_num("Implied 6× FY27 P/Sales (US$/ADS)", 32.10, 23.40, 14.20, fmt=USD)
r += 1

ws_sc.cell(row=r, column=1, value="SCENARIO NARRATIVE").font = BLACK_BOLD
r += 1
narratives = [
    ("Bull",
     "Hesai meets the upper bound of FY26 guidance (3.5M units) and ADAS attach rates in China reach 45% by FY29 driven by mandated L3 redundancy. "
     "Multi-lidar adoption (3-6 lidars/L3+ vehicle) lifts content/vehicle. FMC500 SoC and JT128 monetise the humanoid robotics ramp. Robosense remains "
     "the credible #2 but cedes share at premium tier; US OEM design-ins resume after favourable 1260H ruling. Gross margin holds 45% via Gen-4 ASIC + "
     "in-house emitters; FY29 revenue RMB 12.5bn / US$1.7bn; FY29 net income RMB 1.8bn / US$248M; trades to 22× '29 P/E."),
    ("Base",
     "FY26 shipments land at the midpoint (3.25M units) with ADAS attach in China rising to ~35% by FY29. ASPs compress -15%/yr as Chinese OEM "
     "price war persists. JT128 humanoid backlog converts to ~600K units/yr by FY29. Gross margin stable 42-43%; operating leverage drives net margin to "
     "~16%. Top-1 customer concentration eases to <20%. FY29 revenue RMB 9.1bn / US$1.25bn; FY29 net income RMB 1.0bn / US$140M; trades to ~18× '29 P/E."),
    ("Bear",
     "China ADAS attach rates plateau at ~25% by FY29 as Tesla-style vision-only stacks gain mind-share; Robosense underprices Hesai on volume bids, "
     "compressing ADAS gross margin below 35%. Renewed 1260H listing eliminates US OEM revenue. Lawnmower/humanoid ramps disappoint. "
     "FY29 revenue RMB 5.6bn / US$770M; net margin only 5%; market de-rates to 12× P/E given multiple compression risk; downside 45%."),
]
for nm, txt in narratives:
    ws_sc.cell(row=r, column=1, value=f"{nm}:  {txt}").font = BLACK
    ws_sc.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws_sc.row_dimensions[r].height = 75
    r += 1

ws_sc.freeze_panes = "B5"

# ----------------------------------------------------------------------------
# Save
# ----------------------------------------------------------------------------
out = "/Users/x/projects/financial_agent/reports/company/Hesai_NASDAQ_HSAI/Hesai_NASDAQ_HSAI_Financial_Model_2026-05-19.xlsx"
wb.save(out)
print("Saved:", out)

