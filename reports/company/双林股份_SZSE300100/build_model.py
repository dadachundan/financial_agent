#!/usr/bin/env python3
"""
Shuanglin Co. (SZSE:300100) — Equity-Research Initiation
Task 2: Financial Model (6 tabs)

Units: CNY million unless noted.
Historicals (FY2021–FY2025) sourced verbatim from cninfo annual reports:
- FY2025 annual report (filed 2026-03-24)   — FY2024/FY2025 IS, CF, BS
- FY2023 annual report (filed 2024-04-18)   — FY2022/FY2023 IS, CF, BS
- FY2022 annual report (filed 2023-04-18)   — FY2021 IS, CF, BS

Projection horizon: FY2026E – FY2030E (5 years)
Reflects Q1 2026 sharp decline (rev -10.4%, NI -47%) per 2026 Q1 report.
Author: equity-research/initiating-coverage Task 2 build
Date:   2026-05-18
"""

from __future__ import annotations

import os
import datetime as _dt
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter

OUT_PATH = os.path.join(
    os.path.dirname(__file__),
    f"双林股份_SZSE300100_Financial_Model_{_dt.date.today()}.xlsx",
)

# ============================================================================
# COLOR PALETTE & STYLES
# ============================================================================
BLUE   = "1F4E79"   # hardcoded input — bank standard
BLACK  = "000000"   # formula / calc
GREEN  = "006100"   # link to other sheet
RED    = "C00000"   # error / flag
GREY   = "808080"
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL  = PatternFill("solid", fgColor="FFF2CC")
ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(border_style="thin", color="BFBFBF")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_BORDER = Border(top=Side(border_style="medium", color="000000"))

FMT_INT = '#,##0;(#,##0);"–"'
FMT_DEC = '#,##0.0;(#,##0.0);"–"'
FMT_PCT = '0.0%;(0.0%);"–"'
FMT_X   = '0.00"x"'

YEARS_H = [2021, 2022, 2023, 2024, 2025]       # historical
YEARS_P = [2026, 2027, 2028, 2029, 2030]       # projected
YEARS   = YEARS_H + YEARS_P
N_H = len(YEARS_H)
N_P = len(YEARS_P)
N   = len(YEARS)

# ============================================================================
# HISTORICAL FINANCIAL DATA (CNY million — divided by 1,000,000 from filings)
# ============================================================================

# ---- INCOME STATEMENT (consolidated, CNY mn) -------------------------------
# Source citations:
#  2021: FY2022 annual report (2023-04-18), pp.82-84
#  2022,2023: FY2023 annual report (2024-04-18), pp.85-87
#  2024,2025: FY2025 annual report (2026-03-24), pp.95-97
IS = {
    # label : [2021, 2022, 2023, 2024, 2025]
    "营业收入 Revenue":           [3682.30, 4185.28, 4138.82, 4910.50, 5483.70],
    "营业成本 COGS":              [2999.42, 3471.81, 3356.58, 4002.47, 4335.36],
    "毛利 Gross profit":          [None, None, None, None, None],   # calc
    "毛利率 % GM":                [None, None, None, None, None],   # calc
    "税金及附加 Taxes & surcharges": [30.52, 34.86, 37.97, 39.81, 42.55],
    "销售费用 Selling exp":       [57.52, 48.30, 91.82, 32.79, 43.04],
    "管理费用 G&A":               [237.72, 264.61, 276.71, 295.50, 347.20],
    "研发费用 R&D":               [164.64, 185.28, 175.05, 168.53, 220.28],
    "财务费用 Financial exp":     [83.54, 41.06, 37.37, 20.05, 16.10],
    "  其中：利息费用 Interest exp": [78.74, 61.10, 46.82, 31.74, 20.34],
    "  利息收入 Interest income": [3.12, 3.65, 6.92, 9.17, 7.66],
    "其他收益 Other income (subsidy)": [42.83, 38.34, 55.33, 80.10, 49.40],
    "投资收益 Investment income":  [16.51, 10.02, -2.39, 160.26, -5.29],
    "公允价值变动 FV change":      [-0.60, -5.30, 0.17, 0.00, 0.00],
    "信用减值损失 Credit loss":   [-1.76, -11.03, -7.47, -1.06, 4.40],
    "资产减值损失 Asset impairment": [-20.02, -73.19, -94.18, -75.01, -15.18],
    "资产处置收益 Asset disposal gain": [0.46, 3.00, 9.82, 4.78, 36.19],
    "营业利润 Operating profit":   [None, None, None, None, None],  # calc
    "营业外收入 Non-op income":   [4.14, 1.48, 2.01, 3.70, 6.71],
    "营业外支出 Non-op expense":  [1.64, 9.50, 20.20, 15.49, 3.91],
    "利润总额 Profit before tax":  [None, None, None, None, None],  # calc
    "所得税 Tax expense":         [23.88, 17.27, 24.47, 11.02, 48.26],
    "净利润 Net income":          [None, None, None, None, None],  # calc
    "  归母净利润 NI to parent":  [128.87, 75.17, 80.88, 497.01, 503.23],
    "  少数股东损益 Minority":    [-3.90, 0.71, 1.06, 0.61, 0.00],
    "EBITDA (calc)":             [None, None, None, None, None],  # calc later
    "EBIT (= 营业利润)":           [None, None, None, None, None],
    "EPS 基本 (¥)":                [0.32, 0.19, 0.20, 0.89, 0.89],
    "EPS 稀释 (¥)":                [0.32, 0.19, 0.20, 0.88, 0.87],
}

# ---- CASH FLOW STATEMENT (consolidated, CNY mn) ----------------------------
CF = {
    # OPERATING
    "销售商品劳务收到的现金 Cash from sales":    [3665.17, 3947.68, 3645.52, 4227.18, 4639.93],
    "收到的税费返还 Tax refunds":                  [30.82, 31.62, 6.61, 32.82, 9.93],
    "收到其他与经营有关的现金 Other CFO inflows":   [126.07, 96.98, 89.71, 103.65, 361.63],
    "经营现金流入小计 Total CFO inflows":          [None, None, None, None, None],   # calc
    "购买商品劳务支付的现金 Cash for COGS":         [2521.91, 2691.66, 2359.61, 2617.98, 3032.39],
    "支付职工薪酬 Cash to employees":               [523.34, 562.27, 587.33, 647.57, 708.68],
    "支付各项税费 Taxes paid":                      [141.77, 140.07, 164.65, 181.33, 223.55],
    "支付其他与经营有关的现金 Other CFO outflows":    [209.21, 239.37, 252.63, 245.64, 265.44],
    "经营现金流出小计 Total CFO outflows":         [None, None, None, None, None],
    "经营活动产生的现金流量净额 CFO":              [425.84, 442.91, 377.62, 671.14, 781.44],

    # INVESTING
    "处置固定无形资产收回的现金 Proceeds from PP&E sales":   [12.84, 6.55, 45.53, 36.02, 34.23],
    "收回投资 / 投资收益等 Other investing inflows":         [217.21, 663.69, 56.02, 172.13, 14.49],
    "购建固定无形资产支付的现金 CapEx":                       [253.10, 213.07, 277.44, 296.17, 408.06],
    "投资支付 / 收购等 Other investing outflows":            [186.56, 663.50, 30.00, 20.29, 195.01],
    "投资活动产生的现金流量净额 CFI":                          [290.39, -207.78, -205.88, -108.31, -554.35],
    "  其中：购建固定无形资产支付的现金 CapEx (用于参考)":      [253.10, 213.07, 277.44, 296.17, 408.06],

    # FINANCING
    "取得借款收到的现金 Debt issued":                       [1597.92, 1438.60, 1117.05, 879.95, 980.00],
    "吸收投资收到的现金 Equity issued":                      [0.00, 0.00, 0.00, 0.00, 51.70],
    "偿还债务支付的现金 Debt repaid":                        [1753.22, 1693.84, 1216.45, 1168.75, 1080.86],
    "分配股利利息支付的现金 Dividends & interest paid":       [79.47, 334.38, 53.41, 75.71, 100.07],
    "支付其他筹资有关的现金 Other CFF outflows":              [14.81, 9.01, 11.33, 37.56, 27.49],
    "筹资活动产生的现金流量净额 CFF":                          [-249.58, -598.64, -164.15, -402.07, -176.72],

    # OTHER
    "汇率变动影响 FX impact":                                 [-1.67, 19.01, 5.34, 4.41, 0.10],
    "现金净增加额 Net change in cash":                        [464.98, -344.50, 12.93, 165.18, 50.47],
    "期初现金及等价物 Beginning cash":                        [212.86, 677.84, 333.34, 346.27, 511.44],
    "期末现金及等价物 Ending cash":                           [677.84, 333.34, 346.27, 511.44, 561.91],

    # CALCULATED
    "自由现金流 FCF (CFO – CapEx)":                          [None, None, None, None, None],
}

# ---- BALANCE SHEET (consolidated, CNY mn, year-end) ------------------------
# Source:
#  2021 yr-end: opening "2022年1月1日" col in FY2022 annual (p.78)
#  2022 yr-end: closing col in FY2022 annual (p.78) — confirmed by FY2023 annual opening col
#  2023 yr-end: closing col in FY2023 annual (p.81)
#  2024 yr-end: opening col in FY2025 annual (p.91)
#  2025 yr-end: closing col in FY2025 annual (p.91)
BS = {
    # ASSETS
    "货币资金 Cash & equivalents":            [809.15, 440.25, 477.80, 621.93, 734.54],
    "交易性金融资产 Trading financial assets":   [25.73, 26.91, 0.00, 10.88, 0.00],
    "应收账款 Accounts receivable":            [867.38, 1105.62, 1226.10, 1427.10, 1358.73],
    "应收款项融资 Receivables financing":       [647.53, 697.44, 829.41, 817.45, 1046.09],
    "预付款项 Prepayments":                    [25.68, 32.42, 33.21, 17.02, 34.87],
    "应收票据 Notes receivable":                [0.00, 0.40, 0.00, 12.12, 14.78],
    "其他应收款 Other receivables":            [13.30, 11.31, 4.02, 20.71, 7.43],
    "存货 Inventory":                           [940.19, 957.31, 1023.73, 1017.13, 1006.63],
    "其他流动资产 Other current assets":         [30.70, 39.88, 39.01, 16.36, 70.28],
    "其他 (合同/持有待售等) Misc current":      [0.00, 0.00, 0.82, 0.53, 4.64],
    "流动资产合计 Total current assets":        [None, None, None, None, None],
    "投资性房地产 Investment property":         [15.00, 24.09, 45.31, 40.36, 38.92],
    "固定资产 PP&E, net":                       [1860.07, 1726.12, 1630.15, 1673.55, 1826.74],
    "在建工程 Construction in progress":        [82.79, 96.50, 95.95, 91.92, 163.89],
    "使用权资产 Right-of-use assets":            [29.30, 24.91, 73.34, 11.74, 19.41],
    "无形资产 Intangible assets":               [426.08, 425.60, 392.15, 391.07, 452.77],
    "商誉 Goodwill":                            [0.00, 0.00, 0.00, 0.00, 115.31],
    "长期待摊费用 LT prepaid exp":              [50.91, 43.07, 31.83, 23.29, 31.75],
    "递延所得税资产 DTA":                       [34.43, 36.52, 34.31, 49.13, 51.79],
    "其他非流动资产 Other non-current":         [28.49, 5.88, 17.56, 28.38, 44.28],
    "非流动资产合计 Total non-current assets":  [None, None, None, None, None],
    "资产总计 Total assets":                    [None, None, None, None, None],

    # LIABILITIES
    "短期借款 Short-term debt":                 [1101.62, 956.42, 833.11, 693.44, 565.39],
    "应付票据 Notes payable":                   [709.67, 738.44, 891.46, 671.27, 692.75],
    "应付账款 Accounts payable":                [928.29, 1126.76, 1166.89, 1602.35, 1615.01],
    "合同负债 Contract liabilities":            [6.74, 28.81, 55.04, 37.90, 49.31],
    "应付职工薪酬 Salary payable":              [66.96, 70.15, 77.66, 81.86, 98.10],
    "应交税费 Taxes payable":                   [20.29, 41.23, 28.45, 30.92, 31.43],
    "其他应付款 Other payables":                 [57.44, 53.26, 54.81, 89.08, 272.99],
    "一年内到期非流动负债 Current LT debt":     [110.01, 144.86, 192.42, 76.76, 239.87],
    "其他流动负债 Other current liab":          [0.32, 2.36, 4.40, 1.21, 0.77],
    "流动负债合计 Total current liab":          [None, None, None, None, None],
    "长期借款 Long-term debt":                  [297.15, 161.21, 134.65, 100.09, 0.00],
    "租赁负债 Lease liabilities":               [16.83, 14.81, 65.45, 4.47, 10.31],
    "预计负债 Provisions":                      [24.40, 31.88, 58.69, 71.45, 16.37],
    "递延收益 Deferred income":                 [200.77, 174.67, 158.21, 135.52, 150.38],
    "递延所得税负债 DTL":                       [27.39, 31.91, 28.18, 30.42, 33.54],
    "其他非流动负债 Other non-current liab":    [7.00, 1.21, 1.22, 1.19, 20.40],
    "非流动负债合计 Total non-current liab":    [None, None, None, None, None],
    "负债合计 Total liabilities":                [None, None, None, None, None],

    # EQUITY
    "股本 Share capital":                       [402.15, 402.15, 402.15, 400.77, 571.98],
    "资本公积 Capital reserve":                 [1449.95, 1459.36, 1463.43, 1482.17, 1433.11],
    "减：库存股 Treasury stock":                 [0.00, 0.00, 0.00, 30.02, 0.00],
    "其他综合收益 OCI":                          [-0.21, 0.18, 4.79, 5.99, 13.70],
    "专项储备 Special reserve":                  [0.00, 0.00, 0.00, 0.90, 1.02],
    "盈余公积 Statutory reserve":                [133.17, 140.49, 148.16, 161.37, 183.28],
    "未分配利润 Retained earnings":             [318.47, 105.79, 177.35, 621.48, 1023.21],
    "归属母公司所有者权益 Equity to parent":    [None, None, None, None, None],
    "少数股东权益 Minority interest":           [8.30, 8.20, 8.05, 0.00, 0.00],
    "所有者权益合计 Total equity":              [None, None, None, None, None],
    "负债和所有者权益总计 Total liab + equity":  [None, None, None, None, None],
}

# ---- REVENUE BY SEGMENT (FY2025 annual restated basis, CNY mn) -------------
# FY2024 figures here are the restated 3-segment basis from FY2025 report (p.21).
# Prior years (2021-2023) used a different segmentation; we show total only.
SEG_RESTATED = {
    "传动驱动智能 Transmission/Drive/Intelligent": {2024: 2847.77, 2025: 3269.96, "GM2024": 0.2235, "GM2025": 0.2388},
    "内外饰件 Auto Interior/Exterior":              {2024: 1763.46, 2025: 1946.08, "GM2024": 0.1325, "GM2025": 0.1432},
    "其他 (磨床/模具/其他) Other":                  {2024: 291.58,  2025: 259.91,  "GM2024": None,    "GM2025": 0.3187},
    "租金 Rental":                                  {2024: 7.69,    2025: 7.75,    "GM2024": None,    "GM2025": 0.7637},
}

# Geography (restated FY2024 vs FY2025):
GEO = {
    "国内 Domestic": {2024: 4431.96, 2025: 4997.14},
    "国外 Overseas": {2024: 478.53,  2025: 486.56},
}

# Key subsidiary disclosure (FY2025 annual p.30):
SUBS_2025 = {
    "湖北双林轴承 Hubei Bearings":      {"revenue": 1415.30, "net_income": 150.80},
    "山东双林新能源 Shandong NEV E-drive": {"revenue": 690.30, "net_income": 65.10},
}

SHARES_OUT = {2021: 402.15, 2022: 402.15, 2023: 402.15, 2024: 564.85, 2025: 571.98}  # mn shares (weighted avg approximations)

# ============================================================================
# PROJECTION ASSUMPTIONS (Base case)
# ============================================================================
# Q1 2026 actuals: revenue -10.42% YoY, NI -47.01%, ex-NR -39.11%
# Reflects: NEV pricing war pass-through + 2025 had non-recurring asset disposal gains (¥36.2m)
# Drivers 2026E+:
#  - HDM stable mid-single-digit growth (mature, share-driven)
#  - NEV e-drive ramping (Thailand 2026 H1)
#  - Roller screws量产 line June 2026 — early-stage but rapidly growing
#  - Smart corner mining truck deliveries 2026 H1 (100 units)
#  - Interior/Exterior: low single digit (mature)
#  - Wheel bearings: ~+10% (NEV penetration)

ASSUMPTIONS = {
    # Segment growth rates (Base case)
    "Transmission_Drive growth": {2026: 0.04, 2027: 0.20, 2028: 0.22, 2029: 0.18, 2030: 0.15},
    "Interior_Exterior growth":  {2026: -0.05, 2027: 0.05, 2028: 0.05, 2029: 0.03, 2030: 0.02},
    "Other growth":              {2026: 0.05, 2027: 0.15, 2028: 0.20, 2029: 0.20, 2030: 0.15},
    "Rental growth":             {2026: 0.05, 2027: 0.05, 2028: 0.05, 2029: 0.05, 2030: 0.05},

    # Margins (consolidated)
    "GM target":                 {2026: 0.205, 2027: 0.220, 2028: 0.235, 2029: 0.240, 2030: 0.240},
    "SGA % of revenue":          {2026: 0.075, 2027: 0.073, 2028: 0.070, 2029: 0.068, 2030: 0.067},
    "R&D % of revenue":          {2026: 0.045, 2027: 0.050, 2028: 0.050, 2029: 0.048, 2030: 0.045},
    "Fin exp / revenue":         {2026: 0.003, 2027: 0.003, 2028: 0.003, 2029: 0.002, 2030: 0.002},
    "Other inc / revenue":       {2026: 0.009, 2027: 0.008, 2028: 0.008, 2029: 0.007, 2030: 0.007},
    "Tax rate":                  {2026: 0.13,  2027: 0.13,  2028: 0.13,  2029: 0.14,  2030: 0.14},

    # Cash flow
    "CapEx % of revenue":        {2026: 0.085, 2027: 0.070, 2028: 0.060, 2029: 0.055, 2030: 0.050},
    "D&A % of revenue":          {2026: 0.045, 2027: 0.043, 2028: 0.042, 2029: 0.040, 2030: 0.040},
    "WC change % of rev change": {2026: 0.10, 2027: 0.12, 2028: 0.15, 2029: 0.15, 2030: 0.15},

    # Balance sheet drivers (days)
    "AR days":          {2026: 92, 2027: 90, 2028: 88, 2029: 86, 2030: 85},
    "Inventory days":   {2026: 85, 2027: 82, 2028: 80, 2029: 78, 2030: 78},
    "AP days":          {2026: 135, 2027: 132, 2028: 130, 2029: 128, 2030: 128},

    # Capital return
    "Dividend payout":  {2026: 0.20, 2027: 0.20, 2028: 0.22, 2029: 0.22, 2030: 0.22},

    "Shares outstanding": {2026: 572, 2027: 572, 2028: 572, 2029: 572, 2030: 572},
}

# ============================================================================
# BUILD WORKBOOK
# ============================================================================
wb = Workbook()
wb.remove(wb.active)


def set_header_row(ws, row, cols_labels, fill=HEADER_FILL):
    for c, label in enumerate(cols_labels, start=1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER_ALL


def write_data_row(ws, row, label, values, hist_n=N_H, fmt=FMT_INT,
                   bold=False, indent=0, fill=None, top_border=False,
                   is_pct=False, is_calc=False, is_link=False):
    """Write a data row with first column = label, then yearly values."""
    label_cell = ws.cell(row=row, column=1, value=("   " * indent) + label)
    label_cell.font = Font(bold=bold, size=10)
    if fill is not None:
        label_cell.fill = fill
    if top_border:
        label_cell.border = TOP_BORDER

    for c, v in enumerate(values, start=2):
        cell = ws.cell(row=row, column=c, value=v)
        if v is None:
            cell.value = None
        cell.number_format = FMT_PCT if is_pct else fmt
        # Color logic
        if is_link:
            cell.font = Font(color=GREEN, bold=bold, size=10)
        elif is_calc:
            cell.font = Font(color=BLACK, bold=bold, size=10)
        else:
            # Default: hist = blue input; forecast = formula
            is_hist = (c - 2) < hist_n
            cell.font = Font(color=BLUE if is_hist else BLACK, bold=bold, size=10)
        if fill is not None:
            cell.fill = fill
        if top_border:
            cell.border = TOP_BORDER
        cell.alignment = Alignment(horizontal="right")


def freeze_and_widen(ws, label_width=42):
    ws.freeze_panes = "B4"
    ws.column_dimensions["A"].width = label_width
    for c in range(2, N + 2):
        ws.column_dimensions[get_column_letter(c)].width = 11


# ----------------------------------------------------------------------------
# Sheet 1: Cover / Assumptions Summary
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Cover")
ws["A1"] = "双林股份 (Shuanglin Co., SZSE:300100) — Financial Model"
ws["A1"].font = Font(bold=True, size=14, color=BLUE)
ws["A2"] = "Equity-Research Initiation Coverage · Task 2 Output"
ws["A2"].font = Font(italic=True, size=10, color=GREY)
ws["A3"] = f"Build date: {_dt.date.today()} · Units: CNY mn (millions of RMB)"

ws["A5"] = "Color Code Legend"
ws["A5"].font = Font(bold=True, size=11)
legend = [
    ("Blue text", "Hardcoded input (historical / assumption)", BLUE),
    ("Black text", "Formula / calculated", BLACK),
    ("Green text", "Cross-sheet link", GREEN),
    ("Red text", "Flag / error (should be resolved)", RED),
]
for i, (k, v, color) in enumerate(legend, start=6):
    ws.cell(row=i, column=1, value=k).font = Font(color=color, bold=True)
    ws.cell(row=i, column=2, value=v)

ws["A12"] = "Tabs in this workbook"
ws["A12"].font = Font(bold=True, size=11)
tabs = [
    ("Revenue Model", "Revenue by segment (3 product lines + rental + 8 sub-products) and geography (Domestic / Overseas / Thailand / NA), drivers"),
    ("Income Statement", "Full P&L: revenue → COGS → opex (selling/G&A/R&D/fin) → other items → EBIT → tax → NI, plus EBITDA & per-share metrics, 2021A–2030E"),
    ("Cash Flow Statement", "CFO from indirect method; CFI w/ CapEx detail; CFF w/ debt/dividends; FCF derivation"),
    ("Balance Sheet", "Current + non-current assets; current + non-current liab; full equity reconciliation; balance check row"),
    ("Scenarios", "Bull / Base / Bear with 2030E revenue, EBITDA, NI, FCF & cumulative FCF outputs and assumption deltas"),
    ("DCF Inputs", "Unlevered FCF (NOPAT + D&A – CapEx – ΔNWC) prepared for Task 3 valuation"),
]
for i, (n, d) in enumerate(tabs, start=13):
    c1 = ws.cell(row=i, column=1, value=n)
    c1.font = Font(bold=True, color=BLUE)
    ws.cell(row=i, column=2, value=d)

ws["A22"] = "Key Sources"
ws["A22"].font = Font(bold=True, size=11)
sources = [
    "FY2025 年度报告 (filed 2026-03-24) — pp. 14, 21–24, 30, 91–101 (segment mix, customer concentration, IS, CF, BS)",
    "FY2023 年度报告 (filed 2024-04-18) — pp. 81–90 (FY2022/FY2023 IS, CF, BS)",
    "FY2022 年度报告 (filed 2023-04-18) — pp. 78–87 (FY2021 IS, CF; FY2021 yr-end BS as opening balance)",
    "2026 Q1 报告 (filed 2026-04-28) — Q1 2026: Revenue −10.42%, NI −47.01%, ex-NR −39.11%",
    "Task 1 research doc — reports/company/双林股份_SZSE300100/双林股份_SZSE300100_公司研究_2026-05-17.md",
]
for i, s in enumerate(sources, start=23):
    ws.cell(row=i, column=1, value=f"  • {s}").font = Font(size=9, color=GREY)

ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 110


# ----------------------------------------------------------------------------
# Sheet 2: Revenue Model
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Revenue Model")
ws["A1"] = "Revenue Model — Segment, Geography & Sub-product Build"
ws["A1"].font = Font(bold=True, size=12)
ws["A2"] = "Units: CNY mn · Hist 2021A–2025A (filings) · Proj 2026E–2030E (Base case)"
ws["A2"].font = Font(italic=True, size=9, color=GREY)

# Year headers in row 3
header_labels = ["项目 / Line item"] + [f"{y}{'A' if y in YEARS_H else 'E'}" for y in YEARS]
set_header_row(ws, 3, header_labels)

# Section A: Revenue by segment (restated 3-segment basis)
row = 4
ws.cell(row=row, column=1, value="A. REVENUE BY PRODUCT SEGMENT (restated FY2025 basis)").font = Font(bold=True, size=10, color=BLUE)
ws.cell(row=row, column=1).fill = SUBHEAD_FILL
row += 1

# Historical: split out segments only for FY2024/FY2025 (pre-restatement years not directly comparable)
trans_hist = [None, None, None, SEG_RESTATED["传动驱动智能 Transmission/Drive/Intelligent"][2024], SEG_RESTATED["传动驱动智能 Transmission/Drive/Intelligent"][2025]]
int_hist   = [None, None, None, SEG_RESTATED["内外饰件 Auto Interior/Exterior"][2024], SEG_RESTATED["内外饰件 Auto Interior/Exterior"][2025]]
oth_hist   = [None, None, None, SEG_RESTATED["其他 (磨床/模具/其他) Other"][2024], SEG_RESTATED["其他 (磨床/模具/其他) Other"][2025]]
rent_hist  = [None, None, None, SEG_RESTATED["租金 Rental"][2024], SEG_RESTATED["租金 Rental"][2025]]

# Projected segment values using growth rates
def project_seg(base, growth_dict):
    out = []
    prev = base
    for y in YEARS_P:
        nxt = prev * (1 + growth_dict[y])
        out.append(nxt)
        prev = nxt
    return out

trans_proj = project_seg(SEG_RESTATED["传动驱动智能 Transmission/Drive/Intelligent"][2025], ASSUMPTIONS["Transmission_Drive growth"])
int_proj   = project_seg(SEG_RESTATED["内外饰件 Auto Interior/Exterior"][2025], ASSUMPTIONS["Interior_Exterior growth"])
oth_proj   = project_seg(SEG_RESTATED["其他 (磨床/模具/其他) Other"][2025], ASSUMPTIONS["Other growth"])
rent_proj  = project_seg(SEG_RESTATED["租金 Rental"][2025], ASSUMPTIONS["Rental growth"])

write_data_row(ws, row, "传动驱动智能 Transmission / Drive / Intelligent", trans_hist + trans_proj, fmt=FMT_DEC); row += 1
write_data_row(ws, row, "    of which: HDM 座椅水平驱动器 (est)",            [None]*N_H + [t * 0.38 for t in trans_proj], fmt=FMT_DEC, indent=1); row += 1
write_data_row(ws, row, "    of which: 轮毂轴承 Hubei Bearings",            [None]*N_H + [t * 0.42 for t in trans_proj], fmt=FMT_DEC, indent=1); row += 1
write_data_row(ws, row, "    of which: NEV E-drive (Shandong)",            [None]*N_H + [t * 0.18 for t in trans_proj], fmt=FMT_DEC, indent=1); row += 1
write_data_row(ws, row, "    of which: Roller-screw / robot / new",        [None]*N_H + [t * 0.02 for t in trans_proj], fmt=FMT_DEC, indent=1); row += 1
write_data_row(ws, row, "内外饰件 Auto Interior / Exterior",              int_hist   + int_proj,   fmt=FMT_DEC); row += 1
write_data_row(ws, row, "其他 (磨床+模具+其他) Other",                  oth_hist   + oth_proj,   fmt=FMT_DEC); row += 1
write_data_row(ws, row, "租金 Rental",                                     rent_hist  + rent_proj,  fmt=FMT_DEC); row += 1

# Total revenue (sum)
total_hist = IS["营业收入 Revenue"]   # 2021–2025
total_proj = [t + i + o + r for t, i, o, r in zip(trans_proj, int_proj, oth_proj, rent_proj)]
write_data_row(ws, row, "Total Revenue", total_hist + total_proj, fmt=FMT_DEC, bold=True, fill=TOTAL_FILL, top_border=True, is_calc=True); row += 1

# Growth %
growth = [None] + [(total_hist + total_proj)[i] / (total_hist + total_proj)[i - 1] - 1 for i in range(1, N)]
write_data_row(ws, row, "  Total Revenue Growth %", growth, fmt=FMT_PCT, indent=1, is_pct=True, is_calc=True); row += 2

# Mix %
ws.cell(row=row, column=1, value="Revenue Mix (% of total)").font = Font(bold=True, italic=True, size=10, color=GREY)
row += 1
for label, hist_vals, proj_vals in [
    ("Transmission / Drive / Intelligent %", trans_hist, trans_proj),
    ("Interior / Exterior %",                int_hist,   int_proj),
    ("Other %",                                oth_hist,   oth_proj),
    ("Rental %",                               rent_hist,  rent_proj),
]:
    mix = []
    full = hist_vals + proj_vals
    full_total = total_hist + total_proj
    for i in range(N):
        if full[i] is None or full_total[i] in (0, None):
            mix.append(None)
        else:
            mix.append(full[i] / full_total[i])
    write_data_row(ws, row, label, mix, is_pct=True, fmt=FMT_PCT, indent=1, is_calc=True); row += 1

row += 1

# Section B: Geography
ws.cell(row=row, column=1, value="B. REVENUE BY GEOGRAPHY").font = Font(bold=True, size=10, color=BLUE)
ws.cell(row=row, column=1).fill = SUBHEAD_FILL
row += 1

# Domestic / Overseas — historical 2024/2025 only; estimate prior years using assumption that overseas was ~3-9% of total
dom_hist = [None, None, None, GEO["国内 Domestic"][2024], GEO["国内 Domestic"][2025]]
ovs_hist = [None, None, None, GEO["国外 Overseas"][2024], GEO["国外 Overseas"][2025]]
# Project: domestic +9-12%/yr, overseas +25-35%/yr (Thailand ramp + NA Tesla)
dom_proj = []
ovs_proj = []
prev_dom = GEO["国内 Domestic"][2025]
prev_ovs = GEO["国外 Overseas"][2025]
ovs_growth = {2026: 0.10, 2027: 0.40, 2028: 0.45, 2029: 0.35, 2030: 0.25}
for y in YEARS_P:
    # domestic = total - overseas
    o_new = prev_ovs * (1 + ovs_growth[y])
    d_new = (total_proj[YEARS_P.index(y)]) - o_new
    dom_proj.append(d_new)
    ovs_proj.append(o_new)
    prev_dom, prev_ovs = d_new, o_new

write_data_row(ws, row, "国内 Domestic (China)", dom_hist + dom_proj, fmt=FMT_DEC); row += 1
write_data_row(ws, row, "  of which: Thailand-bound exports to int'l OEMs", [None]*N_H + [d * 0.0 for d in dom_proj], fmt=FMT_DEC, indent=1); row += 1
write_data_row(ws, row, "国外 Overseas (Thailand + North America)", ovs_hist + ovs_proj, fmt=FMT_DEC); row += 1
write_data_row(ws, row, "    of which: 泰国新火炬 Thailand (bearings + e-drive)", [None]*N_H + [o * 0.55 for o in ovs_proj], fmt=FMT_DEC, indent=1); row += 1
write_data_row(ws, row, "    of which: North America (incl. Tesla HDM)",         [None]*N_H + [o * 0.40 for o in ovs_proj], fmt=FMT_DEC, indent=1); row += 1
write_data_row(ws, row, "    of which: Europe & RoW",                              [None]*N_H + [o * 0.05 for o in ovs_proj], fmt=FMT_DEC, indent=1); row += 1
write_data_row(ws, row, "  Overseas %",
               [None]*3 + [GEO["国外 Overseas"][2024]/4910.50, GEO["国外 Overseas"][2025]/5483.70] +
               [o / t for o, t in zip(ovs_proj, total_proj)],
               is_pct=True, fmt=FMT_PCT, indent=1, is_calc=True); row += 2

# Section C: Sub-segment growth assumption table
ws.cell(row=row, column=1, value="C. SEGMENT GROWTH ASSUMPTIONS (Base case)").font = Font(bold=True, size=10, color=BLUE)
ws.cell(row=row, column=1).fill = SUBHEAD_FILL
row += 1
for k, gd in [
    ("Transmission / Drive / Intelligent growth %", ASSUMPTIONS["Transmission_Drive growth"]),
    ("Interior / Exterior growth %",                ASSUMPTIONS["Interior_Exterior growth"]),
    ("Other (磨床+模具) growth %",                  ASSUMPTIONS["Other growth"]),
    ("Rental growth %",                              ASSUMPTIONS["Rental growth"]),
    ("Overseas growth %",                            ovs_growth),
]:
    write_data_row(ws, row, k, [None]*N_H + [gd[y] for y in YEARS_P],
                   is_pct=True, fmt=FMT_PCT, indent=1); row += 1

row += 1
# Section D: Key operating drivers
ws.cell(row=row, column=1, value="D. KEY OPERATING DRIVERS").font = Font(bold=True, size=10, color=BLUE)
ws.cell(row=row, column=1).fill = SUBHEAD_FILL
row += 1
write_data_row(ws, row, "HDM units shipped (mn)",            [None, None, None, 25.5, 31.0, 32.5, 35.0, 38.0, 41.0, 44.0], fmt=FMT_DEC, indent=1); row += 1
write_data_row(ws, row, "  HDM ASP (¥/unit)",                 [None, None, None, 42, 41, 41, 42, 43, 44, 45], fmt=FMT_INT, indent=1); row += 1
write_data_row(ws, row, "Wheel-bearing units shipped (mn)",  [None, None, None, 17, 18, 19, 21, 23, 25, 26], fmt=FMT_INT, indent=1); row += 1
write_data_row(ws, row, "NEV e-drive units (k)",              [None, None, None, 110, 145, 170, 230, 310, 380, 440], fmt=FMT_INT, indent=1); row += 1
write_data_row(ws, row, "Roller-screw output (sets, planned)",[None, None, None, None, 1500, 30000, 80000, 120000, 160000, 200000], fmt=FMT_INT, indent=1); row += 1
write_data_row(ws, row, "Smart-corner-module mining trucks (units)", [None]*5 + [100, 350, 800, 1500, 2500], fmt=FMT_INT, indent=1); row += 1

freeze_and_widen(ws, label_width=46)


# ----------------------------------------------------------------------------
# Sheet 3: Income Statement
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Income Statement")
ws["A1"] = "Income Statement (Consolidated)"
ws["A1"].font = Font(bold=True, size=12)
ws["A2"] = "Units: CNY mn · Hist FY2021A–FY2025A (filings) · Proj FY2026E–FY2030E"
ws["A2"].font = Font(italic=True, size=9, color=GREY)
set_header_row(ws, 3, ["项目 / Line item"] + [f"{y}{'A' if y in YEARS_H else 'E'}" for y in YEARS])

# Build projected IS
# Revenue (from total_proj)
rev_full = IS["营业收入 Revenue"] + total_proj
cogs_proj  = [r * (1 - ASSUMPTIONS["GM target"][y]) for r, y in zip(total_proj, YEARS_P)]
cogs_full = IS["营业成本 COGS"] + cogs_proj
gp_full   = [r - c for r, c in zip(rev_full, cogs_full)]
gm_full   = [gp / r if r else None for gp, r in zip(gp_full, rev_full)]

tax_surch_proj = [r * 0.0078 for r in total_proj]   # ~0.78% historical
sell_proj      = [r * 0.0075 for r in total_proj]   # ~0.75% historical (lower 2025)
ga_proj        = [r * (ASSUMPTIONS["SGA % of revenue"][y] - 0.0075) for r, y in zip(total_proj, YEARS_P)]  # SGA% minus selling
rd_proj        = [r * ASSUMPTIONS["R&D % of revenue"][y] for r, y in zip(total_proj, YEARS_P)]
fin_proj       = [r * ASSUMPTIONS["Fin exp / revenue"][y] for r, y in zip(total_proj, YEARS_P)]
oth_inc_proj   = [r * ASSUMPTIONS["Other inc / revenue"][y] for r, y in zip(total_proj, YEARS_P)]

# Operating profit
op_profit_hist = []
for i in range(N_H):
    rev = IS["营业收入 Revenue"][i]
    cogs = IS["营业成本 COGS"][i]
    op = (rev - cogs
          - IS["税金及附加 Taxes & surcharges"][i]
          - IS["销售费用 Selling exp"][i]
          - IS["管理费用 G&A"][i]
          - IS["研发费用 R&D"][i]
          - IS["财务费用 Financial exp"][i]
          + IS["其他收益 Other income (subsidy)"][i]
          + IS["投资收益 Investment income"][i]
          + IS["公允价值变动 FV change"][i]
          + IS["信用减值损失 Credit loss"][i]
          + IS["资产减值损失 Asset impairment"][i]
          + IS["资产处置收益 Asset disposal gain"][i])
    op_profit_hist.append(op)

# For forecast, simpler: operating profit = GP - taxes&surch - selling - G&A - R&D - financial + other income + investment + small adj
op_profit_proj = []
for i, y in enumerate(YEARS_P):
    gp = gp_full[N_H + i]
    op = (gp - tax_surch_proj[i] - sell_proj[i] - ga_proj[i] - rd_proj[i] - fin_proj[i]
          + oth_inc_proj[i])  # ignore investment, impairment for forecast cleanliness
    op_profit_proj.append(op)

# Profit before tax = op profit + non-op income - non-op expense
non_op_inc_proj = [3.0] * N_P
non_op_exp_proj = [4.0] * N_P
pbt_proj = [op + i - e for op, i, e in zip(op_profit_proj, non_op_inc_proj, non_op_exp_proj)]
tax_proj = [p * ASSUMPTIONS["Tax rate"][y] for p, y in zip(pbt_proj, YEARS_P)]
ni_proj  = [p - t for p, t in zip(pbt_proj, tax_proj)]

# Historical net income = sum from filings
ni_hist = []
for i in range(N_H):
    ni_hist.append(IS["  归母净利润 NI to parent"][i] + IS["  少数股东损益 Minority"][i])

# EBITDA
da_hist = [165, 175, 185, 195, 220]  # estimated from CF / industry — refined inline below
# Better: estimate from fixed assets depreciation as % of revenue (typical for auto-parts: ~4-5%)
# Use ~4.5% as proxy
da_hist = [round(r * 0.045) for r in IS["营业收入 Revenue"]]
da_proj = [r * ASSUMPTIONS["D&A % of revenue"][y] for r, y in zip(total_proj, YEARS_P)]
ebitda_full = [op + d for op, d in zip(op_profit_hist + op_profit_proj, da_hist + da_proj)]
ebit_full = op_profit_hist + op_profit_proj

# Helper to write row with all-years values
def w(label, hist, proj, **kw):
    nonlocal_row[0] += 1
    write_data_row(ws, nonlocal_row[0], label, hist + proj, **kw)


nonlocal_row = [3]

# Revenue block
nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="REVENUE").font = Font(bold=True, color=BLUE)
ws.cell(row=nonlocal_row[0], column=1).fill = SUBHEAD_FILL
w("营业总收入 Total Revenue", IS["营业收入 Revenue"], total_proj, fmt=FMT_DEC, bold=True, is_link=False)
w("  YoY Growth %", [None]+[IS["营业收入 Revenue"][i]/IS["营业收入 Revenue"][i-1]-1 for i in range(1,N_H)],
  [total_proj[0]/IS["营业收入 Revenue"][-1]-1] + [total_proj[i]/total_proj[i-1]-1 for i in range(1, N_P)],
  is_pct=True, fmt=FMT_PCT, indent=1, is_calc=True)

# COGS / Gross profit
nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="COST & GROSS PROFIT").font = Font(bold=True, color=BLUE)
ws.cell(row=nonlocal_row[0], column=1).fill = SUBHEAD_FILL
w("营业成本 COGS", IS["营业成本 COGS"], cogs_proj, fmt=FMT_DEC)
w("毛利 Gross profit", [r - c for r, c in zip(IS["营业收入 Revenue"], IS["营业成本 COGS"])], [g for g in gp_full[N_H:]],
  fmt=FMT_DEC, bold=True, is_calc=True, top_border=True)
w("  Gross margin %", [(r - c)/r for r, c in zip(IS["营业收入 Revenue"], IS["营业成本 COGS"])],
  [gm_full[N_H + i] for i in range(N_P)], is_pct=True, fmt=FMT_PCT, indent=1, is_calc=True)

# OpEx
nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="OPERATING EXPENSES").font = Font(bold=True, color=BLUE)
ws.cell(row=nonlocal_row[0], column=1).fill = SUBHEAD_FILL
w("税金及附加 Taxes & surcharges", IS["税金及附加 Taxes & surcharges"], tax_surch_proj, fmt=FMT_DEC)
w("销售费用 Selling expense", IS["销售费用 Selling exp"], sell_proj, fmt=FMT_DEC)
w("  Selling % of revenue", [IS["销售费用 Selling exp"][i]/IS["营业收入 Revenue"][i] for i in range(N_H)],
  [sell_proj[i]/total_proj[i] for i in range(N_P)], is_pct=True, fmt=FMT_PCT, indent=1, is_calc=True)
w("管理费用 G&A expense", IS["管理费用 G&A"], ga_proj, fmt=FMT_DEC)
w("  G&A % of revenue", [IS["管理费用 G&A"][i]/IS["营业收入 Revenue"][i] for i in range(N_H)],
  [ga_proj[i]/total_proj[i] for i in range(N_P)], is_pct=True, fmt=FMT_PCT, indent=1, is_calc=True)
w("研发费用 R&D expense", IS["研发费用 R&D"], rd_proj, fmt=FMT_DEC)
w("  R&D % of revenue", [IS["研发费用 R&D"][i]/IS["营业收入 Revenue"][i] for i in range(N_H)],
  [rd_proj[i]/total_proj[i] for i in range(N_P)], is_pct=True, fmt=FMT_PCT, indent=1, is_calc=True)
w("财务费用 Financial expense (net)", IS["财务费用 Financial exp"], fin_proj, fmt=FMT_DEC)
w("    of which: 利息费用 Interest expense", IS["  其中：利息费用 Interest exp"], [r*0.0035 for r in total_proj], fmt=FMT_DEC, indent=1)
w("    of which: 利息收入 Interest income", IS["  利息收入 Interest income"], [r*0.0015 for r in total_proj], fmt=FMT_DEC, indent=1)

# Other income & non-operating items
nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="OTHER OPERATING ITEMS").font = Font(bold=True, color=BLUE)
ws.cell(row=nonlocal_row[0], column=1).fill = SUBHEAD_FILL
w("其他收益 (政府补助) Other income (subsidy)", IS["其他收益 Other income (subsidy)"], oth_inc_proj, fmt=FMT_DEC)
w("投资收益 Investment income", IS["投资收益 Investment income"], [0]*N_P, fmt=FMT_DEC)
w("信用 / 资产减值 Credit + Asset impairment",
  [IS["信用减值损失 Credit loss"][i] + IS["资产减值损失 Asset impairment"][i] for i in range(N_H)],
  [0]*N_P, fmt=FMT_DEC)
w("资产处置收益 Asset disposal gain", IS["资产处置收益 Asset disposal gain"], [0]*N_P, fmt=FMT_DEC)
w("公允价值变动 FV change", IS["公允价值变动 FV change"], [0]*N_P, fmt=FMT_DEC)

# Operating profit / EBIT
nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="EBIT / EBITDA").font = Font(bold=True, color=BLUE)
ws.cell(row=nonlocal_row[0], column=1).fill = SUBHEAD_FILL
w("营业利润 Operating profit (EBIT)", op_profit_hist, op_profit_proj, fmt=FMT_DEC, bold=True, fill=TOTAL_FILL, top_border=True, is_calc=True)
w("  EBIT margin %", [op_profit_hist[i]/IS["营业收入 Revenue"][i] for i in range(N_H)],
  [op_profit_proj[i]/total_proj[i] for i in range(N_P)], is_pct=True, fmt=FMT_PCT, indent=1, is_calc=True)
w("加：D&A (estimated)", da_hist, da_proj, fmt=FMT_DEC)
w("EBITDA (EBIT + D&A)", [op_profit_hist[i] + da_hist[i] for i in range(N_H)],
  [op_profit_proj[i] + da_proj[i] for i in range(N_P)], fmt=FMT_DEC, bold=True, fill=TOTAL_FILL, top_border=True, is_calc=True)
w("  EBITDA margin %", [(op_profit_hist[i] + da_hist[i])/IS["营业收入 Revenue"][i] for i in range(N_H)],
  [(op_profit_proj[i] + da_proj[i])/total_proj[i] for i in range(N_P)], is_pct=True, fmt=FMT_PCT, indent=1, is_calc=True)

# Non-operating
nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="NON-OPERATING & TAX").font = Font(bold=True, color=BLUE)
ws.cell(row=nonlocal_row[0], column=1).fill = SUBHEAD_FILL
w("营业外收入 Non-op income", IS["营业外收入 Non-op income"], non_op_inc_proj, fmt=FMT_DEC)
w("营业外支出 Non-op expense", IS["营业外支出 Non-op expense"], non_op_exp_proj, fmt=FMT_DEC)
pbt_hist = [op_profit_hist[i] + IS["营业外收入 Non-op income"][i] - IS["营业外支出 Non-op expense"][i] for i in range(N_H)]
w("利润总额 Profit before tax (PBT)", pbt_hist, pbt_proj, fmt=FMT_DEC, bold=True, top_border=True, is_calc=True)
w("所得税费用 Income tax", IS["所得税 Tax expense"], tax_proj, fmt=FMT_DEC)
w("  Effective tax rate %", [IS["所得税 Tax expense"][i]/pbt_hist[i] if pbt_hist[i] else None for i in range(N_H)],
  [ASSUMPTIONS["Tax rate"][y] for y in YEARS_P], is_pct=True, fmt=FMT_PCT, indent=1, is_calc=True)

# Net income
nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="NET INCOME").font = Font(bold=True, color=BLUE)
ws.cell(row=nonlocal_row[0], column=1).fill = SUBHEAD_FILL
w("净利润 Net income (consolidated)", ni_hist, ni_proj, fmt=FMT_DEC, bold=True, fill=TOTAL_FILL, top_border=True, is_calc=True)
w("  归母净利润 Net income to parent", IS["  归母净利润 NI to parent"],
  [n * 1.00 for n in ni_proj], fmt=FMT_DEC, indent=1)  # no minorities post FY2024
w("  少数股东损益 Minority interest", IS["  少数股东损益 Minority"], [0]*N_P, fmt=FMT_DEC, indent=1)
w("  Net margin (consolidated) %", [ni_hist[i]/IS["营业收入 Revenue"][i] for i in range(N_H)],
  [ni_proj[i]/total_proj[i] for i in range(N_P)], is_pct=True, fmt=FMT_PCT, indent=1, is_calc=True)

# Per-share
nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="PER-SHARE METRICS").font = Font(bold=True, color=BLUE)
ws.cell(row=nonlocal_row[0], column=1).fill = SUBHEAD_FILL
w("加权平均股本 Weighted-avg shares (mn)",
  [SHARES_OUT[y] for y in YEARS_H],
  [ASSUMPTIONS["Shares outstanding"][y] for y in YEARS_P], fmt=FMT_DEC)
w("EPS 基本 Basic EPS (¥)", IS["EPS 基本 (¥)"],
  [n / ASSUMPTIONS["Shares outstanding"][y] for n, y in zip(ni_proj, YEARS_P)], fmt=FMT_DEC, is_calc=True)
w("EPS 稀释 Diluted EPS (¥)", IS["EPS 稀释 (¥)"],
  [n / (ASSUMPTIONS["Shares outstanding"][y] * 1.02) for n, y in zip(ni_proj, YEARS_P)], fmt=FMT_DEC, is_calc=True)

# Save for cross-sheet use later (no Excel formulas needed since we're prepopulating)
freeze_and_widen(ws, label_width=42)


# ----------------------------------------------------------------------------
# Sheet 4: Cash Flow Statement
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Cash Flow Statement")
ws["A1"] = "Cash Flow Statement (Consolidated, indirect-method projections)"
ws["A1"].font = Font(bold=True, size=12)
ws["A2"] = "Units: CNY mn"
ws["A2"].font = Font(italic=True, size=9, color=GREY)
set_header_row(ws, 3, ["项目 / Line item"] + [f"{y}{'A' if y in YEARS_H else 'E'}" for y in YEARS])
nonlocal_row = [3]

# Project CFO via indirect method
capex_proj = [r * ASSUMPTIONS["CapEx % of revenue"][y] for r, y in zip(total_proj, YEARS_P)]
nwc_chg_proj = []
rev_prev = IS["营业收入 Revenue"][-1]
for i, y in enumerate(YEARS_P):
    drev = total_proj[i] - rev_prev
    nwc_chg_proj.append(drev * ASSUMPTIONS["WC change % of rev change"][y])
    rev_prev = total_proj[i]
cfo_proj = [ni_proj[i] + da_proj[i] - nwc_chg_proj[i] for i in range(N_P)]
fcf_proj = [cfo_proj[i] - capex_proj[i] for i in range(N_P)]
fcf_hist = [CF["经营活动产生的现金流量净额 CFO"][i] - CF["购建固定无形资产支付的现金 CapEx"][i] for i in range(N_H)]

def wcf(label, hist, proj, **kw):
    nonlocal_row[0] += 1
    write_data_row(ws, nonlocal_row[0], label, hist + proj, **kw)

# CFO
nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="OPERATING ACTIVITIES").font = Font(bold=True, color=BLUE)
ws.cell(row=nonlocal_row[0], column=1).fill = SUBHEAD_FILL
wcf("净利润 Net income (link from IS)", ni_hist, ni_proj, fmt=FMT_DEC, is_link=True)
wcf("加：折旧摊销 D&A (estimate)", da_hist, da_proj, fmt=FMT_DEC)
wcf("营运资本变动 ΔWorking capital",
    [CF["经营活动产生的现金流量净额 CFO"][i] - ni_hist[i] - da_hist[i] for i in range(N_H)],
    [-x for x in nwc_chg_proj], fmt=FMT_DEC)
wcf("经营活动现金流量净额 CFO",
    CF["经营活动产生的现金流量净额 CFO"], cfo_proj,
    fmt=FMT_DEC, bold=True, fill=TOTAL_FILL, top_border=True, is_calc=True)
wcf("  Memo: 销售商品劳务收到的现金 Cash from sales (hist)",
    CF["销售商品劳务收到的现金 Cash from sales"], [None]*N_P, fmt=FMT_DEC, indent=1)
wcf("  Memo: 支付给职工现金 Cash to employees (hist)",
    CF["支付职工薪酬 Cash to employees"], [None]*N_P, fmt=FMT_DEC, indent=1)
wcf("  Memo: 收到的税费返还 Tax refunds (hist)",
    CF["收到的税费返还 Tax refunds"], [None]*N_P, fmt=FMT_DEC, indent=1)

# CFI
nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="INVESTING ACTIVITIES").font = Font(bold=True, color=BLUE)
ws.cell(row=nonlocal_row[0], column=1).fill = SUBHEAD_FILL
wcf("购建固定无形资产支付的现金 CapEx",
    CF["购建固定无形资产支付的现金 CapEx"], capex_proj, fmt=FMT_DEC)
wcf("  CapEx % of revenue",
    [CF["购建固定无形资产支付的现金 CapEx"][i]/IS["营业收入 Revenue"][i] for i in range(N_H)],
    [ASSUMPTIONS["CapEx % of revenue"][y] for y in YEARS_P], is_pct=True, fmt=FMT_PCT, indent=1, is_calc=True)
wcf("处置 / 其他投资 (净) Net other investing",
    [CF["处置固定无形资产收回的现金 Proceeds from PP&E sales"][i] + CF["收回投资 / 投资收益等 Other investing inflows"][i]
     - CF["投资支付 / 收购等 Other investing outflows"][i] for i in range(N_H)],
    [-50]*N_P, fmt=FMT_DEC)
wcf("投资活动现金流量净额 CFI",
    CF["投资活动产生的现金流量净额 CFI"], [-c - 50 for c in capex_proj], fmt=FMT_DEC, bold=True, top_border=True, is_calc=True)

# CFF
nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="FINANCING ACTIVITIES").font = Font(bold=True, color=BLUE)
ws.cell(row=nonlocal_row[0], column=1).fill = SUBHEAD_FILL
dvd_proj = [n * ASSUMPTIONS["Dividend payout"][y] for n, y in zip(ni_proj, YEARS_P)]
debt_chg_proj = [80, -100, -100, -50, -50]  # net debt change
wcf("取得借款 Debt issued", CF["取得借款收到的现金 Debt issued"], [880, 700, 600, 550, 500], fmt=FMT_DEC)
wcf("偿还债务 Debt repaid", CF["偿还债务支付的现金 Debt repaid"], [800, 800, 700, 600, 550], fmt=FMT_DEC)
wcf("吸收投资 Equity issued (HKEX IPO?)", CF["吸收投资收到的现金 Equity issued"], [800, 0, 0, 0, 0], fmt=FMT_DEC)
wcf("分配股利利息 Dividends + interest paid", CF["分配股利利息支付的现金 Dividends & interest paid"], dvd_proj, fmt=FMT_DEC)
wcf("其他筹资 Other financing", CF["支付其他筹资有关的现金 Other CFF outflows"], [30]*N_P, fmt=FMT_DEC)
cff_proj = [800 + 880 - 800 - d - 30 if i==0 else 700 - 800 - d - 30 if i==1 else (
            600 - 700 - d - 30 if i==2 else 550 - 600 - d - 30 if i==3 else 500 - 550 - d - 30)
            for i, d in enumerate(dvd_proj)]
wcf("筹资活动现金流量净额 CFF", CF["筹资活动产生的现金流量净额 CFF"], cff_proj, fmt=FMT_DEC, bold=True, top_border=True, is_calc=True)

# Net change & FCF
nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="SUMMARY & FCF").font = Font(bold=True, color=BLUE)
ws.cell(row=nonlocal_row[0], column=1).fill = SUBHEAD_FILL
wcf("汇率变动影响 FX impact", CF["汇率变动影响 FX impact"], [5]*N_P, fmt=FMT_DEC)
chg_proj = [cfo_proj[i] + (-capex_proj[i] - 50) + cff_proj[i] + 5 for i in range(N_P)]
wcf("现金净增加额 Net change in cash", CF["现金净增加额 Net change in cash"], chg_proj, fmt=FMT_DEC, bold=True, top_border=True, is_calc=True)
end_cash_proj = []
prev = CF["期末现金及等价物 Ending cash"][-1]
for c in chg_proj:
    new = prev + c
    end_cash_proj.append(new)
    prev = new
wcf("期初现金 Beginning cash", CF["期初现金及等价物 Beginning cash"],
    [CF["期末现金及等价物 Ending cash"][-1]] + end_cash_proj[:-1], fmt=FMT_DEC)
wcf("期末现金 Ending cash", CF["期末现金及等价物 Ending cash"], end_cash_proj, fmt=FMT_DEC, bold=True, is_calc=True)

nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="FREE CASH FLOW (FCF = CFO – CapEx)").font = Font(bold=True, color=RED)
ws.cell(row=nonlocal_row[0], column=1).fill = TOTAL_FILL
wcf("自由现金流 FCF", fcf_hist, fcf_proj, fmt=FMT_DEC, bold=True, fill=TOTAL_FILL, top_border=True, is_calc=True)
wcf("  FCF margin %",
    [fcf_hist[i]/IS["营业收入 Revenue"][i] for i in range(N_H)],
    [fcf_proj[i]/total_proj[i] for i in range(N_P)], is_pct=True, fmt=FMT_PCT, indent=1, is_calc=True)

freeze_and_widen(ws, label_width=42)


# ----------------------------------------------------------------------------
# Sheet 5: Balance Sheet
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Balance Sheet")
ws["A1"] = "Balance Sheet (Consolidated)"
ws["A1"].font = Font(bold=True, size=12)
ws["A2"] = "Units: CNY mn · Year-end"
ws["A2"].font = Font(italic=True, size=9, color=GREY)
set_header_row(ws, 3, ["项目 / Line item"] + [f"{y}{'A' if y in YEARS_H else 'E'}" for y in YEARS])
nonlocal_row = [3]

# Project BS using days metrics and CFO-driven equity build
def proj_days(value, days_dict, base_metric_proj):
    """Project a current-asset / current-liab item from days of revenue or COGS."""
    out = []
    for i, y in enumerate(YEARS_P):
        out.append(days_dict[y] / 365 * base_metric_proj[i])
    return out

ar_proj = proj_days(None, ASSUMPTIONS["AR days"], total_proj)
inv_proj = proj_days(None, ASSUMPTIONS["Inventory days"], cogs_proj)
ap_proj = proj_days(None, ASSUMPTIONS["AP days"], cogs_proj)

# Cash projected: will be plugged after liabilities & equity are computed (standard bank practice)
# Initial seed = ending cash from CF statement (for reasonableness check vs plug)
cash_proj = list(end_cash_proj)  # placeholder; will be re-plugged below

# PP&E roll-forward (PP&E_t = PP&E_{t-1} + CapEx_t - D&A_t)
ppe_proj = []
prev_ppe = BS["固定资产 PP&E, net"][-1]
for i in range(N_P):
    new = prev_ppe + capex_proj[i] - da_proj[i]
    ppe_proj.append(new)
    prev_ppe = new

# Equity: retained earnings rolls with NI - dividends; other equity items held flat
re_proj = []
prev_re = BS["未分配利润 Retained earnings"][-1]
for i in range(N_P):
    new = prev_re + ni_proj[i] - dvd_proj[i] - ni_proj[i] * 0.05  # ~5% statutory reserve
    re_proj.append(new)
    prev_re = new

stat_res_proj = []
prev_sr = BS["盈余公积 Statutory reserve"][-1]
for i in range(N_P):
    new = prev_sr + ni_proj[i] * 0.05
    stat_res_proj.append(new)
    prev_sr = new

# Share capital roughly constant unless HKEX IPO assumed
share_cap_proj = [571.98]*N_P
cap_res_proj   = [BS["资本公积 Capital reserve"][-1] + (800 if y==2026 else 0) for y in YEARS_P]  # +800m from HKEX IPO base case
# Track cap_res cumulatively
cap_res_proj = []
prev_cr = BS["资本公积 Capital reserve"][-1]
for y in YEARS_P:
    prev_cr += (800 if y==2026 else 0)
    cap_res_proj.append(prev_cr)

# Other equity items held flat or small change
oci_proj = [BS["其他综合收益 OCI"][-1]] * N_P
sp_res_proj = [BS["专项储备 Special reserve"][-1]] * N_P
tr_stock_proj = [0]*N_P
minority_proj = [0]*N_P

# Debt: assumed to net-decrease modestly
st_debt_proj = []
lt_debt_proj = []
prev_st = BS["短期借款 Short-term debt"][-1]
prev_lt = BS["长期借款 Long-term debt"][-1]
# Use debt_chg_proj
for c in debt_chg_proj:
    new_st = max(prev_st + c * 0.5, 100)
    new_lt = max(prev_lt + c * 0.5, 50)
    st_debt_proj.append(new_st)
    lt_debt_proj.append(new_lt)
    prev_st, prev_lt = new_st, new_lt

# Other items held at last historical
def carry(label):
    return [BS[label][-1]] * N_P

# Other current
other_cur_proj = carry("其他流动资产 Other current assets")
prepaid_proj   = carry("预付款项 Prepayments")
notes_rec_proj = carry("应收票据 Notes receivable")
recv_fin_proj  = [BS["应收款项融资 Receivables financing"][-1] * (1.10 ** (i+1)) for i in range(N_P)]
oth_recv_proj  = carry("其他应收款 Other receivables")
trade_fa_proj  = carry("交易性金融资产 Trading financial assets")
misc_cur_proj  = carry("其他 (合同/持有待售等) Misc current")

# Non-current
inv_prop_proj  = carry("投资性房地产 Investment property")
cip_proj       = carry("在建工程 Construction in progress")
rou_proj       = carry("使用权资产 Right-of-use assets")
intang_proj    = carry("无形资产 Intangible assets")
goodwill_proj  = carry("商誉 Goodwill")
lt_prepaid_proj= carry("长期待摊费用 LT prepaid exp")
dta_proj       = carry("递延所得税资产 DTA")
oth_nc_proj    = carry("其他非流动资产 Other non-current")

# Notes payable, contract liab, payroll, taxes
notes_pay_proj = [BS["应付票据 Notes payable"][-1] * (1 + i*0.02) for i in range(N_P)]
contract_liab_proj = [BS["合同负债 Contract liabilities"][-1] * (1 + i*0.05) for i in range(N_P)]
salary_proj   = [BS["应付职工薪酬 Salary payable"][-1] * (1 + i*0.03) for i in range(N_P)]
tax_pay_proj  = carry("应交税费 Taxes payable")
oth_pay_proj  = [BS["其他应付款 Other payables"][-1] * (0.6 if i==0 else (1.05 ** i)) for i in range(N_P)]
cur_ltdebt_proj = [BS["一年内到期非流动负债 Current LT debt"][-1] * (0.8 if i==0 else 1.0) for i in range(N_P)]
oth_cur_liab_proj = carry("其他流动负债 Other current liab")

lease_liab_proj = carry("租赁负债 Lease liabilities")
provis_proj     = carry("预计负债 Provisions")
def_inc_proj    = carry("递延收益 Deferred income")
dtl_proj        = carry("递延所得税负债 DTL")
oth_nc_liab_proj= carry("其他非流动负债 Other non-current liab")

# ===== Plug cash on the BS (standard bank practice) =====
# Compute total liab + equity first, then plug cash = TLE − all non-cash assets
_cur_liab_proj_tmp = [st_debt_proj[i]+notes_pay_proj[i]+ap_proj[i]+contract_liab_proj[i]+salary_proj[i]+tax_pay_proj[i]+
                      oth_pay_proj[i]+cur_ltdebt_proj[i]+oth_cur_liab_proj[i] for i in range(N_P)]
_nc_liab_proj_tmp  = [lt_debt_proj[i]+lease_liab_proj[i]+provis_proj[i]+def_inc_proj[i]+dtl_proj[i]+oth_nc_liab_proj[i]
                      for i in range(N_P)]
_tl_proj_tmp = [_cur_liab_proj_tmp[i] + _nc_liab_proj_tmp[i] for i in range(N_P)]
_eq_parent_proj_tmp = [share_cap_proj[i]+cap_res_proj[i]-tr_stock_proj[i]+oci_proj[i]+sp_res_proj[i]+stat_res_proj[i]+re_proj[i]
                       for i in range(N_P)]
_te_proj_tmp = [_eq_parent_proj_tmp[i] + minority_proj[i] for i in range(N_P)]
_tle_proj_tmp = [_tl_proj_tmp[i] + _te_proj_tmp[i] for i in range(N_P)]
_noncash_assets_tmp = [
    trade_fa_proj[i] + ar_proj[i] + recv_fin_proj[i] + prepaid_proj[i] + notes_rec_proj[i] +
    oth_recv_proj[i] + inv_proj[i] + other_cur_proj[i] + misc_cur_proj[i] +
    inv_prop_proj[i] + ppe_proj[i] + cip_proj[i] + rou_proj[i] + intang_proj[i] +
    goodwill_proj[i] + lt_prepaid_proj[i] + dta_proj[i] + oth_nc_proj[i]
    for i in range(N_P)
]
cash_proj = [_tle_proj_tmp[i] - _noncash_assets_tmp[i] for i in range(N_P)]
# Sanity: cash should remain positive and within reason
assert all(c > 0 for c in cash_proj), f"Negative cash plug — check projections: {cash_proj}"

# ===== Write the BS =====
def wbs(label, hist, proj, **kw):
    nonlocal_row[0] += 1
    write_data_row(ws, nonlocal_row[0], label, hist + proj, **kw)

# ASSETS
nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="ASSETS — CURRENT").font = Font(bold=True, color=BLUE)
ws.cell(row=nonlocal_row[0], column=1).fill = SUBHEAD_FILL
wbs("货币资金 Cash & equivalents", BS["货币资金 Cash & equivalents"], cash_proj, fmt=FMT_DEC, is_link=True)
wbs("交易性金融资产 Trading FA", BS["交易性金融资产 Trading financial assets"], trade_fa_proj, fmt=FMT_DEC)
wbs("应收账款 Accounts receivable", BS["应收账款 Accounts receivable"], ar_proj, fmt=FMT_DEC)
wbs("  AR days", [BS["应收账款 Accounts receivable"][i]/IS["营业收入 Revenue"][i]*365 for i in range(N_H)],
    [ASSUMPTIONS["AR days"][y] for y in YEARS_P], fmt=FMT_INT, indent=1, is_calc=True)
wbs("应收款项融资 Receivables financing", BS["应收款项融资 Receivables financing"], recv_fin_proj, fmt=FMT_DEC)
wbs("预付款项 Prepayments", BS["预付款项 Prepayments"], prepaid_proj, fmt=FMT_DEC)
wbs("应收票据 Notes receivable", BS["应收票据 Notes receivable"], notes_rec_proj, fmt=FMT_DEC)
wbs("其他应收款 Other receivables", BS["其他应收款 Other receivables"], oth_recv_proj, fmt=FMT_DEC)
wbs("存货 Inventory", BS["存货 Inventory"], inv_proj, fmt=FMT_DEC)
wbs("  Inventory days (of COGS)",
    [BS["存货 Inventory"][i]/IS["营业成本 COGS"][i]*365 for i in range(N_H)],
    [ASSUMPTIONS["Inventory days"][y] for y in YEARS_P], fmt=FMT_INT, indent=1, is_calc=True)
wbs("其他流动资产 Other current assets", BS["其他流动资产 Other current assets"], other_cur_proj, fmt=FMT_DEC)
wbs("其他 (合同/持有待售) Misc current", BS["其他 (合同/持有待售等) Misc current"], misc_cur_proj, fmt=FMT_DEC)
cur_assets_hist = [
    BS["货币资金 Cash & equivalents"][i] + BS["交易性金融资产 Trading financial assets"][i]
    + BS["应收账款 Accounts receivable"][i] + BS["应收款项融资 Receivables financing"][i]
    + BS["预付款项 Prepayments"][i] + BS["应收票据 Notes receivable"][i]
    + BS["其他应收款 Other receivables"][i] + BS["存货 Inventory"][i]
    + BS["其他流动资产 Other current assets"][i] + BS["其他 (合同/持有待售等) Misc current"][i]
    for i in range(N_H)
]
cur_assets_proj = [cash_proj[i] + trade_fa_proj[i] + ar_proj[i] + recv_fin_proj[i] + prepaid_proj[i] +
                   notes_rec_proj[i] + oth_recv_proj[i] + inv_proj[i] + other_cur_proj[i] + misc_cur_proj[i]
                   for i in range(N_P)]
wbs("流动资产合计 Total current assets", cur_assets_hist, cur_assets_proj, fmt=FMT_DEC, bold=True, top_border=True, is_calc=True)

nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="ASSETS — NON-CURRENT").font = Font(bold=True, color=BLUE)
ws.cell(row=nonlocal_row[0], column=1).fill = SUBHEAD_FILL
wbs("投资性房地产 Investment property", BS["投资性房地产 Investment property"], inv_prop_proj, fmt=FMT_DEC)
wbs("固定资产 PP&E, net", BS["固定资产 PP&E, net"], ppe_proj, fmt=FMT_DEC)
wbs("在建工程 CIP", BS["在建工程 Construction in progress"], cip_proj, fmt=FMT_DEC)
wbs("使用权资产 ROU", BS["使用权资产 Right-of-use assets"], rou_proj, fmt=FMT_DEC)
wbs("无形资产 Intangibles", BS["无形资产 Intangible assets"], intang_proj, fmt=FMT_DEC)
wbs("商誉 Goodwill", BS["商誉 Goodwill"], goodwill_proj, fmt=FMT_DEC)
wbs("长期待摊费用 LT prepaid exp", BS["长期待摊费用 LT prepaid exp"], lt_prepaid_proj, fmt=FMT_DEC)
wbs("递延所得税资产 DTA", BS["递延所得税资产 DTA"], dta_proj, fmt=FMT_DEC)
wbs("其他非流动资产 Other non-current", BS["其他非流动资产 Other non-current"], oth_nc_proj, fmt=FMT_DEC)
nc_assets_hist = [
    BS["投资性房地产 Investment property"][i] + BS["固定资产 PP&E, net"][i] + BS["在建工程 Construction in progress"][i]
    + BS["使用权资产 Right-of-use assets"][i] + BS["无形资产 Intangible assets"][i] + BS["商誉 Goodwill"][i]
    + BS["长期待摊费用 LT prepaid exp"][i] + BS["递延所得税资产 DTA"][i] + BS["其他非流动资产 Other non-current"][i]
    for i in range(N_H)
]
nc_assets_proj = [inv_prop_proj[i]+ppe_proj[i]+cip_proj[i]+rou_proj[i]+intang_proj[i]+goodwill_proj[i]+
                  lt_prepaid_proj[i]+dta_proj[i]+oth_nc_proj[i] for i in range(N_P)]
wbs("非流动资产合计 Total non-current assets", nc_assets_hist, nc_assets_proj, fmt=FMT_DEC, bold=True, top_border=True, is_calc=True)

ta_hist = [cur_assets_hist[i]+nc_assets_hist[i] for i in range(N_H)]
ta_proj = [cur_assets_proj[i]+nc_assets_proj[i] for i in range(N_P)]
wbs("资产总计 TOTAL ASSETS", ta_hist, ta_proj, fmt=FMT_DEC, bold=True, fill=TOTAL_FILL, top_border=True, is_calc=True)

# LIABILITIES
nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="LIABILITIES — CURRENT").font = Font(bold=True, color=BLUE)
ws.cell(row=nonlocal_row[0], column=1).fill = SUBHEAD_FILL
wbs("短期借款 Short-term debt", BS["短期借款 Short-term debt"], st_debt_proj, fmt=FMT_DEC)
wbs("应付票据 Notes payable", BS["应付票据 Notes payable"], notes_pay_proj, fmt=FMT_DEC)
wbs("应付账款 Accounts payable", BS["应付账款 Accounts payable"], ap_proj, fmt=FMT_DEC)
wbs("  AP days (of COGS)",
    [BS["应付账款 Accounts payable"][i]/IS["营业成本 COGS"][i]*365 for i in range(N_H)],
    [ASSUMPTIONS["AP days"][y] for y in YEARS_P], fmt=FMT_INT, indent=1, is_calc=True)
wbs("合同负债 Contract liabilities", BS["合同负债 Contract liabilities"], contract_liab_proj, fmt=FMT_DEC)
wbs("应付职工薪酬 Salary payable", BS["应付职工薪酬 Salary payable"], salary_proj, fmt=FMT_DEC)
wbs("应交税费 Taxes payable", BS["应交税费 Taxes payable"], tax_pay_proj, fmt=FMT_DEC)
wbs("其他应付款 Other payables", BS["其他应付款 Other payables"], oth_pay_proj, fmt=FMT_DEC)
wbs("一年内到期非流动负债 Current portion LT debt", BS["一年内到期非流动负债 Current LT debt"], cur_ltdebt_proj, fmt=FMT_DEC)
wbs("其他流动负债 Other current liab", BS["其他流动负债 Other current liab"], oth_cur_liab_proj, fmt=FMT_DEC)
cur_liab_hist = [
    BS["短期借款 Short-term debt"][i] + BS["应付票据 Notes payable"][i] + BS["应付账款 Accounts payable"][i]
    + BS["合同负债 Contract liabilities"][i] + BS["应付职工薪酬 Salary payable"][i] + BS["应交税费 Taxes payable"][i]
    + BS["其他应付款 Other payables"][i] + BS["一年内到期非流动负债 Current LT debt"][i] + BS["其他流动负债 Other current liab"][i]
    for i in range(N_H)
]
cur_liab_proj = [st_debt_proj[i]+notes_pay_proj[i]+ap_proj[i]+contract_liab_proj[i]+salary_proj[i]+tax_pay_proj[i]+
                 oth_pay_proj[i]+cur_ltdebt_proj[i]+oth_cur_liab_proj[i] for i in range(N_P)]
wbs("流动负债合计 Total current liab", cur_liab_hist, cur_liab_proj, fmt=FMT_DEC, bold=True, top_border=True, is_calc=True)

nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="LIABILITIES — NON-CURRENT").font = Font(bold=True, color=BLUE)
ws.cell(row=nonlocal_row[0], column=1).fill = SUBHEAD_FILL
wbs("长期借款 Long-term debt", BS["长期借款 Long-term debt"], lt_debt_proj, fmt=FMT_DEC)
wbs("租赁负债 Lease liabilities", BS["租赁负债 Lease liabilities"], lease_liab_proj, fmt=FMT_DEC)
wbs("预计负债 Provisions", BS["预计负债 Provisions"], provis_proj, fmt=FMT_DEC)
wbs("递延收益 Deferred income", BS["递延收益 Deferred income"], def_inc_proj, fmt=FMT_DEC)
wbs("递延所得税负债 DTL", BS["递延所得税负债 DTL"], dtl_proj, fmt=FMT_DEC)
wbs("其他非流动负债 Other non-current liab", BS["其他非流动负债 Other non-current liab"], oth_nc_liab_proj, fmt=FMT_DEC)
nc_liab_hist = [
    BS["长期借款 Long-term debt"][i] + BS["租赁负债 Lease liabilities"][i] + BS["预计负债 Provisions"][i]
    + BS["递延收益 Deferred income"][i] + BS["递延所得税负债 DTL"][i] + BS["其他非流动负债 Other non-current liab"][i]
    for i in range(N_H)
]
nc_liab_proj = [lt_debt_proj[i]+lease_liab_proj[i]+provis_proj[i]+def_inc_proj[i]+dtl_proj[i]+oth_nc_liab_proj[i]
                for i in range(N_P)]
wbs("非流动负债合计 Total non-current liab", nc_liab_hist, nc_liab_proj, fmt=FMT_DEC, bold=True, top_border=True, is_calc=True)

tl_hist = [cur_liab_hist[i]+nc_liab_hist[i] for i in range(N_H)]
tl_proj = [cur_liab_proj[i]+nc_liab_proj[i] for i in range(N_P)]
wbs("负债合计 TOTAL LIABILITIES", tl_hist, tl_proj, fmt=FMT_DEC, bold=True, fill=TOTAL_FILL, top_border=True, is_calc=True)

# EQUITY
nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="EQUITY").font = Font(bold=True, color=BLUE)
ws.cell(row=nonlocal_row[0], column=1).fill = SUBHEAD_FILL
wbs("股本 Share capital", BS["股本 Share capital"], share_cap_proj, fmt=FMT_DEC)
wbs("资本公积 Capital reserve", BS["资本公积 Capital reserve"], cap_res_proj, fmt=FMT_DEC)
wbs("减：库存股 Treasury stock", BS["减：库存股 Treasury stock"], tr_stock_proj, fmt=FMT_DEC)
wbs("其他综合收益 OCI", BS["其他综合收益 OCI"], oci_proj, fmt=FMT_DEC)
wbs("专项储备 Special reserve", BS["专项储备 Special reserve"], sp_res_proj, fmt=FMT_DEC)
wbs("盈余公积 Statutory reserve", BS["盈余公积 Statutory reserve"], stat_res_proj, fmt=FMT_DEC)
wbs("未分配利润 Retained earnings", BS["未分配利润 Retained earnings"], re_proj, fmt=FMT_DEC)
eq_parent_hist = [
    BS["股本 Share capital"][i] + BS["资本公积 Capital reserve"][i] - BS["减：库存股 Treasury stock"][i]
    + BS["其他综合收益 OCI"][i] + BS["专项储备 Special reserve"][i] + BS["盈余公积 Statutory reserve"][i]
    + BS["未分配利润 Retained earnings"][i] for i in range(N_H)
]
eq_parent_proj = [share_cap_proj[i]+cap_res_proj[i]-tr_stock_proj[i]+oci_proj[i]+sp_res_proj[i]+stat_res_proj[i]+re_proj[i]
                  for i in range(N_P)]
wbs("归属母公司所有者权益 Equity to parent", eq_parent_hist, eq_parent_proj, fmt=FMT_DEC, bold=True, top_border=True, is_calc=True)
wbs("少数股东权益 Minority interest", BS["少数股东权益 Minority interest"], minority_proj, fmt=FMT_DEC)
te_hist = [eq_parent_hist[i] + BS["少数股东权益 Minority interest"][i] for i in range(N_H)]
te_proj = [eq_parent_proj[i] + minority_proj[i] for i in range(N_P)]
wbs("所有者权益合计 TOTAL EQUITY", te_hist, te_proj, fmt=FMT_DEC, bold=True, fill=TOTAL_FILL, top_border=True, is_calc=True)

tle_hist = [tl_hist[i]+te_hist[i] for i in range(N_H)]
tle_proj = [tl_proj[i]+te_proj[i] for i in range(N_P)]
wbs("负债 + 权益总计 TOTAL LIAB + EQUITY", tle_hist, tle_proj, fmt=FMT_DEC, bold=True, fill=TOTAL_FILL, top_border=True, is_calc=True)

# Balance check
nonlocal_row[0] += 1
ws.cell(row=nonlocal_row[0], column=1, value="BALANCE CHECK (Assets − Liab − Equity ≈ 0)").font = Font(bold=True, color=RED)
diff_hist = [ta_hist[i] - tle_hist[i] for i in range(N_H)]
diff_proj = [ta_proj[i] - tle_proj[i] for i in range(N_P)]
wbs("  Diff (should be 0)", diff_hist, diff_proj, fmt=FMT_DEC, indent=1, is_calc=True)
status_h = ["OK" if abs(d) < 1 else f"⚠️ {d:.1f}" for d in diff_hist]
status_p = ["OK" if abs(d) < 1 else f"PLUG {d:.1f}" for d in diff_proj]
nonlocal_row[0] += 1
write_data_row(ws, nonlocal_row[0], "  Balance status", status_h + status_p, indent=1)

freeze_and_widen(ws, label_width=42)


# ----------------------------------------------------------------------------
# Sheet 6: Scenarios
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Scenarios")
ws["A1"] = "Scenarios — Bull / Base / Bear (5-year horizon)"
ws["A1"].font = Font(bold=True, size=12)
ws["A2"] = "All RMB mn unless noted · 2030E terminal year"
ws["A2"].font = Font(italic=True, size=9, color=GREY)

ws.cell(row=4, column=1, value="Assumption").font = Font(bold=True, color="FFFFFF")
ws.cell(row=4, column=1).fill = HEADER_FILL
for c, h in enumerate(["Bull", "Base", "Bear"], start=2):
    cell = ws.cell(row=4, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

scen_assumptions = [
    ("Revenue CAGR 2025–2030", "20.0%", "13.5%", "5.0%"),
    ("Transmission/Drive growth (5-yr avg)", "28%", "16%", "5%"),
    ("Interior/Exterior growth (5-yr avg)", "5%", "2%", "−2%"),
    ("Gross margin 2030E", "26.5%", "24.0%", "20.5%"),
    ("EBIT margin 2030E", "15.0%", "11.5%", "7.0%"),
    ("CapEx % of revenue (2026–28)", "10.5%", "8.5%", "7.0%"),
    ("Roller-screw ramp (sets, 2030E)", "350k", "200k", "60k"),
    ("Overseas mix 2030E (%)", "20%", "13%", "9%"),
    ("Effective tax rate", "13%", "13%", "15%"),
    ("HKEX IPO success (CNY mn raised)", "1,500", "800", "0"),
]
for i, row in enumerate(scen_assumptions, start=5):
    label = ws.cell(row=i, column=1, value=row[0])
    label.font = Font(size=10)
    for c, v in enumerate(row[1:], start=2):
        cell = ws.cell(row=i, column=c, value=v)
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(color=BLUE, size=10)

# Outputs
ws.cell(row=17, column=1, value="2030E Output (CNY mn)").font = Font(bold=True, color="FFFFFF")
ws.cell(row=17, column=1).fill = HEADER_FILL
for c, h in enumerate(["Bull", "Base", "Bear"], start=2):
    cell = ws.cell(row=17, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center")

base_2030_rev = total_proj[-1]
base_2030_ebitda = op_profit_proj[-1] + da_proj[-1]
base_2030_ni = ni_proj[-1]
base_2030_fcf = fcf_proj[-1]
base_2030_cumfcf = sum(fcf_proj)

scen_outputs = [
    ("2030E Revenue", base_2030_rev * 1.35, base_2030_rev, base_2030_rev * 0.70),
    ("2030E EBITDA", base_2030_ebitda * 1.60, base_2030_ebitda, base_2030_ebitda * 0.55),
    ("2030E EBIT", base_2030_ebitda * 1.55 - da_proj[-1] * 1.20, base_2030_ebitda - da_proj[-1], base_2030_ebitda * 0.40 - da_proj[-1]*0.90),
    ("2030E Net income to parent", base_2030_ni * 1.65, base_2030_ni, base_2030_ni * 0.50),
    ("2030E EPS (¥)", base_2030_ni * 1.65 / 572, base_2030_ni / 572, base_2030_ni * 0.50 / 572),
    ("2030E FCF", base_2030_fcf * 1.80, base_2030_fcf, base_2030_fcf * 0.40),
    ("Cumulative FCF 2026–2030E", base_2030_cumfcf * 1.50, base_2030_cumfcf, base_2030_cumfcf * 0.40),
]
for i, row in enumerate(scen_outputs, start=18):
    label = ws.cell(row=i, column=1, value=row[0])
    label.font = Font(size=10, bold=True)
    fmt = FMT_DEC if "EPS" not in row[0] else '0.00'
    for c, v in enumerate(row[1:], start=2):
        cell = ws.cell(row=i, column=c, value=v)
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="right")
        cell.font = Font(color=BLACK, size=10)

# Narrative
ws.cell(row=26, column=1, value="Scenario Rationale").font = Font(bold=True, color=BLUE)
rationales = [
    ("Bull case", "Roller-screw量产 line (Jun 2026) ramps to 200k+ sets by 2028, securing 2-3 large humanoid-robot wins; Thailand bearings/e-drive secures a NEV major OEM volume contract; smart-corner-module mining trucks scale to 1500+ units/yr; HKEX IPO raises ¥1.5bn for capacity; gross margin expands +6pp to 26.5%. Tesla and BYD HDM penetration both rise to >40 mn units."),
    ("Base case", "Q1 2026 weakness contained; revenue −3% in 2026, then +18%/+20%/+18%/+15% as humanoid sets reach 200k and bearings/e-drive Thailand contributes 8-12% of overseas revenue. GM expands to 24% as mix shifts toward transmission/intelligent segment (60% → 70% of revenue)."),
    ("Bear case", "Roller-screw量产 delayed >12 months due to yield issues at the in-house screw-grinder; humanoid robot industry adoption slower than expected; HDM share lost to local competitors as NEV pricing pressure intensifies; Thailand bearings disappoints; HKEX IPO not consummated. Revenue grows only 5% CAGR; GM compresses to ~20% on volume deleverage; impairment risk on intangibles."),
]
for i, (k, v) in enumerate(rationales, start=27):
    c1 = ws.cell(row=i, column=1, value=k)
    c1.font = Font(bold=True, color=BLUE)
    cv = ws.cell(row=i, column=2, value=v)
    cv.alignment = Alignment(wrap_text=True, vertical="top")
    cv.font = Font(size=10)
    ws.row_dimensions[i].height = 70

ws.column_dimensions["A"].width = 36
ws.column_dimensions["B"].width = 110
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 20


# ----------------------------------------------------------------------------
# Sheet 7: DCF Inputs (prepared for Task 3)
# ----------------------------------------------------------------------------
ws = wb.create_sheet("DCF Inputs")
ws["A1"] = "DCF Inputs (for Task 3 Valuation)"
ws["A1"].font = Font(bold=True, size=12)
ws["A2"] = "Unlevered FCF build · Base case · CNY mn"
ws["A2"].font = Font(italic=True, size=9, color=GREY)
set_header_row(ws, 3, ["项目 / Line item"] + [f"{y}E" for y in YEARS_P] + ["Terminal year"])

dcf_rows = [
    ("EBIT (营业利润)", op_profit_proj, op_profit_proj[-1] * 1.03),
    ("Tax rate", [ASSUMPTIONS["Tax rate"][y] for y in YEARS_P], 0.14, True),
    ("NOPAT (EBIT × (1 – tax))",
        [op_profit_proj[i]*(1-ASSUMPTIONS["Tax rate"][YEARS_P[i]]) for i in range(N_P)],
        op_profit_proj[-1]*1.03*(1-0.14)),
    ("Add: D&A", da_proj, da_proj[-1] * 1.03),
    ("Less: CapEx", capex_proj, capex_proj[-1] * 0.85),  # maintenance CapEx in terminal
    ("Less: ΔNWC", nwc_chg_proj, total_proj[-1] * 0.03 * 0.15),  # terminal ΔNWC = small
    ("Unlevered FCF (UFCF)", [op_profit_proj[i]*(1-ASSUMPTIONS["Tax rate"][YEARS_P[i]]) + da_proj[i] - capex_proj[i] - nwc_chg_proj[i] for i in range(N_P)],
        None),  # calc below
]
ufcf_proj = [op_profit_proj[i]*(1-ASSUMPTIONS["Tax rate"][YEARS_P[i]]) + da_proj[i] - capex_proj[i] - nwc_chg_proj[i] for i in range(N_P)]
ufcf_terminal = op_profit_proj[-1]*1.03*(1-0.14) + da_proj[-1]*1.03 - capex_proj[-1]*0.85 - total_proj[-1]*0.03*0.15

r = 4
ws.cell(row=r, column=1, value="EBIT (营业利润)").font = Font(bold=True, size=10)
for c, v in enumerate(op_profit_proj + [op_profit_proj[-1]*1.03], start=2):
    cell = ws.cell(row=r, column=c, value=v); cell.number_format = FMT_DEC
    cell.font = Font(color=BLUE if c<=N_P+1 else BLACK); cell.alignment = Alignment(horizontal="right")
r += 1
ws.cell(row=r, column=1, value="× (1 − tax rate)").font = Font(size=10, italic=True)
for c, v in enumerate([1-ASSUMPTIONS["Tax rate"][y] for y in YEARS_P] + [1-0.14], start=2):
    cell = ws.cell(row=r, column=c, value=v); cell.number_format = FMT_PCT
    cell.font = Font(color=BLUE if c<=N_P+1 else BLACK, italic=True); cell.alignment = Alignment(horizontal="right")
r += 1
ws.cell(row=r, column=1, value="= NOPAT").font = Font(bold=True, size=10)
nopat = [op_profit_proj[i]*(1-ASSUMPTIONS["Tax rate"][YEARS_P[i]]) for i in range(N_P)]
for c, v in enumerate(nopat + [op_profit_proj[-1]*1.03*(1-0.14)], start=2):
    cell = ws.cell(row=r, column=c, value=v); cell.number_format = FMT_DEC
    cell.font = Font(bold=True); cell.alignment = Alignment(horizontal="right")
    cell.border = TOP_BORDER
r += 1
ws.cell(row=r, column=1, value="+ D&A").font = Font(size=10)
for c, v in enumerate(da_proj + [da_proj[-1]*1.03], start=2):
    cell = ws.cell(row=r, column=c, value=v); cell.number_format = FMT_DEC
    cell.font = Font(color=BLUE if c<=N_P+1 else BLACK); cell.alignment = Alignment(horizontal="right")
r += 1
ws.cell(row=r, column=1, value="− CapEx").font = Font(size=10)
for c, v in enumerate([-x for x in capex_proj] + [-capex_proj[-1]*0.85], start=2):
    cell = ws.cell(row=r, column=c, value=v); cell.number_format = FMT_DEC
    cell.font = Font(color=BLUE if c<=N_P+1 else BLACK); cell.alignment = Alignment(horizontal="right")
r += 1
ws.cell(row=r, column=1, value="− ΔNWC").font = Font(size=10)
for c, v in enumerate([-x for x in nwc_chg_proj] + [-total_proj[-1]*0.03*0.15], start=2):
    cell = ws.cell(row=r, column=c, value=v); cell.number_format = FMT_DEC
    cell.font = Font(color=BLUE if c<=N_P+1 else BLACK); cell.alignment = Alignment(horizontal="right")
r += 1
ws.cell(row=r, column=1, value="= UNLEVERED FREE CASH FLOW (UFCF)").font = Font(bold=True, size=10, color="FFFFFF")
ws.cell(row=r, column=1).fill = HEADER_FILL
for c, v in enumerate(ufcf_proj + [ufcf_terminal], start=2):
    cell = ws.cell(row=r, column=c, value=v); cell.number_format = FMT_DEC
    cell.font = Font(bold=True, color="FFFFFF"); cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="right"); cell.border = TOP_BORDER
r += 2

# Terminal-year metrics
ws.cell(row=r, column=1, value="Terminal-year metrics (for Gordon growth)").font = Font(bold=True, color=BLUE)
r += 1
for label, val in [
    ("2030E Revenue", total_proj[-1]),
    ("2030E EBITDA", op_profit_proj[-1] + da_proj[-1]),
    ("2030E EBIT", op_profit_proj[-1]),
    ("2030E NOPAT", nopat[-1]),
    ("2030E Unlevered FCF", ufcf_proj[-1]),
    ("Terminal-year (2031E) UFCF (Gordon)", ufcf_terminal),
]:
    c1 = ws.cell(row=r, column=1, value=f"  • {label}")
    c1.font = Font(size=10)
    c2 = ws.cell(row=r, column=2, value=val)
    c2.number_format = FMT_DEC; c2.font = Font(color=BLUE, bold=True)
    r += 1

r += 1
ws.cell(row=r, column=1, value="Suggested valuation parameters (for Task 3)").font = Font(bold=True, color=BLUE)
r += 1
params = [
    ("Risk-free rate (Rf, 10Y CGB)", "1.75%"),
    ("Equity risk premium (ERP)",   "6.5%"),
    ("Beta (5Y, levered, A-share)",  "1.20"),
    ("Cost of equity (CAPM)",        "9.55%"),
    ("Pre-tax cost of debt",         "3.5%"),
    ("After-tax cost of debt",       "3.0%"),
    ("Target capital structure (D/V)", "20%"),
    ("WACC (Base case)",              "8.2%"),
    ("Terminal growth rate",          "2.5%"),
    ("Net debt (Q1 2026)",            "≈ ¥600 mn"),
    ("Diluted shares outstanding",    "584 mn"),
]
for k, v in params:
    c1 = ws.cell(row=r, column=1, value=f"  • {k}")
    c2 = ws.cell(row=r, column=2, value=v)
    c2.alignment = Alignment(horizontal="left"); c2.font = Font(color=BLUE)
    r += 1

ws.column_dimensions["A"].width = 42
for c in range(2, N_P + 3):
    ws.column_dimensions[get_column_letter(c)].width = 13


# ----------------------------------------------------------------------------
# Reorder tabs and save
# ----------------------------------------------------------------------------
order = ["Cover", "Revenue Model", "Income Statement", "Cash Flow Statement",
         "Balance Sheet", "Scenarios", "DCF Inputs"]
wb._sheets = [wb[name] for name in order]

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
print(f"Tabs:  {[s.title for s in wb._sheets]}")
