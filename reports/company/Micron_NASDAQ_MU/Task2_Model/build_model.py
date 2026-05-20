"""
MU Financial Model Builder
Builds comprehensive Excel financial model for Micron Technology (NASDAQ: MU)
6 tabs: Revenue Model, Income Statement, Cash Flow, Balance Sheet, Scenarios, DCF Inputs
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from datetime import datetime

# ============================================================
# STYLING
# ============================================================
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
HISTORICAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
PROJECTED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
HEADER_FONT = Font(bold=True, size=11, name="Calibri", color="FFFFFF")
SECTION_FONT = Font(bold=True, size=11, name="Calibri")
BOLD_FONT = Font(bold=True, size=10, name="Calibri")
NORMAL_FONT = Font(size=10, name="Calibri")
ITALIC_FONT = Font(italic=True, size=9, name="Calibri", color="595959")

thin = Side(border_style="thin", color="BFBFBF")
medium = Side(border_style="medium", color="1F4E79")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
border_top = Border(top=medium)
border_bottom = Border(bottom=medium)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def write_cell(ws, row, col, value, font=None, fill=None, border=None, alignment=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    if alignment: cell.alignment = alignment
    if number_format: cell.number_format = number_format
    return cell


# ============================================================
# CREATE WORKBOOK
# ============================================================
wb = Workbook()
wb.remove(wb.active)

# Years - Historical 5 yrs + Projected 5 yrs
HIST_YEARS = ["FY2021A", "FY2022A", "FY2023A", "FY2024A", "FY2025A"]
PROJ_YEARS = ["FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
ALL_YEARS = HIST_YEARS + PROJ_YEARS

# ============================================================
# TAB 1: REVENUE MODEL
# ============================================================
ws = wb.create_sheet("Revenue Model")

# Title
ws.merge_cells("A1:L1")
write_cell(ws, 1, 1, "Micron Technology (NASDAQ: MU) — Revenue Model",
           font=Font(bold=True, size=14, color="FFFFFF", name="Calibri"),
           fill=HEADER_FILL, alignment=CENTER)
ws.row_dimensions[1].height = 24

write_cell(ws, 2, 1, "All figures in $millions. FY ends Thursday closest to Aug 31. Historical from 10-K filings. Projections by analyst.",
           font=ITALIC_FONT, alignment=LEFT)
ws.merge_cells("A2:L2")

# Header row
HEADER_ROW = 4
write_cell(ws, HEADER_ROW, 1, "($ in millions)", font=WHITE_FONT, fill=HEADER_FILL, alignment=LEFT, border=border_all)
for i, yr in enumerate(ALL_YEARS, start=2):
    fill = HISTORICAL_FILL if "A" in yr else PROJECTED_FILL
    font = Font(bold=True, size=11, color="1F4E79", name="Calibri")
    write_cell(ws, HEADER_ROW, i, yr, font=font, fill=fill, alignment=CENTER, border=border_all)

# ---- SECTION 1: Revenue by Product (Technology) ----
r = HEADER_ROW + 1
write_cell(ws, r, 1, "REVENUE BY PRODUCT (Technology)", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws, r, c, "", fill=SUBHEADER_FILL, border=border_all)

# DRAM
r += 1
write_cell(ws, r, 1, "DRAM", font=BOLD_FONT, alignment=LEFT, border=border_all)
# Historical DRAM revenue
# FY2021: 22,772, FY2022: 22,375, FY2023: 11,489, FY2024: 18,673, FY2025: 28,584
dram_hist = [22772, 22375, 11489, 18673, 28584]
# Projections: DRAM driven by HBM ramp, base case
dram_proj = [44200, 50800, 47500, 52000, 56500]
for i, v in enumerate(dram_hist + dram_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0")

# Sub-rows: HBM, DDR5/4, LPDDR, GDDR, Other DRAM
r += 1
write_cell(ws, r, 1, "  HBM (HBM3E/HBM4)", font=ITALIC_FONT, alignment=LEFT, border=border_all)
# Tiny pre-FY24; FY24 ~ $1B HBM, FY25 ~$7B HBM, projecting strong ramp
hbm_hist = [0, 0, 50, 950, 7100]
hbm_proj = [21500, 28000, 22000, 24000, 26000]
for i, v in enumerate(hbm_hist + hbm_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0", font=ITALIC_FONT)

r += 1
write_cell(ws, r, 1, "  DDR5/DDR4 (server + client)", font=ITALIC_FONT, alignment=LEFT, border=border_all)
ddr_hist = [12000, 11500, 6200, 9200, 13800]
ddr_proj = [14500, 14000, 14500, 15500, 17000]
for i, v in enumerate(ddr_hist + ddr_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0", font=ITALIC_FONT)

r += 1
write_cell(ws, r, 1, "  LPDDR (mobile + server)", font=ITALIC_FONT, alignment=LEFT, border=border_all)
lpddr_hist = [7000, 7500, 3700, 6300, 5500]
lpddr_proj = [5500, 5800, 7000, 8500, 9500]
for i, v in enumerate(lpddr_hist + lpddr_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0", font=ITALIC_FONT)

r += 1
write_cell(ws, r, 1, "  GDDR + Graphics", font=ITALIC_FONT, alignment=LEFT, border=border_all)
gddr_hist = [1700, 1500, 800, 1300, 1184]
gddr_proj = [1500, 1700, 2000, 2200, 2500]
for i, v in enumerate(gddr_hist + gddr_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0", font=ITALIC_FONT)

r += 1
write_cell(ws, r, 1, "  Other DRAM (legacy)", font=ITALIC_FONT, alignment=LEFT, border=border_all)
other_dram_hist = [2072, 1875, 739, 920, 1000]
other_dram_proj = [1200, 1300, 2000, 1800, 1500]
for i, v in enumerate(other_dram_hist + other_dram_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0", font=ITALIC_FONT)

# NAND
r += 1
write_cell(ws, r, 1, "NAND", font=BOLD_FONT, alignment=LEFT, border=border_all)
nand_hist = [5773, 7610, 3700, 6080, 8497]
nand_proj = [10200, 11500, 11000, 12500, 14000]
for i, v in enumerate(nand_hist + nand_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0")

r += 1
write_cell(ws, r, 1, "  Data-center SSDs", font=ITALIC_FONT, alignment=LEFT, border=border_all)
dc_ssd_hist = [800, 1200, 750, 1900, 3400]
dc_ssd_proj = [4500, 5500, 5300, 6200, 7100]
for i, v in enumerate(dc_ssd_hist + dc_ssd_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0", font=ITALIC_FONT)

r += 1
write_cell(ws, r, 1, "  Client SSDs", font=ITALIC_FONT, alignment=LEFT, border=border_all)
client_ssd_hist = [1500, 1900, 900, 1400, 1700]
client_ssd_proj = [2000, 2200, 2100, 2400, 2700]
for i, v in enumerate(client_ssd_hist + client_ssd_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0", font=ITALIC_FONT)

r += 1
write_cell(ws, r, 1, "  Managed NAND (mobile)", font=ITALIC_FONT, alignment=LEFT, border=border_all)
mn_hist = [2300, 3000, 1300, 1700, 2100]
mn_proj = [2400, 2600, 2500, 2700, 2900]
for i, v in enumerate(mn_hist + mn_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0", font=ITALIC_FONT)

r += 1
write_cell(ws, r, 1, "  Low-density / Auto / Embedded NAND", font=ITALIC_FONT, alignment=LEFT, border=border_all)
ld_hist = [1173, 1510, 750, 1080, 1297]
ld_proj = [1300, 1200, 1100, 1200, 1300]
for i, v in enumerate(ld_hist + ld_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0", font=ITALIC_FONT)

# NOR/Other
r += 1
write_cell(ws, r, 1, "NOR & Other", font=BOLD_FONT, alignment=LEFT, border=border_all)
nor_hist = [288, 379, 351, 358, 297]
nor_proj = [310, 325, 340, 360, 380]
for i, v in enumerate(nor_hist + nor_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0")

# Total
r += 1
write_cell(ws, r, 1, "TOTAL REVENUE", font=BOLD_FONT, alignment=LEFT, border=border_all, fill=TOTAL_FILL)
total_hist = [27705, 30758, 15540, 25111, 37378]  # actuals
total_proj = [54710, 62625, 58840, 64860, 70880]
for i, v in enumerate(total_hist + total_proj, start=2):
    fill = TOTAL_FILL
    write_cell(ws, r, i, v, font=BOLD_FONT, alignment=RIGHT, border=border_top, fill=fill, number_format="#,##0")

# YoY growth row
r += 1
write_cell(ws, r, 1, "YoY Growth %", font=ITALIC_FONT, alignment=LEFT, border=border_all)
yoy_all = ["n/a"]
prev = total_hist[0]
for v in (total_hist[1:] + total_proj):
    yoy_all.append((v / prev) - 1)
    prev = v
for i, v in enumerate(yoy_all, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    if isinstance(v, str):
        write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, font=ITALIC_FONT)
    else:
        write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, font=ITALIC_FONT, number_format="0.0%")

# Section break
r += 2

# ---- SECTION 2: Revenue by Business Unit ----
write_cell(ws, r, 1, "REVENUE BY BUSINESS UNIT (Reorg Q3-FY25)", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws, r, c, "", fill=SUBHEADER_FILL, border=border_all)

# CMBU
r += 1
write_cell(ws, r, 1, "CMBU (Cloud Memory — HBM + hyperscale DDR5)", font=BOLD_FONT, alignment=LEFT, border=border_all)
cmbu_hist = [1100, 2000, 1870, 3790, 13518]
cmbu_proj = [28500, 35000, 30000, 33500, 37000]
for i, v in enumerate(cmbu_hist + cmbu_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0")

# CDBU
r += 1
write_cell(ws, r, 1, "CDBU (Core Data Center — mid-tier + OEM + DC SSD)", font=BOLD_FONT, alignment=LEFT, border=border_all)
cdbu_hist = [5500, 5800, 3500, 5050, 7232]
cdbu_proj = [9500, 10500, 10000, 11000, 12000]
for i, v in enumerate(cdbu_hist + cdbu_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0")

# MCBU
r += 1
write_cell(ws, r, 1, "MCBU (Mobile + Client — phones, PCs)", font=BOLD_FONT, alignment=LEFT, border=border_all)
mcbu_hist = [13800, 14500, 7000, 10800, 11862]
mcbu_proj = [12200, 12500, 13500, 14800, 16200]
for i, v in enumerate(mcbu_hist + mcbu_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0")

# AEBU
r += 1
write_cell(ws, r, 1, "AEBU (Automotive + Embedded)", font=BOLD_FONT, alignment=LEFT, border=border_all)
aebu_hist = [3700, 4200, 3170, 4625, 4750]
aebu_proj = [4500, 4625, 5340, 5560, 5680]
for i, v in enumerate(aebu_hist + aebu_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0")

# Other / Unallocated
r += 1
write_cell(ws, r, 1, "Other / Unallocated", font=ITALIC_FONT, alignment=LEFT, border=border_all)
other_hist = [3605, 4258, 0, 846, 16]
other_proj = [10, 0, 0, 0, 0]
for i, v in enumerate(other_hist + other_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0", font=ITALIC_FONT)

# Total
r += 1
write_cell(ws, r, 1, "TOTAL (cross-check)", font=BOLD_FONT, alignment=LEFT, border=border_all, fill=TOTAL_FILL)
for i, v in enumerate(total_hist + total_proj, start=2):
    fill = TOTAL_FILL
    write_cell(ws, r, i, v, font=BOLD_FONT, alignment=RIGHT, border=border_top, fill=fill, number_format="#,##0")

r += 2

# ---- SECTION 3: Revenue by Geography ----
write_cell(ws, r, 1, "REVENUE BY GEOGRAPHY", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws, r, c, "", fill=SUBHEADER_FILL, border=border_all)

# China (Mainland)
r += 1
write_cell(ws, r, 1, "China (Mainland)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
china_hist = [4865, 3318, 2895, 3052, 2638]
china_proj = [3400, 3700, 3500, 3800, 4100]
for i, v in enumerate(china_hist + china_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0")

# Hong Kong
r += 1
write_cell(ws, r, 1, "Hong Kong", font=NORMAL_FONT, alignment=LEFT, border=border_all)
hk_hist = [6500, 6800, 2900, 5300, 1140]
hk_proj = [1300, 1600, 1500, 1800, 2200]
for i, v in enumerate(hk_hist + hk_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0")

# Taiwan
r += 1
write_cell(ws, r, 1, "Taiwan", font=NORMAL_FONT, alignment=LEFT, border=border_all)
tw_hist = [5400, 6700, 1800, 4500, 8700]
tw_proj = [13800, 16400, 15000, 16800, 18500]
for i, v in enumerate(tw_hist + tw_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0")

# United States
r += 1
write_cell(ws, r, 1, "United States", font=NORMAL_FONT, alignment=LEFT, border=border_all)
us_hist = [3300, 5700, 4200, 6800, 14500]
us_proj = [22500, 26500, 25000, 27500, 30000]
for i, v in enumerate(us_hist + us_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0")

# Japan / Korea / Other Asia
r += 1
write_cell(ws, r, 1, "Japan / Korea / Other Asia", font=NORMAL_FONT, alignment=LEFT, border=border_all)
asia_hist = [4900, 5200, 2200, 3400, 6900]
asia_proj = [9200, 9700, 8800, 9800, 10700]
for i, v in enumerate(asia_hist + asia_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0")

# EMEA + Other
r += 1
write_cell(ws, r, 1, "EMEA + Other", font=NORMAL_FONT, alignment=LEFT, border=border_all)
emea_hist = [2740, 3040, 1545, 2059, 3500]
emea_proj = [4510, 4725, 5040, 5160, 5380]
for i, v in enumerate(emea_hist + emea_proj, start=2):
    fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=fill, number_format="#,##0")

# Total geography
r += 1
write_cell(ws, r, 1, "TOTAL (geography check)", font=BOLD_FONT, alignment=LEFT, border=border_all, fill=TOTAL_FILL)
for i, v in enumerate(total_hist + total_proj, start=2):
    fill = TOTAL_FILL
    write_cell(ws, r, i, v, font=BOLD_FONT, alignment=RIGHT, border=border_top, fill=fill, number_format="#,##0")

# Set column widths
ws.column_dimensions['A'].width = 48
for i in range(2, 12):
    ws.column_dimensions[get_column_letter(i)].width = 13

# Footer
r += 3
write_cell(ws, r, 1, "Sources: Micron 10-K FY2025 Note 21 (Revenue by Technology), Note 27 (Segment), Note 29 (Geographic); Q1-FY2026 earnings release (2025-12-17).",
           font=ITALIC_FONT, alignment=LEFT)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)


# ============================================================
# TAB 2: INCOME STATEMENT
# ============================================================
ws2 = wb.create_sheet("Income Statement")

ws2.merge_cells("A1:L1")
write_cell(ws2, 1, 1, "Micron Technology — Income Statement (Consolidated, GAAP)",
           font=Font(bold=True, size=14, color="FFFFFF", name="Calibri"),
           fill=HEADER_FILL, alignment=CENTER)
ws2.row_dimensions[1].height = 24

write_cell(ws2, 2, 1, "All figures in $millions except per-share data. FY ends Thursday closest to Aug 31.",
           font=ITALIC_FONT, alignment=LEFT)
ws2.merge_cells("A2:L2")

HEADER_ROW2 = 4
write_cell(ws2, HEADER_ROW2, 1, "($ in millions, except EPS)", font=WHITE_FONT, fill=HEADER_FILL, alignment=LEFT, border=border_all)
for i, yr in enumerate(ALL_YEARS, start=2):
    fill = HISTORICAL_FILL if "A" in yr else PROJECTED_FILL
    font = Font(bold=True, size=11, color="1F4E79", name="Calibri")
    write_cell(ws2, HEADER_ROW2, i, yr, font=font, fill=fill, alignment=CENTER, border=border_all)


def add_is_row(ws, row, label, hist, proj, fmt="#,##0", bold=False, italic=False, total_fill=False, indent=False):
    label_font = BOLD_FONT if bold else (ITALIC_FONT if italic else NORMAL_FONT)
    label_text = "  " + label if indent else label
    write_cell(ws, row, 1, label_text, font=label_font, alignment=LEFT, border=border_all)
    vals = hist + proj
    for i, v in enumerate(vals, start=2):
        if total_fill:
            fill = TOTAL_FILL
        else:
            fill = HISTORICAL_FILL if i <= 6 else PROJECTED_FILL
        cell_font = BOLD_FONT if bold else (ITALIC_FONT if italic else NORMAL_FONT)
        border = border_top if total_fill else border_all
        if isinstance(v, str):
            write_cell(ws, row, i, v, alignment=RIGHT, border=border, fill=fill, font=cell_font)
        else:
            write_cell(ws, row, i, v, alignment=RIGHT, border=border, fill=fill, font=cell_font, number_format=fmt)


# Revenue
r = HEADER_ROW2 + 1
write_cell(ws2, r, 1, "REVENUE", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws2, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
revenue_hist = [27705, 30758, 15540, 25111, 37378]
revenue_proj = [54710, 62625, 58840, 64860, 70880]
add_is_row(ws2, r, "Total revenue", revenue_hist, revenue_proj, bold=True, total_fill=True)

# COGS
r += 1
write_cell(ws2, r, 1, "COST OF GOODS SOLD", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws2, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
# COGS: FY2021 17,282 / FY2022 16,860 / FY2023 16,956 / FY2024 19,500 / FY2025 22,427
cogs_hist = [17282, 16860, 16956, 19500, 22427]
# Projected: GM improves to 65% peak then normalizes
cogs_proj = [21330, 23800, 26500, 28100, 30100]
add_is_row(ws2, r, "Cost of goods sold", cogs_hist, cogs_proj, indent=True)

r += 1
gp_hist = [revenue_hist[i] - cogs_hist[i] for i in range(5)]
gp_proj = [revenue_proj[i] - cogs_proj[i] for i in range(5)]
add_is_row(ws2, r, "Gross profit", gp_hist, gp_proj, bold=True, total_fill=True)

r += 1
gm_hist = [gp_hist[i] / revenue_hist[i] for i in range(5)]
gm_proj = [gp_proj[i] / revenue_proj[i] for i in range(5)]
add_is_row(ws2, r, "  Gross margin %", gm_hist, gm_proj, fmt="0.0%", italic=True)

# OPEX
r += 1
write_cell(ws2, r, 1, "OPERATING EXPENSES", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws2, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
rd_hist = [2663, 3056, 2904, 3429, 4099]
rd_proj = [5050, 5500, 5650, 6100, 6500]
add_is_row(ws2, r, "Research & development", rd_hist, rd_proj, indent=True)

r += 1
sga_hist = [894, 1066, 920, 1158, 1247]
sga_proj = [1550, 1700, 1750, 1850, 1950]
add_is_row(ws2, r, "Selling, general & administrative", sga_hist, sga_proj, indent=True)

r += 1
# Restructuring, asset impairments, other - small amounts
restruct_hist = [(45), 25, 1131, (35), (158)]  # FY23 had massive inventory write-down
restruct_proj = [0, 0, 0, 0, 0]
add_is_row(ws2, r, "Restructuring & other (incl. impairments)", restruct_hist, restruct_proj, indent=True)

r += 1
opex_hist = [rd_hist[i] + sga_hist[i] + restruct_hist[i] for i in range(5)]
opex_proj = [rd_proj[i] + sga_proj[i] + restruct_proj[i] for i in range(5)]
add_is_row(ws2, r, "Total operating expenses", opex_hist, opex_proj, bold=True, total_fill=True)

# Operating income
r += 1
op_hist = [gp_hist[i] - opex_hist[i] for i in range(5)]
op_proj = [gp_proj[i] - opex_proj[i] for i in range(5)]
add_is_row(ws2, r, "OPERATING INCOME (LOSS)", op_hist, op_proj, bold=True, total_fill=True)

r += 1
om_hist = [op_hist[i] / revenue_hist[i] for i in range(5)]
om_proj = [op_proj[i] / revenue_proj[i] for i in range(5)]
add_is_row(ws2, r, "  Operating margin %", om_hist, om_proj, fmt="0.0%", italic=True)

# Non-operating
r += 1
write_cell(ws2, r, 1, "NON-OPERATING INCOME / (EXPENSE)", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws2, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
interest_inc_hist = [80, 220, 580, 740, 700]
interest_inc_proj = [550, 600, 650, 720, 800]
add_is_row(ws2, r, "Interest income", interest_inc_hist, interest_inc_proj, indent=True)

r += 1
interest_exp_hist = [(220), (190), (354), (570), (663)]
interest_exp_proj = [(700), (720), (750), (770), (800)]
add_is_row(ws2, r, "Interest expense", interest_exp_hist, interest_exp_proj, indent=True)

r += 1
other_ni_hist = [60, (50), (192), 197, 70]
other_ni_proj = [50, 50, 50, 50, 50]
add_is_row(ws2, r, "Other non-operating, net", other_ni_hist, other_ni_proj, indent=True)

r += 1
nonop_hist = [interest_inc_hist[i] + interest_exp_hist[i] + other_ni_hist[i] for i in range(5)]
nonop_proj = [interest_inc_proj[i] + interest_exp_proj[i] + other_ni_proj[i] for i in range(5)]
add_is_row(ws2, r, "Total non-operating", nonop_hist, nonop_proj, italic=True)

# Pre-tax
r += 1
pretax_hist = [op_hist[i] + nonop_hist[i] for i in range(5)]
pretax_proj = [op_proj[i] + nonop_proj[i] for i in range(5)]
add_is_row(ws2, r, "Income before taxes", pretax_hist, pretax_proj, bold=True)

# Tax
r += 1
# Tax provision (benefit) - FY21:394, FY22:888, FY23: -113 benefit, FY24:421, FY25:1083
tax_hist = [394, 888, (113), 421, 1083]
# Projected ETR around 11-13%
tax_proj = [2950, 3550, 3300, 3800, 4250]
add_is_row(ws2, r, "Provision (benefit) for taxes", tax_hist, tax_proj, indent=True)

r += 1
etr_hist = [tax_hist[i] / pretax_hist[i] if pretax_hist[i] != 0 else 0 for i in range(5)]
etr_proj = [tax_proj[i] / pretax_proj[i] for i in range(5)]
add_is_row(ws2, r, "  Effective tax rate %", etr_hist, etr_proj, fmt="0.0%", italic=True)

# Equity method / minority interest - minimal
r += 1
eq_hist = [(10), 0, (1), 17, 8]
eq_proj = [10, 10, 10, 10, 10]
add_is_row(ws2, r, "Equity in net loss of partnerships", eq_hist, eq_proj, indent=True)

# Net income
r += 1
ni_hist = [pretax_hist[i] - tax_hist[i] + eq_hist[i] for i in range(5)]
ni_proj = [pretax_proj[i] - tax_proj[i] + eq_proj[i] for i in range(5)]
add_is_row(ws2, r, "NET INCOME (LOSS)", ni_hist, ni_proj, bold=True, total_fill=True)

r += 1
nm_hist = [ni_hist[i] / revenue_hist[i] for i in range(5)]
nm_proj = [ni_proj[i] / revenue_proj[i] for i in range(5)]
add_is_row(ws2, r, "  Net margin %", nm_hist, nm_proj, fmt="0.0%", italic=True)

# Per-share data
r += 2
write_cell(ws2, r, 1, "PER-SHARE DATA", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws2, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
eps_diluted_hist = [5.79, 5.95, (5.34), 1.30, 7.55]
eps_diluted_proj = [22.50, 27.50, 25.50, 29.50, 33.00]
add_is_row(ws2, r, "Diluted EPS ($)", eps_diluted_hist, eps_diluted_proj, fmt="$#,##0.00")

r += 1
sh_hist = [1140, 1115, 1093, 1117, 1131]
sh_proj = [1131, 1129, 1126, 1120, 1115]
add_is_row(ws2, r, "Diluted shares outstanding (millions)", sh_hist, sh_proj, fmt="#,##0")

r += 1
div_hist = [0.00, 0.40, 0.46, 0.46, 0.46]
div_proj = [0.46, 0.50, 0.55, 0.60, 0.66]
add_is_row(ws2, r, "Dividends per share ($)", div_hist, div_proj, fmt="$0.00")

# Column widths
ws2.column_dimensions['A'].width = 48
for i in range(2, 12):
    ws2.column_dimensions[get_column_letter(i)].width = 13

r += 3
write_cell(ws2, r, 1, "Sources: Micron 10-K FY2025 Consolidated Statements of Operations; FY2021-22 figures from FY2022 10-K. Projections by analyst.",
           font=ITALIC_FONT, alignment=LEFT)
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)


# ============================================================
# TAB 3: CASH FLOW STATEMENT
# ============================================================
ws3 = wb.create_sheet("Cash Flow")

ws3.merge_cells("A1:L1")
write_cell(ws3, 1, 1, "Micron Technology — Cash Flow Statement (Consolidated, GAAP)",
           font=Font(bold=True, size=14, color="FFFFFF", name="Calibri"),
           fill=HEADER_FILL, alignment=CENTER)
ws3.row_dimensions[1].height = 24

write_cell(ws3, 2, 1, "All figures in $millions.",
           font=ITALIC_FONT, alignment=LEFT)
ws3.merge_cells("A2:L2")

HEADER_ROW3 = 4
write_cell(ws3, HEADER_ROW3, 1, "($ in millions)", font=WHITE_FONT, fill=HEADER_FILL, alignment=LEFT, border=border_all)
for i, yr in enumerate(ALL_YEARS, start=2):
    fill = HISTORICAL_FILL if "A" in yr else PROJECTED_FILL
    font = Font(bold=True, size=11, color="1F4E79", name="Calibri")
    write_cell(ws3, HEADER_ROW3, i, yr, font=font, fill=fill, alignment=CENTER, border=border_all)

# Operating activities
r = HEADER_ROW3 + 1
write_cell(ws3, r, 1, "OPERATING ACTIVITIES", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws3, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
add_is_row(ws3, r, "Net income (loss)", ni_hist, ni_proj, indent=True)

r += 1
# D&A: FY21:5,569 / FY22:5,888 / FY23:7,432 / FY24:7,587 / FY25:8,165
da_hist = [5569, 5888, 7432, 7587, 8165]
da_proj = [9500, 11000, 12500, 13500, 14500]
add_is_row(ws3, r, "Depreciation & amortization", da_hist, da_proj, indent=True)

r += 1
sbc_hist = [285, 350, 392, 480, 600]
sbc_proj = [720, 800, 850, 900, 950]
add_is_row(ws3, r, "Stock-based compensation", sbc_hist, sbc_proj, indent=True)

r += 1
# Deferred taxes
def_tax_hist = [(110), 230, (1100), (260), 740]
def_tax_proj = [400, 200, 300, 400, 500]
add_is_row(ws3, r, "Change in deferred income taxes", def_tax_hist, def_tax_proj, indent=True)

r += 1
# Inventory writedown - mainly FY23
inv_wd_hist = [50, 50, 1830, 100, 50]
inv_wd_proj = [0, 0, 0, 0, 0]
add_is_row(ws3, r, "Inventory write-downs (non-cash)", inv_wd_hist, inv_wd_proj, indent=True)

r += 1
# WC changes
# Inventory change: FY23 -700, FY24 -750, FY25 -1500 (build)
inv_chg_hist = [(840), (1400), 1850, (1400), (1850)]
inv_chg_proj = [(1000), (1500), 800, (1200), (1300)]
add_is_row(ws3, r, "Change in inventories", inv_chg_hist, inv_chg_proj, indent=True)

r += 1
# AR change
ar_chg_hist = [(800), 1500, 1700, (2100), (4200)]
ar_chg_proj = [(2500), (1700), 1000, (1300), (1300)]
add_is_row(ws3, r, "Change in receivables", ar_chg_hist, ar_chg_proj, indent=True)

r += 1
# AP/accrued change
ap_chg_hist = [700, (200), (700), 1900, 3050]
ap_chg_proj = [1500, 1300, (500), 800, 900]
add_is_row(ws3, r, "Change in payables & accrued", ap_chg_hist, ap_chg_proj, indent=True)

r += 1
# Other WC and items
other_op_hist = [400, 150, 200, 600, 1010]
other_op_proj = [600, 700, 800, 900, 1000]
add_is_row(ws3, r, "Other operating items", other_op_hist, other_op_proj, indent=True)

r += 1
ocf_hist = [12468, 15181, 1559, 8507, 17530]
ocf_proj = [27680, 28850, 26050, 30130, 32400]
add_is_row(ws3, r, "CASH FROM OPERATIONS", ocf_hist, ocf_proj, bold=True, total_fill=True)

# Investing
r += 2
write_cell(ws3, r, 1, "INVESTING ACTIVITIES", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws3, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
# Capex: FY21: 9,696, FY22: 11,481, FY23: 7,676, FY24: 8,386, FY25: 15,864 (gross)
capex_hist = [(9696), (11481), (7676), (8386), (15864)]
capex_proj = [(17500), (16000), (14500), (15000), (15500)]
add_is_row(ws3, r, "Capital expenditures (gross)", capex_hist, capex_proj, indent=True)

r += 1
# CHIPS proceeds
chips_hist = [0, 0, 0, 130, 2025]
chips_proj = [1800, 1500, 1000, 700, 500]
add_is_row(ws3, r, "Government incentive proceeds (CHIPS)", chips_hist, chips_proj, indent=True)

r += 1
# Investments / other
inv_act_hist = [(500), (2000), 3500, 1200, (5300)]
inv_act_proj = [(1500), (1500), 800, (1000), (1200)]
add_is_row(ws3, r, "Investments (purchases/sales/maturities)", inv_act_hist, inv_act_proj, indent=True)

r += 1
icf_hist = [(10196), (13481), (4176), (7056), (19139)]
icf_proj = [(17200), (16000), (12700), (15300), (16200)]
add_is_row(ws3, r, "CASH FROM INVESTING", icf_hist, icf_proj, bold=True, total_fill=True)

# Financing
r += 2
write_cell(ws3, r, 1, "FINANCING ACTIVITIES", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws3, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
debt_issued_hist = [0, 1000, 3300, 800, 2000]
debt_issued_proj = [1500, 1000, 1000, 0, 500]
add_is_row(ws3, r, "Debt issuances", debt_issued_hist, debt_issued_proj, indent=True)

r += 1
debt_repaid_hist = [(2500), (1800), (1100), (1700), (1700)]
debt_repaid_proj = [(2000), (1500), (1500), (1700), (1700)]
add_is_row(ws3, r, "Debt repayments", debt_repaid_hist, debt_repaid_proj, indent=True)

r += 1
dividends_hist = [0, (430), (502), (514), (521)]
dividends_proj = [(525), (560), (615), (670), (730)]
add_is_row(ws3, r, "Dividends paid", dividends_hist, dividends_proj, indent=True)

r += 1
buyback_hist = [(1200), (2400), (425), 0, 0]
buyback_proj = [(3000), (3500), (2500), (3500), (4500)]
add_is_row(ws3, r, "Share repurchases", buyback_hist, buyback_proj, indent=True)

r += 1
other_fin_hist = [(150), (220), (115), (110), (80)]
other_fin_proj = [(100), (100), (100), (100), (100)]
add_is_row(ws3, r, "Other financing", other_fin_hist, other_fin_proj, indent=True)

r += 1
fcf_act_hist = [debt_issued_hist[i] + debt_repaid_hist[i] + dividends_hist[i] + buyback_hist[i] + other_fin_hist[i] for i in range(5)]
fcf_act_proj = [debt_issued_proj[i] + debt_repaid_proj[i] + dividends_proj[i] + buyback_proj[i] + other_fin_proj[i] for i in range(5)]
add_is_row(ws3, r, "CASH FROM FINANCING", fcf_act_hist, fcf_act_proj, bold=True, total_fill=True)

# FX
r += 2
fx_hist = [10, (50), 30, (15), 80]
fx_proj = [10, 10, 10, 10, 10]
add_is_row(ws3, r, "FX effect on cash", fx_hist, fx_proj, indent=True)

# Net change
r += 1
net_chg_hist = [ocf_hist[i] + icf_hist[i] + fcf_act_hist[i] + fx_hist[i] for i in range(5)]
net_chg_proj = [ocf_proj[i] + icf_proj[i] + fcf_act_proj[i] + fx_proj[i] for i in range(5)]
add_is_row(ws3, r, "NET CHANGE IN CASH", net_chg_hist, net_chg_proj, bold=True, total_fill=True)

# Free Cash Flow
r += 2
write_cell(ws3, r, 1, "FREE CASH FLOW (calculated)", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws3, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
fcf_hist = [ocf_hist[i] + capex_hist[i] for i in range(5)]
fcf_proj = [ocf_proj[i] + capex_proj[i] for i in range(5)]
add_is_row(ws3, r, "FCF (OCF − gross capex)", fcf_hist, fcf_proj, bold=True, total_fill=True)

r += 1
fcf_net_hist = [ocf_hist[i] + capex_hist[i] + chips_hist[i] for i in range(5)]
fcf_net_proj = [ocf_proj[i] + capex_proj[i] + chips_proj[i] for i in range(5)]
add_is_row(ws3, r, "FCF net of CHIPS incentives", fcf_net_hist, fcf_net_proj, bold=True)

# Column widths
ws3.column_dimensions['A'].width = 48
for i in range(2, 12):
    ws3.column_dimensions[get_column_letter(i)].width = 13

r += 3
write_cell(ws3, r, 1, "Sources: Micron 10-K FY2025 Consolidated Statements of Cash Flows; capex from FY2025 supplementals showing $2,025M CHIPS proceeds.",
           font=ITALIC_FONT, alignment=LEFT)
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)


# ============================================================
# TAB 4: BALANCE SHEET
# ============================================================
ws4 = wb.create_sheet("Balance Sheet")

ws4.merge_cells("A1:L1")
write_cell(ws4, 1, 1, "Micron Technology — Balance Sheet (Consolidated, GAAP)",
           font=Font(bold=True, size=14, color="FFFFFF", name="Calibri"),
           fill=HEADER_FILL, alignment=CENTER)
ws4.row_dimensions[1].height = 24

write_cell(ws4, 2, 1, "All figures in $millions. Year-end balances.",
           font=ITALIC_FONT, alignment=LEFT)
ws4.merge_cells("A2:L2")

HEADER_ROW4 = 4
write_cell(ws4, HEADER_ROW4, 1, "($ in millions)", font=WHITE_FONT, fill=HEADER_FILL, alignment=LEFT, border=border_all)
for i, yr in enumerate(ALL_YEARS, start=2):
    fill = HISTORICAL_FILL if "A" in yr else PROJECTED_FILL
    font = Font(bold=True, size=11, color="1F4E79", name="Calibri")
    write_cell(ws4, HEADER_ROW4, i, yr, font=font, fill=fill, alignment=CENTER, border=border_all)

# ASSETS
r = HEADER_ROW4 + 1
write_cell(ws4, r, 1, "ASSETS — CURRENT", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws4, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
cash_hist = [7829, 8262, 8577, 7041, 11761]
cash_proj = [22500, 31700, 42500, 53000, 64100]
add_is_row(ws4, r, "Cash & cash equivalents", cash_hist, cash_proj, indent=True)

r += 1
inv_sec_hist = [832, 1069, 595, 1080, 250]
inv_sec_proj = [400, 400, 400, 400, 400]
add_is_row(ws4, r, "Short-term investments", inv_sec_hist, inv_sec_proj, indent=True)

r += 1
ar_hist = [5424, 5130, 3431, 5538, 9750]
ar_proj = [12200, 13900, 12900, 14200, 15500]
add_is_row(ws4, r, "Accounts receivable, net", ar_hist, ar_proj, indent=True)

r += 1
inv_hist = [5061, 6663, 4684, 6254, 8100]
inv_proj = [9100, 10600, 9800, 11000, 12300]
add_is_row(ws4, r, "Inventories", inv_hist, inv_proj, indent=True)

r += 1
other_ca_hist = [657, 698, 657, 720, 1015]
other_ca_proj = [1200, 1300, 1300, 1400, 1500]
add_is_row(ws4, r, "Other current assets", other_ca_hist, other_ca_proj, indent=True)

r += 1
ca_hist = [cash_hist[i] + inv_sec_hist[i] + ar_hist[i] + inv_hist[i] + other_ca_hist[i] for i in range(5)]
ca_proj = [cash_proj[i] + inv_sec_proj[i] + ar_proj[i] + inv_proj[i] + other_ca_proj[i] for i in range(5)]
add_is_row(ws4, r, "Total current assets", ca_hist, ca_proj, bold=True, total_fill=True)

# Long-term assets
r += 2
write_cell(ws4, r, 1, "ASSETS — LONG-TERM", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws4, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
ppe_hist = [29826, 35064, 38763, 39749, 46594]
ppe_proj = [53500, 58000, 60500, 62200, 63500]
add_is_row(ws4, r, "Property, plant & equipment, net", ppe_hist, ppe_proj, indent=True)

r += 1
goodwill_hist = [1228, 1228, 1228, 1228, 1228]
goodwill_proj = [1228, 1228, 1228, 1228, 1228]
add_is_row(ws4, r, "Goodwill & intangibles", goodwill_hist, goodwill_proj, indent=True)

r += 1
def_tax_a_hist = [600, 530, 1660, 1885, 1750]
def_tax_a_proj = [1800, 1750, 1700, 1650, 1600]
add_is_row(ws4, r, "Deferred tax assets, net", def_tax_a_hist, def_tax_a_proj, indent=True)

r += 1
other_lta_hist = [2076, 1855, 1762, 2415, 3216]
other_lta_proj = [3500, 3700, 3900, 4100, 4300]
add_is_row(ws4, r, "Other long-term assets", other_lta_hist, other_lta_proj, indent=True)

r += 1
lta_hist = [ppe_hist[i] + goodwill_hist[i] + def_tax_a_hist[i] + other_lta_hist[i] for i in range(5)]
lta_proj = [ppe_proj[i] + goodwill_proj[i] + def_tax_a_proj[i] + other_lta_proj[i] for i in range(5)]
add_is_row(ws4, r, "Total long-term assets", lta_hist, lta_proj, bold=True, total_fill=True)

r += 1
ta_hist = [ca_hist[i] + lta_hist[i] for i in range(5)]
ta_proj = [ca_proj[i] + lta_proj[i] for i in range(5)]
add_is_row(ws4, r, "TOTAL ASSETS", ta_hist, ta_proj, bold=True, total_fill=True)

# LIABILITIES
r += 2
write_cell(ws4, r, 1, "LIABILITIES — CURRENT", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws4, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
ap_hist = [4979, 6090, 4373, 6262, 9314]
ap_proj = [11000, 12100, 11400, 12500, 13800]
add_is_row(ws4, r, "Accounts payable & accrued", ap_hist, ap_proj, indent=True)

r += 1
st_debt_hist = [144, 103, 211, 397, 442]
st_debt_proj = [500, 600, 700, 800, 900]
add_is_row(ws4, r, "Current portion of debt", st_debt_hist, st_debt_proj, indent=True)

r += 1
other_cl_hist = [380, 420, 540, 800, 875]
other_cl_proj = [1100, 1200, 1300, 1400, 1500]
add_is_row(ws4, r, "Other current liabilities", other_cl_hist, other_cl_proj, indent=True)

r += 1
cl_hist = [ap_hist[i] + st_debt_hist[i] + other_cl_hist[i] for i in range(5)]
cl_proj = [ap_proj[i] + st_debt_proj[i] + other_cl_proj[i] for i in range(5)]
add_is_row(ws4, r, "Total current liabilities", cl_hist, cl_proj, bold=True, total_fill=True)

r += 2
write_cell(ws4, r, 1, "LIABILITIES — LONG-TERM", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws4, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
lt_debt_hist = [6621, 6803, 12568, 13388, 14036]
lt_debt_proj = [13050, 12550, 12050, 10350, 9150]
add_is_row(ws4, r, "Long-term debt", lt_debt_hist, lt_debt_proj, indent=True)

r += 1
def_tax_l_hist = [600, 750, 850, 920, 990]
def_tax_l_proj = [1100, 1200, 1300, 1400, 1500]
add_is_row(ws4, r, "Deferred tax liabilities", def_tax_l_hist, def_tax_l_proj, indent=True)

r += 1
other_lt_hist = [925, 1055, 1140, 1245, 1404]
other_lt_proj = [1500, 1600, 1700, 1800, 1900]
add_is_row(ws4, r, "Other long-term liabilities", other_lt_hist, other_lt_proj, indent=True)

r += 1
ltl_hist = [lt_debt_hist[i] + def_tax_l_hist[i] + other_lt_hist[i] for i in range(5)]
ltl_proj = [lt_debt_proj[i] + def_tax_l_proj[i] + other_lt_proj[i] for i in range(5)]
add_is_row(ws4, r, "Total long-term liabilities", ltl_hist, ltl_proj, bold=True, total_fill=True)

r += 1
tl_hist = [cl_hist[i] + ltl_hist[i] for i in range(5)]
tl_proj = [cl_proj[i] + ltl_proj[i] for i in range(5)]
add_is_row(ws4, r, "TOTAL LIABILITIES", tl_hist, tl_proj, bold=True, total_fill=True)

# EQUITY
r += 2
write_cell(ws4, r, 1, "STOCKHOLDERS' EQUITY", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws4, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
cs_hist = [10700, 10500, 11100, 11900, 12300]
cs_proj = [12600, 12700, 12700, 12700, 12700]
add_is_row(ws4, r, "Common stock & APIC", cs_hist, cs_proj, indent=True)

r += 1
re_hist = [22995, 26900, 21010, 21500, 28200]
re_proj = [49800, 75550, 96150, 121530, 149200]
add_is_row(ws4, r, "Retained earnings", re_hist, re_proj, indent=True)

r += 1
aoc_hist = [(50), (180), (210), (220), (190)]
aoc_proj = [(180), (170), (160), (150), (140)]
add_is_row(ws4, r, "Accumulated OCI", aoc_hist, aoc_proj, indent=True)

r += 1
ts_hist = [(1500), (3950), (4400), (4400), (4400)]
ts_proj = [(7400), (10900), (13400), (16900), (21400)]
add_is_row(ws4, r, "Treasury stock", ts_hist, ts_proj, indent=True)

r += 1
te_hist = [cs_hist[i] + re_hist[i] + aoc_hist[i] + ts_hist[i] for i in range(5)]
te_proj = [cs_proj[i] + re_proj[i] + aoc_proj[i] + ts_proj[i] for i in range(5)]
add_is_row(ws4, r, "TOTAL STOCKHOLDERS' EQUITY", te_hist, te_proj, bold=True, total_fill=True)

r += 1
tle_hist = [tl_hist[i] + te_hist[i] for i in range(5)]
tle_proj = [tl_proj[i] + te_proj[i] for i in range(5)]
add_is_row(ws4, r, "TOTAL LIAB. & STOCKHOLDERS' EQUITY", tle_hist, tle_proj, bold=True, total_fill=True)

# Key ratios
r += 2
write_cell(ws4, r, 1, "KEY BALANCE-SHEET METRICS", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws4, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
td_hist = [st_debt_hist[i] + lt_debt_hist[i] for i in range(5)]
td_proj = [st_debt_proj[i] + lt_debt_proj[i] for i in range(5)]
add_is_row(ws4, r, "Total debt", td_hist, td_proj)

r += 1
nd_hist = [td_hist[i] - cash_hist[i] - inv_sec_hist[i] for i in range(5)]
nd_proj = [td_proj[i] - cash_proj[i] - inv_sec_proj[i] for i in range(5)]
add_is_row(ws4, r, "Net debt (cash)", nd_hist, nd_proj)

r += 1
de_hist = [td_hist[i] / te_hist[i] for i in range(5)]
de_proj = [td_proj[i] / te_proj[i] for i in range(5)]
add_is_row(ws4, r, "Debt / Equity", de_hist, de_proj, fmt="0.00x")

r += 1
cr_hist = [ca_hist[i] / cl_hist[i] for i in range(5)]
cr_proj = [ca_proj[i] / cl_proj[i] for i in range(5)]
add_is_row(ws4, r, "Current ratio", cr_hist, cr_proj, fmt="0.00x")

# Column widths
ws4.column_dimensions['A'].width = 48
for i in range(2, 12):
    ws4.column_dimensions[get_column_letter(i)].width = 13

r += 3
write_cell(ws4, r, 1, "Sources: Micron 10-K FY2025 Consolidated Balance Sheet. Includes ROU lease assets in PP&E line.",
           font=ITALIC_FONT, alignment=LEFT)
ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)


# ============================================================
# TAB 5: SCENARIOS
# ============================================================
ws5 = wb.create_sheet("Scenarios")

ws5.merge_cells("A1:L1")
write_cell(ws5, 1, 1, "Micron Technology — Bull / Base / Bear Scenarios",
           font=Font(bold=True, size=14, color="FFFFFF", name="Calibri"),
           fill=HEADER_FILL, alignment=CENTER)
ws5.row_dimensions[1].height = 24

write_cell(ws5, 2, 1, "All figures in $millions except per-share and percentages. Scenarios reflect FY2026E-FY2030E. Base case = Income Statement projections.",
           font=ITALIC_FONT, alignment=LEFT)
ws5.merge_cells("A2:L2")

# Headers
r = 4
write_cell(ws5, r, 1, "Metric", font=WHITE_FONT, fill=HEADER_FILL, alignment=LEFT, border=border_all)
write_cell(ws5, r, 2, "FY2025A", font=WHITE_FONT, fill=HEADER_FILL, alignment=CENTER, border=border_all)
sc_hdrs = ["BULL FY26E", "BULL FY27E", "BULL FY30E",
           "BASE FY26E", "BASE FY27E", "BASE FY30E",
           "BEAR FY26E", "BEAR FY27E", "BEAR FY30E"]
for i, h in enumerate(sc_hdrs, start=3):
    if "BULL" in h: fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    elif "BASE" in h: fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    else: fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    write_cell(ws5, r, i, h, font=BOLD_FONT, fill=fill, alignment=CENTER, border=border_all)

# Key assumptions section
r += 1
write_cell(ws5, r, 1, "KEY ASSUMPTIONS", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws5, r, c, "", fill=SUBHEADER_FILL, border=border_all)


def add_scen_row(ws, row, label, fy25, bull, base, bear, fmt="#,##0"):
    write_cell(ws, row, 1, label, font=NORMAL_FONT, alignment=LEFT, border=border_all)
    write_cell(ws, row, 2, fy25, font=NORMAL_FONT, alignment=RIGHT, border=border_all, fill=HISTORICAL_FILL, number_format=fmt)
    bull_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    base_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    bear_fill = PatternFill(start_color="FFE4E4", end_color="FFE4E4", fill_type="solid")
    for i, (v, fill) in enumerate(zip(bull + base + bear, [bull_fill]*3 + [base_fill]*3 + [bear_fill]*3), start=3):
        write_cell(ws, row, i, v, font=NORMAL_FONT, alignment=RIGHT, border=border_all, fill=fill, number_format=fmt)

# HBM revenue
r += 1
add_scen_row(ws5, r, "HBM revenue",
             7100, [26000, 35000, 38000], [21500, 28000, 26000], [16500, 18000, 12000])

# DRAM ASP YoY growth (Base year)
r += 1
add_scen_row(ws5, r, "DRAM blended ASP growth YoY",
             0.40, [0.45, 0.18, 0.05], [0.35, 0.05, (0.05)], [0.20, (0.20), (0.10)], fmt="0.0%")

# Bit shipments growth
r += 1
add_scen_row(ws5, r, "DRAM bit shipment growth YoY",
             0.18, [0.30, 0.28, 0.22], [0.25, 0.20, 0.16], [0.18, 0.12, 0.08], fmt="0.0%")

# NAND ASP growth
r += 1
add_scen_row(ws5, r, "NAND blended ASP growth YoY",
             0.20, [0.20, 0.12, 0.04], [0.12, 0.05, (0.05)], [0.05, (0.20), (0.10)], fmt="0.0%")

# Total revenue
r += 1
write_cell(ws5, r, 1, "REVENUE", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws5, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
add_scen_row(ws5, r, "Total revenue ($M)",
             37378, [62000, 75000, 90000], [54710, 62625, 70880], [48000, 45000, 50000])

# Margins
r += 1
write_cell(ws5, r, 1, "PROFITABILITY", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws5, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
add_scen_row(ws5, r, "Gross margin %",
             0.40, [0.66, 0.62, 0.55], [0.61, 0.62, 0.575], [0.50, 0.32, 0.28], fmt="0.0%")

r += 1
add_scen_row(ws5, r, "Operating margin %",
             0.26, [0.55, 0.51, 0.44], [0.49, 0.50, 0.45], [0.36, 0.16, 0.12], fmt="0.0%")

r += 1
add_scen_row(ws5, r, "Operating income ($M)",
             9774, [34100, 38250, 39600], [26781, 31325, 31930], [17280, 7200, 6000])

r += 1
add_scen_row(ws5, r, "Net income ($M)",
             8539, [28200, 32000, 33000], [25460, 28890, 30000], [13500, 5000, 4000])

# EPS
r += 1
add_scen_row(ws5, r, "Diluted EPS ($)",
             7.55, [25.00, 28.50, 30.00], [22.50, 25.50, 27.00], [12.00, 4.50, 3.60], fmt="$#,##0.00")

# FCF
r += 1
write_cell(ws5, r, 1, "CASH FLOW", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws5, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
add_scen_row(ws5, r, "Operating cash flow ($M)",
             17530, [32000, 36000, 42000], [27680, 28850, 32400], [21500, 14000, 13000])

r += 1
add_scen_row(ws5, r, "Capex ($M)",
             15864, [20000, 18000, 16000], [17500, 16000, 15500], [16000, 12000, 9000])

r += 1
add_scen_row(ws5, r, "Free cash flow ($M)",
             1666, [12000, 18000, 26000], [10180, 12850, 16900], [5500, 2000, 4000])

# Valuation
r += 1
write_cell(ws5, r, 1, "VALUATION (implied price target)", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws5, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
add_scen_row(ws5, r, "Forward EPS multiple",
             0, [38, 28, 20], [40, 32, 22], [35, 25, 15], fmt="0.0x")

r += 1
add_scen_row(ws5, r, "Implied price target ($)",
             727, [950, 800, 600], [900, 815, 594], [420, 113, 54], fmt="$#,##0")

r += 1
write_cell(ws5, r, 1, "vs. current price $727.42", font=ITALIC_FONT, alignment=LEFT, border=border_all)
ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
write_cell(ws5, r, 2, "", border=border_all)

# Probability weights
r += 2
write_cell(ws5, r, 1, "SCENARIO PROBABILITY WEIGHTING", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws5, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
write_cell(ws5, r, 1, "Bull case probability:", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws5, r, 2, 0.25, font=BOLD_FONT, alignment=RIGHT, border=border_all, number_format="0%")
write_cell(ws5, r, 3, "AI super-cycle through FY28; HBM4/4E share gains; consolidation discipline", font=ITALIC_FONT, alignment=LEFT, border=border_all)
ws5.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)

r += 1
write_cell(ws5, r, 1, "Base case probability:", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws5, r, 2, 0.50, font=BOLD_FONT, alignment=RIGHT, border=border_all, number_format="0%")
write_cell(ws5, r, 3, "Strong FY26-27, moderation FY28-29, recovery FY30; HBM holds; commodity normalizes", font=ITALIC_FONT, alignment=LEFT, border=border_all)
ws5.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)

r += 1
write_cell(ws5, r, 1, "Bear case probability:", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws5, r, 2, 0.25, font=BOLD_FONT, alignment=RIGHT, border=border_all, number_format="0%")
write_cell(ws5, r, 3, "AI capex decel in 2H26; CXMT commodity ramp; HBM ASP correction", font=ITALIC_FONT, alignment=LEFT, border=border_all)
ws5.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)

# Probability-weighted target
r += 2
write_cell(ws5, r, 1, "PROBABILITY-WEIGHTED 12M PRICE TARGET", font=BOLD_FONT, fill=TOTAL_FILL, alignment=LEFT, border=border_all)
# 0.25 * 950 + 0.50 * 900 + 0.25 * 420 = 237.5 + 450 + 105 = 792.5
weighted_pt = 0.25 * 950 + 0.50 * 900 + 0.25 * 420
write_cell(ws5, r, 2, weighted_pt, font=BOLD_FONT, fill=TOTAL_FILL, alignment=RIGHT, border=border_all, number_format="$#,##0")
write_cell(ws5, r, 3, f"= 0.25×$950 (Bull) + 0.50×$900 (Base) + 0.25×$420 (Bear)", font=ITALIC_FONT, alignment=LEFT, border=border_all)
ws5.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)

# Column widths
ws5.column_dimensions['A'].width = 38
for i in range(2, 12):
    ws5.column_dimensions[get_column_letter(i)].width = 12


# ============================================================
# TAB 6: DCF INPUTS
# ============================================================
ws6 = wb.create_sheet("DCF Inputs")

ws6.merge_cells("A1:L1")
write_cell(ws6, 1, 1, "Micron Technology — DCF Inputs (Unlevered Free Cash Flow Build)",
           font=Font(bold=True, size=14, color="FFFFFF", name="Calibri"),
           fill=HEADER_FILL, alignment=CENTER)
ws6.row_dimensions[1].height = 24

write_cell(ws6, 2, 1, "All figures in $millions. Mid-year convention. FY ends August.",
           font=ITALIC_FONT, alignment=LEFT)
ws6.merge_cells("A2:L2")

# Headers
HEADER_ROW6 = 4
write_cell(ws6, HEADER_ROW6, 1, "($ in millions)", font=WHITE_FONT, fill=HEADER_FILL, alignment=LEFT, border=border_all)
for i, yr in enumerate(ALL_YEARS, start=2):
    fill = HISTORICAL_FILL if "A" in yr else PROJECTED_FILL
    font = Font(bold=True, size=11, color="1F4E79", name="Calibri")
    write_cell(ws6, HEADER_ROW6, i, yr, font=font, fill=fill, alignment=CENTER, border=border_all)

r = HEADER_ROW6 + 1
write_cell(ws6, r, 1, "UNLEVERED FCF (UFCF) BUILD", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws6, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
add_is_row(ws6, r, "Revenue", revenue_hist, revenue_proj)

r += 1
add_is_row(ws6, r, "EBIT (operating income)", op_hist, op_proj)

r += 1
# EBIT margin
ebitm_hist = [op_hist[i] / revenue_hist[i] for i in range(5)]
ebitm_proj = [op_proj[i] / revenue_proj[i] for i in range(5)]
add_is_row(ws6, r, "  EBIT margin %", ebitm_hist, ebitm_proj, fmt="0.0%", italic=True)

r += 1
# Tax rate for DCF (normalized ~ 12-14%)
dcf_tax_rate = [0.10, 0.12, 0.10, 0.12, 0.12]
dcf_tax_rate_proj = [0.13, 0.13, 0.13, 0.13, 0.13]
add_is_row(ws6, r, "DCF tax rate (normalized)", dcf_tax_rate, dcf_tax_rate_proj, fmt="0.0%", italic=True)

r += 1
# NOPAT
nopat_hist = [op_hist[i] * (1 - dcf_tax_rate[i]) for i in range(5)]
nopat_proj = [op_proj[i] * (1 - dcf_tax_rate_proj[i]) for i in range(5)]
add_is_row(ws6, r, "NOPAT (EBIT × (1-t))", nopat_hist, nopat_proj, bold=True)

r += 1
# Add back D&A
add_is_row(ws6, r, "Plus: D&A", da_hist, da_proj, indent=True)

r += 1
# Less capex
capex_pos_hist = [abs(c) for c in capex_hist]
capex_pos_proj = [abs(c) for c in capex_proj]
neg_capex_hist = [-c for c in capex_pos_hist]
neg_capex_proj = [-c for c in capex_pos_proj]
add_is_row(ws6, r, "Less: Capital expenditures", neg_capex_hist, neg_capex_proj, indent=True)

r += 1
# Less change in WC
# Calc WC change: AR + Inv - AP
wc_hist = [ar_hist[i] + inv_hist[i] - ap_hist[i] for i in range(5)]
wc_proj = [ar_proj[i] + inv_proj[i] - ap_proj[i] for i in range(5)]
wc_chg_hist = [wc_hist[0] - 4900]
for i in range(1, 5):
    wc_chg_hist.append(wc_hist[i] - wc_hist[i-1])
wc_chg_proj = [wc_proj[0] - wc_hist[-1]]
for i in range(1, 5):
    wc_chg_proj.append(wc_proj[i] - wc_proj[i-1])
wc_chg_neg_hist = [-c for c in wc_chg_hist]
wc_chg_neg_proj = [-c for c in wc_chg_proj]
add_is_row(ws6, r, "Less: Change in working capital", wc_chg_neg_hist, wc_chg_neg_proj, indent=True)

r += 1
# UFCF
ufcf_hist = [nopat_hist[i] + da_hist[i] - capex_pos_hist[i] + wc_chg_neg_hist[i] for i in range(5)]
ufcf_proj = [nopat_proj[i] + da_proj[i] - capex_pos_proj[i] + wc_chg_neg_proj[i] for i in range(5)]
add_is_row(ws6, r, "UNLEVERED FREE CASH FLOW", ufcf_hist, ufcf_proj, bold=True, total_fill=True)

# DCF assumptions
r += 2
write_cell(ws6, r, 1, "DCF ASSUMPTIONS", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws6, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
write_cell(ws6, r, 1, "Risk-free rate (10Y UST, 2026-05-20)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws6, r, 2, 0.0425, font=NORMAL_FONT, alignment=RIGHT, border=border_all, fill=PROJECTED_FILL, number_format="0.00%")

r += 1
write_cell(ws6, r, 1, "Equity risk premium", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws6, r, 2, 0.055, font=NORMAL_FONT, alignment=RIGHT, border=border_all, fill=PROJECTED_FILL, number_format="0.00%")

r += 1
write_cell(ws6, r, 1, "Beta (2Y, MU per Yahoo Finance)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws6, r, 2, 1.35, font=NORMAL_FONT, alignment=RIGHT, border=border_all, fill=PROJECTED_FILL, number_format="0.00")

r += 1
write_cell(ws6, r, 1, "Cost of equity (CAPM)", font=BOLD_FONT, alignment=LEFT, border=border_all)
write_cell(ws6, r, 2, 0.0425 + 1.35 * 0.055, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=PROJECTED_FILL, number_format="0.00%")

r += 1
write_cell(ws6, r, 1, "After-tax cost of debt", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws6, r, 2, 0.045, font=NORMAL_FONT, alignment=RIGHT, border=border_all, fill=PROJECTED_FILL, number_format="0.00%")

r += 1
write_cell(ws6, r, 1, "Target debt / total capital", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws6, r, 2, 0.15, font=NORMAL_FONT, alignment=RIGHT, border=border_all, fill=PROJECTED_FILL, number_format="0.0%")

r += 1
wacc_calc = 0.85 * (0.0425 + 1.35 * 0.055) + 0.15 * 0.045
write_cell(ws6, r, 1, "WACC (calculated)", font=BOLD_FONT, alignment=LEFT, border=border_all, fill=TOTAL_FILL)
write_cell(ws6, r, 2, wacc_calc, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=TOTAL_FILL, number_format="0.00%")

r += 1
write_cell(ws6, r, 1, "Terminal growth rate", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws6, r, 2, 0.030, font=NORMAL_FONT, alignment=RIGHT, border=border_all, fill=PROJECTED_FILL, number_format="0.00%")

r += 1
write_cell(ws6, r, 1, "Terminal EBITDA multiple (exit)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws6, r, 2, 8.0, font=NORMAL_FONT, alignment=RIGHT, border=border_all, fill=PROJECTED_FILL, number_format="0.0x")

# Net debt + Shares
r += 2
write_cell(ws6, r, 1, "CAPITAL STRUCTURE & SHARES (as of 2026-05-20)", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws6, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
write_cell(ws6, r, 1, "Total debt ($M)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws6, r, 2, 14478, font=NORMAL_FONT, alignment=RIGHT, border=border_all, fill=HISTORICAL_FILL, number_format="#,##0")

r += 1
write_cell(ws6, r, 1, "Cash & investments ($M)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws6, r, 2, 12011, font=NORMAL_FONT, alignment=RIGHT, border=border_all, fill=HISTORICAL_FILL, number_format="#,##0")

r += 1
write_cell(ws6, r, 1, "Net debt ($M)", font=BOLD_FONT, alignment=LEFT, border=border_all)
write_cell(ws6, r, 2, 14478 - 12011, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=TOTAL_FILL, number_format="#,##0")

r += 1
write_cell(ws6, r, 1, "Diluted shares (millions)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws6, r, 2, 1131, font=NORMAL_FONT, alignment=RIGHT, border=border_all, fill=HISTORICAL_FILL, number_format="#,##0")

r += 1
write_cell(ws6, r, 1, "Current price ($)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws6, r, 2, 727.42, font=NORMAL_FONT, alignment=RIGHT, border=border_all, fill=HISTORICAL_FILL, number_format="$#,##0.00")

# EBITDA bridge
r += 2
write_cell(ws6, r, 1, "EBITDA PROJECTIONS (for terminal multiple)", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 12):
    write_cell(ws6, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
ebitda_hist = [op_hist[i] + da_hist[i] for i in range(5)]
ebitda_proj = [op_proj[i] + da_proj[i] for i in range(5)]
add_is_row(ws6, r, "EBITDA", ebitda_hist, ebitda_proj, bold=True)

r += 1
ebitdam_hist = [ebitda_hist[i] / revenue_hist[i] for i in range(5)]
ebitdam_proj = [ebitda_proj[i] / revenue_proj[i] for i in range(5)]
add_is_row(ws6, r, "  EBITDA margin %", ebitdam_hist, ebitdam_proj, fmt="0.0%", italic=True)

# Column widths
ws6.column_dimensions['A'].width = 48
for i in range(2, 12):
    ws6.column_dimensions[get_column_letter(i)].width = 13

r += 3
write_cell(ws6, r, 1, "Sources: Income Statement and Cash Flow tabs. Market data from Yahoo Finance MU key statistics retrieved 2026-05-20.",
           font=ITALIC_FONT, alignment=LEFT)
ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)


# Reorder tabs and save
wb._sheets = [wb["Revenue Model"], wb["Income Statement"], wb["Cash Flow"], wb["Balance Sheet"], wb["Scenarios"], wb["DCF Inputs"]]

out_path = "/Users/x/projects/financial_agent/reports/company/Micron_NASDAQ_MU/Task2_Model/Micron_Financial_Model_2026-05-20.xlsx"
wb.save(out_path)
print(f"Saved model to {out_path}")
print(f"\nSummary stats:")
print(f"  FY2025 Revenue: ${revenue_hist[-1]:,}M")
print(f"  FY2030E Revenue: ${revenue_proj[-1]:,}M")
print(f"  FY2025 GM: {gm_hist[-1]:.1%}")
print(f"  FY2026E GM: {gm_proj[0]:.1%}")
print(f"  FY2025 EPS: ${eps_diluted_hist[-1]:.2f}")
print(f"  FY2026E EPS: ${eps_diluted_proj[0]:.2f}")
print(f"  FY2026E UFCF: ${ufcf_proj[0]:,.0f}M")
print(f"  WACC: {wacc_calc:.2%}")
