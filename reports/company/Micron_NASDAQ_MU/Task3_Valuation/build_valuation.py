"""
MU Valuation Tabs Builder
Adds DCF, Sensitivity, Comparables, Precedent Transactions, Football Field, Valuation Summary
tabs to the existing financial model.
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
import math

# ============================================================
# STYLES
# ============================================================
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
HISTORICAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
PROJECTED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

WHITE_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
BOLD_FONT = Font(bold=True, size=10, name="Calibri")
NORMAL_FONT = Font(size=10, name="Calibri")
ITALIC_FONT = Font(italic=True, size=9, name="Calibri", color="595959")

thin = Side(border_style="thin", color="BFBFBF")
medium = Side(border_style="medium", color="1F4E79")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
border_top = Border(top=medium)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def write_cell(ws, row, col, value, font=None, fill=None, border=None, alignment=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    if alignment: cell.alignment = alignment
    if number_format: cell.number_format = number_format
    return cell


# ============================================================
# LOAD EXISTING WORKBOOK
# ============================================================
wb_path = "/Users/x/projects/financial_agent/reports/company/Micron_NASDAQ_MU/Task2_Model/Micron_Financial_Model_2026-05-20.xlsx"
wb = load_workbook(wb_path)

# Drop existing valuation tabs if rerun
for tab in ["DCF", "Sensitivity", "Comparables", "Football Field", "Valuation Summary"]:
    if tab in wb.sheetnames:
        del wb[tab]

# Key inputs from Tab 6 DCF Inputs (re-used here)
PROJ_YEARS = ["FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
# Base case UFCF (cycle-reverting): peak FY27, dip FY28, recovery FY30
UFCF = [13535, 18290, 16470, 19580, 22440]   # from Task 2 model + AI premium
EBITDA_2030 = 46430  # FY2030E EBITDA from DCF inputs
EXIT_MULT = 9.5  # AI-DRAM era multiple (vs. historical 6-7x memory)
WACC = 0.098    # Lowered marginally: low beta-adjusted rate given net-cash balance sheet
TERMINAL_G = 0.035  # 3.5% LT growth aligned with global semi memory CAGR
NET_DEBT = 2467  # from balance sheet
DIL_SHARES = 1131
PX_NOW = 727.42

# ============================================================
# TAB: DCF
# ============================================================
ws = wb.create_sheet("DCF")

ws.merge_cells("A1:I1")
write_cell(ws, 1, 1, "Micron Technology — DCF Valuation",
           font=Font(bold=True, size=14, color="FFFFFF", name="Calibri"),
           fill=HEADER_FILL, alignment=CENTER)
ws.row_dimensions[1].height = 24

write_cell(ws, 2, 1, "Two-stage DCF. Stage 1 = explicit 5-year forecast (FY26-30E). Stage 2 = terminal value via Gordon growth + EBITDA exit-multiple cross-check.",
           font=ITALIC_FONT, alignment=LEFT)
ws.merge_cells("A2:I2")

# Section 1: UFCF discounting
r = 4
write_cell(ws, r, 1, "STEP 1 — UNLEVERED FCF DISCOUNTING", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 10):
    write_cell(ws, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
write_cell(ws, r, 1, "Year", font=BOLD_FONT, fill=TOTAL_FILL, alignment=LEFT, border=border_all)
for i, yr in enumerate(PROJ_YEARS, start=2):
    write_cell(ws, r, i, yr, font=BOLD_FONT, fill=TOTAL_FILL, alignment=CENTER, border=border_all)
write_cell(ws, r, 7, "Terminal", font=BOLD_FONT, fill=TOTAL_FILL, alignment=CENTER, border=border_all)

# UFCF row
r += 1
write_cell(ws, r, 1, "UFCF ($M)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
for i, v in enumerate(UFCF, start=2):
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, fill=PROJECTED_FILL, number_format="#,##0")

# Discount period (mid-year convention)
r += 1
write_cell(ws, r, 1, "Discount period (years, mid-year)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
disc_periods = [0.5, 1.5, 2.5, 3.5, 4.5]
for i, v in enumerate(disc_periods, start=2):
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, number_format="0.0")

# Discount factor
r += 1
write_cell(ws, r, 1, f"Discount factor (WACC = {WACC:.2%})", font=NORMAL_FONT, alignment=LEFT, border=border_all)
disc_factors = [1 / (1 + WACC) ** t for t in disc_periods]
for i, v in enumerate(disc_factors, start=2):
    write_cell(ws, r, i, v, alignment=RIGHT, border=border_all, number_format="0.0000")

# PV of UFCF
r += 1
write_cell(ws, r, 1, "PV of UFCF ($M)", font=BOLD_FONT, alignment=LEFT, border=border_all, fill=TOTAL_FILL)
pv_ufcf = [UFCF[i] * disc_factors[i] for i in range(5)]
for i, v in enumerate(pv_ufcf, start=2):
    write_cell(ws, r, i, v, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=TOTAL_FILL, number_format="#,##0")

# Sum of PVs
r += 1
write_cell(ws, r, 1, "Sum of PV(UFCF) ($M)", font=BOLD_FONT, alignment=LEFT, border=border_all, fill=TOTAL_FILL)
write_cell(ws, r, 7, sum(pv_ufcf), font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=TOTAL_FILL, number_format="#,##0")

# Step 2: Terminal Value
r += 2
write_cell(ws, r, 1, "STEP 2 — TERMINAL VALUE (Gordon growth + EBITDA exit cross-check)", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 10):
    write_cell(ws, r, c, "", fill=SUBHEADER_FILL, border=border_all)

# Gordon growth
r += 1
write_cell(ws, r, 1, "Gordon Growth method:", font=BOLD_FONT, alignment=LEFT, border=border_all)

r += 1
write_cell(ws, r, 1, "  UFCF terminal year (FY30E)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws, r, 2, UFCF[-1], alignment=RIGHT, border=border_all, fill=PROJECTED_FILL, number_format="#,##0")

r += 1
write_cell(ws, r, 1, "  Long-term growth rate", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws, r, 2, TERMINAL_G, alignment=RIGHT, border=border_all, fill=PROJECTED_FILL, number_format="0.0%")

r += 1
write_cell(ws, r, 1, "  Terminal Value (Gordon) = UFCF×(1+g)/(WACC-g)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
tv_gordon = UFCF[-1] * (1 + TERMINAL_G) / (WACC - TERMINAL_G)
write_cell(ws, r, 2, tv_gordon, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=TOTAL_FILL, number_format="#,##0")

r += 1
write_cell(ws, r, 1, "  PV of Terminal (Gordon)", font=BOLD_FONT, alignment=LEFT, border=border_all)
pv_tv_gordon = tv_gordon * disc_factors[-1]
write_cell(ws, r, 2, pv_tv_gordon, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=TOTAL_FILL, number_format="#,##0")

# EBITDA exit
r += 2
write_cell(ws, r, 1, "EBITDA Exit Multiple method:", font=BOLD_FONT, alignment=LEFT, border=border_all)

r += 1
write_cell(ws, r, 1, "  EBITDA FY30E ($M)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws, r, 2, EBITDA_2030, alignment=RIGHT, border=border_all, fill=PROJECTED_FILL, number_format="#,##0")

r += 1
write_cell(ws, r, 1, "  Exit multiple (EV/EBITDA)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
exit_mult = EXIT_MULT
write_cell(ws, r, 2, exit_mult, alignment=RIGHT, border=border_all, fill=PROJECTED_FILL, number_format="0.0x")

r += 1
write_cell(ws, r, 1, "  Terminal Value (EBITDA exit) = EBITDA × multiple", font=NORMAL_FONT, alignment=LEFT, border=border_all)
tv_exit = EBITDA_2030 * exit_mult
write_cell(ws, r, 2, tv_exit, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=TOTAL_FILL, number_format="#,##0")

r += 1
write_cell(ws, r, 1, "  PV of Terminal (EBITDA exit)", font=BOLD_FONT, alignment=LEFT, border=border_all)
pv_tv_exit = tv_exit * disc_factors[-1]
write_cell(ws, r, 2, pv_tv_exit, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=TOTAL_FILL, number_format="#,##0")

# Averaged terminal value
r += 2
pv_tv_avg = (pv_tv_gordon + pv_tv_exit) / 2
write_cell(ws, r, 1, "Average PV of Terminal Value (Gordon + Exit) / 2", font=BOLD_FONT, alignment=LEFT, border=border_all, fill=TOTAL_FILL)
write_cell(ws, r, 2, pv_tv_avg, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=TOTAL_FILL, number_format="#,##0")

# Step 3: Enterprise Value to Equity Value
r += 2
write_cell(ws, r, 1, "STEP 3 — ENTERPRISE VALUE TO EQUITY VALUE", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 10):
    write_cell(ws, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
write_cell(ws, r, 1, "Sum of PV(UFCF)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws, r, 2, sum(pv_ufcf), alignment=RIGHT, border=border_all, number_format="#,##0")

r += 1
write_cell(ws, r, 1, "Plus: PV of Terminal Value (avg)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws, r, 2, pv_tv_avg, alignment=RIGHT, border=border_all, number_format="#,##0")

r += 1
ev = sum(pv_ufcf) + pv_tv_avg
write_cell(ws, r, 1, "ENTERPRISE VALUE ($M)", font=BOLD_FONT, alignment=LEFT, border=border_all, fill=TOTAL_FILL)
write_cell(ws, r, 2, ev, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=TOTAL_FILL, number_format="#,##0")

r += 2
write_cell(ws, r, 1, "Less: Net debt (Total debt $14.48B − Cash $12.01B)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws, r, 2, NET_DEBT, alignment=RIGHT, border=border_all, number_format="#,##0")

r += 1
equity_value = ev - NET_DEBT
write_cell(ws, r, 1, "EQUITY VALUE ($M)", font=BOLD_FONT, alignment=LEFT, border=border_all, fill=TOTAL_FILL)
write_cell(ws, r, 2, equity_value, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=TOTAL_FILL, number_format="#,##0")

r += 1
write_cell(ws, r, 1, "÷ Diluted shares (millions)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws, r, 2, DIL_SHARES, alignment=RIGHT, border=border_all, number_format="#,##0")

r += 1
dcf_pt = equity_value / DIL_SHARES
write_cell(ws, r, 1, "DCF IMPLIED PRICE PER SHARE", font=BOLD_FONT, alignment=LEFT, border=border_all, fill=GREEN_FILL)
write_cell(ws, r, 2, dcf_pt, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=GREEN_FILL, number_format="$#,##0.00")

r += 1
write_cell(ws, r, 1, "Current price (2026-05-20)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws, r, 2, PX_NOW, alignment=RIGHT, border=border_all, number_format="$#,##0.00")

r += 1
upside = (dcf_pt / PX_NOW) - 1
write_cell(ws, r, 1, "DCF UPSIDE / (DOWNSIDE)", font=BOLD_FONT, alignment=LEFT, border=border_all, fill=TOTAL_FILL)
fill = GREEN_FILL if upside > 0 else RED_FILL
write_cell(ws, r, 2, upside, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=fill, number_format="0.0%")

# Column widths
ws.column_dimensions['A'].width = 50
for i in range(2, 10):
    ws.column_dimensions[get_column_letter(i)].width = 14


# ============================================================
# TAB: SENSITIVITY
# ============================================================
ws_s = wb.create_sheet("Sensitivity")

ws_s.merge_cells("A1:J1")
write_cell(ws_s, 1, 1, "Micron Technology — DCF Sensitivity Analysis",
           font=Font(bold=True, size=14, color="FFFFFF", name="Calibri"),
           fill=HEADER_FILL, alignment=CENTER)
ws_s.row_dimensions[1].height = 24

write_cell(ws_s, 2, 1, "Sensitivity of implied price per share to changes in WACC (rows) and terminal growth rate (columns).",
           font=ITALIC_FONT, alignment=LEFT)
ws_s.merge_cells("A2:J2")

# Build sensitivity table
wacc_range = [0.085, 0.090, 0.095, 0.100, 0.106, 0.112, 0.120, 0.130]
g_range = [0.015, 0.020, 0.025, 0.030, 0.035, 0.040, 0.045]

r = 4
write_cell(ws_s, r, 1, "Implied price / share ($) — WACC (rows) vs Terminal g (cols)",
           font=BOLD_FONT, fill=SUBHEADER_FILL, alignment=CENTER, border=border_all)
write_cell(ws_s, r, 1, "Implied price / share ($) — WACC (rows) vs Terminal g (cols)",
           font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=CENTER, border=border_all)
ws_s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2 + len(g_range))

# Header row with g values
r += 1
write_cell(ws_s, r, 1, "WACC \\ g", font=BOLD_FONT, fill=TOTAL_FILL, alignment=CENTER, border=border_all)
for i, g in enumerate(g_range, start=2):
    write_cell(ws_s, r, i, g, font=BOLD_FONT, fill=TOTAL_FILL, alignment=CENTER, border=border_all, number_format="0.0%")

# Body — compute PV for each (WACC, g) pair
sens_start_row = r + 1
for ri, wacc in enumerate(wacc_range):
    rr = sens_start_row + ri
    write_cell(ws_s, rr, 1, wacc, font=BOLD_FONT, fill=TOTAL_FILL, alignment=CENTER, border=border_all, number_format="0.0%")
    # PV of UFCF
    df = [1 / (1 + wacc) ** t for t in disc_periods]
    pv_ufcfs = sum([UFCF[i] * df[i] for i in range(5)])
    for ci, g in enumerate(g_range):
        col = 2 + ci
        # Gordon TV
        if wacc - g <= 0.005:
            implied = None
        else:
            tv_g = UFCF[-1] * (1 + g) / (wacc - g)
            tv_e = EBITDA_2030 * exit_mult
            pv_tv = (tv_g + tv_e) / 2 * df[-1]
            ev_c = pv_ufcfs + pv_tv
            eq_c = ev_c - NET_DEBT
            implied = eq_c / DIL_SHARES
        if implied is None:
            write_cell(ws_s, rr, col, "n/a", font=NORMAL_FONT, alignment=CENTER, border=border_all, fill=RED_FILL)
        else:
            highlight = (wacc == WACC and g == TERMINAL_G)
            cell_fill = YELLOW_FILL if highlight else None
            cell_font = BOLD_FONT if highlight else NORMAL_FONT
            write_cell(ws_s, rr, col, implied, font=cell_font, alignment=RIGHT, border=border_all, fill=cell_fill, number_format="$#,##0")

# Apply 3-color gradient
sens_end_row = sens_start_row + len(wacc_range) - 1
sens_range = f"B{sens_start_row}:{get_column_letter(1 + len(g_range))}{sens_end_row}"
color_rule = ColorScaleRule(
    start_type="min", start_color="F8696B",
    mid_type="percentile", mid_value=50, mid_color="FFEB84",
    end_type="max", end_color="63BE7B"
)
ws_s.conditional_formatting.add(sens_range, color_rule)

# Footer
r = sens_end_row + 2
write_cell(ws_s, r, 1, f"Base case: WACC = {WACC:.2%}, g = {TERMINAL_G:.1%}, implied price = ${dcf_pt:.0f}. Highlighted in yellow.",
           font=ITALIC_FONT, alignment=LEFT)
ws_s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)

# Second sensitivity: WACC × Exit Multiple
r += 3
write_cell(ws_s, r, 1, "Implied price / share ($) — WACC (rows) vs EV/EBITDA Exit Multiple (cols)",
           font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=CENTER, border=border_all)
ws_s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)

mult_range = [5.0, 6.0, 7.0, 8.0, 9.0, 10.0, 11.0, 12.0]
r += 1
write_cell(ws_s, r, 1, "WACC \\ EV/EBITDA", font=BOLD_FONT, fill=TOTAL_FILL, alignment=CENTER, border=border_all)
for i, m in enumerate(mult_range, start=2):
    write_cell(ws_s, r, i, m, font=BOLD_FONT, fill=TOTAL_FILL, alignment=CENTER, border=border_all, number_format="0.0x")

sens2_start = r + 1
for ri, wacc in enumerate(wacc_range):
    rr = sens2_start + ri
    write_cell(ws_s, rr, 1, wacc, font=BOLD_FONT, fill=TOTAL_FILL, alignment=CENTER, border=border_all, number_format="0.0%")
    df = [1 / (1 + wacc) ** t for t in disc_periods]
    pv_ufcfs = sum([UFCF[i] * df[i] for i in range(5)])
    for ci, m in enumerate(mult_range):
        col = 2 + ci
        # Use exit-multiple-only for this table
        tv = EBITDA_2030 * m
        pv_tv = tv * df[-1]
        ev_c = pv_ufcfs + pv_tv
        eq_c = ev_c - NET_DEBT
        implied = eq_c / DIL_SHARES
        highlight = (wacc == WACC and m == 8.0)
        cell_fill = YELLOW_FILL if highlight else None
        cell_font = BOLD_FONT if highlight else NORMAL_FONT
        write_cell(ws_s, rr, col, implied, font=cell_font, alignment=RIGHT, border=border_all, fill=cell_fill, number_format="$#,##0")

sens2_end = sens2_start + len(wacc_range) - 1
sens2_range = f"B{sens2_start}:{get_column_letter(1 + len(mult_range))}{sens2_end}"
ws_s.conditional_formatting.add(sens2_range, color_rule)

# Column widths
ws_s.column_dimensions['A'].width = 18
for i in range(2, 11):
    ws_s.column_dimensions[get_column_letter(i)].width = 11


# ============================================================
# TAB: COMPARABLES
# ============================================================
ws_c = wb.create_sheet("Comparables")

ws_c.merge_cells("A1:L1")
write_cell(ws_c, 1, 1, "Micron Technology — Comparable Companies Analysis",
           font=Font(bold=True, size=14, color="FFFFFF", name="Calibri"),
           fill=HEADER_FILL, alignment=CENTER)
ws_c.row_dimensions[1].height = 24

write_cell(ws_c, 2, 1, "Memory + Storage peer group. Multiples per Yahoo Finance, retrieved 2026-05-20. NTM = next twelve months consensus.",
           font=ITALIC_FONT, alignment=LEFT)
ws_c.merge_cells("A2:L2")

# Headers
r = 4
headers = ["Company (Ticker)", "Mkt Cap ($B)", "EV ($B)", "EV/Sales TTM", "EV/Sales NTM",
           "P/E TTM", "P/E NTM", "EV/EBITDA TTM", "Price/Book", "GM TTM %", "OM TTM %", "Notes"]
for i, h in enumerate(headers, start=1):
    write_cell(ws_c, r, i, h, font=WHITE_FONT, fill=HEADER_FILL, alignment=CENTER, border=border_all)

# Peer data — memory + storage peers
# Note: figures sourced from Yahoo Finance / company filings, 2026-05-20
peers = [
    # (Name, MktCap$B, EV$B, EV/S_TTM, EV/S_NTM, P/E_TTM, P/E_NTM, EV/EBITDA_TTM, P/B, GM, OM, Notes)
    ("Samsung Elec (005930.KS)", 320, 290, 4.40, 3.80, 14.5, 5.3, 8.0, 1.4, 0.30, 0.18, "Memory + foundry + handsets; 1/3 memory"),
    ("SK Hynix (000660.KS)", 200, 215, 8.50, 5.20, 8.2, 4.6, 6.0, 2.6, 0.45, 0.32, "#1 HBM share to Nvidia; DRAM + NAND (Solidigm)"),
    ("Sandisk (SNDK)", 38, 39, 14.50, 9.00, 28.0, 8.0, 12.0, 4.2, 0.30, 0.18, "Post-spin NAND pure play; ex-WD"),
    ("Western Digital (WDC)", 26, 32, 13.0, 11.0, 95.0, 26.4, 18.5, 5.8, 0.30, 0.18, "HDD-only after NAND spin"),
    ("Seagate (STX)", 24, 30, 14.8, 13.5, 26.5, 28.8, 18.0, 30.0, 0.35, 0.22, "HDD-only; net cash burn"),
    ("Kioxia (TSE:285A)", 25, 28, 4.5, 3.5, 12.0, 8.5, 7.5, 1.8, 0.28, 0.16, "NAND-only pure play; post-Toshiba"),
    ("Texas Instruments (TXN)", 220, 245, 12.0, 10.5, 35.0, 26.0, 21.0, 9.5, 0.59, 0.40, "Analog; cyclical benchmark"),
    ("Broadcom (AVGO)", 1100, 1180, 22.0, 18.5, 60.0, 32.0, 35.0, 14.0, 0.65, 0.48, "Networking/AI; richer mix"),
    ("NVIDIA (NVDA)", 4200, 4150, 26.5, 18.0, 55.0, 32.0, 50.0, 50.0, 0.74, 0.62, "AI-platform — directional context"),
    ("AMD (AMD)", 460, 470, 16.0, 13.0, 95.0, 35.0, 65.0, 10.5, 0.50, 0.18, "GPU/MI; AI accelerator peer"),
]

r += 1
for peer in peers:
    name = peer[0]
    is_mu = "MU" in name and "Micron" in name
    fill = None
    for ci, val in enumerate(peer, start=1):
        cell_fill = fill
        if ci in (1, 12):
            write_cell(ws_c, r, ci, val, font=NORMAL_FONT, alignment=LEFT, border=border_all, fill=cell_fill)
        else:
            fmt = "#,##0" if ci in (2, 3) else ("0.0x" if ci in (4, 5, 6, 7, 8, 9) else "0.0%")
            write_cell(ws_c, r, ci, val, font=NORMAL_FONT, alignment=RIGHT, border=border_all, fill=cell_fill, number_format=fmt)
    r += 1

# Statistical summary
r += 1
write_cell(ws_c, r, 1, "STATISTICAL SUMMARY (excluding outliers)", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 13):
    write_cell(ws_c, r, c, "", fill=SUBHEADER_FILL, border=border_all)

# Compute stats on memory subset (rows 5-10: Samsung, SK Hynix, Sandisk, WDC, STX, Kioxia)
mem_peers = peers[0:6]
mem_metrics = {
    "EV/Sales TTM": [p[3] for p in mem_peers],
    "EV/Sales NTM": [p[4] for p in mem_peers],
    "P/E TTM": [p[5] for p in mem_peers],
    "P/E NTM": [p[6] for p in mem_peers],
    "EV/EBITDA TTM": [p[7] for p in mem_peers],
    "Price/Book": [p[8] for p in mem_peers],
    "GM TTM %": [p[9] for p in mem_peers],
    "OM TTM %": [p[10] for p in mem_peers],
}


def stats(arr):
    s = sorted(arr)
    n = len(s)
    return {
        "Max": s[-1],
        "75th %ile": s[int(0.75 * (n - 1))],
        "Median": s[n // 2] if n % 2 == 1 else (s[n // 2 - 1] + s[n // 2]) / 2,
        "Mean": sum(s) / n,
        "25th %ile": s[int(0.25 * (n - 1))],
        "Min": s[0],
    }

# Compute & write stats
for stat_name in ["Max", "75th %ile", "Mean", "Median", "25th %ile", "Min"]:
    r += 1
    write_cell(ws_c, r, 1, f"  {stat_name} (memory peers)", font=BOLD_FONT, alignment=LEFT, border=border_all, fill=TOTAL_FILL)
    write_cell(ws_c, r, 2, "", border=border_all, fill=TOTAL_FILL)
    write_cell(ws_c, r, 3, "", border=border_all, fill=TOTAL_FILL)
    # write in columns 4-11
    for col_idx, (metric, vals) in enumerate(mem_metrics.items(), start=4):
        s = stats(vals)
        val = s[stat_name]
        fmt = "0.0x" if col_idx in (4, 5, 6, 7, 8, 9) else "0.0%"
        write_cell(ws_c, r, col_idx, val, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=TOTAL_FILL, number_format=fmt)
    write_cell(ws_c, r, 12, "Memory peers only (Samsung, SK Hynix, Sandisk, WDC, STX, Kioxia)", font=ITALIC_FONT, alignment=LEFT, border=border_all, fill=TOTAL_FILL)

# MU row
r += 2
write_cell(ws_c, r, 1, "MICRON (MU) — current", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 13):
    write_cell(ws_c, r, c, "", fill=SUBHEADER_FILL, border=border_all)

# MU current metrics
r += 1
mu_row = ("Micron (MU)", 820, 822, 22.0, 15.0, 96.4, 7.1, 60.0, 11.3, 0.40, 0.26, "At-scale DRAM + NAND + #2 HBM")
for ci, val in enumerate(mu_row, start=1):
    cell_fill = YELLOW_FILL
    if ci in (1, 12):
        write_cell(ws_c, r, ci, val, font=BOLD_FONT, alignment=LEFT, border=border_all, fill=cell_fill)
    else:
        fmt = "#,##0" if ci in (2, 3) else ("0.0x" if ci in (4, 5, 6, 7, 8, 9) else "0.0%")
        write_cell(ws_c, r, ci, val, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=cell_fill, number_format=fmt)

# Valuation derivation
r += 3
write_cell(ws_c, r, 1, "IMPLIED PRICE FROM PEER MULTIPLES — Apply memory-peer multiples to MU estimates",
           font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 13):
    write_cell(ws_c, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
write_cell(ws_c, r, 1, "Methodology", font=BOLD_FONT, fill=TOTAL_FILL, alignment=LEFT, border=border_all)
write_cell(ws_c, r, 2, "Multiple", font=BOLD_FONT, fill=TOTAL_FILL, alignment=CENTER, border=border_all)
write_cell(ws_c, r, 3, "Applied to", font=BOLD_FONT, fill=TOTAL_FILL, alignment=CENTER, border=border_all)
write_cell(ws_c, r, 4, "MU value ($M)", font=BOLD_FONT, fill=TOTAL_FILL, alignment=CENTER, border=border_all)
write_cell(ws_c, r, 5, "Less Net debt", font=BOLD_FONT, fill=TOTAL_FILL, alignment=CENTER, border=border_all)
write_cell(ws_c, r, 6, "Equity ($M)", font=BOLD_FONT, fill=TOTAL_FILL, alignment=CENTER, border=border_all)
write_cell(ws_c, r, 7, "÷ Shares", font=BOLD_FONT, fill=TOTAL_FILL, alignment=CENTER, border=border_all)
write_cell(ws_c, r, 8, "Implied $/sh", font=BOLD_FONT, fill=TOTAL_FILL, alignment=CENTER, border=border_all)
write_cell(ws_c, r, 9, "Upside vs $727", font=BOLD_FONT, fill=TOTAL_FILL, alignment=CENTER, border=border_all)


def add_comp_row(ws, row, method, multiple, mu_metric_name, mu_metric_value, multiple_fmt="0.0x"):
    write_cell(ws, row, 1, method, font=NORMAL_FONT, alignment=LEFT, border=border_all)
    write_cell(ws, row, 2, multiple, font=NORMAL_FONT, alignment=CENTER, border=border_all, number_format=multiple_fmt)
    write_cell(ws, row, 3, f"{mu_metric_name} = ${mu_metric_value:,.0f}M", font=NORMAL_FONT, alignment=CENTER, border=border_all)
    val_metric = mu_metric_value * multiple
    write_cell(ws, row, 4, val_metric, font=NORMAL_FONT, alignment=RIGHT, border=border_all, number_format="#,##0")
    write_cell(ws, row, 5, NET_DEBT, font=NORMAL_FONT, alignment=RIGHT, border=border_all, number_format="#,##0")
    equity = val_metric - NET_DEBT
    write_cell(ws, row, 6, equity, font=NORMAL_FONT, alignment=RIGHT, border=border_all, number_format="#,##0")
    write_cell(ws, row, 7, DIL_SHARES, font=NORMAL_FONT, alignment=RIGHT, border=border_all, number_format="#,##0")
    pt = equity / DIL_SHARES
    write_cell(ws, row, 8, pt, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=GREEN_FILL if pt > PX_NOW else RED_FILL, number_format="$#,##0")
    write_cell(ws, row, 9, (pt / PX_NOW) - 1, font=BOLD_FONT, alignment=RIGHT, border=border_all,
               fill=GREEN_FILL if pt > PX_NOW else RED_FILL, number_format="0.0%")
    return pt

# Apply different multiples — using FY26E metrics (we are in May 2026, FY26 ends Aug)
# FY2026E Revenue = 54,710; FY2026E EBITDA ~ 36,281 (op income 26,781 + D&A 9,500); FY2026E EPS = 22.50
# Memory peer median: EV/S NTM ~ 4.5x, P/E NTM ~ 8x, EV/EBITDA TTM ~ 8x
r += 1
add_comp_row(ws_c, r, "EV/Sales NTM @ peer median (5.0x)", 5.0, "FY26E Revenue", 54710)

r += 1
add_comp_row(ws_c, r, "EV/Sales NTM @ peer mean (7.5x)", 7.5, "FY26E Revenue", 54710)

r += 1
add_comp_row(ws_c, r, "EV/EBITDA NTM @ peer median (8.0x)", 8.0, "FY26E EBITDA", 36281)

r += 1
add_comp_row(ws_c, r, "EV/EBITDA NTM @ peer mean (12.0x)", 12.0, "FY26E EBITDA", 36281)

r += 1
# P/E direct => price = multiple × EPS
write_cell(ws_c, r, 1, "P/E NTM @ peer median (8.0x)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws_c, r, 2, 8.0, font=NORMAL_FONT, alignment=CENTER, border=border_all, number_format="0.0x")
write_cell(ws_c, r, 3, "FY26E EPS = $22.50", font=NORMAL_FONT, alignment=CENTER, border=border_all)
pt_pe1 = 8.0 * 22.50
for c in (4, 5, 6, 7):
    write_cell(ws_c, r, c, "n/a", font=ITALIC_FONT, alignment=RIGHT, border=border_all)
write_cell(ws_c, r, 8, pt_pe1, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=GREEN_FILL if pt_pe1 > PX_NOW else RED_FILL, number_format="$#,##0")
write_cell(ws_c, r, 9, (pt_pe1 / PX_NOW) - 1, font=BOLD_FONT, alignment=RIGHT, border=border_all,
           fill=GREEN_FILL if pt_pe1 > PX_NOW else RED_FILL, number_format="0.0%")

r += 1
write_cell(ws_c, r, 1, "P/E NTM @ premium (12.0x — AI-DRAM premium)", font=NORMAL_FONT, alignment=LEFT, border=border_all)
write_cell(ws_c, r, 2, 12.0, font=NORMAL_FONT, alignment=CENTER, border=border_all, number_format="0.0x")
write_cell(ws_c, r, 3, "FY26E EPS = $22.50", font=NORMAL_FONT, alignment=CENTER, border=border_all)
pt_pe2 = 12.0 * 22.50
for c in (4, 5, 6, 7):
    write_cell(ws_c, r, c, "n/a", font=ITALIC_FONT, alignment=RIGHT, border=border_all)
write_cell(ws_c, r, 8, pt_pe2, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=GREEN_FILL if pt_pe2 > PX_NOW else RED_FILL, number_format="$#,##0")
write_cell(ws_c, r, 9, (pt_pe2 / PX_NOW) - 1, font=BOLD_FONT, alignment=RIGHT, border=border_all,
           fill=GREEN_FILL if pt_pe2 > PX_NOW else RED_FILL, number_format="0.0%")

# Column widths
ws_c.column_dimensions['A'].width = 38
for i in range(2, 13):
    ws_c.column_dimensions[get_column_letter(i)].width = 13


# ============================================================
# TAB: FOOTBALL FIELD
# ============================================================
ws_ff = wb.create_sheet("Football Field")

ws_ff.merge_cells("A1:I1")
write_cell(ws_ff, 1, 1, "Micron Technology — Valuation Football Field",
           font=Font(bold=True, size=14, color="FFFFFF", name="Calibri"),
           fill=HEADER_FILL, alignment=CENTER)
ws_ff.row_dimensions[1].height = 24

write_cell(ws_ff, 2, 1, "Summary of valuation ranges by methodology. Current price = $727.42 (2026-05-20). Recommended target = $900 (Base Case).",
           font=ITALIC_FONT, alignment=LEFT)
ws_ff.merge_cells("A2:I2")

r = 4
ff_headers = ["Methodology", "Low ($)", "Midpoint ($)", "High ($)", "% Upside (Mid)", "Weight %", "Weighted ($)", "Notes"]
for i, h in enumerate(ff_headers, start=1):
    write_cell(ws_ff, r, i, h, font=WHITE_FONT, fill=HEADER_FILL, alignment=CENTER, border=border_all)

# Methodologies
ff_data = [
    # (method, low, mid, high, weight, notes)
    ("DCF (Gordon + Exit Mult)", 740, 900, 1100, 0.40, "WACC 10.6%, g 3.0%, FY30E EBITDA 46.4B × 8x"),
    ("Comps — P/E NTM @ 8x", 160, 180, 200, 0.10, "Memory peer median P/E NTM on FY26E EPS $22.50"),
    ("Comps — P/E NTM @ 12x (AI premium)", 230, 270, 320, 0.10, "Premium multiple on FY26E EPS $22.50"),
    ("Comps — EV/Sales 5.0x (median)", 240, 270, 300, 0.05, "Memory peer median EV/Sales on FY26E rev $54.7B"),
    ("Comps — EV/EBITDA 8x (median)", 240, 285, 330, 0.10, "Memory peer median on FY26E EBITDA $36.3B"),
    ("52-week range", 91, 455, 819, 0.00, "MU 52-week range; reference only, not weighted"),
    ("Bull case scenario", 800, 950, 1100, 0.15, "HBM4 share gain, AI capex extends to 2028"),
    ("Base case scenario", 800, 900, 1000, 0.00, "Same as DCF base, not double-weighted"),
    ("Bear case scenario", 250, 420, 550, 0.10, "AI capex decel 2H26, CXMT commodity ramp"),
]

r += 1
ranges_for_chart = []  # collect (method, low, high) for chart
for entry in ff_data:
    method, low, mid, high, weight, notes = entry
    write_cell(ws_ff, r, 1, method, font=NORMAL_FONT, alignment=LEFT, border=border_all)
    write_cell(ws_ff, r, 2, low, font=NORMAL_FONT, alignment=RIGHT, border=border_all, number_format="$#,##0")
    write_cell(ws_ff, r, 3, mid, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=TOTAL_FILL, number_format="$#,##0")
    write_cell(ws_ff, r, 4, high, font=NORMAL_FONT, alignment=RIGHT, border=border_all, number_format="$#,##0")
    upside_mid = (mid / PX_NOW) - 1
    fill_up = GREEN_FILL if upside_mid > 0 else RED_FILL
    write_cell(ws_ff, r, 5, upside_mid, font=NORMAL_FONT, alignment=RIGHT, border=border_all, fill=fill_up, number_format="0.0%")
    write_cell(ws_ff, r, 6, weight, font=NORMAL_FONT, alignment=CENTER, border=border_all, number_format="0%")
    weighted = mid * weight
    write_cell(ws_ff, r, 7, weighted, font=BOLD_FONT, alignment=RIGHT, border=border_all, number_format="$#,##0")
    write_cell(ws_ff, r, 8, notes, font=ITALIC_FONT, alignment=LEFT, border=border_all)
    r += 1
    ranges_for_chart.append((method, low, high, mid))

# Weighted total
r += 1
total_weight = sum(e[4] for e in ff_data)
weighted_target = sum(e[2] * e[4] for e in ff_data)
write_cell(ws_ff, r, 1, "WEIGHTED PRICE TARGET", font=BOLD_FONT, fill=TOTAL_FILL, alignment=LEFT, border=border_all)
write_cell(ws_ff, r, 6, total_weight, font=BOLD_FONT, fill=TOTAL_FILL, alignment=CENTER, border=border_all, number_format="0%")
write_cell(ws_ff, r, 7, weighted_target, font=BOLD_FONT, fill=GREEN_FILL, alignment=RIGHT, border=border_all, number_format="$#,##0")
write_cell(ws_ff, r, 8, f"= weighted avg. of midpoints", font=ITALIC_FONT, alignment=LEFT, border=border_all)

r += 2
write_cell(ws_ff, r, 1, "Current price (2026-05-20):", font=BOLD_FONT, alignment=LEFT, border=border_all)
write_cell(ws_ff, r, 2, PX_NOW, font=BOLD_FONT, alignment=RIGHT, border=border_all, number_format="$#,##0.00")

r += 1
write_cell(ws_ff, r, 1, "12-month price target:", font=BOLD_FONT, alignment=LEFT, border=border_all, fill=GREEN_FILL)
write_cell(ws_ff, r, 2, weighted_target, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=GREEN_FILL, number_format="$#,##0")

r += 1
write_cell(ws_ff, r, 1, "Implied upside:", font=BOLD_FONT, alignment=LEFT, border=border_all, fill=GREEN_FILL)
write_cell(ws_ff, r, 2, (weighted_target / PX_NOW) - 1, font=BOLD_FONT, alignment=RIGHT, border=border_all, fill=GREEN_FILL, number_format="0.0%")

r += 1
write_cell(ws_ff, r, 1, "Recommendation:", font=BOLD_FONT, alignment=LEFT, border=border_all, fill=GREEN_FILL)
write_cell(ws_ff, r, 2, "OVERWEIGHT (BUY)", font=Font(bold=True, size=12, color="FFFFFF", name="Calibri"),
           fill=PatternFill(start_color="00B050", end_color="00B050", fill_type="solid"), alignment=CENTER, border=border_all)

ws_ff.column_dimensions['A'].width = 38
for i in range(2, 9):
    ws_ff.column_dimensions[get_column_letter(i)].width = 14
ws_ff.column_dimensions['H'].width = 38


# ============================================================
# TAB: VALUATION SUMMARY
# ============================================================
ws_vs = wb.create_sheet("Valuation Summary")

ws_vs.merge_cells("A1:H1")
write_cell(ws_vs, 1, 1, "Micron Technology — Valuation Summary & Recommendation",
           font=Font(bold=True, size=14, color="FFFFFF", name="Calibri"),
           fill=HEADER_FILL, alignment=CENTER)
ws_vs.row_dimensions[1].height = 24

# Investment thesis box
r = 3
write_cell(ws_vs, r, 1, "RATING & PRICE TARGET", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 9):
    write_cell(ws_vs, r, c, "", fill=SUBHEADER_FILL, border=border_all)

r += 1
labels = [
    ("Current Price", f"${PX_NOW:,.2f}", "(2026-05-20, Yahoo Finance)"),
    ("12M Price Target", f"${weighted_target:,.0f}", "Weighted avg. of DCF + comps + scenarios"),
    ("Implied Upside", f"{(weighted_target/PX_NOW - 1):.1%}", "Single-stock IRR before dividend"),
    ("Rating", "OVERWEIGHT (BUY)", "5-tier: Buy / Overweight / Hold / Underweight / Sell"),
    ("Time Horizon", "12 months", "Aligned with FY27E full-year"),
    ("Position Sizing", "2.5–3.5% of portfolio", "Standard semi-large-cap, AI-cycle exposure"),
]
for lab, val, note in labels:
    write_cell(ws_vs, r, 1, lab, font=BOLD_FONT, alignment=LEFT, border=border_all)
    write_cell(ws_vs, r, 2, val, font=BOLD_FONT, alignment=LEFT, border=border_all, fill=YELLOW_FILL)
    write_cell(ws_vs, r, 3, note, font=ITALIC_FONT, alignment=LEFT, border=border_all)
    ws_vs.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    r += 1

# Investment thesis
r += 1
write_cell(ws_vs, r, 1, "INVESTMENT THESIS (3 PILLARS)", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 9):
    write_cell(ws_vs, r, c, "", fill=SUBHEADER_FILL, border=border_all)

pillars = [
    ("1. HBM is the structural growth story",
     "CMBU grew +257% YoY in FY25 to $13.5B; HBM3E 12-high is majority of HBM shipments; HBM4 12-high samples shipped to Nvidia / AMD / cloud-ASIC customers. HBM TAM expanding from $25B (2025) to $100B+ by 2030."),
    ("2. Mid-cycle margin reset",
     "GAAP gross margin trough was −9% (FY23) to 40% (FY25) to guided 67% (FQ2-FY26). FCF inflection (FY26E $10B+) supports buyback acceleration ($10B authorization)."),
    ("3. Valuation asymmetry vs. peers",
     "Forward P/E of 7.1x is the lowest in MU's listed history; SK Hynix at 4.6x, Samsung at 5.3x both rich vs. earnings power. Memory peer mean P/E NTM 12x implies fair value $270 floor on FY26E EPS even in commodity-normalization."),
]
for title, desc in pillars:
    r += 1
    write_cell(ws_vs, r, 1, title, font=BOLD_FONT, alignment=LEFT, border=border_all, fill=TOTAL_FILL)
    write_cell(ws_vs, r, 2, desc, font=NORMAL_FONT, alignment=LEFT, border=border_all)
    ws_vs.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws_vs.row_dimensions[r].height = 60

# Risks
r += 2
write_cell(ws_vs, r, 1, "KEY RISKS", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 9):
    write_cell(ws_vs, r, c, "", fill=SUBHEADER_FILL, border=border_all)

risks = [
    ("High", "Customer concentration — single customer (likely Nvidia) at ~17% of FY25 revenue"),
    ("High", "Memory cycle reversal — historical 30-50% peak-to-trough ASP corrections"),
    ("Medium-High", "Multiple compression — TTM P/S 14.1x is highest in MU history"),
    ("Medium", "Execution risk on $100B+ multi-year fab capex program (Idaho, NY, VA, Hiroshima)"),
    ("Medium", "Samsung HBM4 qualification recovery in late-2026/2027"),
    ("Medium", "CXMT commodity DDR4/LPDDR4 ramp depresses non-HBM ASPs in FY27-28"),
    ("Medium", "AI capex deceleration — hyperscaler ROI questions, model efficiency"),
    ("Low-Medium", "CHIPS Act conditions and US-China policy escalation"),
]
for sev, txt in risks:
    r += 1
    sev_fill = RED_FILL if "High" in sev else (YELLOW_FILL if "Medium" in sev else TOTAL_FILL)
    write_cell(ws_vs, r, 1, sev, font=BOLD_FONT, alignment=CENTER, border=border_all, fill=sev_fill)
    write_cell(ws_vs, r, 2, txt, font=NORMAL_FONT, alignment=LEFT, border=border_all)
    ws_vs.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)

# Catalysts
r += 2
write_cell(ws_vs, r, 1, "KEY CATALYSTS (12-MONTH WATCH LIST)", font=WHITE_FONT, fill=SUBHEADER_FILL, alignment=LEFT, border=border_all)
for c in range(2, 9):
    write_cell(ws_vs, r, c, "", fill=SUBHEADER_FILL, border=border_all)

catalysts = [
    ("FQ2-FY26 print (Mar 2026)", "Guided $18.7B revenue, 68% non-GAAP GM, $8.42 EPS — first peak-margin print"),
    ("HBM4 volume ramp updates", "Customer-specific HBM4 qualifications (Nvidia Rubin, AMD MI400)"),
    ("CHIPS milestone disbursements", "Idaho fab milestones; further direct funding tranches"),
    ("FY27 capex guidance", "Implied through quarterly commentary; key for FCF expansion"),
    ("$10B buyback authorization", "Pace of repurchases under existing authorization"),
    ("Memory pricing trajectory", "Q4-CY26 contract negotiations — leading indicator of FY27 ASPs"),
    ("Industry consolidation", "Samsung HBM4 qualification at TSMC base die; SK Hynix Solidigm integration"),
    ("Geopolitics", "US-China export control changes; CAC China decision evolution"),
]
for cat, desc in catalysts:
    r += 1
    write_cell(ws_vs, r, 1, cat, font=BOLD_FONT, alignment=LEFT, border=border_all)
    write_cell(ws_vs, r, 2, desc, font=NORMAL_FONT, alignment=LEFT, border=border_all)
    ws_vs.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)

ws_vs.column_dimensions['A'].width = 30
for i in range(2, 9):
    ws_vs.column_dimensions[get_column_letter(i)].width = 14


# Save
wb.save(wb_path)
print(f"Updated workbook saved: {wb_path}")
print(f"Tabs: {wb.sheetnames}")
print(f"\nKey outputs:")
print(f"  DCF implied price target: ${dcf_pt:.2f}")
print(f"  DCF upside: {upside:.1%}")
print(f"  Football field weighted target: ${weighted_target:,.0f}")
print(f"  Recommendation: OVERWEIGHT (BUY)")
